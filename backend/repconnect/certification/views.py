import base64
import io
import os
import re

from django.db import transaction
from django.http import HttpResponse
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from userProfile.models import workInformation
from .models import Certificate, CertificateCategory, CertificateView
from .serializers import (
    CertificateCategorySerializer,
    CertificateEditSerializer,
    CertificateSerializer,
)

_FILENAME_RE = re.compile(r'^(\d+)_(.+)\.pdf$', re.IGNORECASE)

# ── helpers ───────────────────────────────────────────────────────────────────

def _require_hr_admin(request) -> Response | None:
    """Return 403 Response unless the user has admin=True or hr=True."""
    u = request.user
    if not (getattr(u, 'admin', False) or getattr(u, 'hr', False)):
        return Response({'detail': 'Admin or HR permission required.'}, status=status.HTTP_403_FORBIDDEN)
    return None


def _send_certificate_notification(certificate: Certificate) -> None:
    """Create an in-app notification for the employee on transaction commit."""
    recipient_id = certificate.employee.pk
    title_text   = certificate.title
    cert_id      = certificate.pk

    def _create() -> None:
        try:
            from activityLog.models import Notification
            Notification.objects.create(
                recipient_id=recipient_id,
                notification_type='certificate_issued',
                title='Certificate Issued',
                message=f'A new certificate "{title_text}" has been issued to you.',
                module='certification',
                related_object_id=cert_id,
            )
        except Exception:
            import logging
            logging.getLogger(__name__).exception('Failed to create certificate notification')

    try:
        transaction.on_commit(_create)
    except Exception:
        _create()


def _build_error_excel(failures: list[dict]) -> str:
    """Build an error-report xlsx and return base64-encoded bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Upload Errors'

    def _side():
        return Side(style='thin', color='FF000000')

    thin = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    hdr_fill = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')

    headers = ['Filename', 'Reason']
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFFFF')
        cell.fill      = hdr_fill
        cell.border    = thin
        cell.alignment = Alignment(horizontal='center')

    for row_num, item in enumerate(failures, 2):
        ws.cell(row=row_num, column=1, value=item.get('filename', '')).border = thin
        ws.cell(row=row_num, column=2, value=item.get('reason',   '')).border = thin

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


# ── Views ─────────────────────────────────────────────────────────────────────

class CertificateCategoryListView(APIView):
    """GET /api/certificates/categories — list all categories (auth required)."""
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        qs = CertificateCategory.objects.all()
        return Response(CertificateCategorySerializer(qs, many=True).data)


class CertificateUploadView(APIView):
    """
    POST /api/certificates/admin/upload
    Multi-file upload — each file must be named {idnumber}_{fullname}.pdf.
    Also accepts: category (pk), title, objective.
    Returns: { uploaded: int, failed: int, results: [...], error_report_b64: str | null }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request) -> Response:
        err = _require_hr_admin(request)
        if err:
            return err

        files     = request.FILES.getlist('files')
        category_id = request.data.get('category')
        title     = (request.data.get('title') or '').strip()
        objective = (request.data.get('objective') or '').strip()

        # Validate shared metadata.
        if not files:
            return Response({'detail': 'No files provided.'}, status=status.HTTP_400_BAD_REQUEST)
        if not category_id:
            return Response({'detail': 'Category is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not title:
            return Response({'detail': 'Title is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not objective:
            return Response({'detail': 'Objective is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            category = CertificateCategory.objects.get(pk=category_id)
        except CertificateCategory.DoesNotExist:
            return Response({'detail': 'Category not found.'}, status=status.HTTP_400_BAD_REQUEST)

        results  = []
        failures = []

        for f in files:
            filename = os.path.basename(f.name)
            match    = _FILENAME_RE.match(filename)

            if not match:
                failures.append({
                    'filename': filename,
                    'reason':   'Filename must follow the pattern {idnumber}_{fullname}.pdf',
                })
                results.append({'filename': filename, 'status': 'failed'})
                continue

            idnumber = match.group(1)

            from userLogin.models import loginCredentials
            try:
                employee = loginCredentials.objects.get(idnumber=idnumber)
            except loginCredentials.DoesNotExist:
                failures.append({
                    'filename': filename,
                    'reason':   f'No employee found with ID number "{idnumber}".',
                })
                results.append({'filename': filename, 'status': 'failed'})
                continue

            if f.content_type not in ('application/pdf', 'application/octet-stream'):
                # Accept octet-stream too in case of browser inconsistency.
                suffix = os.path.splitext(filename)[1].lower()
                if suffix != '.pdf':
                    failures.append({'filename': filename, 'reason': 'Only PDF files are accepted.'})
                    results.append({'filename': filename, 'status': 'failed'})
                    continue

            with transaction.atomic():
                cert = Certificate.objects.create(
                    employee=employee,
                    category=category,
                    title=title,
                    objective=objective,
                    file=f,
                    original_filename=filename,
                    uploaded_by=request.user,
                )
            _send_certificate_notification(cert)
            results.append({'filename': filename, 'status': 'success', 'id': cert.pk})

        error_b64 = _build_error_excel(failures) if failures else None

        return Response({
            'uploaded':         len([r for r in results if r['status'] == 'success']),
            'failed':           len(failures),
            'results':          results,
            'error_report_b64': error_b64,
        })


class CertificateAdminListView(APIView):
    """
    GET /api/certificates/admin/list
    Returns a paginated list of employees who have certificates,
    with a `certificates` array nested inside each entry.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_hr_admin(request)
        if err:
            return err

        from userLogin.models import loginCredentials
        from django.db.models import Prefetch, Count, Q

        search        = request.GET.get('search',        '').strip()
        sort_by       = request.GET.get('sort_by',        'idnumber')
        sort_dir      = request.GET.get('sort_dir',       'asc')
        department_id = request.GET.get('department_id',  '').strip()
        line_id       = request.GET.get('line_id',        '').strip()

        if sort_by not in {'idnumber', 'name', 'cert_count', 'department', 'line'}:
            sort_by = 'idnumber'
        prefix = '-' if sort_dir == 'desc' else ''

        # Employees who own at least one certificate.
        qs = (
            loginCredentials.objects
            .filter(certificates__isnull=False)
            .annotate(cert_count=Count('certificates', distinct=True))
            .prefetch_related(
                Prefetch('certificates', queryset=Certificate.objects.select_related('category').order_by('-created_at')),
                Prefetch('workinformation_set', queryset=workInformation.objects.select_related('department', 'line')),
            )
            .distinct()
        )

        if search:
            qs = qs.filter(
                Q(idnumber__icontains=search)
                | Q(firstname__icontains=search)
                | Q(lastname__icontains=search)
            )

        if department_id:
            try:
                qs = qs.filter(workinformation__department_id=int(department_id))
            except (ValueError, TypeError):
                pass

        if line_id:
            try:
                qs = qs.filter(workinformation__line_id=int(line_id))
            except (ValueError, TypeError):
                pass

        if sort_by == 'name':
            qs = qs.order_by(f'{prefix}lastname', f'{prefix}firstname')
        elif sort_by == 'cert_count':
            qs = qs.order_by(f'{prefix}cert_count')
        elif sort_by == 'department':
            qs = qs.order_by(f'{prefix}workinformation__department__name')
        elif sort_by == 'line':
            qs = qs.order_by(f'{prefix}workinformation__line__name')
        else:
            qs = qs.order_by(f'{prefix}idnumber')

        # Paginate at the employee level.
        paginator = PageNumberPagination()
        paginator.page_size = 10
        page_qs = paginator.paginate_queryset(qs, request) or []

        def _work_info(emp):
            wi = emp.workinformation_set.first() if hasattr(emp, '_prefetched_objects_cache') else None
            return {
                'department': wi.department.name if wi and wi.department else '',
                'line':       wi.line.name       if wi and wi.line       else '',
            }

        serializer_context = {'request': request}
        data = []
        for emp in page_qs:
            certs  = CertificateSerializer(emp.certificates.all(), many=True, context=serializer_context).data
            wi_data = _work_info(emp)
            data.append({
                'idnumber':     emp.idnumber,
                'firstname':    emp.firstname or '',
                'lastname':     emp.lastname  or '',
                'department':   wi_data['department'],
                'line':         wi_data['line'],
                'certificates': certs,
            })

        response = paginator.get_paginated_response(data)
        if paginator.page is not None:
            assert response.data is not None
            response.data['total_pages'] = paginator.page.paginator.num_pages
        return response


class CertificateUserListView(APIView):
    """GET /api/certificates/my — returns the current user's certificates."""
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        qs = Certificate.objects.filter(employee=request.user).select_related('category').order_by('-created_at')
        serializer = CertificateSerializer(qs, many=True, context={'request': request})
        return Response({'results': serializer.data, 'count': len(serializer.data)})


class CertificateDetailView(APIView):
    """
    GET    /api/certificates/admin/<pk>   — view one certificate
    PATCH  /api/certificates/admin/<pk>   — edit (replace file/metadata)
    DELETE /api/certificates/admin/<pk>   — delete with confirmation
    """
    permission_classes = [IsAuthenticated]

    def _get_cert(self, pk: int) -> Certificate | None:
        try:
            return Certificate.objects.select_related('category', 'employee').get(pk=pk)
        except Certificate.DoesNotExist:
            return None

    def get(self, request, pk: int) -> Response:
        err = _require_hr_admin(request)
        if err:
            return err
        cert = self._get_cert(pk)
        if not cert:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(CertificateSerializer(cert, context={'request': request}).data)

    def patch(self, request, pk: int) -> Response:
        err = _require_hr_admin(request)
        if err:
            return err
        cert = self._get_cert(pk)
        if not cert:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = CertificateEditSerializer(cert, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            if 'file' in request.FILES:
                cert.file.delete(save=False)
                assert isinstance(serializer.validated_data, dict)
                serializer.validated_data['original_filename'] = os.path.basename(
                    request.FILES['file'].name
                )
            serializer.save(uploaded_by=request.user)

        return Response(CertificateSerializer(cert, context={'request': request}).data)

    def delete(self, request, pk: int) -> Response:
        err = _require_hr_admin(request)
        if err:
            return err
        cert = self._get_cert(pk)
        if not cert:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        with transaction.atomic():
            cert.file.delete(save=False)
            cert.delete()

        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Email copy helpers ─────────────────────────────────────────────────────────

_CERT_SUBJECT_PREFIX: dict[str, str] = {
    'award':       'Certificate of Recognition',
    'compliance':  'Certificate of Compliance',
    'training':    'Certificate of Training',
    'health':      'Certificate of Completion',
    'performance': 'Certificate of Excellence',
    'graduation':  'Certificate of Completion',
    'safety':      'Certificate of Achievement',
}

_CERT_OPENING: dict[str, str] = {
    'award': (
        'Congratulations on your outstanding leadership. '
        'Your dedication to guiding and developing those around you has been formally recognized.'
    ),
    'compliance': (
        'Congratulations on meeting the quality and compliance standards required for this certification. '
        'Your commitment to precision and excellence is commendable.'
    ),
    'training': (
        'Congratulations on completing your technical training. '
        'Your investment in skills development strengthens both your expertise and the team.'
    ),
    'health': (
        'Congratulations on completing your occupational health certification. '
        'Your dedication to workplace safety and well-being is truly valued.'
    ),
    'performance': (
        'Congratulations on this achievement in production and operations. '
        'Your consistent performance and efficiency have earned this well-deserved recognition.'
    ),
    'graduation': (
        'Congratulations on your professional development milestone. '
        'Your commitment to continuous learning reflects the highest standard of excellence.'
    ),
    'safety': (
        'Congratulations on completing this safety certification. '
        'Your awareness and dedication to a safe working environment make a real difference every day.'
    ),
}


def _build_cert_email(cert) -> tuple[str, str]:
    """Return (subject, body) dynamically tailored to the certificate category."""
    icon          = cert.category.icon_key if cert.category else ''
    category_name = cert.category.name     if cert.category else 'Achievement'
    first_name    = cert.employee.firstname or cert.employee.idnumber

    prefix  = _CERT_SUBJECT_PREFIX.get(icon, 'Certificate of Achievement')
    opening = _CERT_OPENING.get(
        icon,
        'Congratulations on this achievement. Your hard work and dedication have been formally recognized.',
    )

    subject = f'{prefix} \u2013 {category_name} | {cert.title}'

    body = (
        f'Dear {first_name},\n\n'
        f'{opening}\n\n'
        f'Your certificate "{cert.title}" has been issued and is attached to this email.\n\n'
        f'You may also view and download it anytime through the REPConnect portal '
        f'under My Certificates.\n\n'
        f'Best regards,\n'
        f'REPConnect'
    )

    return subject, body


class CertificateSendEmailView(APIView):
    """POST /api/certificates/<pk>/send-email — send this certificate PDF to the employee."""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int) -> Response:
        # Both user (their own cert) and admin can send.
        try:
            if getattr(request.user, 'admin', False) and getattr(request.user, 'hr', False):
                cert = Certificate.objects.select_related('employee', 'category').get(pk=pk)
            else:
                cert = Certificate.objects.select_related('employee', 'category').get(
                    pk=pk, employee=request.user
                )
        except Certificate.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Load email config.
        from generalsettings.models import EmailConfiguration
        try:
            config = EmailConfiguration.objects.get(pk=1)
        except EmailConfiguration.DoesNotExist:
            return Response({'detail': 'Email configuration not set.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        recipient_email = cert.employee.email
        if not recipient_email:
            return Response({'detail': 'Employee has no email address.'}, status=status.HTTP_400_BAD_REQUEST)

        _default_subject, _default_body = _build_cert_email(cert)
        subject = request.data.get('subject', _default_subject)
        body    = request.data.get('body',    _default_body)

        import smtplib
        from email.mime.application import MIMEApplication
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        msg = MIMEMultipart()
        from_addr = f'{config.from_name} <{config.username}>' if config.from_name else config.username
        msg['From']    = from_addr
        msg['To']      = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        # Attach the PDF.
        try:
            with cert.file.open('rb') as pdf_file:
                part = MIMEApplication(pdf_file.read(), Name=cert.original_filename or 'certificate.pdf')
            part['Content-Disposition'] = f'attachment; filename="{cert.original_filename or "certificate.pdf"}"'
            msg.attach(part)
        except Exception as exc:
            return Response({'detail': f'Could not read certificate file: {exc}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            # Port 465 always requires a direct SSL connection (SMTP_SSL).
            # Port 587 always requires a plain connection upgraded via STARTTLS.
            # For any other port, honour the saved flags.
            port = config.smtp_port
            if port == 465:
                use_ssl, use_tls = True, False
            elif port == 587:
                use_ssl, use_tls = False, True
            else:
                use_ssl, use_tls = bool(config.use_ssl), bool(config.use_tls)

            _cls = smtplib.SMTP_SSL if use_ssl else smtplib.SMTP
            with _cls(config.smtp_host, port, timeout=15) as smtp:
                if use_tls and not use_ssl:
                    smtp.starttls()
                smtp.login(config.username, config.password)
                smtp.sendmail(config.username, recipient_email, msg.as_string())
        except Exception as exc:
            return Response({'detail': f'Email sending failed: {exc}'}, status=status.HTTP_502_BAD_GATEWAY)

        return Response({'detail': 'Email sent successfully.'})


class CertificateMarkViewedView(APIView):
    """
    POST /api/certificates/<pk>/mark-viewed
    Records that the authenticated user has opened a certificate, clearing its
    "New" status.  Idempotent — does nothing if the record already exists.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int) -> Response:
        try:
            cert = Certificate.objects.get(pk=pk, employee=request.user)
        except Certificate.DoesNotExist:
            # Allow admins (hr+admin) to mark any certificate as viewed.
            if getattr(request.user, 'admin', False) and getattr(request.user, 'hr', False):
                try:
                    cert = Certificate.objects.get(pk=pk)
                except Certificate.DoesNotExist:
                    return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        CertificateView.objects.get_or_create(
            certificate=cert,
            viewer=request.user,
        )
        return Response({'detail': 'Marked as viewed.'}, status=status.HTTP_200_OK)


class CertificateFiltersView(APIView):
    """
    GET /api/certificates/admin/filters
    Returns departments and lines available for filtering the admin list.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_hr_admin(request)
        if err:
            return err

        from generalsettings.models import Department, Line

        departments = list(
            Department.objects
            .order_by('office__name', 'name')
            .values('id', 'name')
        )
        lines = list(
            Line.objects
            .order_by('department__name', 'name')
            .values('id', 'name', 'department_id')
        )
        return Response({'departments': departments, 'lines': lines})

