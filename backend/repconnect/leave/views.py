"""
Leave Management views.

All mutating endpoints use transaction.atomic() + select_for_update() for
concurrency safety.  Email and in-app notifications are dispatched via
transaction.on_commit() so they are never fired on a rolled-back transaction.
"""

import datetime
import logging
import smtplib
from decimal import Decimal
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO

from django.db import IntegrityError, transaction
from django.db.models import Case, DecimalField, ExpressionWrapper, F, IntegerField, OuterRef, Prefetch, Q, Subquery, Sum, Value, When
from django.db.models.functions import Greatest
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from generalsettings.workdays import get_configured_hours_per_day, get_configured_weekday_durations, get_configured_workdays
from systemCalendar.models import CalendarEvent

from .models import (
    LeaveApprovalStep,
    LeaveBalance,
    LeaveReason,
    LeaveRequest,
    LeaveSubreason,
    LeaveType,
    SundayExemption,
)
from .routing import build_approval_chain, can_act_on_step
from .serializers import (
    LeaveApprovalActionSerializer,
    LeaveApprovalStepSerializer,
    LeaveBalanceSerializer,
    LeaveReasonAdminSerializer,
    LeaveReasonSerializer,
    LeaveRequestCreateSerializer,
    LeaveRequestDetailSerializer,
    LeaveRequestListSerializer,
    LeaveSubreasonAdminSerializer,
    LeaveTypeAdminSerializer,
    LeaveTypeSerializer,
    LeaveBalanceAdminUpdateSerializer,
)

logger = logging.getLogger(__name__)
PAGE_SIZE = 10


BALANCE_ORDERING_MAP = {
    'employee_name': ['employee__lastname', 'employee__firstname', '-id'],
    '-employee_name': ['-employee__lastname', '-employee__firstname', '-id'],
    'department': ['employee__workinformation__department__name', '-id'],
    '-department': ['-employee__workinformation__department__name', '-id'],
    'leave_type': ['leave_type__name', '-id'],
    '-leave_type': ['-leave_type__name', '-id'],
    'period': ['period_start', 'period_end', '-id'],
    '-period': ['-period_start', '-period_end', '-id'],
    'balance': ['entitled_leave', '-id'],
    '-balance': ['-entitled_leave', '-id'],
    'used': ['used_leave', '-id'],
    '-used': ['-used_leave', '-id'],
    'remaining': ['remaining_hours', '-id'],
    '-remaining': ['-remaining_hours', '-id'],
}


# ── Permission helpers ─────────────────────────────────────────────────────────

def _require_admin_or_hr(request):
    if not (getattr(request.user, 'admin', False) or getattr(request.user, 'hr', False)):
        return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
    return None


def _require_admin(request):
    if not getattr(request.user, 'admin', False):
        return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
    return None


class LeaveRoutingRuleListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from .models import LeaveRoutingRule

        rules = (
            LeaveRoutingRule.objects
            .prefetch_related('positions', 'departments', 'steps__target_positions')
            .order_by('description')
        )

        return Response([
            {
                'id': rule.id,
                'description': rule.description,
                'is_active': rule.is_active,
                'positions': [position.name for position in rule.positions.all()],
                'departments': [department.name for department in rule.departments.all()],
                'steps': [
                    {
                        'step_order': step.step_order,
                        'position_ids': list(step.target_positions.values_list('id', flat=True)),
                    }
                    for step in rule.steps.all()
                ],
            }
            for rule in rules
        ])

    def post(self, request):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from .models import LeaveRoutingRule, LeaveRoutingStep
        from generalsettings.models import Position, Department

        description = (request.data.get('description') or '').strip()
        if not description:
            return Response({'detail': 'description is required.'}, status=http_status.HTTP_400_BAD_REQUEST)

        position_ids = request.data.get('position_ids') or []
        department_ids = request.data.get('department_ids') or []
        steps_data = request.data.get('steps') or []

        with transaction.atomic():
            rule = LeaveRoutingRule.objects.create(description=description, is_active=True)
            rule.positions.set(Position.objects.filter(id__in=position_ids))
            rule.departments.set(Department.objects.filter(id__in=department_ids))
            for i, step in enumerate(steps_data, 1):
                step_obj = LeaveRoutingStep.objects.create(rule=rule, step_order=i)
                step_position_ids = step.get('position_ids') or []
                if step_position_ids:
                    step_obj.target_positions.set(Position.objects.filter(id__in=step_position_ids))

        rule.refresh_from_db()
        return Response(
            {
                'id': rule.id,
                'description': rule.description,
                'is_active': rule.is_active,
                'positions': [p.name for p in rule.positions.all()],
                'departments': [d.name for d in rule.departments.all()],
                'steps': [
                    {
                        'step_order': s.step_order,
                        'position_ids': list(s.target_positions.values_list('id', flat=True)),
                    }
                    for s in rule.steps.all()
                ],
            },
            status=http_status.HTTP_201_CREATED,
        )

    def put(self, request, pk):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from .models import LeaveRoutingRule, LeaveRoutingStep
        from generalsettings.models import Position, Department

        try:
            rule = LeaveRoutingRule.objects.get(pk=pk)
        except LeaveRoutingRule.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        description = (request.data.get('description') or '').strip()
        if not description:
            return Response({'detail': 'description is required.'}, status=http_status.HTTP_400_BAD_REQUEST)

        position_ids = request.data.get('position_ids') or []
        department_ids = request.data.get('department_ids') or []
        steps_data = request.data.get('steps') or []

        with transaction.atomic():
            rule.description = description
            rule.save(update_fields=['description', 'updated_at'])
            rule.positions.set(Position.objects.filter(id__in=position_ids))
            rule.departments.set(Department.objects.filter(id__in=department_ids))
            rule.steps.all().delete()
            for i, step in enumerate(steps_data, 1):
                step_obj = LeaveRoutingStep.objects.create(rule=rule, step_order=i)
                step_position_ids = step.get('position_ids') or []
                if step_position_ids:
                    step_obj.target_positions.set(Position.objects.filter(id__in=step_position_ids))

        return Response(
            {
                'id': rule.id,
                'description': rule.description,
                'is_active': rule.is_active,
                'positions': [p.name for p in rule.positions.all()],
                'departments': [d.name for d in rule.departments.all()],
                'steps': [
                    {
                        'step_order': s.step_order,
                        'position_ids': list(s.target_positions.values_list('id', flat=True)),
                    }
                    for s in rule.steps.all()
                ],
            }
        )

    def delete(self, request, pk):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from .models import LeaveRoutingRule

        try:
            rule = LeaveRoutingRule.objects.get(pk=pk)
        except LeaveRoutingRule.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        rule.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


# ── Holiday Range View ─────────────────────────────────────────────────────────

class HolidayRangeView(APIView):
    """
    GET /api/leave/holidays?date_start=YYYY-MM-DD&date_end=YYYY-MM-DD

    Returns holidays (CalendarEvent with holiday event_type) and Sunday
    exemptions that fall within the requested date range.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_start_str = request.query_params.get('date_start')
        date_end_str = request.query_params.get('date_end')
        if not date_start_str or not date_end_str:
            return Response(
                {'detail': 'date_start and date_end query params are required.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        try:
            date_start = datetime.date.fromisoformat(date_start_str)
            date_end = datetime.date.fromisoformat(date_end_str)
        except ValueError:
            return Response(
                {'detail': 'Invalid date format. Use YYYY-MM-DD.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        if date_end < date_start:
            return Response(
                {'detail': 'date_end must be on or after date_start.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        holidays = list(
            CalendarEvent.objects
            .filter(
                event_type__in=['legal', 'special', 'day_off', 'company'],
                date__range=[date_start, date_end],
            )
            .values('date', 'title', 'event_type')
            .distinct()
        )

        sunday_exemptions = list(
            SundayExemption.objects
            .filter(date__range=[date_start, date_end])
            .values_list('date', flat=True)
        )

        hours_per_day = get_configured_hours_per_day()
        half_day_hours = (hours_per_day / Decimal('2')).quantize(Decimal('0.1'))
        weekday_durations = get_configured_weekday_durations()

        return Response({
            'holidays': [
                {
                    'date': h['date'].isoformat(),
                    'title': h['title'],
                    'event_type': h['event_type'],
                }
                for h in holidays
            ],
            'sunday_exemptions': [d.isoformat() for d in sunday_exemptions],
            'workdays': get_configured_workdays(),
            'hours_per_day': str(hours_per_day),
            'half_day_hours': str(half_day_hours),
            'weekday_durations': {str(day): str(hours) for day, hours in weekday_durations.items()},
        })


# ── Email helpers ──────────────────────────────────────────────────────────────

def _format_days_hours(hours: Decimal, hours_per_day: Decimal) -> str:
    days = (hours / hours_per_day).quantize(Decimal('0.01'))
    hours_text = hours.quantize(Decimal('0.1'))
    return f'{days} days ({hours_text} hours)'


def _build_leave_balance_notice(lr, deducted_hours: Decimal, remaining_after: Decimal) -> str | None:
    if deducted_hours <= Decimal('0') and remaining_after == Decimal('0'):
        # Balance was already zero and no deduction occurred.
        total_hours = Decimal(str(getattr(lr, 'total_hours', lr.hours)))
        if total_hours <= Decimal('0'):
            return None
        hours_per_day = get_configured_hours_per_day()
        covered_text = _format_days_hours(Decimal('0'), hours_per_day)
        excess_text = _format_days_hours(total_hours, hours_per_day)
        return (
            f'Please note that your {lr.leave_type.name} balance was only sufficient to cover {covered_text} of this leave request. '
            f'A deduction of {covered_text} has been applied and your remaining balance is now 0 days (0 hours). '
            f'The remaining {excess_text} of this leave were approved without a corresponding balance deduction.'
        )

    total_hours = Decimal(str(getattr(lr, 'total_hours', lr.hours)))
    hours_per_day = get_configured_hours_per_day()

    if deducted_hours == total_hours and remaining_after == Decimal('0'):
        return (
            f'Please note that this leave approval has fully consumed your remaining {lr.leave_type.name} balance. '
            f'Your current balance is now 0 days (0 hours).'
        )

    if deducted_hours < total_hours:
        covered_text = _format_days_hours(deducted_hours, hours_per_day)
        remaining_text = _format_days_hours(total_hours - deducted_hours, hours_per_day)
        return (
            f'Please note that your {lr.leave_type.name} balance was only sufficient to cover {covered_text} of this leave request. '
            f'A deduction of {covered_text} has been applied and your remaining balance is now 0 days (0 hours). '
            f'The remaining {remaining_text} of this leave were approved without a corresponding balance deduction.'
        )

    return None


def _build_leave_email(lr, final_status: str, balance_notice: str | None = None):
    """Return (subject, body) tailored to final_status."""
    employee = lr.employee
    first_name = employee.firstname or employee.idnumber
    control_no = lr.control_number
    date_from = lr.date_start.strftime('%B %d, %Y')
    date_to = lr.date_end.strftime('%B %d, %Y')

    if final_status == 'approved':
        subject = 'Leave Request Approved'
        body = (
            f'Dear {first_name},\n\n'
            f'We are pleased to inform you that your leave request '
            f'(Control No: {control_no}) has been successfully approved.\n\n'
            f'Your requested leave schedule from {date_from} to {date_to} has been confirmed. '
            f'Please ensure proper coordination with your team prior to your leave dates.\n\n'
        )
        if balance_notice:
            body += f'{balance_notice}\n\n'
        body += (
            f'Thank you.\n\n'
            f'RepConnect'
        )
    else:
        subject = 'Leave Request Update'
        body = (
            f'Dear {first_name},\n\n'
            f'We regret to inform you that your leave request '
            f'(Control No: {control_no}) has been disapproved during the approval process.\n\n'
            f'You may review the request details and routing history in the system '
            f'for further information. If needed, you may coordinate with your approver '
            f'for clarification.\n\n'
            f'Thank you.\n\n'
            f'RepConnect'
        )
    return subject, body


def _send_leave_email_task(leave_request_pk: int, final_status: str, balance_notice: str | None = None) -> None:
    """Send final-status email to the employee. Called inside on_commit — non-blocking."""
    try:
        from generalsettings.models import EmailConfiguration
        config = EmailConfiguration.objects.get(pk=1)
    except Exception:
        logger.warning('leave: Email configuration not found, skipping email.')
        return

    try:
        lr = LeaveRequest.objects.select_related('employee').get(pk=leave_request_pk)
    except LeaveRequest.DoesNotExist:
        return

    recipient_email = lr.employee.email
    if not recipient_email:
        logger.warning('leave: Employee %s has no email address.', lr.employee.idnumber)
        return

    subject, body = _build_leave_email(lr, final_status, balance_notice)

    msg = MIMEMultipart()
    from_addr = (
        f'{config.from_name} <{config.username}>' if config.from_name else config.username
    )
    msg['From'] = from_addr
    msg['To'] = recipient_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    try:
        if config.use_ssl:
            smtp = smtplib.SMTP_SSL(config.smtp_host, config.smtp_port)
        else:
            smtp = smtplib.SMTP(config.smtp_host, config.smtp_port)
            if config.use_tls:
                smtp.starttls()
        smtp.login(config.username, config.password)
        smtp.sendmail(config.username, recipient_email, msg.as_string())
        smtp.quit()
    except Exception:
        logger.exception('leave: Failed to send email to %s', recipient_email)


# ── In-app notification helpers ────────────────────────────────────────────────

def _schedule_leave_notification(leave_request_pk: int, final_status: str) -> None:
    """Create in-app notification for the employee after HR finalises."""
    def _create() -> None:
        try:
            from activityLog.models import Notification
            lr = LeaveRequest.objects.select_related('employee').get(pk=leave_request_pk)
            notif_type = 'leave_approved' if final_status == 'approved' else 'leave_disapproved'
            title_text = (
                'Leave Request Approved'
                if final_status == 'approved'
                else 'Leave Request Disapproved'
            )
            message = (
                f'Your leave request {lr.control_number} has been {final_status}.'
            )
            Notification.objects.create(
                recipient=lr.employee,
                notification_type=notif_type,
                title=title_text,
                message=message,
                module='leave',
                related_object_id=lr.pk,
            )
        except Exception:
            logger.exception(
                'leave: Failed to create status notification for pk=%d', leave_request_pk
            )

    try:
        transaction.on_commit(_create)
    except Exception:
        _create()


def _notify_next_approver(leave_request_pk: int) -> None:
    """Notify every member of the next pending approval group."""
    try:
        from activityLog.models import Notification
        from userLogin.models import loginCredentials

        lr = LeaveRequest.objects.select_related('employee').get(pk=leave_request_pk)
        step = (
            lr.approval_steps.filter(status='pending').order_by('sequence').first()
        )
        if not step:
            return

        if step.role_group == 'clinic':
            approvers = loginCredentials.objects.filter(clinic=True, is_active=True)
        elif step.role_group == 'iad':
            approvers = loginCredentials.objects.filter(iad=True, is_active=True)
        elif step.role_group == 'hr':
            approvers = loginCredentials.objects.filter(hr=True, is_active=True)
        else:
            approvers = (
                loginCredentials.objects.filter(pk=step.approver_id)
                if step.approver_id else loginCredentials.objects.none()
            )

        employee_name = (
            f'{lr.employee.firstname or ""} {lr.employee.lastname or ""}'.strip()
            or lr.employee.idnumber
        )
        for approver in approvers:
            Notification.objects.create(
                recipient=approver,
                notification_type='leave_pending_approval',
                title='Leave Request Pending Approval',
                message=(
                    f'{employee_name} has submitted a leave request '
                    f'({lr.control_number}) pending your approval.'
                ),
                module='leave',
                related_object_id=lr.pk,
            )
    except Exception:
        logger.exception(
            'leave: Failed to notify approvers for leave_request_pk=%d', leave_request_pk
        )


def _notify_leave_cancelled(leave_request_pk: int) -> None:
    """Notify employee when HR cancels their leave."""
    try:
        from activityLog.models import Notification

        lr = LeaveRequest.objects.select_related('employee').get(pk=leave_request_pk)
        Notification.objects.create(
            recipient=lr.employee,
            notification_type='leave_cancelled',
            title='Leave Request Cancelled',
            message=(
                f'Your leave request {lr.control_number} has been cancelled by HR.'
            ),
            module='leave',
            related_object_id=lr.pk,
        )
    except Exception:
        logger.exception(
            'leave: Failed to notify employee of cancellation for pk=%d', leave_request_pk
        )


# ── Balance helpers ────────────────────────────────────────────────────────────

def _deduct_leave_balance(leave_request) -> tuple[Decimal, Decimal] | None:
    """Deduct filed hours from the matching LeaveBalance. Never goes below 0.

    Returns a tuple of (deducted_hours, remaining_balance_after_deduction), or None
    if no balance record is found."""
    balance = (
        LeaveBalance.objects.select_for_update()
        .filter(
            employee=leave_request.employee,
            leave_type=leave_request.leave_type,
            period_start__lte=leave_request.date_start,
            period_end__gte=leave_request.date_start,
        )
        .first()
    )
    if not balance:
        logger.error(
            'leave: No balance found for employee_id=%d, leave_type_id=%d, date=%s',
            leave_request.employee_id,
            leave_request.leave_type_id,
            leave_request.date_start,
        )
        return None

    # Skip deduction for leaves approved before the balance was uploaded.
    # The admin already accounted for those in the uploaded used_leave value.
    if balance.uploaded_at and leave_request.hr_approved_at:
        if leave_request.hr_approved_at < balance.uploaded_at:
            return None

    deduct_hours = Decimal(str(getattr(leave_request, 'total_hours', leave_request.hours)))
    remaining = max(balance.entitled_leave - balance.used_leave, Decimal('0'))
    deduct = min(deduct_hours, remaining)
    if deduct > 0:
        balance.used_leave = balance.used_leave + deduct
        balance.save(update_fields=['used_leave'])

    return deduct, max(remaining - deduct, Decimal('0'))


def _restore_leave_balance(leave_request) -> None:
    """Restore previously deducted hours back to LeaveBalance."""
    balance = (
        LeaveBalance.objects.select_for_update()
        .filter(
            employee=leave_request.employee,
            leave_type=leave_request.leave_type,
            period_start__lte=leave_request.date_start,
            period_end__gte=leave_request.date_start,
        )
        .first()
    )
    if not balance:
        return

    restore_hours = Decimal(str(getattr(leave_request, 'total_hours', leave_request.hours)))
    balance.used_leave = max(balance.used_leave - restore_hours, Decimal('0'))
    balance.save(update_fields=['used_leave'])


# ── Finalization ───────────────────────────────────────────────────────────────

def _finalize_leave(leave_request) -> None:
    """
    Called when the last approval step is completed.

    Determines the final outcome via three guards (ANY one is enough to disapprove):
      1. lr.status == 'disapproved'  — set only by HR disapproval or legacy data
      2. lr.manager_disapproved      — set when any manager-role step was disapproved
      3. any step recorded as 'disapproved' — catches clinic / IAD disapprovals

    Even if HR approves, conditions 2 or 3 override the outcome to 'disapproved'.
    No balance deduction is applied when the final status is 'disapproved'.

    Must be called inside a transaction.atomic() block.
    """
    lr = LeaveRequest.objects.select_for_update().get(pk=leave_request.pk)

    # Triple guard: any of these conditions forces a 'disapproved' outcome.
    any_step_disapproved = lr.approval_steps.filter(status='disapproved').exists()
    already_disapproved = (
        lr.status == 'disapproved'
        or lr.manager_disapproved
        or any_step_disapproved
    )
    final_status = 'disapproved' if already_disapproved else 'approved'

    update_fields = ['updated_at']

    if lr.status != final_status:
        lr.status = final_status
        update_fields.append('status')

    if final_status in ('approved', 'disapproved'):
        lr.seen = False
        update_fields.append('seen')

    balance_notice = None
    if final_status == 'approved':
        lr.hr_approved_at = timezone.now()
        update_fields.append('hr_approved_at')
        if lr.is_deductible and lr.leave_type.has_balance:
            result = _deduct_leave_balance(lr)
            if result is not None:
                deducted_hours, remaining_after = result
                balance_notice = _build_leave_balance_notice(lr, deducted_hours, remaining_after)

    lr.save(update_fields=update_fields)

    lr_pk = lr.pk
    transaction.on_commit(lambda: _send_leave_email_task(lr_pk, final_status, balance_notice))
    transaction.on_commit(lambda: _schedule_leave_notification(lr_pk, final_status))


# ── Chain progression ──────────────────────────────────────────────────────────

def _activate_next_step_or_finalize(lr, completed_step) -> None:
    """
    After a step is acted upon, activate the next pending step or finalize the request.

    Status transitions: pending → routing (chain in progress) → approved/disapproved (final).
    'routing' is set here whenever the chain advances so that subsequent approvers
    can find the request by filtering on pending/routing — even if an earlier step
    was disapproved.  The final approved/disapproved outcome is determined solely
    by _finalize_leave, which checks manager_disapproved and any_step_disapproved.
    """
    next_step = (
        LeaveApprovalStep.objects
        .filter(leave_request=lr, status='pending')
        .order_by('sequence')
        .first()
    )

    if next_step:
        # Always advance to 'routing' when the chain has more steps.
        # This overrides any premature 'disapproved' set by an earlier step's branch
        # (e.g. manager disapproval) so subsequent approvers can still find the request.
        if lr.status != 'routing':
            lr.status = 'routing'
            lr.save(update_fields=['status', 'updated_at'])
        next_step.activated_at = timezone.now()
        next_step.save(update_fields=['activated_at'])
        transaction.on_commit(lambda: _notify_next_approver(lr.pk))
    else:
        _finalize_leave(lr)


# ── Standard leave views ───────────────────────────────────────────────────────

class LeaveTypeListView(APIView):
    """GET /api/leave/types/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        types = LeaveType.objects.filter(is_active=True).order_by('name')
        return Response(LeaveTypeSerializer(types, many=True).data)


class LeaveReasonListView(APIView):
    """GET /api/leave/reasons/?leave_type=<id>"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        leave_type_id = request.query_params.get('leave_type')
        if not leave_type_id:
            return Response(
                {'detail': 'leave_type query param is required.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        reasons = (
            LeaveReason.objects.filter(leave_types__pk=leave_type_id)
            .prefetch_related('subreasons')
            .order_by('title')
            .distinct()
        )
        return Response(LeaveReasonSerializer(reasons, many=True).data)


class LeaveBalanceListView(APIView):
    """GET /api/leave/balances/ — returns only active (non-expired) balances for the user."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = datetime.date.today()
        balances = (
            LeaveBalance.objects.filter(
                employee=request.user,
                period_end__gte=today,
            )
            .select_related('leave_type')
            .order_by('-period_start', 'leave_type__name')
        )
        return Response(LeaveBalanceSerializer(balances, many=True).data)


class LeaveRequestListCreateView(APIView):
    """GET / POST /api/leave/requests/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = (
            LeaveRequest.objects.filter(employee=request.user)
            .select_related('leave_type', 'reason', 'subreason')
            .order_by('-date_prepared', '-created_at')
        )

        leave_type_ids = request.query_params.getlist('leave_type')
        if leave_type_ids:
            qs = qs.filter(leave_type_id__in=leave_type_ids)

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        try:
            page = max(int(request.query_params.get('page', 1)), 1)
        except ValueError:
            page = 1

        total = qs.count()
        offset = (page - 1) * PAGE_SIZE
        results = qs[offset: offset + PAGE_SIZE]

        return Response({
            'count': total,
            'total_pages': max(1, -(-total // PAGE_SIZE)),
            'results': LeaveRequestListSerializer(
                results, many=True, context={'request': request}
            ).data,
        })

    @transaction.atomic
    def post(self, request):
        # Idempotency guard
        idempotency_key = request.headers.get('X-Idempotency-Key')
        cache_key: str | None = None
        if idempotency_key:
            from django.core.cache import cache
            cache_key = f'leave_submit_{request.user.pk}_{idempotency_key}'
            if cache.get(cache_key):
                return Response(
                    {'detail': 'Duplicate request.'},
                    status=http_status.HTTP_409_CONFLICT,
                )

        ser = LeaveRequestCreateSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)

        d = ser.validated_data
        leave_type = d['leave_type']
        date_start = d['date_start']
        date_end = d['date_end']

        # Overlap check
        overlap = LeaveRequest.objects.filter(
            employee=request.user,
            status__in=['pending', 'routing', 'approved'],
            date_start__lte=date_end,
            date_end__gte=date_start,
        ).exists()
        if overlap:
            return Response(
                {'detail': 'You already have a leave request overlapping the selected dates.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        control_number = LeaveRequest.generate_control_number()

        lr = LeaveRequest.objects.create(
            employee=request.user,
            leave_type=leave_type,
            reason=d['reason'],
            subreason=d.get('subreason'),
            date_start=date_start,
            date_end=date_end,
            total_hours=d['total_hours'],
            total_days=d['total_days'],
            hours=d['hours'],
            days_count=d['days_count'],
            is_deductible=leave_type.deductible,
            status='pending',
            control_number=control_number,
            remarks=d.get('remarks', ''),
            seen=True,
        )

        # Build routing — raises ValidationError if pre-flight fails (rolls back)
        build_approval_chain(lr)

        if cache_key:
            from django.core.cache import cache
            cache.set(cache_key, True, 86400)

        transaction.on_commit(lambda: _notify_next_approver(lr.pk))

        return Response(
            LeaveRequestDetailSerializer(lr, context={'request': request}).data,
            status=http_status.HTTP_201_CREATED,
        )


class LeaveRequestDetailView(APIView):
    """GET /api/leave/requests/<pk>/"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        qs = LeaveRequest.objects.select_related(
            'leave_type', 'reason', 'subreason', 'employee'
        ).prefetch_related(
            'approval_steps__approver', 'approval_steps__acted_by'
        )

        try:
            lr = qs.get(pk=pk)
        except LeaveRequest.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        user = request.user
        is_owner = lr.employee_id == user.pk
        is_admin_hr = getattr(user, 'admin', False) or getattr(user, 'hr', False)

        if not is_admin_hr and not is_owner:
            # Check if user has a step on this request
            has_manager_step = lr.approval_steps.filter(approver=user).exists()
            has_role_step = (
                (getattr(user, 'clinic', False) and lr.approval_steps.filter(role_group='clinic').exists()) or
                (getattr(user, 'iad', False) and lr.approval_steps.filter(role_group='iad').exists())
            )
            if not has_manager_step and not has_role_step:
                return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        return Response(LeaveRequestDetailSerializer(lr, context={'request': request}).data)


def _annotated_with_active_step(qs):
    active_step_rg = Subquery(
        LeaveApprovalStep.objects
        .filter(leave_request=OuterRef('pk'), status='pending')
        .order_by('sequence')
        .values('role_group')[:1]
    )
    active_step_approver = Subquery(
        LeaveApprovalStep.objects
        .filter(leave_request=OuterRef('pk'), status='pending')
        .order_by('sequence')
        .values('approver_id')[:1]
    )
    return qs.annotate(
        _active_step_rg=active_step_rg,
        _active_step_approver=active_step_approver,
    )


class LeaveMyUnseenCountView(APIView):
    """GET /api/leave/requests/unseen-count"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        unseen_count = LeaveRequest.objects.filter(
            employee=request.user,
            status__in=['approved', 'disapproved'],
            seen=False,
        ).count()
        return Response({'unseen_count': unseen_count})


class LeaveMyPendingApprovalCountView(APIView):
    """GET /api/leave/approval/pending-count"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        base_qs = _annotated_with_active_step(
            LeaveRequest.objects.exclude(employee=user)
        )

        if getattr(user, 'admin', False) or getattr(user, 'hr', False):
            count = base_qs.filter(_active_step_rg='hr').count()
            return Response({'pending_count': count})

        is_clinic = getattr(user, 'clinic', False)
        is_iad = getattr(user, 'iad', False)

        # Clinic/IAD badges must reflect only requests currently waiting on
        # their respective active role-group step.
        if is_clinic or is_iad:
            filters = Q()
            if is_clinic:
                filters |= Q(_active_step_rg='clinic')
            if is_iad:
                filters |= Q(_active_step_rg='iad')
        else:
            filters = Q(_active_step_approver=user.pk)

        count = base_qs.filter(filters).count()
        return Response({'pending_count': count})


class LeaveRequestMarkSeenView(APIView):
    """PATCH /api/leave/requests/<pk>/mark-seen"""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk):
        try:
            lr = LeaveRequest.objects.select_for_update().get(pk=pk, employee=request.user)
        except LeaveRequest.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if not lr.seen:
            lr.seen = True
            lr.save(update_fields=['seen', 'updated_at'])

        return Response({'seen': True})


class LeaveRequestCancelView(APIView):
    """PATCH /api/leave/requests/<pk>/cancel/"""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk):
        try:
            lr = LeaveRequest.objects.select_for_update().select_related(
                'leave_type', 'employee'
            ).get(pk=pk)
        except LeaveRequest.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        user = request.user
        is_hr = getattr(user, 'hr', False)
        is_owner = lr.employee_id == user.pk

        if not is_hr and not is_owner:
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        if lr.status in ('cancelled', 'disapproved'):
            return Response(
                {'detail': 'This request cannot be cancelled.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        if not is_hr:
            if lr.status == 'pending':
                pass  # Allowed — no routing has started yet
            elif lr.status == 'approved' and lr.hr_approved_at:
                delta = timezone.now() - lr.hr_approved_at
                if delta.total_seconds() > 3 * 24 * 3600:
                    return Response(
                        {'detail': 'Cancellation window has expired. The 3-day cancellation period has passed.'},
                        status=http_status.HTTP_403_FORBIDDEN,
                    )
            else:
                return Response(
                    {'detail': 'Cancellation is not allowed at this stage.'},
                    status=http_status.HTTP_403_FORBIDDEN,
                )

        # Restore balance if the leave was already approved and balance was deducted
        if lr.status == 'approved' and lr.is_deductible and lr.leave_type.has_balance:
            _restore_leave_balance(lr)

        lr.status = 'cancelled'
        lr.cancelled_at = timezone.now()
        lr.cancelled_by_id = request.user.pk
        lr.save(update_fields=['status', 'cancelled_at', 'cancelled_by_id', 'updated_at'])

        # Preserve only the routing step that corresponds to the cancellation action.
        # Delete any other pending routing entries for this leave request.
        cancellation_step = None
        if is_hr and not is_owner:
            cancellation_step = LeaveApprovalStep.objects.filter(
                leave_request=lr,
                role_group='hr',
                status='pending',
            ).order_by('sequence').first()

        if cancellation_step:
            cancellation_step.status = 'cancelled'
            cancellation_step.acted_by = request.user
            cancellation_step.acted_at = timezone.now()
            if not cancellation_step.remarks:
                cancellation_step.remarks = 'Cancelled by HR'
            cancellation_step.save(update_fields=['status', 'acted_by', 'acted_at', 'remarks'])
            LeaveApprovalStep.objects.filter(
                leave_request=lr,
                status='pending'
            ).exclude(pk=cancellation_step.pk).delete()
        else:
            LeaveApprovalStep.objects.filter(
                leave_request=lr,
                status='pending'
            ).delete()

        if is_hr and not is_owner:
            lr_pk = lr.pk
            transaction.on_commit(lambda: _notify_leave_cancelled(lr_pk))

        return Response(
            LeaveRequestListSerializer(lr, context={'request': request}).data
        )


class LeaveRequestEditView(APIView):
    """PATCH /api/leave/requests/<pk>/edit/ — edit a pending leave request (owner only)"""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk):
        try:
            lr = LeaveRequest.objects.select_for_update().select_related(
                'leave_type', 'reason', 'subreason'
            ).get(pk=pk)
        except LeaveRequest.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if lr.employee_id != request.user.pk:
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        if lr.status != 'pending':
            return Response(
                {'detail': 'Only pending leave requests can be edited.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        ser = LeaveRequestCreateSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)

        d = ser.validated_data
        leave_type = d['leave_type']

        # lr.status stays 'pending' for the entire routing lifecycle (it only
        # changes to approved/disapproved/cancelled at the very end), so it
        # cannot be used to tell whether any approver has acted yet. Check the
        # steps directly: once at least one has been approved/disapproved,
        # the leave type is locked (changing it would require re-routing,
        # e.g. adding/removing the Clinic step, which would discard real
        # approval history). Other fields (dates, reason, remarks) remain
        # editable regardless of routing progress.
        leave_type_changed = leave_type.pk != lr.leave_type_id
        routing_started = leave_type_changed and lr.approval_steps.exclude(status='pending').exists()
        if routing_started:
            return Response(
                {'leave_type': 'Leave type can no longer be changed because at least one approver has '
                               'already reviewed this request.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        # Overlap check — exclude the request being edited
        overlap = LeaveRequest.objects.filter(
            employee=request.user,
            status__in=['pending', 'routing', 'approved'],
            date_start__lte=d['date_end'],
            date_end__gte=d['date_start'],
        ).exclude(pk=pk).exists()
        if overlap:
            return Response(
                {'detail': 'You already have a leave request overlapping the selected dates.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        lr.leave_type = leave_type
        lr.reason = d['reason']
        lr.subreason = d.get('subreason')
        lr.date_start = d['date_start']
        lr.date_end = d['date_end']
        lr.total_hours = d['total_hours']
        lr.total_days = d['total_days']
        lr.hours = d['hours']
        lr.days_count = d['days_count']
        lr.is_deductible = leave_type.deductible
        lr.remarks = d.get('remarks', '')
        lr.save(update_fields=[
            'leave_type', 'reason', 'subreason', 'date_start', 'date_end',
            'total_hours', 'total_days', 'hours', 'days_count', 'is_deductible', 'remarks', 'updated_at',
        ])

        if leave_type_changed:
            # Rebuild the approval chain so it reflects the new leave_type —
            # e.g. switching to a leave type that requires_clinic_approval
            # must add the Clinic step, which the stale chain built at
            # original submission time would otherwise be missing. Safe
            # because routing_started (checked above) was False here.
            lr.approval_steps.all().delete()
            build_approval_chain(lr)
            transaction.on_commit(lambda: _notify_next_approver(lr.pk))

        return Response(
            LeaveRequestDetailSerializer(lr, context={'request': request}).data
        )


class LeaveRequestCalendarView(APIView):
    """GET /api/leave/requests/calendar — date, status and leave-type for calendar/chart display"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = (
            LeaveRequest.objects
            .filter(employee=request.user)
            .values('date_start', 'date_end', 'status', 'leave_type__name')
            .order_by('date_start')
        )
        data = [
            {
                'date_start': str(r['date_start']),
                'date_end': str(r['date_end']),
                'status': r['status'],
                'leave_type_name': r['leave_type__name'] or '',
            }
            for r in qs
        ]
        return Response(data)


class LeaveApprovalView(APIView):
    """PATCH /api/leave/requests/<pk>/action/ — approve or disapprove"""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk):
        try:
            lr = LeaveRequest.objects.select_for_update().select_related('leave_type').get(pk=pk)
        except LeaveRequest.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        # Allow action when pending, routing, OR disapproved (manager disapproved — audit continues)
        if lr.status not in ('pending', 'routing', 'disapproved'):
            return Response(
                {'detail': 'This request is not awaiting approval.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        ser = LeaveApprovalActionSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)

        action = ser.validated_data['action']
        remarks = ser.validated_data.get('remarks', '')

        # Get current active step (lowest-sequence pending)
        step = (
            LeaveApprovalStep.objects.select_for_update()
            .filter(leave_request=lr, status='pending')
            .order_by('sequence')
            .first()
        )
        if not step:
            return Response(
                {'detail': 'No pending approval step found.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        if not can_act_on_step(request.user, step):
            return Response(
                {'detail': 'You are not authorised to act on this step.'},
                status=http_status.HTTP_403_FORBIDDEN,
            )

        # Record the action on the step
        step.status = action
        step.remarks = remarks
        step.acted_by = request.user
        step.acted_at = timezone.now()
        step.save(update_fields=['status', 'remarks', 'acted_by', 'acted_at'])

        if action == 'disapproved' and step.role_group == 'manager':
            # Set the manager_disapproved flag so _finalize_leave produces a
            # 'disapproved' outcome at the end of the chain.  Do NOT set
            # lr.status = 'disapproved' here — _activate_next_step_or_finalize
            # will set 'routing' so subsequent approvers (e.g. HR) can still see
            # and act on the request.
            lr.manager_disapproved = True
            lr.seen = False
            lr.save(update_fields=['manager_disapproved', 'seen', 'updated_at'])
        elif action == 'disapproved' and step.role_group == 'hr':
            # HR disapproval — immediately set overall status and skip all remaining steps
            lr.status = 'disapproved'
            lr.seen = False
            lr.save(update_fields=['status', 'seen', 'updated_at'])
            LeaveApprovalStep.objects.filter(
                leave_request=lr, status='pending'
            ).update(status='skipped')

        _activate_next_step_or_finalize(lr, step)

        lr.refresh_from_db()
        return Response(
            LeaveRequestDetailSerializer(lr, context={'request': request}).data
        )


# ── Approval Queue helpers ─────────────────────────────────────────────────────

def _get_approval_queue_qs(user):
    """
    Return a LeaveRequest queryset scoped to requests this user should see in
    the approval queue.

    Routing-based scope — a request is included only when the logged-in user
    appears as one of the designated approvers in the request's routing chain,
    regardless of whether it is currently their turn to act:

      - HR / Admin  : all requests, because HR is always the final step in
                      every routing sequence (i.e. every request has an HR step).
      - Clinic      : requests whose routing chain contains a 'clinic' step.
      - IAD         : requests whose routing chain contains an 'iad' step.
      - Any user    : requests where they appear as the approver FK on a
                      manager step (standard / rule-based approver).

    Multiple roles are combined with OR so that, e.g., an IAD user who also
    appears as a manager-step approver on some requests sees both sets.

    Everyone's own submitted requests are excluded so they cannot act on them.
    """
    base_qs = (
        LeaveRequest.objects
        .select_related('leave_type', 'reason', 'subreason', 'employee')
        .exclude(employee=user)
    )

    is_admin = getattr(user, 'admin', False)
    is_hr = getattr(user, 'hr', False)
    is_clinic = getattr(user, 'clinic', False)
    is_iad = getattr(user, 'iad', False)

    # Build routing-based filters.  Start with an empty Q and OR in each role
    # the user participates in.  HR / admin use the 'hr' step filter which
    # covers every request (HR is always the last step), so this is equivalent
    # to "see all" but is expressed through the routing chain for consistency.
    filters = Q()

    if is_hr or is_admin:
        filters |= Q(approval_steps__role_group='hr')

    if is_clinic:
        filters |= Q(approval_steps__role_group='clinic')

    if is_iad:
        filters |= Q(approval_steps__role_group='iad')

    # Always include requests where this user is the designated approver FK
    # on any manager step (rule-based or default manager chain).
    filters |= Q(approval_steps__approver=user)

    return base_qs.filter(filters).distinct()


class ApprovalQueueView(APIView):
    """GET /api/leave/approval-queue/ — requests where the current user can or has acted."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        qs = _get_approval_queue_qs(user)

        # ── Filters ────────────────────────────────────────────────────────
        search_q = request.query_params.get('search', '').strip()
        if search_q:
            qs = qs.filter(
                Q(employee__idnumber__icontains=search_q) |
                Q(employee__firstname__icontains=search_q) |
                Q(employee__lastname__icontains=search_q) |
                Q(control_number__icontains=search_q) |
                Q(leave_type__name__icontains=search_q)
            )

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        leave_type_filter = request.query_params.get('leave_type')
        if leave_type_filter:
            qs = qs.filter(leave_type_id=leave_type_filter)

        # ── "My turn first" sort annotation ────────────────────────────────
        # Subquery: role_group of the lowest-sequence pending step for each request
        active_step_rg = Subquery(
            LeaveApprovalStep.objects
            .filter(leave_request=OuterRef('pk'), status='pending')
            .order_by('sequence')
            .values('role_group')[:1]
        )
        active_step_approver = Subquery(
            LeaveApprovalStep.objects
            .filter(leave_request=OuterRef('pk'), status='pending')
            .order_by('sequence')
            .values('approver_id')[:1]
        )

        # First pass: materialise the subquery annotations so the second
        # annotate call can reference them by name in When() conditions.
        qs = qs.annotate(
            _active_step_rg=active_step_rg,
            _active_step_approver=active_step_approver,
        )

        # Build a boolean-like integer: 0 = my turn, 1 = not my turn
        if getattr(user, 'hr', False) or getattr(user, 'admin', False):
            is_my_turn_case = Case(
                When(_active_step_rg='hr', then=Value(0)),
                When(_active_step_approver=user.pk, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        elif getattr(user, 'clinic', False):
            is_my_turn_case = Case(
                When(_active_step_rg='clinic', then=Value(0)),
                When(_active_step_approver=user.pk, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        elif getattr(user, 'iad', False):
            is_my_turn_case = Case(
                When(_active_step_rg='iad', then=Value(0)),
                When(_active_step_approver=user.pk, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        else:
            # Standard approver: match by approver FK on the active step
            is_my_turn_case = Case(
                When(_active_step_approver=user.pk, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )

        qs = qs.annotate(_is_my_turn=is_my_turn_case)
        ordering = request.query_params.get('ordering', '').strip()
        if ordering:
            qs = qs.order_by('_is_my_turn', ordering)
        else:
            qs = qs.order_by('_is_my_turn', '-date_prepared', '-created_at')

        qs = qs.prefetch_related(
            Prefetch(
                'approval_steps',
                queryset=LeaveApprovalStep.objects.order_by('sequence'),
            )
        )

        try:
            page = max(int(request.query_params.get('page', 1)), 1)
        except ValueError:
            page = 1

        total = qs.count()
        offset = (page - 1) * PAGE_SIZE
        results = list(qs[offset: offset + PAGE_SIZE])

        return Response({
            'count': total,
            'total_pages': max(1, -(-total // PAGE_SIZE)),
            'results': LeaveRequestListSerializer(
                results, many=True, context={'request': request}
            ).data,
        })


# ── Admin views ────────────────────────────────────────────────────────────────

class AdminLeaveRequestListView(APIView):
    """GET /api/leave/admin/requests/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        qs = (
            LeaveRequest.objects.select_related(
                'leave_type', 'reason', 'subreason', 'employee'
            )
            .order_by('-created_at')
        )

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        leave_type_filter = request.query_params.get('leave_type')
        if leave_type_filter:
            qs = qs.filter(leave_type_id=leave_type_filter)

        try:
            page = max(int(request.query_params.get('page', 1)), 1)
        except ValueError:
            page = 1

        total = qs.count()
        offset = (page - 1) * PAGE_SIZE
        results = qs[offset: offset + PAGE_SIZE]

        return Response({
            'count': total,
            'total_pages': max(1, -(-total // PAGE_SIZE)),
            'results': LeaveRequestListSerializer(
                results, many=True, context={'request': request}
            ).data,
        })


class AdminLeaveTypeView(APIView):
    """GET/POST /api/leave/admin/types/ and PATCH/DELETE /api/leave/admin/types/<pk>/"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        err = _require_admin_or_hr(request)
        if err:
            return err
        if pk:
            try:
                obj = LeaveType.objects.get(pk=pk)
            except LeaveType.DoesNotExist:
                return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
            return Response(LeaveTypeAdminSerializer(obj).data)
        return Response(LeaveTypeAdminSerializer(LeaveType.objects.all().order_by('name'), many=True).data)

    def post(self, request):
        err = _require_admin(request)
        if err:
            return err
        ser = LeaveTypeAdminSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)
        obj = ser.save()
        return Response(LeaveTypeAdminSerializer(obj).data, status=http_status.HTTP_201_CREATED)

    def patch(self, request, pk):
        err = _require_admin(request)
        if err:
            return err
        try:
            obj = LeaveType.objects.get(pk=pk)
        except LeaveType.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        ser = LeaveTypeAdminSerializer(obj, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)
        obj = ser.save()
        return Response(LeaveTypeAdminSerializer(obj).data)

    def delete(self, request, pk):
        err = _require_admin(request)
        if err:
            return err
        try:
            obj = LeaveType.objects.get(pk=pk)
        except LeaveType.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if obj.requests.exists():
            obj.is_active = False
            obj.save(update_fields=['is_active'])
            return Response({
                'detail': 'Leave type has existing requests and has been deactivated instead of deleted.'
            })
        obj.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminLeaveReasonView(APIView):
    """GET/POST /api/leave/admin/reasons/ and PATCH/DELETE /api/leave/admin/reasons/<pk>/"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        err = _require_admin_or_hr(request)
        if err:
            return err
        if pk:
            try:
                obj = LeaveReason.objects.prefetch_related('subreasons', 'leave_types').get(pk=pk)
            except LeaveReason.DoesNotExist:
                return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
            return Response(LeaveReasonAdminSerializer(obj).data)
        leave_type_id = request.query_params.get('leave_type')
        qs = LeaveReason.objects.prefetch_related('subreasons', 'leave_types').order_by('title')
        if leave_type_id:
            qs = qs.filter(leave_types__pk=leave_type_id).distinct()
        return Response(LeaveReasonAdminSerializer(qs, many=True).data)

    def post(self, request):
        err = _require_admin(request)
        if err:
            return err
        ser = LeaveReasonAdminSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)
        obj = ser.save()
        return Response(LeaveReasonAdminSerializer(obj).data, status=http_status.HTTP_201_CREATED)

    def patch(self, request, pk):
        err = _require_admin(request)
        if err:
            return err
        try:
            obj = LeaveReason.objects.get(pk=pk)
        except LeaveReason.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        ser = LeaveReasonAdminSerializer(obj, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)
        obj = ser.save()
        return Response(LeaveReasonAdminSerializer(obj).data)

    def delete(self, request, pk):
        err = _require_admin(request)
        if err:
            return err
        try:
            obj = LeaveReason.objects.get(pk=pk)
        except LeaveReason.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if obj.requests.exists():
            return Response(
                {'detail': 'Cannot delete a reason with existing leave requests.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        obj.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminLeaveSubreasonView(APIView):
    """GET/POST /api/leave/admin/subreasons/ and PATCH/DELETE /api/leave/admin/subreasons/<pk>/"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        err = _require_admin_or_hr(request)
        if err:
            return err
        if pk:
            try:
                obj = LeaveSubreason.objects.get(pk=pk)
            except LeaveSubreason.DoesNotExist:
                return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
            return Response(LeaveSubreasonAdminSerializer(obj).data)
        reason_id = request.query_params.get('reason')
        qs = LeaveSubreason.objects.all().order_by('reason__title', 'title')
        if reason_id:
            qs = qs.filter(reason_id=reason_id)
        return Response(LeaveSubreasonAdminSerializer(qs, many=True).data)

    def post(self, request):
        err = _require_admin(request)
        if err:
            return err
        ser = LeaveSubreasonAdminSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)
        obj = ser.save()
        return Response(LeaveSubreasonAdminSerializer(obj).data, status=http_status.HTTP_201_CREATED)

    def patch(self, request, pk):
        err = _require_admin(request)
        if err:
            return err
        try:
            obj = LeaveSubreason.objects.get(pk=pk)
        except LeaveSubreason.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        ser = LeaveSubreasonAdminSerializer(obj, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=http_status.HTTP_400_BAD_REQUEST)
        obj = ser.save()
        return Response(LeaveSubreasonAdminSerializer(obj).data)

    def delete(self, request, pk):
        err = _require_admin(request)
        if err:
            return err
        try:
            obj = LeaveSubreason.objects.get(pk=pk)
        except LeaveSubreason.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if obj.requests.exists():
            return Response(
                {'detail': 'Cannot delete a sub-reason with existing leave requests.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        obj.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminLeaveBalanceListView(APIView):
    """GET /api/leave/admin/balances — paginated leave balances across all employees."""
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _employee_name(emp) -> str:
        last = (getattr(emp, 'lastname', '') or '').strip()
        first = (getattr(emp, 'firstname', '') or '').strip()
        if last and first:
            return f'{last}, {first}'
        return getattr(emp, 'idnumber', '') or ''

    @staticmethod
    def _employee_department(emp) -> str:
        if not emp:
            return ''
        wi = emp.workinformation_set.select_related('department').order_by('-created_at').first()
        return wi.department.name if wi and wi.department else ''

    def get(self, request):
        try:
            err = _require_admin_or_hr(request)
            if err:
                return err

            base_qs = (
                LeaveBalance.objects
                .select_related('employee', 'leave_type')
                .prefetch_related('employee__workinformation_set__department')
                .annotate(
                    remaining_hours=Case(
                        When(entitled_leave__gt=F('used_leave'), then=ExpressionWrapper(
                            F('entitled_leave') - F('used_leave'),
                            output_field=DecimalField(max_digits=8, decimal_places=1),
                        )),
                        default=Value(0),
                        output_field=DecimalField(max_digits=8, decimal_places=1),
                    )
                )
            )

            search_q = request.query_params.get('search', '').strip()
            if search_q:
                base_qs = base_qs.filter(
                    Q(employee__idnumber__icontains=search_q)
                    | Q(employee__firstname__icontains=search_q)
                    | Q(employee__lastname__icontains=search_q)
                    | Q(leave_type__name__icontains=search_q)
                    | Q(employee__workinformation__department__name__icontains=search_q)
                )

            leave_type_filter = request.query_params.get('leave_type', '').strip()
            if leave_type_filter:
                base_qs = base_qs.filter(leave_type_id=leave_type_filter)

            department_filter = request.query_params.get('department', '').strip()
            if department_filter:
                base_qs = base_qs.filter(
                    employee__workinformation__department__name=department_filter
                )

            ordering = request.query_params.get('ordering', '').strip()
            base_qs = base_qs.order_by(*(BALANCE_ORDERING_MAP.get(ordering) or ['-created_at', '-id']))

            try:
                page = max(int(request.query_params.get('page', 1)), 1)
            except ValueError:
                page = 1

            total = base_qs.count()
            offset = (page - 1) * PAGE_SIZE
            rows = base_qs[offset: offset + PAGE_SIZE]

            hours_per_day = get_configured_hours_per_day()
            results = []
            for b in rows:
                employee = getattr(b, 'employee', None)
                leave_type = getattr(b, 'leave_type', None)
                entitled_hours = Decimal(str(getattr(b, 'entitled_leave', 0))).quantize(Decimal('0.1'))
                used_hours = Decimal(str(getattr(b, 'used_leave', 0))).quantize(Decimal('0.1'))
                remaining_hours = Decimal(str(getattr(b, 'remaining_hours', 0))).quantize(Decimal('0.1'))

                entitled_days = (entitled_hours / hours_per_day).quantize(Decimal('0.01'))
                used_days = (used_hours / hours_per_day).quantize(Decimal('0.01'))
                remaining_days = (remaining_hours / hours_per_day).quantize(Decimal('0.01'))

                results.append({
                    'id': b.id,
                    'employee_name': self._employee_name(employee),
                    'employee_id_number': getattr(employee, 'idnumber', ''),
                    'department': self._employee_department(employee),
                    'leave_type': getattr(leave_type, 'name', ''),
                    'leave_type_id': getattr(b, 'leave_type_id', None),
                    'period_start': b.period_start.isoformat(),
                    'period_end': b.period_end.isoformat(),
                    'balance_days': str(entitled_days),
                    'balance_hours': str(entitled_hours),
                    'used_days': str(used_days),
                    'used_hours': str(used_hours),
                    'remaining_days': str(remaining_days),
                    'remaining_hours': str(remaining_hours),
                })

            return Response({
                'count': total,
                'total_pages': max(1, -(-total // PAGE_SIZE)),
                'results': results,
            })
        except Exception:
            logger.exception('Admin leave balances list failed')
            return Response({'detail': 'Could not load leave balances.'}, status=http_status.HTTP_500_INTERNAL_SERVER_ERROR)


class AdminLeaveBalanceDetailView(APIView):
    """GET/PATCH/DELETE /api/leave/admin/balances/<pk> — edit or delete a balance record."""
    permission_classes = [IsAuthenticated]

    def _get_balance(self, pk: int):
        return (
            LeaveBalance.objects
            .select_related('employee', 'leave_type')
            .get(pk=pk)
        )

    def get(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err
        try:
            balance = self._get_balance(pk)
        except LeaveBalance.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        return Response({
            'id': balance.id,
            'employee_id': balance.employee_id,
            'employee_name': AdminLeaveBalanceListView._employee_name(balance.employee),
            'employee_id_number': getattr(balance.employee, 'idnumber', ''),
            'department': AdminLeaveBalanceListView._employee_department(balance.employee),
            'leave_type_id': balance.leave_type_id,
            'leave_type': balance.leave_type.name if balance.leave_type else '',
            'period_start': balance.period_start.isoformat(),
            'period_end': balance.period_end.isoformat(),
            'balance_days': str((Decimal(str(balance.entitled_leave)) / get_configured_hours_per_day()).quantize(Decimal('0.01'))),
            'balance_hours': str(Decimal(str(balance.entitled_leave)).quantize(Decimal('0.1'))),
        })

    @transaction.atomic
    def patch(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err
        try:
            balance = LeaveBalance.objects.select_for_update().select_related('employee', 'leave_type').get(pk=pk)
        except LeaveBalance.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        serializer = LeaveBalanceAdminUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)

        balance.leave_type_id = serializer.validated_data['leave_type_id']
        balance.period_start = serializer.validated_data['period_start']
        balance.period_end = serializer.validated_data['period_end']
        balance.entitled_leave = serializer.validated_data['balance_hours'].quantize(Decimal('0.1'))
        balance.used_leave = serializer.validated_data['used_hours'].quantize(Decimal('0.1'))
        try:
            balance.save()
        except IntegrityError:
            return Response({'detail': 'A balance record already exists for this employee, leave type, and period.'}, status=http_status.HTTP_400_BAD_REQUEST)

        return Response(self.get(request, pk).data)

    @transaction.atomic
    def delete(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err
        try:
            balance = LeaveBalance.objects.select_for_update().get(pk=pk)
        except LeaveBalance.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        balance.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminLeaveBalanceTemplateView(APIView):
    """GET /api/leave/admin/balance-template — downloadable Excel template for balance imports."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            assert ws is not None
            ws.title = 'Leave Balances'

            headers = [
                'ID Number',
                'Employee Name',
                'Leave Type',
                'Period Start',
                'Period End',
                'Balance (Days)',
                'Used Leave (Days)',
            ]
            sample_row = ['10001', 'Surname, Firstname', 'Vacation Leave', '2026-01-01', '2026-12-31', '10', '0']
            header_fill = PatternFill(start_color='2845D6', end_color='2845D6', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
            )

            leave_type_names = list(LeaveType.objects.filter(is_active=True).order_by('name').values_list('name', flat=True))

            for col_idx, value in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=value)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Example data row — italic grey so users recognise it as a sample
            example_type = leave_type_names[0] if leave_type_names else 'Vacation Leave'
            sample_row[2] = example_type
            for col_idx, value in enumerate(sample_row, start=1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.border = border
                cell.font = Font(color='FF888888', italic=True)

            widths = [16, 24, 20, 14, 14, 16, 16]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

            if leave_type_names:
                from openpyxl.worksheet.datavalidation import DataValidation
                formula = '"' + ','.join(leave_type_names) + '"'
                dv = DataValidation(
                    type='list',
                    formula1=formula,
                    sqref='C2:C5000',
                    allow_blank=True,
                    showDropDown=False,
                )
                dv.error       = 'Please select a valid leave type from the dropdown list.'
                dv.errorTitle  = 'Invalid Leave Type'
                dv.prompt      = 'Select a leave type from the list.'
                dv.promptTitle = 'Leave Type'
                ws.add_data_validation(dv)

            buf = BytesIO()
            wb.save(buf)
            response = HttpResponse(
                buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="leave_balance_template.xlsx"'
            return response
        except Exception:
            logger.exception('Failed to generate leave balance template')
            return Response({'detail': 'Failed to generate template.'}, status=http_status.HTTP_500_INTERNAL_SERVER_ERROR)


class AdminBulkBalanceUploadView(APIView):
    """POST /api/leave/admin/balance-upload/ — all-or-nothing bulk balance import."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=http_status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True, read_only=True)
            ws = wb.active
            assert ws is not None
        except Exception:
            return Response(
                {'detail': 'Invalid or corrupt Excel file.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        from userLogin.models import loginCredentials

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'detail': 'File is empty.'}, status=http_status.HTTP_400_BAD_REQUEST)

        headers = [str(h or '').strip() for h in rows[0]]
        expected_headers = [
            'ID Number',
            'Employee Name',
            'Leave Type',
            'Period Start',
            'Period End',
            'Balance (Days)',
            'Used Leave (Days)',
        ]

        def col(name, alternates=None):
            alternates = alternates or []
            for idx, header in enumerate(headers):
                if header == name or header in alternates:
                    return idx
            return -1

        col_id = col('ID Number')
        col_lt = col('Leave Type')
        col_ps = col('Period Start')
        col_pe = col('Period End')
        col_el = col('Balance (Days)', alternates=['Entitled Leave (days)', 'Entitled Leave'])
        col_ul = col('Used Leave (Days)')  # optional column

        if any(c == -1 for c in [col_id, col_lt, col_ps, col_pe, col_el]):
            return Response(
                {'detail': f'Invalid template. Expected columns: {", ".join(expected_headers)}'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        import datetime as dt
        hours_per_day = get_configured_hours_per_day()
        errors = []
        valid_rows = []

        for row_num, row in enumerate(rows[1:], start=2):
            row = list(row)
            while len(row) < len(headers):
                row.append(None)

            id_number = str(row[col_id] or '').strip()
            lt_name = str(row[col_lt] or '').strip()
            period_start_raw = row[col_ps]
            period_end_raw = row[col_pe]
            entitled_raw = row[col_el]
            used_raw = row[col_ul] if col_ul != -1 and col_ul < len(row) else None

            row_errors = []
            employee = None
            leave_type_obj = None
            period_start = None
            period_end = None
            entitled = None
            used_days = Decimal('0')

            # ID Number
            if not id_number:
                row_errors.append('ID Number is required.')
            else:
                try:
                    employee = loginCredentials.objects.get(
                        idnumber=id_number,
                        is_active=True,
                        admin=False,
                        hr=False,
                    )
                except loginCredentials.DoesNotExist:
                    row_errors.append(f'Employee ID "{id_number}" not found, inactive, or not eligible for leave balances.')

            # Leave Type
            if not lt_name:
                row_errors.append('Leave Type is required.')
            else:
                try:
                    leave_type_obj = LeaveType.objects.get(name__iexact=lt_name, is_active=True)
                    if not leave_type_obj.has_balance:
                        row_errors.append(f'Leave type "{lt_name}" does not support balance tracking.')
                except LeaveType.DoesNotExist:
                    row_errors.append(f'Leave type "{lt_name}" not found or inactive.')

            # Period Start
            if isinstance(period_start_raw, dt.datetime):
                period_start = period_start_raw.date()
            elif isinstance(period_start_raw, dt.date):
                period_start = period_start_raw
            elif isinstance(period_start_raw, str) and period_start_raw.strip():
                try:
                    period_start = dt.datetime.strptime(period_start_raw.strip(), '%Y-%m-%d').date()
                except ValueError:
                    row_errors.append('Period Start must be in YYYY-MM-DD format.')
            else:
                row_errors.append('Period Start is required.')

            # Period End
            if isinstance(period_end_raw, dt.datetime):
                period_end = period_end_raw.date()
            elif isinstance(period_end_raw, dt.date):
                period_end = period_end_raw
            elif isinstance(period_end_raw, str) and period_end_raw.strip():
                try:
                    period_end = dt.datetime.strptime(period_end_raw.strip(), '%Y-%m-%d').date()
                except ValueError:
                    row_errors.append('Period End must be in YYYY-MM-DD format.')
            else:
                row_errors.append('Period End is required.')

            if period_start and period_end and period_end < period_start:
                row_errors.append('Period End cannot be before Period Start.')

            # Entitled Leave — treat None and blank string as missing, but allow 0
            try:
                if entitled_raw is None or (isinstance(entitled_raw, str) and not entitled_raw.strip()):
                    raise ValueError('missing')
                entitled = Decimal(str(entitled_raw)).quantize(Decimal('0.1'))
                if entitled < Decimal('0'):
                    row_errors.append('Entitled Leave must not be negative.')
            except Exception:
                row_errors.append('Balance must be a valid number (0 or greater).')

            # Used Leave — optional, defaults to 0; negative values are rejected
            if used_raw is not None and not (isinstance(used_raw, str) and not str(used_raw).strip()):
                try:
                    used_days = Decimal(str(used_raw)).quantize(Decimal('0.1'))
                    if used_days < Decimal('0'):
                        row_errors.append('Used Leave (Days) must not be negative.')
                    elif entitled is not None and used_days > entitled:
                        row_errors.append('Used Leave (Days) cannot exceed Balance (Days).')
                except Exception:
                    row_errors.append('Used Leave (Days) must be a valid number (0 or greater).')

            if row_errors:
                errors.append({'row': row_num, 'errors': '; '.join(row_errors), 'data': row})
            else:
                valid_rows.append({
                    'employee': employee,
                    'leave_type': leave_type_obj,
                    'period_start': period_start,
                    'period_end': period_end,
                    'entitled_leave_days': entitled,
                    'used_leave_days': used_days,
                    'row': row_num,
                })

        # Within-file duplicate check
        seen: set = set()
        for r in valid_rows:
            key = (r['employee'].pk, r['leave_type'].pk, r['period_start'], r['period_end'])
            if key in seen:
                errors.append({
                    'row': r['row'],
                    'errors': 'Duplicate entry in file (same employee, leave type, and period).',
                    'data': [],
                })
            seen.add(key)

        if errors:
            error_bytes = self._build_error_report(expected_headers, errors)
            from django.http import HttpResponse
            response = HttpResponse(
                error_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                status=http_status.HTTP_400_BAD_REQUEST,
            )
            response['Content-Disposition'] = (
                'attachment; filename="leave_balance_upload_errors.xlsx"'
            )
            return response

        # All rows valid — save
        upload_time = timezone.now()
        for r in valid_rows:
            entitled_hours = (r['entitled_leave_days'] * hours_per_day).quantize(Decimal('0.1'))
            used_hours = (r['used_leave_days'] * hours_per_day).quantize(Decimal('0.1'))
            LeaveBalance.objects.update_or_create(
                employee=r['employee'],
                leave_type=r['leave_type'],
                period_start=r['period_start'],
                period_end=r['period_end'],
                defaults={
                    'entitled_leave': entitled_hours,
                    'used_leave': used_hours,
                    'uploaded_at': upload_time,
                },
            )

        return Response({'detail': f'{len(valid_rows)} balance record(s) saved successfully.'})

    @staticmethod
    def _build_error_report(headers: list, errors: list) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = 'Errors'

        all_headers = headers + ['Remarks']
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color='2845D6', end_color='2845D6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        red_font = Font(color='CC0000')

        for col_idx, h in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for err in errors:
            row_data = list(err.get('data', []))
            while len(row_data) < len(headers):
                row_data.append('')
            row_data.append(err['errors'])
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=err['row'], column=col_idx, value=str(val) if val is not None else '')
                cell.font = red_font
                cell.border = border

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


class AdminLeaveRequestExportView(APIView):
    """GET /api/leave/admin/export/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        qs = (
            LeaveRequest.objects.select_related(
                'leave_type', 'reason', 'subreason', 'employee'
            )
            .order_by('-created_at')
        )
        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = 'Leave Requests'

        col_headers = [
            'Control No.', 'Employee ID', 'Employee Name', 'Leave Type',
            'Reason', 'Date From', 'Date To', 'Days', 'Hours', 'Status', 'Date Filed',
        ]
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color='2845D6', end_color='2845D6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, h in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_num, lr in enumerate(qs, 2):
            name = (
                f'{lr.employee.firstname or ""} {lr.employee.lastname or ""}'.strip()
                or lr.employee.idnumber
            )
            reason = lr.reason.title
            if lr.subreason:
                reason += f' – {lr.subreason.title}'
            elif lr.remarks:
                reason += f' – {lr.remarks}'
            row_data = [
                lr.control_number,
                lr.employee.idnumber,
                name,
                lr.leave_type.name,
                reason,
                lr.date_start.strftime('%B %d, %Y'),
                lr.date_end.strftime('%B %d, %Y'),
                str(lr.total_days),
                str(lr.total_hours),
                lr.get_status_display(),
                lr.date_prepared.strftime('%B %d, %Y'),
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = border

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="leave_requests_export.xlsx"'
        return response


# ── Approval queue chart view ────────────────────────────────────────────

class ApprovalQueueChartView(APIView):
    """
    GET /api/leave/approval-queue/chart

    Returns chart data grouped by date period for the approval queue.
    Data scope uses the same role-based scoping as ApprovalQueueView.

    Query params:
      view_type   : 'fiscal' | 'monthly' | 'weekly'
      fy_start    : int  (fiscal year start, e.g. 2025 for Jul 2025 – Jun 2026)
      month_year  : 'YYYY-M'  (e.g. '2026-4')
      week_start  : 'YYYY-MM-DD'  (ISO Monday of the target week)
      search      : string
      status      : status value
      leave_type  : leave type PK

    Returns:
      { data: [{label, <type_name>: days, ...}, ...], categories: [{key, label, color}] }
    """
    permission_classes = [IsAuthenticated]

    # Colour palette (cycles if more leave types than colours)
    _COLORS = [
        '#2845D6', '#10B981', '#F59E0B', '#8B5CF6', '#EC4899',
        '#0EA5E9', '#F97316', '#14B8A6', '#F43F5E', '#6366F1',
    ]

    _LIGHT_COLORS = [
        '#5B78E8', '#34D399', '#FCD34D', '#C084FC', '#F9A8D4',
        '#7DD3FC', '#FDBA74', '#5EEAD4', '#FB7185', '#A5B4FC',
    ]

    def get(self, request):
        user = request.user
        qs = _get_approval_queue_qs(user)

        # ── Apply same filters as the list view ──────────────────────────────
        search_q = request.query_params.get('search', '').strip()
        if search_q:
            qs = qs.filter(
                Q(employee__idnumber__icontains=search_q) |
                Q(employee__firstname__icontains=search_q) |
                Q(employee__lastname__icontains=search_q) |
                Q(control_number__icontains=search_q) |
                Q(leave_type__name__icontains=search_q)
            )

        status_filter = request.query_params.get('status', '').strip()
        if status_filter:
            qs = qs.filter(status=status_filter)

        lt_filter = request.query_params.get('leave_type', '').strip()
        if lt_filter:
            qs = qs.filter(leave_type_id=lt_filter)

        # ── Determine date window ────────────────────────────────────────
        view_type = request.query_params.get('view_type', 'fiscal')
        today = datetime.date.today()

        if view_type == 'monthly':
            month_year_param = request.query_params.get('month_year', '')
            try:
                year_s, month_s = month_year_param.split('-')
                year, month = int(year_s), int(month_s)
            except (ValueError, AttributeError):
                year, month = today.year, today.month
            # All days in the month
            import calendar as _cal
            days_in_month = _cal.monthrange(year, month)[1]
            date_start = datetime.date(year, month, 1)
            date_end = datetime.date(year, month, days_in_month)
            # Bucket: day-of-month label
            def _bucket(req_date):
                return str(req_date.day)
            buckets = [str(d) for d in range(1, days_in_month + 1)]

        elif view_type == 'weekly':
            week_start_param = request.query_params.get('week_start', '')
            try:
                date_start = datetime.date.fromisoformat(week_start_param)
            except (ValueError, TypeError):
                # Default: Monday of current week
                date_start = today - datetime.timedelta(days=today.weekday())
            date_end = date_start + datetime.timedelta(days=6)
            DAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            def _bucket(req_date):
                return DAYS[req_date.weekday()]
            buckets = DAYS

        else:  # fiscal (default)
            try:
                fy_start = int(request.query_params.get('fy_start', today.year if today.month >= 7 else today.year - 1))
            except (ValueError, TypeError):
                fy_start = today.year if today.month >= 7 else today.year - 1
            date_start = datetime.date(fy_start, 7, 1)
            date_end = datetime.date(fy_start + 1, 6, 30)
            MONTH_LABELS = [
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec',
                'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
            ]
            def _bucket(req_date):
                # Map month number to fiscal position
                m = req_date.month
                if m >= 7:
                    return MONTH_LABELS[m - 7]
                else:
                    return MONTH_LABELS[m + 5]
            buckets = MONTH_LABELS

        # ── Query ──────────────────────────────────────────────────────
        requests_qs = (
            qs.filter(date_start__gte=date_start, date_start__lte=date_end)
            .values('date_start', 'leave_type__name', 'total_days')
        )

        # Accumulate total_days per (bucket_label, leave_type_name)
        # Keys: leave_type name → {bucket_label: total_days}
        type_bucket_map: dict[str, dict[str, Decimal]] = {}
        for row in requests_qs:
            lt_name = row['leave_type__name'] or 'Unknown'
            bucket = _bucket(row['date_start'])
            days = Decimal(str(row['total_days']))
            if lt_name not in type_bucket_map:
                type_bucket_map[lt_name] = {}
            type_bucket_map[lt_name][bucket] = type_bucket_map[lt_name].get(bucket, Decimal('0')) + days

        # Build categories
        leave_type_names = sorted(type_bucket_map.keys())
        categories = [
            {
                'key': name.replace(' ', '_').lower(),
                'label': name,
                'color': self._COLORS[i % len(self._COLORS)],
                'gradId': f'grad_{i}',
                'lightColor': self._LIGHT_COLORS[i % len(self._LIGHT_COLORS)],
            }
            for i, name in enumerate(leave_type_names)
        ]

        # Build data rows
        data = []
        for bucket in buckets:
            row: dict = {'label': bucket}
            has_any = False
            for name in leave_type_names:
                key = name.replace(' ', '_').lower()
                val = type_bucket_map.get(name, {}).get(bucket, Decimal('0'))
                row[key] = float(val)
                if val > 0:
                    has_any = True
            # Only include buckets that exist within the window
            if view_type == 'fiscal':
                data.append(row)  # always include all 12 months
            else:
                data.append(row)

        return Response({'data': data, 'categories': categories})


# ── Approval Queue Export ─────────────────────────────────────────────────────

class ApprovalQueueExportView(APIView):
    """
    GET /api/leave/approval-queue/export
        ?period_start=YYYY-MM-DD&period_end=YYYY-MM-DD

    Role-based multi-sheet export:
      • hr=True or admin=True → 4 sheets (Overview, Leave Ranking, Leave Summary, Leave Routing)
      • clinic=True or iad=True → 1 sheet (Leave Summary)
    """
    permission_classes = [IsAuthenticated]

    # ── formatting helpers ────────────────────────────────────────────────────

    @staticmethod
    def _fmt_range(d1: datetime.date, d2: datetime.date) -> str:
        """Format a date range according to the spec."""
        if d1.year == d2.year and d1.month == d2.month:
            return f"{d1.strftime('%B')} {d1.day}–{d2.day}, {d1.year}"
        if d1.year == d2.year:
            return f"{d1.strftime('%B')} {d1.day:02d} – {d2.strftime('%B')} {d2.day:02d}, {d1.year}"
        return (
            f"{d1.strftime('%B')} {d1.day}, {d1.year} – "
            f"{d2.strftime('%B')} {d2.day}, {d2.year}"
        )

    @staticmethod
    def _fmt_date(d: datetime.date) -> str:
        return d.strftime('%B %d, %Y')

    @staticmethod
    def _person_name(user) -> str:
        last = (user.lastname or '').strip()
        first = (user.firstname or '').strip()
        if last and first:
            return f"{last}, {first}"
        return user.idnumber

    @staticmethod
    def _dept_of(emp) -> str:
        wi = (
            emp.workinformation_set
            .select_related('department')
            .order_by('-created_at')
            .first()
        )
        return wi.department.name if wi and wi.department else ''

    # ── shared openpyxl styling ───────────────────────────────────────────────

    @staticmethod
    def _apply_header_style(ws, header_row: int, num_cols: int):
        import openpyxl.styles as xls
        thin = xls.Side(style='thin')
        border = xls.Border(left=thin, right=thin, top=thin, bottom=thin)
        yellow_fill = xls.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = xls.Font(bold=True)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.border = border

    @staticmethod
    def _apply_data_border(ws, data_row: int, num_cols: int):
        import openpyxl.styles as xls
        thin = xls.Side(style='thin')
        border = xls.Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in range(1, num_cols + 1):
            ws.cell(row=data_row, column=col).border = border

    @staticmethod
    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # ── status colour ─────────────────────────────────────────────────────────

    @staticmethod
    def _status_fill(status: str):
        import openpyxl.styles as xls
        mapping = {
            'approved':    ('C6EFCE', 'C6EFCE'),
            'pending':     ('FFFF00', 'FFFF00'),
            'disapproved': ('FFCCCC', 'FFCCCC'),
        }
        colours = mapping.get(status, ('FFFFFF', 'FFFFFF'))
        return xls.PatternFill(start_color=colours[0], end_color=colours[1], fill_type='solid')

    # ── colour ranks within a column ─────────────────────────────────────────

    @staticmethod
    def _rank_fills():
        import openpyxl.styles as xls
        return [
            xls.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),   # 1st – red
            xls.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),   # 2nd – orange
            xls.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),   # 3rd – yellow
        ]

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _build_summary_sheet(self, ws, qs, period_start, period_end):
        """Shared Leave Summary sheet used by both HR and clinic/IAD exports."""
        import openpyxl.styles as xls

        ws['A1'] = 'RYONAN ELECTRIC PHILIPPINES CORPORATION'
        ws['A2'] = f'Leave Request Summary for {self._fmt_range(period_start, period_end)}'

        headers = [
            'Control Number', 'Date Prepared', 'ID Number', 'Employee Name',
            'Department', 'Leave Type', 'Reason Category', 'Reason',
            'Duration', 'Days', 'Hours', 'Status',
        ]
        HEADER_ROW = 4
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        self._apply_header_style(ws, HEADER_ROW, len(headers))

        thin = xls.Side(style='thin')
        border = xls.Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_num, lr in enumerate(qs, HEADER_ROW + 1):
            dept = self._dept_of(lr.employee)
            values = [
                lr.control_number,
                self._fmt_date(lr.date_prepared),
                lr.employee.idnumber,
                self._person_name(lr.employee),
                dept,
                lr.leave_type.name,
                lr.reason.title,
                lr.subreason.title if lr.subreason else (lr.remarks or ''),
                self._fmt_range(lr.date_start, lr.date_end),
                str(lr.total_days),
                str(lr.total_hours),
                lr.get_status_display(),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = border
                if col_idx == len(headers):  # Status column
                    cell.fill = self._status_fill(lr.status)

        self._auto_width(ws)

    def _build_routing_sheet(self, ws, qs, period_start, period_end):
        """Leave Routing sheet — only includes requests not yet HR-approved."""
        import openpyxl.styles as xls

        ws['A1'] = 'RYONAN ELECTRIC PHILIPPINES CORPORATION'
        ws['A2'] = f'Leave Routing Summary for {self._fmt_range(period_start, period_end)}'

        headers = [
            'Control Number', 'Date Prepared', 'ID Number', 'Employee Name',
            'Department', 'Leave Type', 'Reason Category', 'Reason',
            'Duration', 'Current Approver',
        ]
        HEADER_ROW = 4
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        self._apply_header_style(ws, HEADER_ROW, len(headers))

        thin = xls.Side(style='thin')
        border = xls.Border(left=thin, right=thin, top=thin, bottom=thin)

        # Exclude requests where HR has already acted, and exclude cancelled requests
        routing_qs = qs.exclude(
            approval_steps__role_group='hr',
            approval_steps__status__in=['approved', 'disapproved'],
        ).exclude(
            status='cancelled',
        ).prefetch_related(
            'approval_steps__acted_by',
            'approval_steps__approver',
        )

        for row_num, lr in enumerate(routing_qs, HEADER_ROW + 1):
            dept = self._dept_of(lr.employee)
            # Active pending step's approver
            active_step = lr.approval_steps.filter(status='pending').order_by('sequence').first()
            current_approver = ''
            if active_step:
                if active_step.approver:
                    current_approver = self._person_name(active_step.approver)
                else:
                    current_approver = active_step.get_role_group_display()

            values = [
                lr.control_number,
                self._fmt_date(lr.date_prepared),
                lr.employee.idnumber,
                self._person_name(lr.employee),
                dept,
                lr.leave_type.name,
                lr.reason.title,
                lr.subreason.title if lr.subreason else (lr.remarks or ''),
                self._fmt_range(lr.date_start, lr.date_end),
                current_approver,
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = border

        self._auto_width(ws)

    def _build_overview_sheet(self, ws, qs, period_start, period_end):
        """Sheet 1 — Overview (HR only)."""
        import openpyxl.styles as xls
        from collections import defaultdict

        ws['A1'] = 'RYONAN ELECTRIC PHILIPPINES CORPORATION'
        ws['A2'] = f'Leave Overview for {self._fmt_range(period_start, period_end)}'

        # Fetch all leave requests in range with department info
        requests = list(
            qs.select_related('leave_type', 'employee')
            .prefetch_related('employee__workinformation_set__department')
        )

        # Build: {leave_type: {dept: count}}
        lt_dept: dict = defaultdict(lambda: defaultdict(int))
        lt_total: dict = defaultdict(int)
        all_depts: set = set()

        for lr in requests:
            dept = self._dept_of(lr.employee)
            lt_dept[lr.leave_type.name][dept] += 1
            lt_total[lr.leave_type.name] += 1
            if dept:
                all_depts.add(dept)

        sorted_depts = sorted(all_depts)
        total_all = sum(lt_total.values()) or 1

        headers = ['Leave Type', 'Leave Count', 'Leave Percentage'] + sorted_depts
        HEADER_ROW = 4
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        self._apply_header_style(ws, HEADER_ROW, len(headers))

        thin = xls.Side(style='thin')
        border = xls.Border(left=thin, right=thin, top=thin, bottom=thin)
        rank_fills = self._rank_fills()

        dept_col_start = 4  # 1-indexed column where dept columns begin

        for row_num, lt_name in enumerate(sorted(lt_dept.keys()), HEADER_ROW + 1):
            count = lt_total[lt_name]
            pct = f"{(count / total_all * 100):.1f}%"
            row_values = [lt_name, count, pct] + [lt_dept[lt_name].get(d, 0) for d in sorted_depts]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = border

        # Apply rank highlighting across each dept column
        for dept_idx, dept in enumerate(sorted_depts):
            col_num = dept_col_start + dept_idx
            # Collect (row_num, value) for this dept column
            col_data = []
            for row_num, lt_name in enumerate(sorted(lt_dept.keys()), HEADER_ROW + 1):
                col_data.append((row_num, lt_dept[lt_name].get(dept, 0)))
            col_data_sorted = sorted(col_data, key=lambda x: x[1], reverse=True)
            for rank, (row_num, val) in enumerate(col_data_sorted[:3]):
                if val > 0:
                    ws.cell(row=row_num, column=col_num).fill = rank_fills[rank]

        self._auto_width(ws)

    def _build_ranking_sheet(self, ws, qs, period_start: datetime.date):
        """Sheet 2 — Leave Ranking (HR only). Fiscal year May–April."""
        import openpyxl.styles as xls
        from collections import defaultdict

        # Fiscal year: May of period_start's year → April of next year
        fy_start_year = period_start.year if period_start.month >= 5 else period_start.year - 1
        fy_start = datetime.date(fy_start_year, 5, 1)
        fy_end = datetime.date(fy_start_year + 1, 4, 30)

        ws['A1'] = 'RYONAN ELECTRIC PHILIPPINES CORPORATION'
        ws['A2'] = f'Leave Ranking for FY {fy_start_year}–{fy_start_year + 1}'

        # Fiscal months: May … April
        fiscal_months = []
        for m in range(5, 13):   # May–Dec of fy_start_year
            fiscal_months.append((fy_start_year, m))
        for m in range(1, 5):    # Jan–Apr of fy_start_year+1
            fiscal_months.append((fy_start_year + 1, m))

        month_labels = [
            datetime.date(y, m, 1).strftime('%B %Y') for y, m in fiscal_months
        ]

        headers = ['Leave Sub Reason'] + month_labels
        HEADER_ROW = 4
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        self._apply_header_style(ws, HEADER_ROW, len(headers))

        thin = xls.Side(style='thin')
        border = xls.Border(left=thin, right=thin, top=thin, bottom=thin)

        # Fetch fiscal-year requests
        fy_qs = (
            LeaveRequest.objects
            .filter(date_start__gte=fy_start, date_start__lte=fy_end)
            .select_related('subreason', 'reason')
        )

        # {subreason_title: {(year, month): count}}
        sr_month: dict = defaultdict(lambda: defaultdict(int))
        for lr in fy_qs:
            sr_title = lr.subreason.title if lr.subreason else f'({lr.reason.title})'
            key = (lr.date_start.year, lr.date_start.month)
            sr_month[sr_title][key] += 1

        all_subreasons = sorted(sr_month.keys())

        rank_fills = self._rank_fills()

        for row_num, sr in enumerate(all_subreasons, HEADER_ROW + 1):
            row_values = [sr] + [sr_month[sr].get(ym, 0) for ym in fiscal_months]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = border

        # Apply rank highlighting per month column
        for m_idx, ym in enumerate(fiscal_months):
            col_num = 2 + m_idx
            col_data = []
            for row_num, sr in enumerate(all_subreasons, HEADER_ROW + 1):
                col_data.append((row_num, sr_month[sr].get(ym, 0)))
            col_data_sorted = sorted(col_data, key=lambda x: x[1], reverse=True)
            for rank, (row_num, val) in enumerate(col_data_sorted[:3]):
                if val > 0:
                    ws.cell(row=row_num, column=col_num).fill = rank_fills[rank]

        self._auto_width(ws)

    # ── Main handler ──────────────────────────────────────────────────────────

    def get(self, request):
        user = request.user
        is_hr_or_admin = getattr(user, 'hr', False) or getattr(user, 'admin', False)
        is_clinic_or_iad = getattr(user, 'clinic', False) or getattr(user, 'iad', False)

        if not (is_hr_or_admin or is_clinic_or_iad):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        # Parse date range
        try:
            period_start = datetime.date.fromisoformat(request.query_params.get('period_start', ''))
            period_end = datetime.date.fromisoformat(request.query_params.get('period_end', ''))
        except (ValueError, TypeError):
            return Response({'detail': 'period_start and period_end are required (YYYY-MM-DD).'}, status=http_status.HTTP_400_BAD_REQUEST)

        if period_end < period_start:
            return Response({'detail': 'period_end must be on or after period_start.'}, status=http_status.HTTP_400_BAD_REQUEST)

        # Base queryset: leave requests in the selected period
        base_qs = (
            LeaveRequest.objects
            .filter(date_start__gte=period_start, date_start__lte=period_end)
            .select_related('leave_type', 'reason', 'subreason', 'employee')
            .prefetch_related(
                'employee__workinformation_set__department',
                'approval_steps__approver',
                'approval_steps__acted_by',
            )
            .order_by('date_prepared', 'control_number')
        )

        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        if is_hr_or_admin:
            ws1 = wb.create_sheet('Overview')
            self._build_overview_sheet(ws1, base_qs, period_start, period_end)

            ws2 = wb.create_sheet('Leave Ranking')
            self._build_ranking_sheet(ws2, base_qs, period_start)

            ws3 = wb.create_sheet('Leave Summary')
            self._build_summary_sheet(ws3, base_qs, period_start, period_end)

            ws4 = wb.create_sheet('Leave Routing')
            self._build_routing_sheet(ws4, base_qs, period_start, period_end)
        else:
            # clinic or iad only
            ws = wb.create_sheet('Leave Summary')
            self._build_summary_sheet(ws, base_qs, period_start, period_end)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        range_str = self._fmt_range(period_start, period_end).replace('–', '-').replace(' ', '_')
        filename = f'leave_export_{range_str}.xlsx'

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
