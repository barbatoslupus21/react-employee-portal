import calendar as _calendar
import datetime
import io
import logging
from typing import cast

from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import CalendarEvent, Timelogs
from .serializers import CalendarEventSerializer
from userLogin.models import loginCredentials

logger = logging.getLogger(__name__)

# ── Shared styling helpers ────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill('solid', fgColor=Color('0D1A63'))   # deep navy
_HEADER_FONT   = Font(bold=True, color='FFFFFF', size=11)
_ERROR_FILL    = PatternFill('solid', fgColor=Color('FFCCCC'))    # light red
_ERROR_FONT    = Font(bold=True, color='CC0000', size=10)
_NORMAL_FONT   = Font(size=10)
_CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT          = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def _thin_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

_BORDER = _thin_border()

_DATE_FORMAT_STR = '%m/%d/%Y %I:%M:%S %p'   # 08/15/2025 9:00:00 AM


def _build_template_wb() -> Workbook:
    """Return a fresh template workbook."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Timelogs'

    headers = ['ID Number', 'Employee', 'Date and Time', 'Entry']
    col_widths = [18, 30, 28, 12]

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Sample data row
    sample = ['EMP-001', 'Juan Dela Cruz', '08/15/2025 9:00:00 AM', 'IN']
    for col_idx, value in enumerate(sample, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font      = _NORMAL_FONT
        cell.alignment = _LEFT
        cell.border    = _BORDER

    # Data validation on Entry column (col D = 4), row 2 onwards
    dv = DataValidation(
        type='list',
        formula1='"IN,OUT"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle='Invalid entry',
        error='Value must be IN or OUT.',
    )
    ws.add_data_validation(dv)
    dv.sqref = 'D2:D10000'

    return wb


# ── Shared timelog pairing helper ────────────────────────────────────────────

def _pair_timelogs(local_logs: list[dict]) -> dict[datetime.date, dict]:
    """
    Given a list of {'dt': datetime, 'entry': 'IN'|'OUT'} dicts ordered by time,
    pair each IN with the next unused OUT within 16 hours.
    Orphan OUTs before 12:00 are attributed to the previous calendar day.
    Returns a dict of { date: {'in': datetime|None, 'out': datetime|None} }.
    """
    MAX_PAIR_HOURS = 16
    work_days: dict[datetime.date, dict] = {}
    used = [False] * len(local_logs)

    for i, rec in enumerate(local_logs):
        if rec['entry'] != 'IN' or used[i]:
            continue
        work_date = rec['dt'].date()
        slot = work_days.setdefault(work_date, {'in': None, 'out': None, 'in_flag': False, 'out_flag': False})
        if not slot['in_flag']:
            slot['in'] = rec['dt']
            slot['in_flag'] = True
        used[i] = True
        for j in range(i + 1, len(local_logs)):
            if used[j] or local_logs[j]['entry'] != 'OUT':
                continue
            gap_h = (local_logs[j]['dt'] - rec['dt']).total_seconds() / 3600
            if gap_h > MAX_PAIR_HOURS:
                break
            if not slot['out_flag']:
                slot['out'] = local_logs[j]['dt']
                slot['out_flag'] = True
            used[j] = True
            break

    for i, rec in enumerate(local_logs):
        if rec['entry'] != 'OUT' or used[i]:
            continue
        work_date = (
            rec['dt'].date() - datetime.timedelta(days=1)
            if rec['dt'].hour < 12
            else rec['dt'].date()
        )
        slot = work_days.setdefault(work_date, {'in': None, 'out': None, 'in_flag': False, 'out_flag': False})
        if not slot['out_flag']:
            slot['out'] = rec['dt']
            slot['out_flag'] = True
        used[i] = True

    return work_days


def _compute_week_completeness(local_logs: list[dict], working_days: list[datetime.date]) -> int:
    """Return an integer 0–100 representing completeness for the given working days."""
    if not working_days:
        return 100
    work_days = _pair_timelogs(local_logs)
    complete = sum(1 for d in working_days if work_days.get(d, {}).get('in') and work_days.get(d, {}).get('out'))
    return round((complete / len(working_days)) * 100)


# ── Timelogs views ─────────────────────────────────────────────────────────────

class TimelogsCompletenessView(APIView):
    """
    GET /api/timelogs/completeness
    Returns all active employees who are not admin, HR, or accounting,
    along with their timelog completeness percentage for the current week.
    Only accessible by HR, HR managers, or admins.

    Completeness = (days with both Time-In and Time-Out) / (working days in
    the current week up to today, Mon–Sat excluding Sun) × 100.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if not (getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'hr_manager', False)):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        today = timezone.localdate()
        # Monday of the current week
        week_start = today - datetime.timedelta(days=today.weekday())

        # Working days in the current week up to today (Mon–Sat, no Sunday)
        working_days: list[datetime.date] = []
        d = week_start
        while d <= today:
            if d.weekday() != 6:
                working_days.append(d)
            d += datetime.timedelta(days=1)

        employees = list(
            loginCredentials.objects
            .filter(active=True, admin=False, hr=False, hr_manager=False, accounting=False)
            .order_by('lastname', 'firstname')
        )

        if not working_days:
            data = [
                {
                    'idnumber': e.idnumber,
                    'firstname': e.firstname or '',
                    'lastname': e.lastname or '',
                    'department': '',
                    'line': '',
                    'completeness': 100,
                }
                for e in employees
            ]
            return Response(data)

        # Fetch all timelogs for all target employees for the week (+ 1-day buffer)
        tz = timezone.get_current_timezone()
        range_start = timezone.make_aware(
            datetime.datetime.combine(week_start - datetime.timedelta(days=1), datetime.time.min), tz)
        range_end = timezone.make_aware(
            datetime.datetime.combine(today + datetime.timedelta(days=1), datetime.time.max), tz)

        all_logs = (
            Timelogs.objects
            .filter(employee__in=employees, time__range=(range_start, range_end))
            .order_by('employee_id', 'time')
            .values('employee_id', 'time', 'entry')
        )

        # Group logs by employee PK
        from collections import defaultdict
        emp_logs: dict[int, list[dict]] = defaultdict(list)
        for log in all_logs:
            emp_logs[log['employee_id']].append({
                'dt': timezone.localtime(log['time'], tz),
                'entry': log['entry'],
            })

        data = []
        for emp in employees:
            completeness = _compute_week_completeness(emp_logs.get(emp.pk, []), working_days)
            data.append({
                'idnumber': emp.idnumber,
                'firstname': emp.firstname or '',
                'lastname': emp.lastname or '',
                'department': '',
                'line': '',
                'completeness': completeness,
            })
        return Response(data)


class UserTimelogsView(APIView):
    """
    GET /api/timelogs/user-logs?idnumber=<idnumber>
    Returns the current week's daily timelog summary for a specific employee.
    Response: list of { date, time_in, time_out, is_complete } for each
    working day (Mon–Sat, no Sunday) in the current week up to today.
    Only accessible by HR, HR managers, or admins.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if not (getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'hr_manager', False)):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        idnumber = request.GET.get('idnumber', '').strip()
        if not idnumber:
            return Response({'detail': 'idnumber is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = loginCredentials.objects.get(idnumber=idnumber, active=True)
        except loginCredentials.DoesNotExist:
            return Response({'detail': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        today = timezone.localdate()
        week_start = today - datetime.timedelta(days=today.weekday())

        working_days: list[datetime.date] = []
        d = week_start
        while d <= today:
            if d.weekday() != 6:
                working_days.append(d)
            d += datetime.timedelta(days=1)

        tz = timezone.get_current_timezone()
        range_start = timezone.make_aware(
            datetime.datetime.combine(week_start - datetime.timedelta(days=1), datetime.time.min), tz)
        range_end = timezone.make_aware(
            datetime.datetime.combine(today + datetime.timedelta(days=1), datetime.time.max), tz)

        raw_logs = list(
            Timelogs.objects
            .filter(employee=employee, time__range=(range_start, range_end))
            .order_by('time')
            .values('time', 'entry')
        )

        local_logs = [
            {'dt': timezone.localtime(log['time'], tz), 'entry': log['entry']}
            for log in raw_logs
        ]

        work_days = _pair_timelogs(local_logs)

        result = []
        for day in working_days:
            slot = work_days.get(day, {})
            time_in_dt = slot.get('in')
            time_out_dt = slot.get('out')
            is_complete = time_in_dt is not None and time_out_dt is not None
            result.append({
                'date': day.isoformat(),
                'time_in': time_in_dt.strftime('%I:%M %p') if time_in_dt else None,
                'time_out': time_out_dt.strftime('%I:%M %p') if time_out_dt else None,
                'is_complete': is_complete,
            })

        return Response(result)


class TimelogsTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if not (getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'hr_manager', False)):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        wb = _build_template_wb()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="timelogs_template.xlsx"'
        return response


@method_decorator(csrf_protect, name='dispatch')
class TimelogsUploadView(APIView):
    """
    POST /api/timelogs/upload
    Accepts a multipart/form-data file (xlsx/xls/csv).

    Validation rules:
      - All four columns required (no blank cells).
      - Date and Time must match MM/DD/YYYY H:MM:SS AM/PM.
      - ID Number must exist in active user records.
      - Entry must be exactly IN or OUT (case-insensitive, stored uppercase).

    On error  → HTTP 422, returns an xlsx with invalid cells highlighted red.
    On success → HTTP 200, all rows saved to Timelogs, returns { saved: N }.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        user = request.user
        if not (getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'hr_manager', False)):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = uploaded.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
            return Response({'detail': 'Unsupported file type. Use .xlsx, .xls, or .csv.'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Load workbook ──────────────────────────────────────────────────
        try:
            wb_in = load_workbook(uploaded, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read file. Make sure it is a valid Excel file.'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb_in.active
        assert ws is not None

        # Build set of valid idnumbers for O(1) lookup
        valid_ids = set(
            loginCredentials.objects
            .filter(active=True)
            .values_list('idnumber', flat=True)
        )

        # ── Collect rows (skip header) ─────────────────────────────────────
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return Response({'detail': 'The file contains no data rows.'}, status=status.HTTP_400_BAD_REQUEST)

        errors: list[tuple[int, int, str]] = []   # (row, col, reason)

        parsed: list[dict] = []   # valid rows ready to insert

        for row_idx, row in enumerate(rows, start=2):
            id_val   = str(row[0]).strip() if row[0] is not None else ''
            emp_val  = str(row[1]).strip() if row[1] is not None else ''
            dt_val   = row[2]
            ent_val  = str(row[3]).strip().upper() if row[3] is not None else ''

            row_errors = False

            # ── Col A: ID Number ──
            if not id_val:
                errors.append((row_idx, 1, 'ID Number is required.'))
                row_errors = True
            elif id_val not in valid_ids:
                errors.append((row_idx, 1, f'ID Number "{id_val}" not found in the system.'))
                row_errors = True

            # ── Col B: Employee ──
            if not emp_val:
                errors.append((row_idx, 2, 'Employee name is required.'))
                row_errors = True

            # ── Col C: Date and Time ──
            parsed_dt = None
            if dt_val is None or str(dt_val).strip() == '':
                errors.append((row_idx, 3, 'Date and Time is required.'))
                row_errors = True
            elif isinstance(dt_val, (datetime.datetime, datetime.date)):
                # openpyxl already parsed the cell as a datetime object
                if isinstance(dt_val, datetime.date) and not isinstance(dt_val, datetime.datetime):
                    dt_val = datetime.datetime(dt_val.year, dt_val.month, dt_val.day)
                parsed_dt = dt_val
            else:
                dt_str = str(dt_val).strip()
                try:
                    parsed_dt = datetime.datetime.strptime(dt_str, _DATE_FORMAT_STR)
                except ValueError:
                    errors.append((row_idx, 3, f'Invalid format "{dt_str}". Expected MM/DD/YYYY H:MM:SS AM/PM.'))
                    row_errors = True

            # ── Col D: Entry ──
            if not ent_val:
                errors.append((row_idx, 4, 'Entry is required.'))
                row_errors = True
            elif ent_val not in ('IN', 'OUT'):
                errors.append((row_idx, 4, f'Entry must be IN or OUT, got "{ent_val}".'))
                row_errors = True

            if not row_errors and parsed_dt is not None:
                parsed.append({
                    'idnumber': id_val,
                    'time':     parsed_dt,
                    'entry':    ent_val,
                })

        # ── If errors → return highlighted workbook ────────────────────────
        if errors:
            wb_out = _build_error_workbook(ws, errors)
            buf = io.BytesIO()
            wb_out.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                status=200,
            )
            response['Content-Disposition'] = 'attachment; filename="timelogs_errors.xlsx"'
            response['X-Validation-Errors'] = 'true'
            return response

        # ── Save to DB ────────────────────────────────────────────────────
        with transaction.atomic():
            user_map = {
                u.idnumber: u
                for u in loginCredentials.objects.filter(
                    idnumber__in=[r['idnumber'] for r in parsed],
                    active=True,
                )
            }
            Timelogs.objects.bulk_create([
                Timelogs(
                    employee=user_map[r['idnumber']],
                    time=timezone.make_aware(r['time']) if timezone.is_naive(r['time']) else r['time'],
                    entry=r['entry'],
                )
                for r in parsed
                if r['idnumber'] in user_map
            ])

        logger.info('Timelogs uploaded: %d rows by user=%s', len(parsed), request.user.pk)
        return Response({'saved': len(parsed)}, status=status.HTTP_200_OK)


def _build_error_workbook(source_ws, errors: list[tuple[int, int, str]]) -> Workbook:
    """
    Copy the source worksheet into a new workbook and mark each invalid cell
    with a red fill + bold red font.  The header row is re-styled normally.
    """
    wb_out = _build_template_wb()
    ws_out = wb_out.active
    assert ws_out is not None

    # Map (row, col) → error reason for fast lookup
    error_map: dict[tuple[int, int], str] = {(r, c): msg for r, c, msg in errors}

    headers = ['ID Number', 'Employee', 'Date and Time', 'Entry']

    # Re-write all data rows from source (skip header = row 1)
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        ws_out.append([''] * 4)   # placeholder — we'll set values below

    for row_idx, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
        for col_idx, value in enumerate(row[:4], start=1):
            cell = ws_out.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = _BORDER
            key = (row_idx, col_idx)
            if key in error_map:
                cell.fill  = _ERROR_FILL
                cell.font  = _ERROR_FONT
                cell.alignment = _LEFT
                cell.comment = None   # no openpyxl comment needed; colour is enough
            else:
                cell.font      = _NORMAL_FONT
                cell.alignment = _LEFT

    return wb_out


# ── Timelog daily status ──────────────────────────────────────────────────────

class TimelogDailyStatusView(APIView):
    """
    GET /api/timelogs/daily-status?year=Y&month=M

    Returns a mapping of { "YYYY-MM-DD": "<status>" } for every day in the
    requested month that has a non-complete timelog status for the current user.

    Statuses:
        no_time_out  — IN exists but no paired OUT
        no_time_in   — OUT exists but no paired IN
        absent       — no records (and not a holiday / Sunday / future date)

    Excluded users (returns {} immediately):
        admin, hr, accounting
    """
    permission_classes = [IsAuthenticated]

    # Holiday event_types that suppress Absent for the event members/owner
    _HOLIDAY_TYPES = {'legal', 'special', 'day_off', 'company'}

    def get(self, request):
        user = request.user

        # ── Privilege check ────────────────────────────────────────────────
        if (
            getattr(user, 'admin', False)
            or getattr(user, 'hr', False)
            or getattr(user, 'accounting', False)
        ):
            return Response({})

        # ── Parse query params ─────────────────────────────────────────────
        try:
            year  = int(request.GET['year'])
            month = int(request.GET['month'])
            if not (1 <= month <= 12):
                raise ValueError
        except (KeyError, ValueError):
            return Response({'detail': 'year and month (1-12) are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Date range ─────────────────────────────────────────────────────
        today         = timezone.localdate()
        first_day     = datetime.date(year, month, 1)
        last_in_month = datetime.date(year, month, _calendar.monthrange(year, month)[1])
        last_day      = min(last_in_month, today)

        if first_day > today:
            return Response({})

        # ── Fetch timelogs ──────────────────────────────────────────────────
        # Extend range by 1 day on each side so night-shift OUTs that fall
        # just after midnight on the first/last day of the month are captured.
        tz          = timezone.get_current_timezone()
        fetch_start = first_day - datetime.timedelta(days=1)
        fetch_end   = last_day  + datetime.timedelta(days=1)
        range_start = timezone.make_aware(
            datetime.datetime.combine(fetch_start, datetime.time.min), tz)
        range_end   = timezone.make_aware(
            datetime.datetime.combine(fetch_end,   datetime.time.max), tz)

        raw_logs = list(
            Timelogs.objects
            .filter(employee=user, time__range=(range_start, range_end))
            .order_by('time')
            .values('time', 'entry')
        )

        local_logs = [
            {'dt': timezone.localtime(log['time'], tz), 'entry': log['entry']}
            for log in raw_logs
        ]

        # ── Phase 1: Forward-pair each IN with the next unused OUT (≤ 16 h) ─
        # Works for both day and night shifts without needing shift configuration.
        # The work-day date is anchored to the IN record's local date.
        MAX_PAIR_HOURS = 16
        work_days: dict[datetime.date, dict] = {}
        used = [False] * len(local_logs)

        for i, rec in enumerate(local_logs):
            if rec['entry'] != 'IN' or used[i]:
                continue
            work_date = rec['dt'].date()
            slot = work_days.setdefault(work_date, {'in': False, 'out': False})
            slot['in'] = True
            used[i] = True
            # Find the next unused OUT within MAX_PAIR_HOURS
            for j in range(i + 1, len(local_logs)):
                if used[j] or local_logs[j]['entry'] != 'OUT':
                    continue
                gap_h = (local_logs[j]['dt'] - rec['dt']).total_seconds() / 3600
                if gap_h > MAX_PAIR_HOURS:
                    break  # logs are chronological — no closer OUT exists
                work_days[work_date]['out'] = True
                used[j] = True
                break

        # ── Phase 2: Handle orphan OUTs (no matched IN within 16 h before) ─
        # Early-morning OUTs (before 12:00) are attributed to the previous
        # calendar day — the employee clocked out after midnight (night shift).
        # Later OUTs are attributed to their own date (missed clock-in).
        for i, rec in enumerate(local_logs):
            if rec['entry'] != 'OUT' or used[i]:
                continue
            if rec['dt'].hour < 12:
                work_date = rec['dt'].date() - datetime.timedelta(days=1)
            else:
                work_date = rec['dt'].date()
            slot = work_days.setdefault(work_date, {'in': False, 'out': False})
            slot['out'] = True
            used[i] = True

        # ── Collect holiday dates visible to this user ─────────────────────
        holiday_events = CalendarEvent.objects.filter(
            Q(owner=user) | Q(members=user),
            event_type__in=self._HOLIDAY_TYPES,
        ).distinct()

        # Expand recurring holiday events into actual dates within the month
        holiday_dates: set[datetime.date] = set()
        for ev in holiday_events:
            d = first_day
            while d <= last_in_month:
                ev_date = ev.date
                d_str   = d.isoformat()
                if d_str < ev_date.isoformat():
                    d += datetime.timedelta(days=1)
                    continue
                match = False
                rep = ev.repetition
                if   rep == 'once':    match = (d == ev_date)
                elif rep == 'daily':   match = (d.weekday() != 6)
                elif rep == 'weekly':  match = (d.weekday() == ev_date.weekday())
                elif rep == 'monthly': match = (d.day == ev_date.day)
                elif rep == 'yearly':  match = (d.month == ev_date.month and d.day == ev_date.day)
                else:                  match = (d == ev_date)
                if match:
                    holiday_dates.add(d)
                d += datetime.timedelta(days=1)

        # ── Evaluate each calendar day ──────────────────────────────────────
        result: dict[str, str] = {}

        d = first_day
        while d <= last_day:
            # Skip Sunday
            if d.weekday() == 6:
                d += datetime.timedelta(days=1)
                continue
            # Skip holidays
            if d in holiday_dates:
                d += datetime.timedelta(days=1)
                continue

            slot = work_days.get(d)
            if slot is None or (not slot['in'] and not slot['out']):
                result[d.isoformat()] = 'absent'
            elif slot['in'] and slot['out']:
                pass  # complete — no pill
            elif slot['in']:
                result[d.isoformat()] = 'no_time_out'
            else:
                result[d.isoformat()] = 'no_time_in'

            d += datetime.timedelta(days=1)

        return Response(result)


@method_decorator(csrf_protect, name='dispatch')
class CalendarEventListCreateView(APIView):
    """
    GET  /api/calendar/events          — list events for the authenticated user
    POST /api/calendar/events          — create a new event

    Query params for GET:
        year:  4-digit integer   (optional, filters by year)
        month: 1-12 integer      (optional, filters by month)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = CalendarEvent.objects.filter(
            Q(owner=request.user) | Q(members=request.user)
        ).distinct().select_related('owner').prefetch_related('members')

        year_param  = request.GET.get('year')
        month_param = request.GET.get('month')

        year_int  = None
        month_int = None

        if year_param:
            try:
                year_int = int(year_param)
            except ValueError:
                return Response({'detail': 'Invalid year.'}, status=status.HTTP_400_BAD_REQUEST)

        if month_param:
            try:
                month_int = int(month_param)
                if not 1 <= month_int <= 12:
                    raise ValueError
            except ValueError:
                return Response({'detail': 'Invalid month.'}, status=status.HTTP_400_BAD_REQUEST)

        if year_int is not None and month_int is not None:
            # Last day of the requested month — recurring events that started on or
            # before this day can potentially appear within this month.
            last_day  = _calendar.monthrange(year_int, month_int)[1]
            month_end = datetime.date(year_int, month_int, last_day)
            qs = qs.filter(
                Q(date__year=year_int, date__month=month_int) |
                (~Q(repetition='once') & Q(date__lte=month_end))
            )
        elif year_int is not None:
            qs = qs.filter(Q(date__year=year_int) | ~Q(repetition='once'))
        elif month_int is not None:
            qs = qs.filter(Q(date__month=month_int) | ~Q(repetition='once'))

        serializer = CalendarEventSerializer(qs, many=True)
        return Response(serializer.data)

    @transaction.atomic
    def post(self, request):
        serializer = CalendarEventSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        # serializer.save() returns CalendarEvent instance; ignore uncertain return type
        event = serializer.save(owner=request.user)  # type: ignore[assignment]
        logger.info(
            'CalendarEvent created id=%d by user=%s',
            cast(CalendarEvent, event).pk,
            request.user.pk,
        )
        return Response(CalendarEventSerializer(event).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_protect, name='dispatch')
class CalendarEventDetailView(APIView):
    """
    GET    /api/calendar/events/<pk>   — retrieve a single event
    PUT    /api/calendar/events/<pk>   — full update
    PATCH  /api/calendar/events/<pk>   — partial update
    DELETE /api/calendar/events/<pk>   — delete
    """
    permission_classes = [IsAuthenticated]

    def _get_object(self, pk: int, user):
        try:
            return CalendarEvent.objects.select_related('owner').prefetch_related('members').get(pk=pk, owner=user)
        except CalendarEvent.DoesNotExist:
            return None

    def get(self, request, pk: int):
        event = self._get_object(pk, request.user)
        if not event:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(CalendarEventSerializer(event).data)

    @transaction.atomic
    def put(self, request, pk: int):
        event = self._get_object(pk, request.user)
        if not event:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CalendarEventSerializer(event, data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data)

    @transaction.atomic
    def patch(self, request, pk: int):
        event = self._get_object(pk, request.user)
        if not event:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CalendarEventSerializer(event, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data)

    @transaction.atomic
    def delete(self, request, pk: int):
        event = self._get_object(pk, request.user)
        if not event:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        event.delete()
        logger.info('CalendarEvent deleted id=%d by user=%s', pk, request.user.pk)
        return Response(status=status.HTTP_204_NO_CONTENT)
