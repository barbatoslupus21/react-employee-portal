import calendar
import datetime
import io
from typing import Any, Dict, cast

from django.db import transaction
from django.http import HttpResponse
from django.db.models import Q, Count, Case, When, Value, IntegerField
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from decimal import Decimal

from finance.models import Loan, LoanType, Allowance, AllowanceType

from .models import PRFRequest, EmergencyLoan, MedicineAllowance
from .serializers import PRFRequestSerializer, EmergencyLoanCreateSerializer, MedicineAllowanceCreateSerializer, PRFAdminSerializer, PRFAdminActionSerializer


# ── Notification helper ────────────────────────────────────────────────────────────

_NOTIF_MAP = {
    'approved':    ('prf_approved',    'PRF Request Approved'),
    'disapproved': ('prf_disapproved', 'PRF Request Disapproved'),
    'cancelled':   ('prf_cancelled',   'PRF Request Cancelled'),
}

_NOTIF_MESSAGES = {
    'approved':    'Your PRF request {prf_no} has been approved.',
    'disapproved': 'Your PRF request {prf_no} has been disapproved.{remarks_part}',
    'cancelled':   'Your PRF request {prf_no} has been cancelled by an administrator.',
}


def _schedule_prf_notification(prf: PRFRequest, new_status: str, admin_remarks: str = '') -> None:
    """Schedule a Notification insert on transaction commit (non-blocking)."""
    entry = _NOTIF_MAP.get(new_status)
    if not entry:
        return
    notif_type, title = entry
    remarks_part = f' Remarks: {admin_remarks}' if admin_remarks else ''
    message = _NOTIF_MESSAGES[new_status].format(
        prf_no=prf.prf_control_number or str(prf.pk),
        remarks_part=remarks_part,
    )
    recipient_id = prf.employee_id
    prf_id       = prf.pk
    prf_no       = prf.prf_control_number or ''

    def _create() -> None:
        try:
            from activityLog.models import Notification
            Notification.objects.create(
                recipient_id=recipient_id,
                notification_type=notif_type,
                title=title,
                message=message,
                module='pr-form',
                related_object_id=prf_id,
            )
        except Exception:
            import logging
            logging.getLogger(__name__).exception('Failed to create PRF notification')

    try:
        transaction.on_commit(_create)
    except Exception:
        _create()


class PRFRequestListCreateView(APIView):
    """
    GET  /api/prform/requests/  — paginated list of the current user's PRF requests
    POST /api/prform/requests/  — create a new PRF request
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = PRFRequest.objects.filter(employee=request.user)

        # ── Filtering ──────────────────────────────────────────────────────
        prf_type    = request.GET.get('prf_type', '').strip()
        status_param = request.GET.get('status',   '').strip()
        if prf_type:
            qs = qs.filter(prf_type=prf_type)
        if status_param:
            qs = qs.filter(status=status_param)

        # ── Search (cross-field, all pages) ────────────────────────────────
        search_q = request.GET.get('search', '').strip()
        if search_q:
            qs = qs.filter(
                Q(prf_control_number__icontains=search_q) |
                Q(prf_type__icontains=search_q) |
                Q(purpose__icontains=search_q) |
                Q(status__icontains=search_q) |
                Q(control_number__icontains=search_q)
            )

        # ── Sorting ────────────────────────────────────────────────────────
        sort_by  = request.GET.get('sort_by',  'created_at')
        sort_dir = request.GET.get('sort_dir', 'desc')
        allowed_sorts = {'prf_control_number', 'prf_type', 'purpose', 'status', 'created_at'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        order_prefix = '' if sort_dir == 'asc' else '-'
        qs = qs.order_by(f'{order_prefix}{sort_by}')

        # ── Pagination ─────────────────────────────────────────────────────
        try:
            page = max(1, int(request.GET.get('page', 1)))
        except (ValueError, TypeError):
            page = 1
        page_size = 10
        total     = qs.count()
        start     = (page - 1) * page_size
        qs_page   = qs[start: start + page_size]

        serializer = PRFRequestSerializer(qs_page, many=True)
        return Response({
            'results':     serializer.data,
            'count':       total,
            'page':        page,
            'page_size':   page_size,
            'total_pages': max(1, -(-total // page_size)),   # ceiling division
        })

    @transaction.atomic
    def post(self, request):
        if getattr(request.user, 'admin', False):
            return Response(
                {'detail': 'Admins use the admin panel to manage PRF requests.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = PRFRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        prf = serializer.save(employee=request.user)
        return Response(PRFRequestSerializer(prf).data, status=status.HTTP_201_CREATED)


class PRFMetaView(APIView):
    """
    GET /api/prform/meta/
    Returns available PRF types, categories, and statuses for dropdown population.
    No privilege required — any authenticated user may call this.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response({
            'prf_types': [
                {'value': v, 'label': l} for v, l in PRFRequest.PRF_TYPES
            ],
            'prf_categories': [
                {'value': v, 'label': l} for v, l in PRFRequest.PRF_CATEGORIES
            ],
            'statuses': [
                {'value': v, 'label': l} for v, l in PRFRequest.STATUS_CHOICES
            ],
        })


class PRFRequestDetailView(APIView):
    """
    GET   /api/prform/requests/<pk>/  — get a single PRF request
    PATCH /api/prform/requests/<pk>/  — edit a pending PRF request
    """
    permission_classes = [IsAuthenticated]

    def _get_own(self, pk, user):
        try:
            return PRFRequest.objects.get(pk=pk, employee=user)
        except PRFRequest.DoesNotExist:
            return None

    def get(self, request, pk):
        prf = self._get_own(pk, request.user)
        if prf is None:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(PRFRequestSerializer(prf).data)

    @transaction.atomic
    def patch(self, request, pk):
        prf = self._get_own(pk, request.user)
        if prf is None:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if prf.status != 'pending':
            return Response(
                {'detail': 'Only pending requests can be edited.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = PRFRequestSerializer(prf, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data)


class PRFRequestCancelView(APIView):
    """
    POST /api/prform/requests/<pk>/cancel/  — cancel a pending PRF request
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, pk):
        try:
            prf = PRFRequest.objects.get(pk=pk, employee=request.user)
        except PRFRequest.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if prf.status != 'pending':
            return Response(
                {'detail': 'Only pending requests can be cancelled.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        prf.status = 'cancelled'
        prf.save()
        return Response(PRFRequestSerializer(prf).data)


class EmergencyLoanCreateView(APIView):
    """
    POST /api/prform/emergency-loan  — atomically create a PRFRequest + EmergencyLoan
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        if getattr(request.user, 'admin', False):
            return Response(
                {'detail': 'Admins use the admin panel to manage PRF requests.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = EmergencyLoanCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = cast(Dict[str, Any], serializer.validated_data)

        # ── Check 1: No existing pending emergency loan PRF ────────────────
        has_pending = PRFRequest.objects.select_for_update().filter(
            employee=request.user,
            prf_type='emergency_loan',
            status='pending',
        ).exists()
        if has_pending:
            return Response(
                {'detail': 'You already have a pending emergency loan request that has not yet been processed.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Check 2: Stacking eligibility ──────────────────────────────────
        el_loan_type, _ = LoanType.objects.get_or_create(
            name='Emergency Loan',
            defaults={'color': '#2845D6', 'stackable': False},
        )
        active_loan = (
            Loan.objects
            .select_for_update()
            .filter(
                employee=request.user,
                loan_type=el_loan_type,
                current_balance__gt=Decimal('0'),
            )
            .first()
        )
        stacking = False
        if active_loan is not None:
            if active_loan.monthly_deduction is None:
                return Response(
                    {
                        'detail': (
                            'You have an active emergency loan but no deduction schedule is set. '
                            'Please contact the Finance department before applying for a new loan.'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            on_final = abs(active_loan.current_balance - active_loan.monthly_deduction) <= Decimal('0.01')
            if not on_final:
                return Response(
                    {
                        'detail': (
                            'You currently have an active emergency loan. '
                            'A new loan can only be stacked when only one deduction remains on your existing balance.'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            today_day = timezone.localdate().day
            in_window = (3 <= today_day <= 8) or (18 <= today_day <= 23)
            if not in_window:
                return Response(
                    {
                        'detail': (
                            'Your existing loan is on its final deduction. '
                            'You may submit a new emergency loan request between the 3rd\u20138th or 18th\u201323rd of the month.'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            stacking = True

        # ── Backend name cross-validation ──────────────────────────────────
        first = (getattr(request.user, 'firstname', None) or '').strip()
        last  = (getattr(request.user, 'lastname',  None) or '').strip()
        expected_name = f'{first} {last}'.strip()
        if expected_name:
            submitted_name = data.get('employee_full_name', '').strip()
            if submitted_name.lower() != expected_name.lower():
                return Response(
                    {'employee_full_name': ['Name does not match your account records.']},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # ── Generate control number & persist PRFRequest ───────────────────
        el_control_number = EmergencyLoan.generate_control_number()

        prf = PRFRequest.objects.create(
            employee=request.user,
            prf_category=data.get('prf_category', ''),
            prf_type='emergency_loan',
            purpose=data.get('purpose', ''),
            control_number=el_control_number,
        )

        EmergencyLoan.objects.create(
            prf_request=prf,
            amount=data.get('amount', 0),
            number_of_cutoff=data.get('number_of_cutoff', 0),
            starting_date=data.get('starting_date'),
            employee_full_name=data.get('employee_full_name', ''),
        )

        return Response(PRFRequestSerializer(prf).data, status=status.HTTP_201_CREATED)


class EmergencyLoanPreCheckView(APIView):
    """
    POST /api/prform/emergency-loan/check
    Pre-flight eligibility check for Emergency Loan PRF.
    Runs Check 1 (no pending EL PRF) and Check 2 (stacking eligibility).
    Returns 200 OK with empty body on success.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        if getattr(request.user, 'admin', False):
            return Response(
                {'detail': 'Admins use the admin panel to manage PRF requests.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check 1: No existing pending emergency loan PRF ──────────────────
        has_pending = PRFRequest.objects.select_for_update().filter(
            employee=request.user,
            prf_type='emergency_loan',
            status='pending',
        ).exists()
        if has_pending:
            return Response(
                {
                    'detail': (
                        'You already have a pending Emergency Loan request that has not yet been processed. '
                        'Please wait for it to be reviewed before submitting a new one.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check 2: Stacking eligibility ────────────────────────────────────
        el_loan_type = LoanType.objects.filter(name='Emergency Loan').first()
        if el_loan_type is not None:
            active_loan = (
                Loan.objects
                .select_for_update()
                .filter(
                    employee=request.user,
                    loan_type=el_loan_type,
                    current_balance__gt=Decimal('0'),
                )
                .first()
            )
            if active_loan is not None:
                if active_loan.monthly_deduction is None:
                    return Response(
                        {
                            'detail': (
                                'You have an active Emergency Loan but no deduction schedule is set. '
                                'Please contact the Finance department before applying for a new loan.'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                on_final = abs(active_loan.current_balance - active_loan.monthly_deduction) <= Decimal('0.01')
                today_day = timezone.localdate().day
                in_window = (3 <= today_day <= 8) or (18 <= today_day <= 23)
                if not on_final or not in_window:
                    return Response(
                        {
                            'detail': (
                                'A new Emergency Loan can only be filed when your existing loan is on its final deduction '
                                'and the request falls within the eligible date window '
                                '(3rd\u20138th or 18th\u201323rd of the month).'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        return Response({}, status=status.HTTP_200_OK)


class MedicineAllowancePreCheckView(APIView):
    """
    POST /api/prform/medicine-allowance/check
    Pre-flight eligibility check for Medicine Allowance PRF.
    Returns the available Finance Allowance balance and covered period on success.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        if getattr(request.user, 'admin', False):
            return Response(
                {'detail': 'Admins use the admin panel to manage PRF requests.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check 1: No existing pending medicine allowance PRF ───────────────
        has_pending = PRFRequest.objects.select_for_update().filter(
            employee=request.user,
            prf_type='medicine_allowance',
            status='pending',
        ).exists()
        if has_pending:
            return Response(
                {'detail': 'You already have a pending medicine allowance request that has not yet been processed.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check 2: Has an available Finance Allowance balance ──────────────
        ma_type = AllowanceType.objects.filter(name__iexact='medicine allowance').first()
        if ma_type is None:
            return Response(
                {
                    'detail': (
                        'Medicine Allowance is not configured in the Finance system. '
                        'Please contact the Finance department.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        allowance = (
            Allowance.objects
            .select_for_update()
            .filter(
                employee=request.user,
                allowance_type=ma_type,
                deposited_date__isnull=True,
            )
            .order_by('-created_at')
            .first()
        )
        if allowance is None or allowance.amount <= Decimal('0'):
            return Response(
                {
                    'detail': (
                        'You have no available medicine allowance balance. '
                        'Please contact the Finance department.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'amount':         str(allowance.amount),
            'covered_period': allowance.covered_period,
        })


class MedicineAllowanceCreateView(APIView):
    """
    POST /api/prform/medicine-allowance
    Atomically creates a PRFRequest + MedicineAllowance record.
    Re-verifies eligibility inside the transaction before committing.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        if getattr(request.user, 'admin', False):
            return Response(
                {'detail': 'Admins use the admin panel to manage PRF requests.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = MedicineAllowanceCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = cast(Dict[str, Any], serializer.validated_data)

        # ── Re-verify Check 1: No pending medicine allowance PRF ──────────
        has_pending = PRFRequest.objects.select_for_update().filter(
            employee=request.user,
            prf_type='medicine_allowance',
            status='pending',
        ).exists()
        if has_pending:
            return Response(
                {'detail': 'You already have a pending medicine allowance request that has not yet been processed.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Re-verify Check 2: Available Finance Allowance balance ────────
        ma_type = AllowanceType.objects.filter(name__iexact='medicine allowance').first()
        if ma_type is None:
            return Response(
                {
                    'detail': (
                        'Medicine Allowance is not configured in the Finance system. '
                        'Please contact the Finance department.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        allowance = (
            Allowance.objects
            .select_for_update()
            .filter(
                employee=request.user,
                allowance_type=ma_type,
                deposited_date__isnull=True,
            )
            .order_by('-created_at')
            .first()
        )
        if allowance is None or allowance.amount <= Decimal('0'):
            return Response(
                {
                    'detail': (
                        'You have no available medicine allowance balance. '
                        'Please contact the Finance department.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        requested_amount = data['amount']
        if requested_amount > allowance.amount:
            return Response(
                {'amount': [f'Entered amount exceeds your available balance of ₱{allowance.amount:,.2f}.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Generate control number & persist PRFRequest ──────────────────
        ma_control_number = MedicineAllowance.generate_control_number()

        prf = PRFRequest.objects.create(
            employee=request.user,
            prf_category=data['prf_category'],
            prf_type='medicine_allowance',
            purpose=data['purpose'],
            control_number=ma_control_number,
        )

        MedicineAllowance.objects.create(
            prf_request=prf,
            start_date=data['start_date'],
            end_date=data['end_date'],
            amount=requested_amount,
        )

        return Response(PRFRequestSerializer(prf).data, status=status.HTTP_201_CREATED)


# ── Admin views ────────────────────────────────────────────────────────────────────────────────

def _require_admin(request):
    """Returns None if admin; returns a 403 Response otherwise."""
    if not getattr(request.user, 'admin', False):
        return Response({'detail': 'Admin access required.'}, status=status.HTTP_403_FORBIDDEN)
    return None


class PRFAdminListView(APIView):
    """
    GET /api/prform/admin/requests/
    Returns a paginated list of ALL users' PRF requests for admin review.
    Supports search, status filter, prf_type filter, and sort.
    Requires admin=True on the requesting user.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin(request)
        if err:
            return err

        qs = PRFRequest.objects.select_related('employee').all()

        # ── Search ─────────────────────────────────────────────────────────
        search_q = request.GET.get('search', '').strip()
        if search_q:
            qs = qs.filter(
                Q(prf_control_number__icontains=search_q) |
                Q(prf_type__icontains=search_q) |
                Q(purpose__icontains=search_q) |
                Q(status__icontains=search_q) |
                Q(control_number__icontains=search_q) |
                Q(employee__idnumber__icontains=search_q) |
                Q(employee__firstname__icontains=search_q) |
                Q(employee__lastname__icontains=search_q)
            )

        # ── Filtering ──────────────────────────────────────────────────────
        prf_type_f   = request.GET.get('prf_type', '').strip()
        status_f     = request.GET.get('status',   '').strip()
        if prf_type_f:
            qs = qs.filter(prf_type=prf_type_f)
        if status_f:
            qs = qs.filter(status=status_f)

        # ── Sorting ────────────────────────────────────────────────────────
        sort_by  = request.GET.get('sort_by',  'created_at')
        sort_dir = request.GET.get('sort_dir', 'desc')
        allowed_sorts = {
            'prf_control_number', 'prf_type', 'purpose', 'status', 'created_at',
            'employee__idnumber',
        }
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        order_prefix = '' if sort_dir == 'asc' else '-'
        # Always surface pending requests first, then apply the requested sort within each group
        pending_first = Case(
            When(status='pending', then=Value(0)),
            default=Value(1),
            output_field=IntegerField(),
        )
        qs = qs.order_by(pending_first, f'{order_prefix}{sort_by}')
        if request.GET.get('export') == '1':
            total    = qs.count()
            qs_all   = qs[:500]
            serializer = PRFAdminSerializer(qs_all, many=True)
            return Response({'results': serializer.data, 'count': min(total, 500)})

        # ── Pagination ─────────────────────────────────────────────────────
        try:
            page = max(1, int(request.GET.get('page', 1)))
        except (ValueError, TypeError):
            page = 1
        page_size = 10
        total     = qs.count()
        start     = (page - 1) * page_size
        qs_page   = qs[start: start + page_size]

        serializer = PRFAdminSerializer(qs_page, many=True)
        return Response({
            'results':     serializer.data,
            'count':       total,
            'page':        page,
            'page_size':   page_size,
            'total_pages': max(1, -(-total // page_size)),
        })


class PRFAdminUpdateView(APIView):
    """
    PATCH /api/prform/admin/requests/<pk>/
    Allows admin to approve, disapprove, or cancel any PRF request.
    Remarks are required when disapproving.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk):
        err = _require_admin(request)
        if err:
            return err

        try:
            prf = PRFRequest.objects.select_related('employee').select_for_update().get(pk=pk)
        except PRFRequest.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = PRFAdminActionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        validated  = serializer.validated_data
        new_status = validated['status']
        new_remarks = validated.get('admin_remarks', '').strip()

        old_status = prf.status
        prf.status       = new_status
        prf.admin_remarks = new_remarks or None
        prf.processed_by  = request.user

        # Pre-validate Medicine Allowance approval before committing the status change.
        if new_status == 'approved' and old_status != 'approved' and prf.prf_type == 'medicine_allowance':
            try:
                ma = prf.medicine_allowance
                ma_type = AllowanceType.objects.filter(name__iexact='medicine allowance').first()
                if not ma_type:
                    return Response(
                        {'detail': 'Medicine Allowance type is not configured in the Finance system.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                allowance = (
                    Allowance.objects
                    .select_for_update()
                    .filter(
                        employee=prf.employee,
                        allowance_type=ma_type,
                        deposited_date__isnull=True,
                    )
                    .order_by('-created_at')
                    .first()
                )
                if allowance is None:
                    return Response(
                        {
                            'detail': (
                                'The Finance Allowance record for this employee is no longer available. '
                                'Approval cannot be processed.'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if allowance.amount < Decimal(str(ma.amount)):
                    return Response(
                        {
                            'detail': (
                                f'Insufficient balance. '
                                f'Available: ₱{allowance.amount:,.2f}, '
                                f'Requested: ₱{ma.amount:,.2f}.'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            except MedicineAllowance.DoesNotExist:
                pass

        prf.save()

        # If an emergency loan PRF transitions to approved, persist to Finance Loan table.
        if new_status == 'approved' and old_status != 'approved' and prf.prf_type == 'emergency_loan':
            try:
                el = prf.emergency_loan
                el_loan_type, _ = LoanType.objects.get_or_create(
                    name='Emergency Loan',
                    defaults={'color': '#2845D6', 'stackable': False},
                )
                amount_d = Decimal(str(el.amount))
                cutoff_d = Decimal(str(el.number_of_cutoff))
                active_loan = (
                    Loan.objects
                    .select_for_update()
                    .filter(
                        employee=prf.employee,
                        loan_type=el_loan_type,
                        current_balance__gt=Decimal('0'),
                    )
                    .first()
                )
                if active_loan is not None:
                    active_loan.current_balance  += amount_d
                    active_loan.principal_amount += amount_d
                    active_loan.monthly_deduction = active_loan.current_balance / cutoff_d
                    active_loan.save(
                        update_fields=['current_balance', 'principal_amount', 'monthly_deduction', 'updated_at']
                    )
                else:
                    Loan.objects.create(
                        employee=prf.employee,
                        loan_type=el_loan_type,
                        principal_amount=amount_d,
                        current_balance=amount_d,
                        monthly_deduction=amount_d / cutoff_d,
                    )
            except EmergencyLoan.DoesNotExist:
                pass

        # If a medicine allowance PRF transitions to approved, deduct from Finance Allowance.
        if new_status == 'approved' and old_status != 'approved' and prf.prf_type == 'medicine_allowance':
            try:
                ma = prf.medicine_allowance
                ma_type = AllowanceType.objects.filter(name__iexact='medicine allowance').first()
                if ma_type:
                    allowance = (
                        Allowance.objects
                        .select_for_update()
                        .filter(
                            employee=prf.employee,
                            allowance_type=ma_type,
                            deposited_date__isnull=True,
                        )
                        .order_by('-created_at')
                        .first()
                    )
                    if allowance:
                        allowance.amount -= Decimal(str(ma.amount))
                        if allowance.amount <= Decimal('0'):
                            allowance.delete()
                        else:
                            allowance.save(update_fields=['amount'])
            except MedicineAllowance.DoesNotExist:
                pass

        # Fire notification only when status actually changes to a notifiable state.
        if new_status != old_status and new_status in _NOTIF_MAP:
            _schedule_prf_notification(prf, new_status, new_remarks)

        return Response(PRFAdminSerializer(prf).data)


class PRFAdminExportView(APIView):
    """
    GET /api/prform/admin/export
    Generates and streams an xlsx file of PRF requests within a given date range.
    Optional filters: prf_category, prf_type, status.
    Requires admin=True.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin(request)
        if err:
            return err

        # ── Parse date range ──────────────────────────────────────────
        today = datetime.date.today()
        try:
            date_from = datetime.date.fromisoformat(
                request.GET.get('date_from', str(today))
            )
        except (ValueError, TypeError):
            date_from = today
        try:
            date_to = datetime.date.fromisoformat(
                request.GET.get('date_to', str(today))
            )
        except (ValueError, TypeError):
            date_to = today
        if date_to < date_from:
            date_from, date_to = date_to, date_from

        # ── Queryset ──────────────────────────────────────────────────
        qs = PRFRequest.objects.select_related('employee').filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        )
        cat_f  = request.GET.get('prf_category', '').strip()
        type_f = request.GET.get('prf_type',     '').strip()
        stat_f = request.GET.get('status',       '').strip()
        if cat_f:
            qs = qs.filter(prf_category=cat_f)
        if type_f:
            qs = qs.filter(prf_type=type_f)
        if stat_f:
            qs = qs.filter(status=stat_f)
        qs = qs.order_by('created_at')

        # ── Build workbook ──────────────────────────────────────────
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'PRF Requests'

        def _side():
            return Side(style='thin', color='FF000000')

        thin_border = Border(
            left=_side(), right=_side(), top=_side(), bottom=_side()
        )

        MONTHS = [
            '', 'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December',
        ]

        def _fmt_date(d):
            return f'{MONTHS[d.month]} {d.day}, {d.year}'

        # ── Row 1: company name ───────────────────────────────────────
        ws['A1'] = 'Ryonan Electric Philippines Corporation'
        ws['A1'].font = Font(bold=True, size=13)

        # ── Row 2: report title ──────────────────────────────────────
        ws['A2'] = f'PRF Request for {_fmt_date(date_from)} to {_fmt_date(date_to)}'
        ws['A2'].font = Font(size=11, italic=True)

        # ── Row 4: column headers ───────────────────────────────────
        HEADERS = [
            'Date Requested', 'PRF Number', 'ID Number', 'Employee Name',
            'PRF Category', 'PRF Type', 'Control Number',
            'Purpose of Request', 'Status', 'Remarks',
        ]
        HEADER_FILL = PatternFill(
            start_color='FF2845D6', end_color='FF2845D6', fill_type='solid'
        )
        STATUS_FILL = {
            'approved':    PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid'),
            'disapproved': PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid'),
            'pending':     PatternFill(start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid'),
            'cancelled':   PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid'),
        }
        STATUS_FONT_COLOR = {
            'approved':    'FF375623',
            'disapproved': 'FF9C0006',
            'pending':     'FF9C6500',
            'cancelled':   'FF595959',
        }

        for col, h in enumerate(HEADERS, start=1):
            cell           = ws.cell(row=4, column=col, value=h)
            cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill      = HEADER_FILL
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # ── Data rows ─────────────────────────────────────────────────
        for row_num, prf in enumerate(qs, start=5):
            emp      = prf.employee
            parts    = [x for x in [emp.lastname or '', emp.firstname or ''] if x]
            emp_name = ', '.join(parts)
            row_vals = [
                _fmt_date(prf.created_at.date()),
                prf.prf_control_number or '',
                emp.idnumber           or '',
                emp_name,
                prf.get_prf_category_display(),
                prf.get_prf_type_display(),
                prf.control_number     or '-',
                prf.purpose            or '',
                prf.get_status_display(),
                prf.admin_remarks      or '',
            ]
            for col, val in enumerate(row_vals, start=1):
                cell           = ws.cell(row=row_num, column=col, value=val)
                cell.border    = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=(col in (8, 10)))
                if col == 9:
                    s_fill  = STATUS_FILL.get(prf.status)
                    s_color = STATUS_FONT_COLOR.get(prf.status)
                    if s_fill:
                        cell.fill = s_fill
                    if s_color:
                        cell.font = Font(color=s_color)
                elif col == 10:
                    cell.font = Font(color='FFFF0000')

        # ── Column widths ─────────────────────────────────────────────
        col_widths = [18, 18, 15, 28, 26, 30, 20, 50, 18, 35]
        for idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # ── Stream xlsx response ───────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'prf-requests-{date_from}-to-{date_to}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PRFAdminChartView(APIView):
    """
    GET /api/prform/admin/chart/
    Returns aggregated PRF submission counts for the admin chart.

    Query params:
      view  — 'fiscal' | 'monthly' | 'weekly' | 'daily'  (default: fiscal)
      year  — fiscal year start (e.g. 2025 means May 2025 – Apr 2026, default: current FY)
      month — required for weekly/daily views (1–12)
      week  — ISO week number, required for daily view
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin(request)
        if err:
            return err

        view_type = request.GET.get('view', 'fiscal')
        now = timezone.now()

        # ── Fiscal year start ─────────────────────────────────────────────────
        current_fy_start_year = now.year if now.month >= 5 else now.year - 1
        try:
            fy_start_year = int(request.GET.get('year', current_fy_start_year))
        except (ValueError, TypeError):
            fy_start_year = current_fy_start_year

        fy_start = datetime.datetime(fy_start_year,     5, 1, tzinfo=datetime.timezone.utc)
        fy_end   = datetime.datetime(fy_start_year + 1, 5, 1, tzinfo=datetime.timezone.utc)

        base_qs    = PRFRequest.objects.filter(created_at__gte=fy_start, created_at__lt=fy_end)
        CATEGORIES = ['government', 'banking', 'hr_payroll']

        def build_lookup(rows, key_fn):
            lookup: dict = {}
            for r in rows:
                k = key_fn(r)
                if k not in lookup:
                    lookup[k] = {c: 0 for c in CATEGORIES}
                cat = r['prf_category']
                if cat in CATEGORIES:
                    lookup[k][cat] = r['count']
            return lookup

        def cat_entry(label, cats):
            return {
                'label': label,
                'government': cats.get('government', 0),
                'banking':    cats.get('banking',    0),
                'hr_payroll': cats.get('hr_payroll', 0),
            }

        # ── Fiscal: 12 monthly buckets (May → April), split by category ────────
        if view_type == 'fiscal':
            rows = (
                base_qs
                .annotate(period=TruncMonth('created_at'))
                .values('period', 'prf_category')
                .annotate(count=Count('id'))
                .order_by('period', 'prf_category')
            )
            lookup = build_lookup(rows, lambda r: r['period'].strftime('%Y-%m'))
            MONTH_LABELS = ['May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr']
            data = []
            for i, label in enumerate(MONTH_LABELS):
                month_num = 5 + i if i < 8 else i - 7
                year      = fy_start_year if month_num >= 5 else fy_start_year + 1
                key       = f'{year}-{month_num:02d}'
                data.append(cat_entry(label, lookup.get(key, {})))
            return Response({'view': 'fiscal', 'fy_start': fy_start_year, 'data': data})

        # ── Monthly: daily data for a selected month, split by category ────────
        if view_type == 'monthly':
            try:
                month = int(request.GET.get('month', now.month))
            except (ValueError, TypeError):
                month = now.month
            year    = fy_start_year if month >= 5 else fy_start_year + 1
            m_start = datetime.datetime(year, month, 1, tzinfo=datetime.timezone.utc)
            m_end   = (datetime.datetime(year + 1, 1, 1, tzinfo=datetime.timezone.utc)
                       if month == 12
                       else datetime.datetime(year, month + 1, 1, tzinfo=datetime.timezone.utc))
            days_in_month = calendar.monthrange(year, month)[1]
            rows = (
                PRFRequest.objects
                .filter(created_at__gte=m_start, created_at__lt=m_end)
                .annotate(period=TruncDate('created_at'))
                .values('period', 'prf_category')
                .annotate(count=Count('id'))
                .order_by('period', 'prf_category')
            )
            lookup = build_lookup(rows, lambda r: r['period'].day)
            data   = [cat_entry(str(day), lookup.get(day, {})) for day in range(1, days_in_month + 1)]
            return Response({'view': 'monthly', 'fy_start': fy_start_year, 'month': month, 'data': data})

        # ── Weekly: Mon–Sun of a specific week, split by category ─────────────
        if view_type == 'weekly':
            week_start_str = request.GET.get('week_start', '')
            try:
                ws         = datetime.datetime.strptime(week_start_str, '%Y-%m-%d')
                week_start = ws.replace(tzinfo=datetime.timezone.utc)
            except (ValueError, TypeError, AttributeError):
                today      = now.date()
                monday     = today - datetime.timedelta(days=today.weekday())
                week_start = datetime.datetime(monday.year, monday.month, monday.day,
                                               tzinfo=datetime.timezone.utc)
            week_end = week_start + datetime.timedelta(days=7)
            rows = (
                PRFRequest.objects
                .filter(created_at__gte=week_start, created_at__lt=week_end)
                .annotate(period=TruncDate('created_at'))
                .values('period', 'prf_category')
                .annotate(count=Count('id'))
                .order_by('period', 'prf_category')
            )
            lookup     = build_lookup(rows, lambda r: r['period'])
            DAY_LABELS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            data = []
            for i in range(7):
                day_date = (week_start + datetime.timedelta(days=i)).date()
                data.append(cat_entry(DAY_LABELS[i], lookup.get(day_date, {})))
            return Response({'view': 'weekly', 'fy_start': fy_start_year, 'data': data})

        return Response({'detail': 'Invalid view type.'}, status=status.HTTP_400_BAD_REQUEST)
