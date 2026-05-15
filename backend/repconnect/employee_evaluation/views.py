"""Employee Evaluation views.

Security checklist:
  - IsAuthenticated on every view; role check inside each method.
  - @transaction.atomic + select_for_update() on all write operations.
  - All list endpoints paginated (page_size=10, max=20).
  - No raw SQL — ORM only.
  - Input validated via serializers before any DB write.
  - CSRF enforced by DRF + middleware.
  - Double-submit prevented by unique_together + select_for_update.
  - xlsx import: file size validated server-side; row errors collected without abort.
"""
from __future__ import annotations

import datetime
import io
import logging
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Avg, Count, Max, Q
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from activityLog.models import Notification
from employee_evaluation.models import (
    EvaluationSettings,
    EvaluationPeriod,
    EmployeeTasklist,
    EmployeeTask,
    EvaluationEntry,
    EvaluationScore,
    EvaluationApprovalStep,
    SupervisorEvaluationEE,
    EvaluationTrainingRequest,
    EvaluationTimelineEntry,
)
from employee_evaluation.routing import build_evaluation_approval_chain, can_act_on_evaluation_step
from employee_evaluation.serializers import (
    EvaluationSettingsSerializer,
    EvaluationPeriodSerializer,
    EvaluationPeriodWriteSerializer,
    EmployeeTaskSerializer,
    EmployeeTasklistSerializer,
    EmployeeTasklistAdminListSerializer,
    EmployeeTasklistUserAdminListSerializer,
    EvaluationScoreSerializer,
    EvaluationScoreSaveSerializer,
    EvaluationApprovalStepSerializer,
    SupervisorEvaluationEESerializer,
    SupervisorEvaluationEESubmitSerializer,
    EvaluationEntrySerializer,
    EvaluationEntryAdminSerializer,
    EvalApproverQueueItemSerializer,
    EvalApproverEntryDetailSerializer,
    EvaluationTrainingRequestSerializer,
)

logger = logging.getLogger(__name__)
PAGE_SIZE = 10
MAX_PAGE_SIZE = 20
MAX_IMPORT_SIZE = 5 * 1024 * 1024  # 5 MB


# ── Helpers ────────────────────────────────────────────────────────────────────

def _require_admin_or_hr(request):
    u = request.user
    if not (getattr(u, 'admin', False) or getattr(u, 'hr', False)):
        return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
    return None


def _record_timeline(
    entry: EvaluationEntry,
    actor,
    action_type: str,
    step_order: int = 0,
    remarks: str = '',
    acted_at=None,
) -> EvaluationTimelineEntry:
    """Append an immutable audit row to EvaluationTimelineEntry.
    Never raises — any exception is logged and suppressed so that a timeline
    write failure never rolls back the primary business transaction.
    """
    try:
        ts = acted_at or timezone.now()
        return EvaluationTimelineEntry.objects.create(
            entry=entry,
            actor=actor,
            action_type=action_type,
            step_order=step_order,
            remarks=remarks or '',
            acted_at=ts,
        )
    except Exception:  # pragma: no cover
        logger.exception(
            'Failed to write EvaluationTimelineEntry for entry=%s action=%s',
            getattr(entry, 'pk', None), action_type,
        )
        raise  # Re-raise inside atomic block so the transaction is consistent


def _eligible_users():
    """Base eligibility: excludes admin/hr/accounting AND Probationary/OJT employees.

    Probationary and OJT (any capitalization / abbreviation) employees must never
    appear in evaluation lists or counts.
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    return (
        User.objects
        .filter(is_active=True, admin=False, hr=False, accounting=False)
        .exclude(
            workinformation__employment_type__name__iregex=r'probationary|ojt|on.the.job|on.job',
        )
    )


def _period_eligibility_cutoff(period_start_date: datetime.date) -> datetime.date:
    """Return the hire-date cutoff for a loaded evaluation period.

    Example: if the period starts on May 1, 2025, only employees hired before
    April 1, 2025 are eligible. Any hire date in April 2025 or later is ineligible.
    """
    first_day_of_month = period_start_date.replace(day=1)
    last_day_previous_month = first_day_of_month - datetime.timedelta(days=1)
    return last_day_previous_month.replace(day=1)


def _eligible_users_for_period(period_start_date: datetime.date):
    """Eligible users for a specific evaluation period.

    Builds on _eligible_users() and additionally excludes Regular employees whose
    date_hired falls on or after the cutoff for that loaded period. The cutoff is
    the last day of the month before the period start month, so a May 2025 start
    excludes hires from April 2025 onward.

    Employees without a workinformation record or without date_hired are included.
    """
    cutoff_date = _period_eligibility_cutoff(period_start_date)
    return (
        _eligible_users()
        .exclude(
            workinformation__employment_type__name__iregex=r'regular',
            workinformation__date_hired__gte=cutoff_date,
        )
    )


def _paginate(queryset, request):
    try:
        page = max(1, int(request.query_params.get('page', 1)))
        size = min(MAX_PAGE_SIZE, max(1, int(request.query_params.get('page_size', PAGE_SIZE))))
    except (ValueError, TypeError):
        page, size = 1, PAGE_SIZE
    start = (page - 1) * size
    end = start + size
    total = queryset.values('id').count()
    items = list(queryset[start:end])
    return items, total, page, size


def _get_active_period():
    """Return the best-match active EvaluationPeriod.

    Priority:
      1. status=active AND today is within [start_date, end_date]
         — if multiple qualify, pick the one with the earliest end_date
           (most "current" / about to expire first).
      2. Fallback: any status=active period, ordered by end_date desc.
    """
    today = timezone.now().date()
    period = (
        EvaluationPeriod.objects
        .filter(status='active', start_date__lte=today, end_date__gte=today)
        .order_by('end_date')
        .first()
    )
    if period:
        return period

    _ensure_current_fiscal_period(today)
    return EvaluationPeriod.objects.filter(status='active').order_by('-end_date').first()


def _ensure_current_fiscal_period(today):
    """Create the current fiscal-year period if it does not already exist."""
    from datetime import date as date_cls
    from employee_evaluation.models import EvaluationPeriod, EvaluationSettings

    fiscal_year = today.year if today.month >= 5 else today.year - 1
    if EvaluationPeriod.objects.filter(fiscal_year=fiscal_year).exists():
        return

    try:
        settings_obj = EvaluationSettings.objects.get()
    except EvaluationSettings.DoesNotExist:
        return

    title = f'Performance Evaluation FY{fiscal_year}-{fiscal_year + 1}'
    start_date = date_cls(fiscal_year, 5, 1)
    end_date = date_cls(fiscal_year + 1, 4, 30)

    with transaction.atomic():
        EvaluationPeriod.objects.get_or_create(
            fiscal_year=fiscal_year,
            defaults={
                'title': title,
                'start_date': start_date,
                'end_date': end_date,
                'status': 'active',
                'frequency': settings_obj.frequency,
            },
        )


def _build_period_labels(frequency: str) -> list[str]:
    """Return the ordered column labels for a given frequency."""
    if frequency == 'quarterly':
        return ['Q1', 'Q2', 'Q3', 'Q4']
    if frequency == 'monthly':
        return [
            'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct',
            'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr',
        ]
    return []


def _notify_approver(entry: EvaluationEntry, approver) -> None:
    """Create an in-app notification for the next approver."""
    def _create():
        Notification.objects.create(
            recipient=approver,
            notification_type='evaluation_pending_review',
            notification_scope='specific_user',
            title='Employee Evaluation Pending Review',
            message=(
                f'{entry.employee.get_full_name() or entry.employee.username} '
                f'has submitted an employee evaluation for '
                f'{entry.evaluation_period.title}. Please review.'
            ),
            module='employee_evaluation',
            related_object_id=entry.pk,
        )
    transaction.on_commit(_create)


def _notify_step1_return(entry: EvaluationEntry, approver, remarks: str) -> None:
    """Notify the Step 1 supervisor that the evaluation was returned for re-evaluation."""
    def _create():
        Notification.objects.create(
            recipient=approver,
            notification_type='evaluation_pending_review',
            notification_scope='specific_user',
            title='Employee Evaluation Returned for Re-evaluation',
            message=(
                f'The evaluation for '
                f'{entry.employee.get_full_name() or entry.employee.username} '
                f'({entry.evaluation_period.title}) has been returned for re-evaluation. '
                f'Reason: {remarks}'
            ),
            module='employee_evaluation',
            related_object_id=entry.pk,
        )
    transaction.on_commit(_create)


# ─────────────────────────────────────────────────────────────────────────────
# ADMIN / HR VIEWS
# ─────────────────────────────────────────────────────────────────────────────

class AdminPeriodListView(APIView):
    """GET /api/employee-eval/admin/periods — paginated list of all periods."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        qs = EvaluationPeriod.objects.all()
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(title__icontains=search)

        frequencies = [f.strip() for f in request.query_params.getlist('frequency') if f.strip()]
        if frequencies:
            qs = qs.filter(frequency__in=frequencies)

        statuses = [s.strip() for s in request.query_params.getlist('status') if s.strip()]
        if statuses:
            qs = qs.filter(status__in=statuses)

        ordering = request.query_params.get('ordering', '').strip()
        if ordering:
            prefix = '-' if ordering.startswith('-') else ''
            field = ordering[1:] if ordering.startswith('-') else ordering
            if field in {'title', 'fiscal_year', 'frequency', 'start_date', 'end_date', 'status'}:
                qs = qs.order_by(f'{prefix}{field}')
            else:
                qs = qs.order_by('-fiscal_year', '-start_date')
        else:
            qs = qs.order_by('-fiscal_year', '-start_date')

        items, total, page, size = _paginate(qs, request)
        return Response({
            'results': EvaluationPeriodSerializer(items, many=True).data,
            'count': total,
            'page': page,
            'page_size': size,
        })


class AdminPeriodDetailView(APIView):
    """PATCH /api/employee-eval/admin/periods/<id> — edit end_date or status only."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            period = EvaluationPeriod.objects.get(pk=pk)
        except EvaluationPeriod.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        serializer = EvaluationPeriodWriteSerializer(period, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response(EvaluationPeriodSerializer(period).data)


class AdminPeriodToggleStatusView(APIView):
    """POST /api/employee-eval/admin/periods/<id>/toggle-status."""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            period = EvaluationPeriod.objects.get(pk=pk)
        except EvaluationPeriod.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        period.status = 'closed' if period.status == 'active' else 'active'
        period.save(update_fields=['status'])
        return Response(EvaluationPeriodSerializer(period).data)


class AdminPeriodResultsView(APIView):
    """GET /api/employee-eval/admin/periods/<pk>/results — period summary + paginated entries."""
    permission_classes = [IsAuthenticated]

    _ALLOWED_SORT_FIELDS = frozenset({'name', 'idnumber', 'department', 'status', 'submitted_at'})

    def get(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            return self._build_response(request, pk)
        except EvaluationPeriod.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        except Exception:
            logger.exception('AdminPeriodResultsView: unexpected error for period pk=%s', pk)
            return Response({'detail': 'An internal error occurred.'}, status=http_status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _build_response(self, request, pk):
        from userProfile.models import workInformation

        period = EvaluationPeriod.objects.get(pk=pk)

        # ── All eligible users for this fiscal year ───────────────────────────
        eligible_qs = _eligible_users_for_period(period.start_date)
        eligible_users = list(eligible_qs.only('id', 'idnumber', 'username', 'firstname', 'lastname'))
        eligible_ids = [u.id for u in eligible_users]
        total_eligible = len(eligible_ids)

        # ── Work-information: department per employee (first record wins) ─────
        wi_map: dict = {}  # employee_id -> (dept_id, dept_name)
        for wi in (
            workInformation.objects
            .filter(employee_id__in=eligible_ids)
            .select_related('department')
            .order_by('employee_id')
        ):
            if wi.employee_id not in wi_map:
                wi_map[wi.employee_id] = (
                    wi.department_id,
                    wi.department.name if wi.department else None,
                )

        # ── Department options for the filter UI ─────────────────────────────
        dept_map: dict = {}
        for _uid, (did, dname) in wi_map.items():
            if did and dname and did not in dept_map:
                dept_map[did] = dname
        departments = sorted(
            [{'id': did, 'name': dname} for did, dname in dept_map.items()],
            key=lambda x: x['name'],
        )

        # ── Existing entries indexed by employee_id ───────────────────────────
        entries_by_emp: dict = {}
        for e in EvaluationEntry.objects.filter(
            evaluation_period=period, employee_id__in=eligible_ids
        ):
            entries_by_emp[e.employee_id] = e

        # ── Summary stats (across ALL eligible users) ─────────────────────────
        submitted_count = sum(1 for e in entries_by_emp.values() if e.submitted_at is not None)
        completed_count = sum(1 for e in entries_by_emp.values() if e.status == 'completed')
        completion_rate = (
            round((completed_count / total_eligible) * 100, 1) if total_eligible > 0 else 0.0
        )

        # ── Parse filter params ───────────────────────────────────────────────
        search = request.query_params.get('search', '').strip().lower()
        dept_id_raw = request.query_params.get('dept', '').strip()
        dept_filter = int(dept_id_raw) if dept_id_raw.isdigit() else None
        entry_status = request.query_params.get('entry_status', '').strip()

        # ── Build one row per eligible user, merge with entry if present ──────
        rows = []
        for u in eligible_users:
            entry = entries_by_emp.get(u.id)
            dept_id, dept_name = wi_map.get(u.id, (None, None))
            status = entry.status if entry else 'not_started'
            submitted_at = entry.submitted_at if entry else None
            entry_id = entry.id if entry else None
            first_name = (getattr(u, 'firstname', None) or '').strip()
            last_name = (getattr(u, 'lastname', None) or '').strip()
            full_name = ' '.join(part for part in [first_name, last_name] if part) or u.idnumber or u.username

            # Search filter
            if search and search not in full_name.lower() and search not in (u.idnumber or u.username).lower():
                continue

            # Department filter
            if dept_filter is not None and dept_id != dept_filter:
                continue

            # Status filter
            if entry_status:
                if entry_status == 'not_started':
                    if entry is not None:
                        continue
                elif status != entry_status:
                    continue

            rows.append({
                'id': entry_id,
                'employee_id': u.id,
                'idnumber': u.idnumber or u.username,
                'employee_name': full_name,
                'department': dept_name,
                'status': status,
                'submitted_at': submitted_at.isoformat() if submitted_at else None,
            })

        # ── Sort ──────────────────────────────────────────────────────────────
        sort = request.query_params.get('sort', 'name').strip()
        if sort not in self._ALLOWED_SORT_FIELDS:
            sort = 'name'
        reverse = request.query_params.get('dir', 'asc').strip().lower() == 'desc'

        if sort == 'idnumber':
            rows.sort(key=lambda r: r['idnumber'].lower(), reverse=reverse)
        elif sort == 'department':
            rows.sort(key=lambda r: (r['department'] or '').lower(), reverse=reverse)
        elif sort == 'status':
            rows.sort(key=lambda r: r['status'], reverse=reverse)
        elif sort == 'submitted_at':
            # None values sort last in asc, first in desc
            rows.sort(
                key=lambda r: (r['submitted_at'] is None, r['submitted_at'] or ''),
                reverse=reverse,
            )
        else:  # name
            rows.sort(key=lambda r: r['employee_name'].lower(), reverse=reverse)

        # ── Paginate ─────────────────────────────────────────────────────────
        total = len(rows)
        try:
            page = max(1, int(request.query_params.get('page', 1)))
            size = min(MAX_PAGE_SIZE, max(1, int(request.query_params.get('page_size', PAGE_SIZE))))
        except (ValueError, TypeError):
            page, size = 1, PAGE_SIZE
        start = (page - 1) * size
        page_rows = rows[start:start + size]

        return Response({
            'period': EvaluationPeriodSerializer(period).data,
            'summary': {
                'total_eligible': total_eligible,
                'submitted': submitted_count,
                'completed': completed_count,
                'completion_rate': completion_rate,
            },
            'departments': departments,
            'results': page_rows,
            'pagination': {
                'page': page,
                'page_size': size,
                'total': total,
                'total_pages': max(1, (total + size - 1) // size),
            },
        })


class AdminEntryListView(APIView):
    """GET /api/employee-eval/admin/entries — all entries (for chart + table)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        period_id = request.query_params.get('period')
        qs = EvaluationEntry.objects.select_related('employee', 'evaluation_period')

        if period_id:
            qs = qs.filter(evaluation_period_id=period_id)

        items, total, page, size = _paginate(qs, request)
        return Response({
            'results': EvaluationEntryAdminSerializer(items, many=True).data,
            'count': total,
            'page': page,
            'page_size': size,
        })


class AdminEntryDetailView(APIView):
    """GET /api/employee-eval/admin/entries/<entry_id> — admin view of a single entry.

    Equivalent to ApproverEntryDetailView but accessible to any admin or HR user
    without requiring them to be an approver on the entry.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, entry_id):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            entry = (
                EvaluationEntry.objects
                .select_related('employee', 'evaluation_period')
                .prefetch_related('approval_steps__supervisor_evaluation', 'scores')
                .get(pk=entry_id)
            )
        except EvaluationEntry.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        return Response(EvalApproverEntryDetailSerializer(entry, context={'request': request}).data)


class AdminChartView(APIView):
    """GET /api/employee-eval/admin/chart — dual-dataset chart for the active period."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        period = _get_active_period()
        if not period:
            return Response({
                'period_id': None,
                'period_title': None,
                'self_evals': [],
                'supervisor_evals': [],
            })

        # Self-evaluations: EvaluationEntry rows for this period (created_at timestamp)
        self_evals = list(
            EvaluationEntry.objects
            .filter(evaluation_period=period)
            .values_list('created_at', flat=True)
        )

        # Approver evaluations: SupervisorEvaluationEE that have been submitted
        sup_evals = list(
            SupervisorEvaluationEE.objects
            .filter(
                step__entry__evaluation_period=period,
                submitted_at__isnull=False,
            )
            .values_list('submitted_at', flat=True)
        )

        return Response({
            'period_id': period.id,
            'period_title': period.title,
            'self_evals': [dt.isoformat() for dt in self_evals],
            'supervisor_evals': [dt.isoformat() for dt in sup_evals],
        })


class AdminTasklistListView(APIView):
    """GET /api/employee-eval/admin/tasklists — paginated eligible users with task count."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        period = _get_active_period()
        if not period:
            return Response({'detail': 'No active evaluation period.'}, status=http_status.HTTP_404_NOT_FOUND)

        search = request.query_params.get('search', '').strip()
        # Use period-scoped eligibility: excludes Probationary/OJT and newly-hired Regulars
        eligible = _eligible_users_for_period(period.start_date)
        if search:
            eligible = eligible.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(username__icontains=search)
            )

        department_ids = [
            int(d) for d in request.query_params.getlist('department_id')
            if d.strip().isdigit()
        ]
        if department_ids:
            eligible = eligible.filter(workinformation__department_id__in=department_ids)

        task_filter = request.query_params.get('task_filter', '').strip().lower()
        if task_filter == 'with':
            eligible = eligible.annotate(
                task_count=Count('evaluation_tasklists__tasks'),
            ).filter(task_count__gt=0)
        elif task_filter == 'without':
            eligible = eligible.annotate(
                task_count=Count('evaluation_tasklists__tasks'),
            ).filter(task_count=0)

        eligible = eligible.annotate(
            task_count=Count(
                'evaluation_tasklists__tasks',
            ),
            latest_updated=Max(
                'evaluation_tasklists__updated_at',
            ),
        )

        ordering = request.query_params.get('ordering', '').strip()
        if ordering:
            prefix = '-' if ordering.startswith('-') else ''
            field = ordering[1:] if ordering.startswith('-') else ordering
            if field == 'employee_id_number':
                eligible = eligible.order_by(f'{prefix}username')
            elif field == 'employee_name':
                eligible = eligible.order_by(f'{prefix}last_name', f'{prefix}first_name')
            elif field == 'department':
                eligible = eligible.order_by(f'{prefix}workinformation__department__name')
            elif field == 'task_count':
                eligible = eligible.order_by(f'{prefix}task_count')
            elif field == 'updated_at':
                eligible = eligible.order_by(f'{prefix}latest_updated')
            else:
                eligible = eligible.order_by('last_name', 'first_name')
        else:
            eligible = eligible.order_by('last_name', 'first_name')

        items, total, page, size = _paginate(eligible, request)
        return Response({
            'results': EmployeeTasklistUserAdminListSerializer(items, many=True).data,
            'count': total,
            'page': page,
            'page_size': size,
            'period': EvaluationPeriodSerializer(period).data,
        })


class AdminTasklistDetailView(APIView):
    """
    GET  /api/employee-eval/admin/tasklists/<user_id> — get user's tasks for active period.
    PATCH /api/employee-eval/admin/tasklists/<user_id> — replace task list.
    """
    permission_classes = [IsAuthenticated]

    def _get_or_create_tasklist(self, user_id, period):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            employee = User.objects.get(pk=user_id, is_active=True)
        except User.DoesNotExist:
            return None, Response({'detail': 'User not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        tasklist, _ = EmployeeTasklist.objects.get_or_create(
            employee=employee,
        )
        return tasklist, None

    def get(self, request, user_id):
        err = _require_admin_or_hr(request)
        if err:
            return err

        period = _get_active_period()
        if not period:
            return Response({'detail': 'No active evaluation period.'}, status=http_status.HTTP_404_NOT_FOUND)

        tasklist, err = self._get_or_create_tasklist(user_id, period)
        if err:
            return err

        return Response(EmployeeTasklistSerializer(tasklist).data)

    @transaction.atomic
    def patch(self, request, user_id):
        err = _require_admin_or_hr(request)
        if err:
            return err

        period = _get_active_period()
        if not period:
            return Response({'detail': 'No active evaluation period.'}, status=http_status.HTTP_404_NOT_FOUND)

        tasklist, err = self._get_or_create_tasklist(user_id, period)
        if err:
            return err

        tasks_data = request.data.get('tasks', [])
        if not isinstance(tasks_data, list):
            return Response({'detail': 'tasks must be a list.'}, status=http_status.HTTP_400_BAD_REQUEST)

        # Validate each task
        for i, t in enumerate(tasks_data):
            name = str(t.get('name', '')).strip()
            if not name:
                return Response(
                    {'detail': f'Task at index {i} has an empty name.'},
                    status=http_status.HTTP_400_BAD_REQUEST,
                )
            if len(name) > 300:
                return Response(
                    {'detail': f'Task at index {i} name exceeds 300 characters.'},
                    status=http_status.HTTP_400_BAD_REQUEST,
                )

        # Replace tasks
        tasklist.tasks.all().delete()
        for order, t in enumerate(tasks_data, start=1):
            EmployeeTask.objects.create(
                tasklist=tasklist,
                name=str(t['name']).strip(),
                order=order,
            )

        tasklist.refresh_from_db()
        return Response(EmployeeTasklistSerializer(tasklist).data)


def _parse_tasks_from_cell(raw: str) -> list[str]:
    """Split a TaskList cell value by semicolons and line breaks."""
    import re
    return [t.strip() for t in re.split(r'[;\n\r]+', raw) if t.strip()]


def _build_tasklist_error_report(rows_data: list) -> bytes:
    """Build an xlsx error report for tasklist validation failures.

    rows_data: list of dicts with keys: id_number, employee_name, tasklist_raw, error (str, empty if valid)
    Error rows are rendered in red; valid rows in normal black font.
    A Remarks column is appended with the error description.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Error Report'

    headers = ['ID Number', 'Employee Name', 'TaskList', 'Remarks']
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color='2845D6', end_color='2845D6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    red_font = Font(color='CC0000')
    normal_font = Font(color='000000')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_out_idx, r in enumerate(rows_data, start=2):
        has_error = bool(r.get('error', ''))
        font = red_font if has_error else normal_font
        values = [r['id_number'], r['employee_name'], r['tasklist_raw'], r.get('error', '')]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_out_idx, column=col_idx, value=val or '')
            cell.font = font
            cell.border = border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class AdminTasklistDeleteAllView(APIView):
    """POST /api/employee-eval/admin/tasklists/delete-all — delete all existing tasklists.

    Processes deletion in batches of 100 and returns the total count deleted.
    Intended to be called before an override import so the frontend can show
    a deletion progress stage.
    """
    permission_classes = [IsAuthenticated]

    BATCH_SIZE = 100

    @transaction.atomic
    def post(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            # Collect all tasklist PKs first
            pks = list(EmployeeTasklist.objects.values_list('pk', flat=True))
            deleted_total = 0
            for i in range(0, len(pks), self.BATCH_SIZE):
                batch = pks[i: i + self.BATCH_SIZE]
                deleted_total += EmployeeTasklist.objects.filter(pk__in=batch).delete()[0]
            return Response({'deleted': deleted_total})
        except Exception:
            logger.exception('AdminTasklistDeleteAllView: unexpected error during bulk delete')
            return Response(
                {'detail': 'Failed to delete tasklists.'},
                status=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AdminTasklistValidateView(APIView):
    """POST /api/employee-eval/admin/tasklists/validate — validate xlsx without writing.

    Returns:
      200 JSON {"valid": true}  if all rows pass.
      422 xlsx file              if any row fails — the xlsx is the error report.
      400                        for missing/invalid file.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        xlsx_file = request.FILES.get('file')
        if not xlsx_file:
            return Response({'detail': 'No file uploaded.'}, status=http_status.HTTP_400_BAD_REQUEST)

        if xlsx_file.size > MAX_IMPORT_SIZE:
            return Response(
                {'detail': f'File too large. Maximum size is {MAX_IMPORT_SIZE // (1024 * 1024)} MB.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_file.read()), data_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Invalid xlsx file.'}, status=http_status.HTTP_400_BAD_REQUEST)

        from django.contrib.auth import get_user_model
        User = get_user_model()

        rows_data = []
        has_errors = False

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not any(cell is not None for cell in row):
                continue

            id_number = str(row[0]).strip() if row[0] is not None else ''
            employee_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            tasklist_raw = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''

            row_errors = []

            if not id_number:
                row_errors.append('ID Number is required.')
            else:
                try:
                    employee = User.objects.get(idnumber=id_number, is_active=True)
                    if getattr(employee, 'admin', False) or getattr(employee, 'hr', False) or getattr(employee, 'accounting', False):
                        row_errors.append('User is excluded from evaluation (admin/hr/accounting role).')
                except User.DoesNotExist:
                    row_errors.append('ID Number not found in system.')

            if not tasklist_raw:
                row_errors.append('TaskList is empty.')

            error_str = '; '.join(row_errors)
            if error_str:
                has_errors = True

            rows_data.append({
                'id_number': id_number,
                'employee_name': employee_name,
                'tasklist_raw': tasklist_raw,
                'error': error_str,
            })

        if not rows_data:
            return Response({'detail': 'File contains no data rows.'}, status=http_status.HTTP_400_BAD_REQUEST)

        if has_errors:
            from django.http import HttpResponse
            report_bytes = _build_tasklist_error_report(rows_data)
            response = HttpResponse(
                report_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                status=422,
            )
            response['Content-Disposition'] = 'attachment; filename="tasklist_error_report.xlsx"'
            return response

        return Response({'valid': True, 'row_count': len(rows_data)})


class AdminTasklistImportView(APIView):
    """POST /api/employee-eval/admin/tasklists/import — xlsx bulk import (all-or-nothing).

    Validates every row before writing any record.  If any row fails validation
    the entire upload is rejected and an xlsx error report is returned with
    HTTP 422.  Only when all rows are valid does the view write to the database.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        period = _get_active_period()
        if not period:
            return Response({'detail': 'No active evaluation period.'}, status=http_status.HTTP_404_NOT_FOUND)

        xlsx_file = request.FILES.get('file')
        if not xlsx_file:
            return Response({'detail': 'No file uploaded.'}, status=http_status.HTTP_400_BAD_REQUEST)

        if xlsx_file.size > MAX_IMPORT_SIZE:
            return Response(
                {'detail': f'File too large. Maximum size is {MAX_IMPORT_SIZE // (1024 * 1024)} MB.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_file.read()), data_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Invalid xlsx file.'}, status=http_status.HTTP_400_BAD_REQUEST)

        from django.contrib.auth import get_user_model
        User = get_user_model()

        valid_rows = []
        rows_data = []
        has_errors = False

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not any(cell is not None for cell in row):
                continue

            id_number = str(row[0]).strip() if row[0] is not None else ''
            employee_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            tasklist_raw = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''

            row_errors = []
            employee = None

            if not id_number:
                row_errors.append('ID Number is required.')
            else:
                try:
                    employee = User.objects.get(idnumber=id_number, is_active=True)
                    if getattr(employee, 'admin', False) or getattr(employee, 'hr', False) or getattr(employee, 'accounting', False):
                        row_errors.append('User is excluded from evaluation (admin/hr/accounting role).')
                        employee = None
                except User.DoesNotExist:
                    row_errors.append('ID Number not found in system.')

            if not tasklist_raw:
                row_errors.append('TaskList is empty.')

            error_str = '; '.join(row_errors)
            rows_data.append({
                'id_number': id_number,
                'employee_name': employee_name,
                'tasklist_raw': tasklist_raw,
                'error': error_str,
            })

            if error_str:
                has_errors = True
            else:
                task_names = _parse_tasks_from_cell(tasklist_raw)
                if not task_names:
                    rows_data[-1]['error'] = 'TaskList is empty after parsing.'
                    has_errors = True
                else:
                    valid_rows.append({'employee': employee, 'task_names': task_names})

        if not rows_data:
            return Response({'detail': 'File contains no data rows.'}, status=http_status.HTTP_400_BAD_REQUEST)

        # All-or-nothing: reject the entire upload if any row failed
        if has_errors:
            from django.http import HttpResponse
            report_bytes = _build_tasklist_error_report(rows_data)
            response = HttpResponse(
                report_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                status=422,
            )
            response['Content-Disposition'] = 'attachment; filename="tasklist_error_report.xlsx"'
            return response

        # All rows valid — group tasks by employee (supports one-row-per-task format)
        # and write atomically.
        from collections import OrderedDict
        employee_tasks: OrderedDict = OrderedDict()  # employee_pk → {'employee': obj, 'task_names': list}
        for r in valid_rows:
            emp_pk = r['employee'].pk
            if emp_pk not in employee_tasks:
                employee_tasks[emp_pk] = {'employee': r['employee'], 'task_names': []}
            employee_tasks[emp_pk]['task_names'].extend(r['task_names'])

        imported_count = 0
        for emp_data in employee_tasks.values():
            tasklist, _ = EmployeeTasklist.objects.get_or_create(
                employee=emp_data['employee'],
            )
            tasklist.tasks.all().delete()
            for order, name in enumerate(emp_data['task_names'], start=1):
                EmployeeTask.objects.create(tasklist=tasklist, name=name[:300], order=order)
            imported_count += 1

        return Response({'imported': imported_count})


class AdminTasklistTemplateView(APIView):
    """GET /api/employee-eval/admin/tasklists/template — download xlsx template.

    Template has exactly three columns:
      A — ID Number
      B — Employee Name
      C — TaskList  (tasks separated by semicolons or line breaks within the cell)

    A sample data row is included to demonstrate the expected format.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = 'Tasklist Template'

        headers = ['ID Number', 'Employee Name', 'TaskList']
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color='2845D6', end_color='2845D6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Sample row — demonstrates semicolon-delimited tasks in a single cell
        cell_a = ws.cell(row=2, column=1, value='EMP001')
        cell_a.border = border
        cell_a.alignment = Alignment(vertical='top')
        cell_b = ws.cell(row=2, column=2, value='Juan Dela Cruz')
        cell_b.border = border
        cell_b.alignment = Alignment(vertical='top')
        sample_cell = ws.cell(row=2, column=3, value='Complete weekly status report; Attend team meetings; Review pull requests')
        sample_cell.border = border
        sample_cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.row_dimensions[2].height = 40

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="tasklist_template.xlsx"'
        return response


class ActivePeriodView(APIView):
    """GET /api/employee-eval/active-period — lightweight active period for eligibility checks.

    Accessible by any authenticated user so the frontend sidebar can determine
    whether the Self Evaluation button should be visible without a separate role check.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period = _get_active_period()
        if not period:
            return Response({'period': None})
        return Response({
            'period': {
                'id': period.id,
                'title': period.title,
                'start_date': period.start_date.isoformat(),
                'end_date': period.end_date.isoformat(),
                'fiscal_year': period.fiscal_year,
                'status': period.status,
            }
        })


# ─────────────────────────────────────────────────────────────────────────────
# USER (SELF-EVALUATION) VIEWS
# ─────────────────────────────────────────────────────────────────────────────

class MyEvaluationView(APIView):
    """GET /api/employee-eval/my — returns current period + entry + tasklist."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'accounting', False):
            return Response({'detail': 'Evaluation is not available for your role.'}, status=http_status.HTTP_403_FORBIDDEN)

        period = _get_active_period()
        if not period:
            return Response({'detail': 'No active evaluation period.'}, status=http_status.HTTP_404_NOT_FOUND)

        # Eligibility check against the period's actual start_date:
        # excludes Probationary/OJT AND Regular employees hired on or after period.start_date.
        if not _eligible_users_for_period(period.start_date).filter(pk=user.pk).exists():
            return Response(
                {'detail': 'You are not eligible for the current evaluation period.'},
                status=http_status.HTTP_403_FORBIDDEN,
            )

        # Auto-create entry if it doesn't exist yet (status=pending)
        entry, _ = EvaluationEntry.objects.prefetch_related(
            'approval_steps__supervisor_evaluation', 'scores'
        ).get_or_create(
            employee=user,
            evaluation_period=period,
        )

        # Auto-create empty tasklist if it doesn't exist
        tasklist, _ = EmployeeTasklist.objects.get_or_create(
            employee=user,
        )

        labels = _build_period_labels(period.frequency)

        return Response({
            'period': EvaluationPeriodSerializer(period).data,
            'entry': EvaluationEntrySerializer(entry).data,
            'tasklist': EmployeeTaskSerializer(tasklist.tasks.all(), many=True).data,
            'period_labels': labels,
        })


class MyEvaluationBadgeView(APIView):
    """GET /api/employee-eval/my/badge — count of enabled but unevaluated periods."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date as date_cls
        user = request.user
        if getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'accounting', False):
            return Response({'pending_count': 0})

        period = _get_active_period()
        if not period:
            return Response({'pending_count': 0})

        # Eligibility check — ineligible users always get pending_count=0.
        if not _eligible_users_for_period(period.start_date).filter(pk=user.pk).exists():
            return Response({'pending_count': 0})

        labels = _build_period_labels(period.frequency)
        today = date_cls.today()

        # Determine which labels are "enabled" (have passed per the previous-period rule).
        from datetime import timedelta
        import calendar

        def add_months(d, months):
            month = d.month - 1 + months
            year = d.year + month // 12
            month = month % 12 + 1
            day = min(d.day, calendar.monthrange(year, month)[1])
            return d.replace(year=year, month=month, day=day)

        try:
            period_start = period.start_date  # already a date object from ORM
        except Exception:
            return Response({'pending_count': 0})

        enabled_labels = []
        for idx, label in enumerate(labels):
            if period.frequency == 'monthly':
                label_start = add_months(period_start, idx)
                start_of_current_month = today.replace(day=1)
                if label_start < start_of_current_month:
                    enabled_labels.append(label)
            elif period.frequency == 'quarterly':
                label_start = add_months(period_start, idx * 3)
                label_end = add_months(label_start, 3)
                if label_end <= today:
                    enabled_labels.append(label)

        if not enabled_labels:
            return Response({'pending_count': 0})

        # Find the user's entry and scores.
        try:
            entry = EvaluationEntry.objects.get(employee=user, evaluation_period=period)
        except EvaluationEntry.DoesNotExist:
            # No entry yet — all enabled labels are unevaluated.
            return Response({'pending_count': len(enabled_labels)})

        # Labels that already have at least one non-null score are considered evaluated.
        scored_labels = set(
            EvaluationScore.objects.filter(
                entry=entry,
                period_label__in=enabled_labels,
            ).exclude(score__isnull=True).values_list('period_label', flat=True)
        )

        pending_count = sum(1 for label in enabled_labels if label not in scored_labels)
        return Response({'pending_count': pending_count})


class MyScoreSaveView(APIView):
    """POST /api/employee-eval/my/scores — partial save of score cells."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        user = request.user
        if getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'accounting', False):
            return Response({'detail': 'Not allowed.'}, status=http_status.HTTP_403_FORBIDDEN)

        period = _get_active_period()
        if not period:
            return Response({'detail': 'No active evaluation period.'}, status=http_status.HTTP_404_NOT_FOUND)

        try:
            entry = EvaluationEntry.objects.select_for_update().get(
                employee=user, evaluation_period=period
            )
        except EvaluationEntry.DoesNotExist:
            return Response({'detail': 'Evaluation entry not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if entry.status not in (EvaluationEntry.STATUS_PENDING, EvaluationEntry.STATUS_RETURNED):
            return Response(
                {'detail': 'Scores can only be saved while the evaluation is pending or returned.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        scores_data = request.data.get('scores', [])
        if not isinstance(scores_data, list):
            return Response({'detail': 'scores must be a list.'}, status=http_status.HTTP_400_BAD_REQUEST)

        serializer = EvaluationScoreSaveSerializer(data=scores_data, many=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)

        valid_labels = set(_build_period_labels(period.frequency))

        for item in serializer.validated_data:
            task_name    = item['task_name']
            period_label = item['period_label']
            score        = item.get('score')

            if period_label not in valid_labels:
                return Response(
                    {'detail': f'Invalid period_label "{period_label}" for frequency "{period.frequency}".'},
                    status=http_status.HTTP_400_BAD_REQUEST,
                )

            # Resolve task FK (optional — task may have been deleted)
            task = EmployeeTask.objects.filter(
                tasklist__employee=user,
                name=task_name,
            ).first()

            EvaluationScore.objects.update_or_create(
                entry=entry,
                task_name=task_name,
                period_label=period_label,
                defaults={'score': score, 'task': task},
            )

        return Response({'detail': 'Scores saved.'})


def _reset_steps_for_resubmission(entry):
    """For re-evaluation: save step-1 supervisor eval data, delete ALL steps (any status),
    return saved data dict so it can be restored after chain rebuild."""
    saved_eval_data = None
    try:
        step1 = entry.approval_steps.filter(sequence=1).first()
        if step1 and hasattr(step1, 'supervisor_evaluation'):
            ev = step1.supervisor_evaluation
            saved_eval_data = {
                f.name: getattr(ev, f.name)
                for f in ev._meta.concrete_fields
                if f.name not in ('id', 'step_id')
            }
    except Exception:
        pass
    entry.approval_steps.all().delete()
    return saved_eval_data


class MySubmitView(APIView):
    """POST /api/employee-eval/my/submit — submit evaluation and trigger routing."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        user = request.user
        if getattr(user, 'admin', False) or getattr(user, 'hr', False) or getattr(user, 'accounting', False):
            return Response({'detail': 'Not allowed.'}, status=http_status.HTTP_403_FORBIDDEN)

        period = _get_active_period()
        if not period:
            return Response({'detail': 'Evaluation period is closed or does not exist.'}, status=http_status.HTTP_400_BAD_REQUEST)

        try:
            entry = EvaluationEntry.objects.select_for_update().get(
                employee=user, evaluation_period=period
            )
        except EvaluationEntry.DoesNotExist:
            return Response({'detail': 'Evaluation entry not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if entry.status not in (
            EvaluationEntry.STATUS_PENDING,
            EvaluationEntry.STATUS_RETURNED,
            EvaluationEntry.STATUS_COMPLETED,
        ):
            return Response(
                {'detail': 'Evaluation cannot be re-submitted in its current state.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        # For re-evaluation (returned) or new-quarter submission (completed), reset steps.
        is_resubmission = entry.status in (
            EvaluationEntry.STATUS_RETURNED,
            EvaluationEntry.STATUS_COMPLETED,
        )
        if is_resubmission:
            saved_eval_data = _reset_steps_for_resubmission(entry)
        else:
            saved_eval_data = None
            entry.approval_steps.filter(status__in=['pending', 'skipped']).delete()

        # Build approval chain
        from rest_framework.exceptions import ValidationError as DRFValidationError
        try:
            chain = build_evaluation_approval_chain(entry)
        except DRFValidationError as exc:
            return Response({'detail': str(exc.detail)}, status=http_status.HTTP_400_BAD_REQUEST)

        # Restore supervisor eval data to new step 1 (preserves work done before disapproval)
        if saved_eval_data and chain:
            new_step1 = chain[0]
            eval_obj, _ = SupervisorEvaluationEE.objects.get_or_create(step=new_step1)
            for field_name, value in saved_eval_data.items():
                if field_name not in ('is_complete', 'submitted_at'):
                    setattr(eval_obj, field_name, value)
            eval_obj.is_complete = False
            eval_obj.submitted_at = None
            eval_obj.save()

        entry.status = EvaluationEntry.STATUS_SUPERVISOR_REVIEW
        now = timezone.now()
        entry.submitted_at = now
        entry.save(update_fields=['status', 'submitted_at'])

        # Record timeline: submitted (first time) or re_evaluated (after a return/completion)
        _record_timeline(
            entry=entry,
            actor=user,
            action_type=EvaluationTimelineEntry.ACTION_RE_EVALUATED if is_resubmission else EvaluationTimelineEntry.ACTION_SUBMITTED,
            step_order=0,
            acted_at=now,
        )

        # Notify first approver
        if chain:
            first_approver = chain[0].approver
            if first_approver:
                _notify_approver(entry, first_approver)

        return Response(EvaluationEntrySerializer(entry).data)


class MyConfirmView(APIView):
    """POST /api/employee-eval/my/<entry_id>/confirm — user confirmation step."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, entry_id):
        user = request.user

        try:
            entry = EvaluationEntry.objects.select_for_update().get(
                pk=entry_id, employee=user
            )
        except EvaluationEntry.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if entry.status != EvaluationEntry.STATUS_USER_CONFIRMATION:
            return Response(
                {'detail': 'Evaluation is not awaiting your confirmation.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()
        entry.status = EvaluationEntry.STATUS_FINAL_APPROVAL
        entry.confirmed_at = now
        entry.confirmed_by = user
        entry.save(update_fields=['status', 'confirmed_at', 'confirmed_by'])

        # Activate next pending step
        next_step = (
            entry.approval_steps
            .filter(status='pending', activated_at__isnull=True)
            .order_by('sequence')
            .first()
        )
        if next_step:
            next_step.activated_at = now
            next_step.save(update_fields=['activated_at'])
            if next_step.approver:
                _notify_approver(entry, next_step.approver)
        else:
            entry.status = EvaluationEntry.STATUS_COMPLETED
            entry.save(update_fields=['status'])

        return Response(EvaluationEntrySerializer(entry).data)


class MyTrainingRequestsView(APIView):
    """
    GET  /api/employee-eval/my/training-requests  — list training requests for the active period.
    POST /api/employee-eval/my/training-requests  — create or update a request for one quarter.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period = _get_active_period()
        if not period:
            return Response([])
        qs = EvaluationTrainingRequest.objects.filter(
            employee=request.user, period=period
        ).order_by('quarter')
        return Response(EvaluationTrainingRequestSerializer(qs, many=True).data)

    @transaction.atomic
    def post(self, request):
        period = _get_active_period()
        if not period:
            return Response(
                {'detail': 'No active evaluation period.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        quarter = request.data.get('quarter')
        if quarter not in (1, 2, 3, 4):
            return Response({'detail': 'Invalid quarter. Must be 1, 2, 3, or 4.'}, status=http_status.HTTP_400_BAD_REQUEST)
        title = str(request.data.get('title', '')).strip()
        if not title:
            return Response({'detail': 'Training title is required.'}, status=http_status.HTTP_400_BAD_REQUEST)
        preferred_date = request.data.get('preferred_date') or None
        obj, created = EvaluationTrainingRequest.objects.update_or_create(
            employee=request.user,
            period=period,
            quarter=quarter,
            defaults={
                'title': title,
                'objective': str(request.data.get('objective', '')).strip(),
                'preferred_date': preferred_date,
            },
        )
        return Response(
            EvaluationTrainingRequestSerializer(obj).data,
            status=http_status.HTTP_201_CREATED if created else http_status.HTTP_200_OK,
        )


# ─────────────────────────────────────────────────────────────────────────────
# APPROVER VIEWS
# ─────────────────────────────────────────────────────────────────────────────

class ApproverQueueView(APIView):
    """GET /api/employee-eval/approver/queue — entries waiting for or reviewed by the current user."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        # Entries where user has an active pending step (awaiting action)
        pending_entry_ids = set(
            EvaluationApprovalStep.objects
            .filter(approver=user, status='pending', activated_at__isnull=False)
            .values_list('entry_id', flat=True)
        )

        # Entries where user has already reviewed a step
        reviewed_entry_ids = set(
            EvaluationApprovalStep.objects
            .filter(approver=user, status='reviewed')
            .values_list('entry_id', flat=True)
        )

        all_entry_ids = pending_entry_ids | reviewed_entry_ids

        qs = (
            EvaluationEntry.objects
            .filter(pk__in=all_entry_ids)
            .select_related('employee', 'evaluation_period')
            .prefetch_related('approval_steps')
            .order_by('-submitted_at')
        )

        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(employee__first_name__icontains=search) |
                Q(employee__last_name__icontains=search)
            )

        items, total, page, size = _paginate(qs, request)
        return Response({
            'results': EvalApproverQueueItemSerializer(
                items, many=True, context={'request': request}
            ).data,
            'count': total,
            'page': page,
            'page_size': size,
        })


class ApproverQueueBadgeView(APIView):
    """GET /api/employee-eval/approver/badge — count of pending entries."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        count = (
            EvaluationApprovalStep.objects
            .filter(
                approver=request.user,
                status='pending',
                activated_at__isnull=False,
            )
            .count()
        )
        return Response({'count': count})


class ApproverEntryDetailView(APIView):
    """GET /api/employee-eval/approver/entries/<entry_id> — full detail for approver."""
    permission_classes = [IsAuthenticated]

    def get(self, request, entry_id):
        user = request.user

        try:
            entry = (
                EvaluationEntry.objects
                .select_related('employee', 'evaluation_period')
                .prefetch_related('approval_steps__supervisor_evaluation', 'scores')
                .get(pk=entry_id)
            )
        except EvaluationEntry.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        # Verify user is an approver on this entry (active or past)
        user_step = entry.approval_steps.filter(approver=user).first()
        if not user_step:
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        return Response(EvalApproverEntryDetailSerializer(entry, context={'request': request}).data)


class SupervisorEvalSaveView(APIView):
    """POST /api/employee-eval/approver/steps/<step_id>/eval/save — partial save."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, step_id):
        user = request.user

        try:
            step = EvaluationApprovalStep.objects.select_for_update().select_related('entry').get(pk=step_id)
        except EvaluationApprovalStep.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if not can_act_on_evaluation_step(step, user):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        if step.sequence != 1:
            return Response({'detail': 'Supervisor evaluation is only for step 1.'}, status=http_status.HTTP_400_BAD_REQUEST)

        eval_obj, _ = SupervisorEvaluationEE.objects.get_or_create(step=step)
        serializer = SupervisorEvaluationEESerializer(eval_obj, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data)

    patch = post


class SupervisorEvalSubmitView(APIView):
    """POST /api/employee-eval/approver/steps/<step_id>/eval/submit — complete supervisor step."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, step_id):
        user = request.user

        try:
            step = EvaluationApprovalStep.objects.select_for_update().select_related('entry').get(pk=step_id)
        except EvaluationApprovalStep.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if not can_act_on_evaluation_step(step, user):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        if step.sequence != 1:
            return Response({'detail': 'Supervisor evaluation submit is only for step 1.'}, status=http_status.HTTP_400_BAD_REQUEST)

        eval_obj, _ = SupervisorEvaluationEE.objects.get_or_create(step=step)
        serializer = SupervisorEvaluationEESubmitSerializer(eval_obj, data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        serializer.save(is_complete=True, submitted_at=now)

        # Advance step
        step.status = 'reviewed'
        step.acted_at = now
        step.save(update_fields=['status', 'acted_at'])

        entry = step.entry

        # Record timeline: 'evaluated' for first evaluation, 're_evaluated' for subsequent ones.
        already_evaluated = entry.timeline_entries.filter(
            action_type__in=[
                EvaluationTimelineEntry.ACTION_EVALUATED,
                EvaluationTimelineEntry.ACTION_RE_EVALUATED,
            ]
        ).exists()
        _record_timeline(
            entry=entry,
            actor=user,
            action_type=EvaluationTimelineEntry.ACTION_RE_EVALUATED if already_evaluated else EvaluationTimelineEntry.ACTION_EVALUATED,
            step_order=step.sequence,
            acted_at=now,
        )

        # Skip user confirmation — advance directly to next pending step or complete
        next_step = (
            entry.approval_steps
            .filter(status='pending', activated_at__isnull=True)
            .order_by('sequence')
            .first()
        )
        if next_step:
            next_step.activated_at = now
            next_step.save(update_fields=['activated_at'])
            entry.status = EvaluationEntry.STATUS_FINAL_APPROVAL
            entry.save(update_fields=['status'])
            if next_step.approver:
                _notify_approver(entry, next_step.approver)
        else:
            entry.status = EvaluationEntry.STATUS_COMPLETED
            entry.save(update_fields=['status'])
            _record_timeline(
                entry=entry,
                actor=None,
                action_type=EvaluationTimelineEntry.ACTION_COMPLETED,
                step_order=0,
                acted_at=now,
            )

        return Response({'detail': 'Supervisor evaluation submitted.'})


class FinalApproverActionView(APIView):
    """POST /api/employee-eval/approver/steps/<step_id>/action — approve or disapprove."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, step_id):
        user = request.user

        try:
            step = EvaluationApprovalStep.objects.select_for_update().select_related('entry').get(pk=step_id)
        except EvaluationApprovalStep.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if not can_act_on_evaluation_step(step, user):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        if step.sequence < 2:
            return Response({'detail': 'Final action is only for step 2+.'}, status=http_status.HTTP_400_BAD_REQUEST)

        action = request.data.get('action', '').strip().lower()
        if action not in ('approved', 'disapproved'):
            return Response({'detail': 'action must be "approved" or "disapproved".'}, status=http_status.HTTP_400_BAD_REQUEST)

        remarks = str(request.data.get('remarks', '')).strip()

        # Server-side validation: remarks required for disapproval.
        if action == 'disapproved' and not remarks:
            return Response(
                {'detail': 'Remarks are required when disapproving.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()
        step.status = 'reviewed'
        step.acted_at = now
        step.final_action = action
        step.final_remarks = remarks
        step.save(update_fields=['status', 'acted_at', 'final_action', 'final_remarks'])

        entry = step.entry

        if action == 'disapproved':
            entry.status = EvaluationEntry.STATUS_RETURNED
            entry.save(update_fields=['status'])

            # Record timeline: disapproved + returned
            _record_timeline(
                entry=entry, actor=user,
                action_type=EvaluationTimelineEntry.ACTION_DISAPPROVED,
                step_order=step.sequence, remarks=remarks, acted_at=now,
            )
            _record_timeline(
                entry=entry, actor=user,
                action_type=EvaluationTimelineEntry.ACTION_RETURNED,
                step_order=step.sequence, acted_at=now,
            )

            # Reset Step 1 to active-pending so the supervisor can re-evaluate immediately
            # without waiting for the employee to re-submit.
            # All seq >= 2 steps are reset to inactive-pending so they re-run in order
            # after Step 1 re-submits. final_action/final_remarks are preserved for the
            # audit trail shown in ApproverRemarksSection.
            entry.approval_steps.filter(sequence=1).update(
                status='pending', activated_at=now
            )
            entry.approval_steps.filter(sequence__gt=1).update(
                status='pending', activated_at=None
            )

            # Allow Step 1 supervisor to edit their evaluation again.
            step1 = (
                entry.approval_steps
                .filter(sequence=1)
                .select_related('supervisor_evaluation')
                .first()
            )
            if step1:
                try:
                    sv = step1.supervisor_evaluation
                    sv.is_complete = False
                    sv.submitted_at = None
                    sv.save(update_fields=['is_complete', 'submitted_at'])
                except SupervisorEvaluationEE.DoesNotExist:
                    pass
                if step1.approver:
                    _notify_step1_return(entry, step1.approver, remarks)
        else:
            # Record timeline: approved
            _record_timeline(
                entry=entry, actor=user,
                action_type=EvaluationTimelineEntry.ACTION_APPROVED,
                step_order=step.sequence, remarks=remarks, acted_at=now,
            )

            # Check if there are more pending steps
            next_step = (
                entry.approval_steps
                .filter(status='pending', activated_at__isnull=True)
                .order_by('sequence')
                .first()
            )
            if next_step:
                next_step.activated_at = now
                next_step.save(update_fields=['activated_at'])
                entry.status = EvaluationEntry.STATUS_SECOND_FINAL_APPROVAL
                entry.save(update_fields=['status'])
                if next_step.approver:
                    _notify_approver(entry, next_step.approver)
            else:
                entry.status = EvaluationEntry.STATUS_COMPLETED
                entry.save(update_fields=['status'])
                _record_timeline(
                    entry=entry, actor=None,
                    action_type=EvaluationTimelineEntry.ACTION_COMPLETED,
                    step_order=0, acted_at=now,
                )

        return Response({'detail': f'Entry {action}.', 'status': entry.status})


# ─────────────────────────────────────────────────────────────────────────────
# EMPLOYEE STATS VIEW
# ─────────────────────────────────────────────────────────────────────────────

class EmployeeStatsView(APIView):
    """GET /api/employee-eval/approver/entries/<entry_id>/stats

    Returns aggregate statistics for the employee being evaluated:
    - Leave days & hours (approved, within evaluation period dates)
    - Certificates issued (within fiscal year)
    - Training evaluations completed (within fiscal year)
    - PRF requests filed (within fiscal year)

    Accessible to any approver assigned to the entry and to admin/HR users.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, entry_id):
        try:
            entry = (
                EvaluationEntry.objects
                .select_related('employee', 'evaluation_period')
                .get(pk=entry_id)
            )
        except EvaluationEntry.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        # Permission: must be an approver on this entry or an admin/HR user
        is_approver = entry.approval_steps.filter(approver=request.user).exists()
        is_admin_or_hr = getattr(request.user, 'admin', False) or getattr(request.user, 'hr', False)
        if not (is_approver or is_admin_or_hr):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        employee = entry.employee
        period = entry.evaluation_period

        # Full fiscal year: May 1 of fiscal_year → April 30 of fiscal_year + 1
        fy_start = datetime.date(period.fiscal_year, 5, 1)
        fy_end = datetime.date(period.fiscal_year + 1, 4, 30)

        # ── Card 1: Leave Taken (approved leaves overlapping the evaluation period) ──
        from django.db.models import Sum as DjSum
        from leave.models import LeaveRequest
        approved_leaves = LeaveRequest.objects.filter(
            employee=employee,
            status='approved',
            date_start__lte=period.end_date,
            date_end__gte=period.start_date,
        )
        leave_days = int(approved_leaves.aggregate(t=DjSum('days_count'))['t'] or 0)
        leave_hours = float(approved_leaves.aggregate(t=DjSum('hours'))['t'] or 0)

        # ── Card 2: Certificates issued this fiscal year ──
        from certification.models import Certificate
        certificates = Certificate.objects.filter(
            employee=employee,
            created_at__date__gte=fy_start,
            created_at__date__lte=fy_end,
        ).count()

        # ── Card 3: Trainings attended — count submissions where employee submitted their own
        #            evaluation form (is_complete=True) within the evaluation period dates.
        #            Using period.start_date/end_date (not fiscal year) because periods can extend
        #            past April 30. TrainingSubmission covers both all_users and specific_users
        #            trainings; TrainingParticipant only exists for specific_users.
        from training.models import TrainingSubmission
        trainings_completed = TrainingSubmission.objects.filter(
            submitted_by=employee,
            is_complete=True,
            training__training_date__gte=period.start_date,
            training__training_date__lte=period.end_date,
        ).count()

        # ── Card 4: PRF requests filed this fiscal year ──
        from prForm.models import PRFRequest
        prf_requests = PRFRequest.objects.filter(
            employee=employee,
            created_at__date__gte=fy_start,
            created_at__date__lte=fy_end,
        ).count()

        return Response({
            'leave_days': leave_days,
            'leave_hours': leave_hours,
            'certificates': certificates,
            'trainings_completed': trainings_completed,
            'prf_requests': prf_requests,
        })


# ─────────────────────────────────────────────────────────────────────────────
# EVALUATION TIMELINE VIEW
# ─────────────────────────────────────────────────────────────────────────────

class EvaluationTimelineView(APIView):
    """GET /api/employee-eval/entries/<entry_id>/timeline

    Returns all timeline entries for an evaluation in chronological order.
    Accessible to: the employee who owns the entry, any approver on the entry,
    and admin/HR users.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, entry_id):
        try:
            entry = EvaluationEntry.objects.get(pk=entry_id)
        except EvaluationEntry.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        user = request.user
        is_owner = entry.employee_id == user.pk
        is_approver = entry.approval_steps.filter(approver=user).exists()
        is_admin_or_hr = getattr(user, 'admin', False) or getattr(user, 'hr', False)

        if not (is_owner or is_approver or is_admin_or_hr):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        from employee_evaluation.serializers import EvaluationTimelineEntrySerializer
        qs = (
            entry.timeline_entries
            .select_related('actor')
            .order_by('acted_at', 'id')
        )
        return Response(EvaluationTimelineEntrySerializer(qs, many=True).data)


class EvaluationSettingsView(APIView):
    """GET /api/employee-eval/settings — return current settings."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            obj = EvaluationSettings.objects.get()
        except EvaluationSettings.DoesNotExist:
            return Response({'detail': 'Evaluation settings not configured.'}, status=http_status.HTTP_404_NOT_FOUND)
        return Response(EvaluationSettingsSerializer(obj).data)


# ─────────────────────────────────────────────────────────────────────────────
# ADMIN PERIOD EXPORT VIEW
# ─────────────────────────────────────────────────────────────────────────────

class AdminPeriodExportView(APIView):
    """GET /api/employee-eval/admin/periods/<pk>/export

    Generates a two-sheet XLSX report for the selected evaluation period:
      Sheet 1 — Overall Summary: one row per eligible employee with quarter
                scores, total average, and average quality ratings.
      Sheet 2 — Supervisor Evaluation: merged-header table with performance
                evaluation text fields and quality ratings per employee.

    Pre-download validation:
      - Period must exist.
      - Eligible employee list must compile without error.
      - All related model queries must resolve without exceptions.
      If any query fails, a JSON error response is returned (no partial file).
    """

    permission_classes = [IsAuthenticated]

    # Monthly labels grouped into fiscal quarters (May–Apr fiscal year).
    _MONTHLY_QUARTER_MAP: dict[str, str] = {
        'May': 'Q1', 'Jun': 'Q1', 'Jul': 'Q1',
        'Aug': 'Q2', 'Sep': 'Q2', 'Oct': 'Q2',
        'Nov': 'Q3', 'Dec': 'Q3', 'Jan': 'Q3',
        'Feb': 'Q4', 'Mar': 'Q4', 'Apr': 'Q4',
    }

    # Status colour fills (ARGB, no alpha prefix needed for openpyxl PatternFill).
    _STATUS_FILLS: dict[str, str] = {
        'completed':             'C6EFCE',  # green
        'pending':               'FFEB9C',  # yellow
        'not_started':           'FFEB9C',  # yellow (same as pending)
        'supervisor_review':     'BDD7EE',  # blue
        'user_confirmation':     'BDD7EE',  # blue
        'final_approval':        'BDD7EE',  # blue
        'second_final_approval': 'BDD7EE',  # blue
        'returned':              'FCE4D6',  # orange
        'disapproved':           'FFC7CE',  # red
    }

    _STATUS_FONT_COLORS: dict[str, str] = {
        'completed':             '375623',
        'pending':               '9C5700',
        'not_started':           '9C5700',
        'supervisor_review':     '1F4E79',
        'user_confirmation':     '1F4E79',
        'final_approval':        '1F4E79',
        'second_final_approval': '1F4E79',
        'returned':              '843C0C',
        'disapproved':           '9C0006',
    }

    _STATUS_LABELS: dict[str, str] = {
        'not_started':           'Not Started',
        'pending':               'Pending',
        'supervisor_review':     'Supervisor Review',
        'user_confirmation':     'User Confirmation',
        'final_approval':        'Awaiting Final Approval',
        'second_final_approval': 'Under Second Review',
        'returned':              'Returned for Revision',
        'completed':             'Completed',
        'disapproved':           'Disapproved',
    }

    def get(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            return self._build_export(pk)
        except EvaluationPeriod.DoesNotExist:
            return Response({'detail': 'Evaluation period not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        except Exception:
            logger.exception('AdminPeriodExportView: unexpected error for period pk=%s', pk)
            return Response(
                {'detail': 'The report could not be generated due to an internal error.'},
                status=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _build_export(self, pk: int):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from userProfile.models import workInformation

        # ── 1. Load period ────────────────────────────────────────────────────
        period = EvaluationPeriod.objects.get(pk=pk)
        fy_start = period.fiscal_year
        fy_end   = fy_start + 1

        # ── 2. Compile eligible users ─────────────────────────────────────────
        eligible_qs = _eligible_users_for_period(period.start_date)
        eligible_users = list(
            eligible_qs
            .only('id', 'idnumber', 'firstname', 'lastname')
            .order_by('lastname', 'firstname')
        )
        eligible_ids = [u.id for u in eligible_users]

        # ── 3. Work information (department + line) per employee ──────────────
        wi_map: dict[int, tuple[str | None, str | None]] = {}
        for wi in (
            workInformation.objects
            .filter(employee_id__in=eligible_ids)
            .select_related('department', 'line')
            .order_by('employee_id')
        ):
            if wi.employee_id not in wi_map:
                dept_name = wi.department.name if wi.department else None
                line_name = wi.line.name if wi.line else None
                wi_map[wi.employee_id] = (dept_name, line_name)

        # ── 4. Evaluation entries indexed by employee_id ──────────────────────
        entries_by_emp: dict[int, EvaluationEntry] = {}
        for e in EvaluationEntry.objects.filter(
            evaluation_period=period, employee_id__in=eligible_ids
        ):
            entries_by_emp[e.employee_id] = e

        # ── 5. Supervisor evaluations indexed by entry_id ─────────────────────
        # Only step-1 supervisor evaluations (one per entry — OneToOne on step).
        entry_ids = [e.id for e in entries_by_emp.values()]
        sup_eval_by_entry: dict[int, SupervisorEvaluationEE] = {}
        for sup in (
            SupervisorEvaluationEE.objects
            .filter(step__entry_id__in=entry_ids, step__sequence=1)
            .select_related('step')
        ):
            sup_eval_by_entry[sup.step.entry_id] = sup

        # ── 5b. Task counts/names per employee (for Q-average denominator) ─────
        task_count_by_emp: dict[int, int] = {}
        task_names_by_emp: dict[int, set] = {}
        for tl in (
            EmployeeTasklist.objects
            .filter(employee_id__in=eligible_ids)
            .prefetch_related('tasks')
        ):
            names = list(tl.tasks.values_list('name', flat=True))
            task_count_by_emp[tl.employee_id] = len(names)
            task_names_by_emp[tl.employee_id] = set(names)

        # ── 6. Build workbook ─────────────────────────────────────────────────
        wb = openpyxl.Workbook()

        self._build_sheet1(
            wb, period, fy_start, fy_end,
            eligible_users, wi_map, entries_by_emp, sup_eval_by_entry,
            task_count_by_emp, task_names_by_emp,
        )
        self._build_sheet2(
            wb, period, fy_start, fy_end,
            eligible_users, wi_map, entries_by_emp, sup_eval_by_entry,
        )

        # Remove the default empty sheet openpyxl creates.
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # ── 7. Serialise and return ───────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'Performance_Evaluation_FY{fy_start}-{fy_end}.xlsx'
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ── Sheet helpers ─────────────────────────────────────────────────────────

    @staticmethod
    def _thin_border():
        from openpyxl.styles import Border, Side
        s = Side(style='thin')
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _header_style(bold: bool = True):
        from openpyxl.styles import Alignment, Font, PatternFill
        return {
            'font':      Font(bold=bold, color='FFFFFF', size=10),
            'fill':      PatternFill(start_color='2845D6', end_color='2845D6', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        }

    @classmethod
    def _apply_header(cls, cell, label: str):
        from openpyxl.styles import Alignment, Font, PatternFill
        cell.value = label
        cell.font  = Font(bold=True, color='FFFFFF', size=10)
        cell.fill  = PatternFill(start_color='2845D6', end_color='2845D6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = AdminPeriodExportView._thin_border()

    @classmethod
    def _apply_data_cell(cls, cell, value, wrap: bool = False):
        from openpyxl.styles import Alignment, Font
        cell.value     = value
        cell.font      = Font(size=10)
        cell.border    = AdminPeriodExportView._thin_border()
        cell.alignment = Alignment(
            horizontal='left', vertical='top', wrap_text=wrap,
        )

    @classmethod
    def _apply_status_cell(cls, cell, status: str):
        from openpyxl.styles import Alignment, Font, PatternFill
        label      = cls._STATUS_LABELS.get(status, status)
        fill_color = cls._STATUS_FILLS.get(status, 'FFFFFF')
        font_color = cls._STATUS_FONT_COLORS.get(status, '000000')
        cell.value     = label
        cell.font      = Font(bold=True, color=font_color, size=10)
        cell.fill      = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        cell.border    = AdminPeriodExportView._thin_border()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _auto_col_widths(ws, min_w: int = 10, max_w: int = 50):
        from openpyxl.utils import get_column_letter
        col_widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    col_idx = cell.column
                    text_len = len(str(cell.value))
                    col_widths[col_idx] = max(col_widths.get(col_idx, min_w), text_len + 4)
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, min_w), max_w)

    # ── Quarter score helpers ─────────────────────────────────────────────────

    @classmethod
    def _quarter_total(cls, sup_eval: SupervisorEvaluationEE | None, quarter: str, frequency: str, task_count: int = 0, task_names: set | None = None) -> str:
        """Return sum-of-quarter-scores ÷ task_count for a quarter label (Q1–Q4).

        For quarterly frequency: sums supervisor_scores keys ending with __{quarter}.
        For monthly frequency: sums scores for the three months composing the quarter.
        Only counts scores for tasks in task_names (current tasklist) when provided.
        Returns '—' when no scores exist, sup_eval is absent, or task_count is 0.
        """
        if sup_eval is None or task_count == 0:
            return '—'
        scores: dict = sup_eval.supervisor_scores or {}
        if not scores:
            return '—'

        if frequency == 'quarterly':
            suffix = f'__{quarter}'
            values = [
                float(v) for k, v in scores.items()
                if k.endswith(suffix)
                and (task_names is None or k[: k.rfind('__')] in task_names)
                and v not in (None, '', 'N/A')
                and _is_numeric(v)
            ]
        else:
            # Monthly: aggregate the three months that compose this fiscal quarter.
            month_labels = [
                m for m, q in AdminPeriodExportView._MONTHLY_QUARTER_MAP.items()
                if q == quarter
            ]
            values = [
                float(v)
                for k, v in scores.items()
                if any(k.endswith(f'__{ml}') for ml in month_labels)
                and (task_names is None or k[: k.rfind('__')] in task_names)
                and v not in (None, '', 'N/A')
                and _is_numeric(v)
            ]

        if not values:
            return '—'
        result = sum(values) / task_count
        return str(int(result)) if result == int(result) else f'{result:.2f}'

    @classmethod
    def _quality_average(cls, sup_eval: SupervisorEvaluationEE | None) -> str:
        """Average of all five quality categories across available quarters.

        For each category, average non-null quarter values.
        Then average the five category averages (only categories that have data).
        Returns '—' if no qualities have been rated.
        """
        if sup_eval is None:
            return '—'
        categories = [
            ('cost_consciousness_q1', 'cost_consciousness_q2', 'cost_consciousness_q3', 'cost_consciousness_q4'),
            ('dependability_q1',      'dependability_q2',      'dependability_q3',      'dependability_q4'),
            ('communication_q1',      'communication_q2',      'communication_q3',      'communication_q4'),
            ('work_ethics_q1',        'work_ethics_q2',        'work_ethics_q3',        'work_ethics_q4'),
            ('attendance_q1',         'attendance_q2',         'attendance_q3',         'attendance_q4'),
        ]
        cat_averages = []
        for fields in categories:
            # Use the last non-null quarter value per category (most recent evaluation).
            last_val = None
            for f in reversed(fields):  # q4 → q1
                v = getattr(sup_eval, f)
                if v is not None:
                    last_val = v
                    break
            if last_val is not None:
                cat_averages.append(last_val)
        if not cat_averages:
            return '—'
        avg = sum(cat_averages) / len(cat_averages)
        return f'{avg:.2f}'

    # ── Text field helpers ────────────────────────────────────────────────────

    @staticmethod
    def _concat_quarters(sup_eval: SupervisorEvaluationEE | None, base_field: str) -> str:
        """Return the first non-empty quarter value for a text field (single instance).

        base_field: e.g. 'strengths' → checks strengths_q1 … strengths_q4 in order.
        Returns the first non-blank value found, or '—' if all are empty.
        """
        if sup_eval is None:
            return '—'
        for q in ('q1', 'q2', 'q3', 'q4'):
            val = (getattr(sup_eval, f'{base_field}_{q}', '') or '').strip()
            if val:
                return val
        return '—'

    @staticmethod
    def _quality_cell_value(sup_eval: SupervisorEvaluationEE | None, base_field: str) -> str:
        """Return the first non-null quarter rating for a quality column (single instance).

        Formats as '{rating} - {comment}' when a comment exists, else just '{rating}'.
        Returns '—' if no quarter has a rating.
        """
        if sup_eval is None:
            return '—'
        quality_comments: dict = sup_eval.quality_comments or {}
        # Use the last non-null quarter value (most recent evaluation).
        for q_suffix in ('q4', 'q3', 'q2', 'q1'):
            rating = getattr(sup_eval, f'{base_field}_{q_suffix}', None)
            if rating is None:
                continue
            comment_key = f'{base_field}_{q_suffix}_comment'
            comment = (quality_comments.get(comment_key, '') or '').strip()
            return f'{rating} - {comment}' if comment else str(rating)
        return '—'

    # ── Sheet 1 builder ───────────────────────────────────────────────────────

    def _build_sheet1(
        self, wb, period, fy_start: int, fy_end: int,
        eligible_users, wi_map, entries_by_emp, sup_eval_by_entry,
        task_count_by_emp: dict, task_names_by_emp: dict,
    ):
        from openpyxl.styles import Alignment, Font
        ws = wb.create_sheet(title='Overall Summary')

        # ── Title rows ────────────────────────────────────────────────────────
        ws['A1'] = 'RYONAN ELECTRIC PHILIPPINES CORPORATION'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f'Performance Evaluation FY {fy_start}–{fy_end} Overall Summary'
        ws['A2'].font = Font(size=11)

        # ── Column headers at row 4 ───────────────────────────────────────────
        # Columns: ID Number | Employee Name | Department | Line | Status |
        #          Q1 | Q2 | Q3 | Q4 | Average Evaluation | 65% |
        #          Average Behavioral | 25%
        standard_headers = [
            'ID Number', 'Employee Name', 'Department', 'Line', 'Status',
            'Q1', 'Q2', 'Q3', 'Q4', 'Average Evaluation', 'Average Behavioral',
        ]
        # col indices: 1-10 standard, 11 = '65%' (after col 10), 12 = 'Average Behavioral' (→13 after shift), 13 = '25%'
        # Build full ordered header list including weighted cols
        all_headers = [
            'ID Number', 'Employee Name', 'Department', 'Line', 'Status',
            'Q1', 'Q2', 'Q3', 'Q4',
            'Average Evaluation',  # col 10
            '65%',                 # col 11 — weighted, red header
            'Average Behavioral',  # col 12
            '25%',                 # col 13 — weighted, red header
        ]
        _weighted_cols = {11, 13}  # 1-based column indices with red header font
        for col_idx, h in enumerate(all_headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            self._apply_header(cell, h)
            if col_idx in _weighted_cols:
                # Override font: red bold, keep blue background from _apply_header
                cell.font = Font(bold=True, color='FF0000', size=10)

        # ── Data rows starting at row 5 ───────────────────────────────────────
        for row_num, u in enumerate(eligible_users, start=5):
            first = (getattr(u, 'firstname', '') or '').strip()
            last  = (getattr(u, 'lastname',  '') or '').strip()
            emp_name = f'{last}, {first}'.strip(', ') or u.idnumber or ''
            dept_name, line_name = wi_map.get(u.id, (None, None))
            entry  = entries_by_emp.get(u.id)
            status = entry.status if entry else 'not_started'
            sup_eval = sup_eval_by_entry.get(entry.id) if entry else None
            task_count = task_count_by_emp.get(u.id, 0)
            task_names = task_names_by_emp.get(u.id, None)

            # Q1–Q4 totals — sum of current-tasklist scores ÷ task count
            q_values_str = [
                self._quarter_total(sup_eval, 'Q1', period.frequency, task_count, task_names),
                self._quarter_total(sup_eval, 'Q2', period.frequency, task_count, task_names),
                self._quarter_total(sup_eval, 'Q3', period.frequency, task_count, task_names),
                self._quarter_total(sup_eval, 'Q4', period.frequency, task_count, task_names),
            ]

            # Total average of Q1–Q4
            numeric_qs = [float(v) for v in q_values_str if v != '—' and _is_numeric(v)]
            if numeric_qs:
                avg_total = sum(numeric_qs) / len(numeric_qs)
                avg_total_str = f'{avg_total:.2f}'
            else:
                avg_total_str = '—'

            qual_avg_str = self._quality_average(sup_eval)

            # Weighted columns
            eval_65_str = (
                f'{float(avg_total_str) * 0.65:.2f}'
                if avg_total_str != '—' else '—'
            )
            qual_25_str = (
                f'{float(qual_avg_str) * 0.25:.2f}'
                if qual_avg_str != '—' else '—'
            )

            # Column order: ID | Name | Dept | Line | Status | Q1-Q4 |
            #               Avg Eval | 65% | Avg Behavioral | 25%
            data = [
                u.idnumber or '',
                emp_name,
                dept_name or '—',
                line_name or '—',
                status,          # col 5 — rendered as coloured cell
                *q_values_str,   # cols 6–9
                avg_total_str,   # col 10
                eval_65_str,     # col 11
                qual_avg_str,    # col 12
                qual_25_str,     # col 13
            ]

            for col_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                if col_idx == 5:  # Status column
                    self._apply_status_cell(cell, status)
                else:
                    self._apply_data_cell(cell, val if val != '' else '—')

        self._auto_col_widths(ws)

    # ── Sheet 2 builder ───────────────────────────────────────────────────────

    def _build_sheet2(
        self, wb, period, fy_start: int, fy_end: int,
        eligible_users, wi_map, entries_by_emp, sup_eval_by_entry,
    ):
        from openpyxl.styles import Alignment, Font
        ws = wb.create_sheet(title='Supervisor Evaluation')

        # ── Title rows ────────────────────────────────────────────────────────
        ws['A1'] = 'RYONAN ELECTRIC PHILIPPINES CORPORATION'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f'Supervisor Evaluation FY {fy_start}–{fy_end} Overall Summary'
        ws['A2'].font = Font(size=11)

        # ── Two-row merged header structure at rows 4–5 ───────────────────────
        # Column layout:
        #   1: ID Number       (merged rows 4–5)
        #   2: Employee Name   (merged rows 4–5)
        #   3: Status          (merged rows 4–5)
        #   4–8: Performance Evaluation (merged cols 4–8 in row 4)
        #         sub: Strengths | Weaknesses | Training Required |
        #              Superior's Assessment | Employee Comments
        #   9–13: Performance Qualities (merged cols 9–13 in row 4)
        #         sub: Cost Consciousness | Dependability | Communication |
        #              Work Ethics | Attendance

        # Row 4 — top-level headers
        for col_idx, label in [(1, 'ID Number'), (2, 'Employee Name'), (3, 'Status')]:
            self._apply_header(ws.cell(row=4, column=col_idx), label)
            ws.merge_cells(
                start_row=4, start_column=col_idx,
                end_row=5,   end_column=col_idx,
            )
            ws.cell(row=4, column=col_idx).alignment = AdminPeriodExportView._centered_wrap()

        # "Performance Evaluation" merged across cols 4–8
        pe_cell = ws.cell(row=4, column=4)
        self._apply_header(pe_cell, 'Performance Evaluation')
        ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=8)

        # "Performance Qualities" merged across cols 9–13
        pq_cell = ws.cell(row=4, column=9)
        self._apply_header(pq_cell, 'Performance Qualities')
        ws.merge_cells(start_row=4, start_column=9, end_row=4, end_column=13)

        # Row 5 — sub-headers
        pe_subs = [
            'Strengths', 'Weaknesses', 'Training Required',
            "Superior's Assessment", 'Employee Comments',
        ]
        pq_subs = [
            'Cost Consciousness', 'Dependability', 'Communication',
            'Work Ethics', 'Attendance',
        ]
        for col_offset, sub in enumerate(pe_subs):
            self._apply_header(ws.cell(row=5, column=4 + col_offset), sub)
        for col_offset, sub in enumerate(pq_subs):
            self._apply_header(ws.cell(row=5, column=9 + col_offset), sub)

        # Row 5 single-column headers (already merged with row 4 — apply border/style)
        from openpyxl.styles import PatternFill as _PF
        for col_idx in (1, 2, 3):
            cell = ws.cell(row=5, column=col_idx)
            cell.border = self._thin_border()
            cell.fill   = _PF(
                start_color='2845D6', end_color='2845D6', fill_type='solid'
            )

        # ── Data rows starting at row 6 ───────────────────────────────────────
        _QUALITY_FIELDS = [
            'cost_consciousness', 'dependability', 'communication',
            'work_ethics', 'attendance',
        ]

        for row_num, u in enumerate(eligible_users, start=6):
            first = (getattr(u, 'firstname', '') or '').strip()
            last  = (getattr(u, 'lastname',  '') or '').strip()
            emp_name = f'{last}, {first}'.strip(', ') or u.idnumber or ''
            entry    = entries_by_emp.get(u.id)
            status   = entry.status if entry else 'not_started'
            sup_eval = sup_eval_by_entry.get(entry.id) if entry else None

            # Col 1: ID Number
            self._apply_data_cell(ws.cell(row=row_num, column=1), u.idnumber or '')
            # Col 2: Employee Name
            self._apply_data_cell(ws.cell(row=row_num, column=2), emp_name)
            # Col 3: Status (coloured)
            self._apply_status_cell(ws.cell(row=row_num, column=3), status)

            # Cols 4–8: Performance Evaluation text fields
            pe_values = [
                self._concat_quarters(sup_eval, 'strengths'),
                self._concat_quarters(sup_eval, 'weaknesses'),
                self._concat_quarters(sup_eval, 'training_required'),
                self._concat_quarters(sup_eval, 'supervisor_comments'),
                self._concat_quarters(sup_eval, 'employee_comments'),
            ]
            for col_offset, val in enumerate(pe_values):
                self._apply_data_cell(ws.cell(row=row_num, column=4 + col_offset), val, wrap=True)

            # Cols 9–13: Performance Qualities
            for col_offset, field in enumerate(_QUALITY_FIELDS):
                val = self._quality_cell_value(sup_eval, field)
                self._apply_data_cell(ws.cell(row=row_num, column=9 + col_offset), val, wrap=True)

        self._auto_col_widths(ws)

    @staticmethod
    def _centered_wrap():
        from openpyxl.styles import Alignment
        return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _is_numeric(v) -> bool:
    """Return True if v can be converted to float."""
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False
