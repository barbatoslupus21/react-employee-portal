"""Finance module views.

All admin endpoints require ``admin=True AND accounting=True`` on the user.

Endpoints
---------
GET  admin/types                       — all 4 type lists
POST admin/types/create                — create a type
PUT  admin/types/<pk>                  — update a type
DEL  admin/types/<pk>                  — delete a type
GET  admin/chart                       — record-count chart aggregates
GET  admin/employees                   — paginated, annotated employee list
POST admin/import                      — xlsx import (record_type param)
GET  admin/export                      — xlsx export (record_type + date range)
"""
from __future__ import annotations

import base64
import calendar
import datetime
import io
from decimal import Decimal, InvalidOperation

from django.conf import settings as django_settings
from django.db import transaction
from django.db.models import Count, Exists, OuterRef, Prefetch, Q, Sum
from django.http import HttpResponse
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from activityLog.models import Notification

from .models import (
    Allowance, AllowanceType,
    Deduction,
    Loan, LoanSettings, LoanType,
    OfficeFinanceRate,
    Payslip, PayslipType,
    Savings, SavingsType,
)
from .serializers import (
    AllowanceSerializer,
    AllowanceTypeSerializer,
    FinanceEmployeeRowSerializer,
    LoanSerializer,
    LoanSettingsSerializer,
    LoanTypeSerializer,
    OfficeFinanceRateSerializer,
    PayslipSerializer,
    PayslipTypeSerializer,
    SavingsSerializer,
    SavingsTypeSerializer,
)


# ── Auth helper ───────────────────────────────────────────────────────────────

def _require_accounting_admin(request) -> Response | None:
    """Return 403 unless user has admin=True OR accounting=True."""
    u = request.user
    if not (getattr(u, 'admin', False) or getattr(u, 'accounting', False)):
        return Response(
            {'detail': 'Admin or Accounting permission required.'},
            status=status.HTTP_403_FORBIDDEN,
        )
    return None


# ── Error-report builder ──────────────────────────────────────────────────────

def _build_error_excel(failures: list[dict]) -> str:
    """Build an error-report xlsx and return base64-encoded bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Import Errors'

    def _side():
        return Side(style='thin', color='FF000000')

    thin = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    hdr_fill = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')

    headers = ['Row', 'Field / Column', 'Reason']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')

    for row_num, item in enumerate(failures, 2):
        ws.cell(row=row_num, column=1, value=item.get('row', '')).border = thin
        ws.cell(row=row_num, column=2, value=item.get('field', '')).border = thin
        reason_cell = ws.cell(row=row_num, column=3, value=item.get('reason', ''))
        reason_cell.border = thin
        reason_cell.font   = Font(color='FFFF0000')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def _build_deduction_error_excel(original_rows: list[list], failures: list[dict]) -> str:
    """
    Build a deduction-specific error report that preserves the original file structure.
    Output columns: ID Number | Loan Type | Deduction | Remarks
    Red header row; erroneous rows rendered in red font.

    original_rows[0] = header row (skipped)
    original_rows[i] = Excel row i+1 (first data row is i=1, Excel row 2)
    failures entries have 'row' = Excel row number (2-based), 'reason'.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Deduction Errors'

    def _side():
        return Side(style='thin', color='FF000000')

    thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    hdr_fill = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')
    red_font = Font(color='FFFF0000')
    hdr_font = Font(bold=True, color='FFFFFFFF')

    display_headers = ['ID Number', 'Employee Name', 'Loan Type', 'Deduction', 'Remarks']
    col_widths      = [16, 28, 28, 18, 60]

    for col, (h, w) in enumerate(zip(display_headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = thin
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = w

    # Build error map: excel_row_num → combined reason string
    error_map: dict[int, str] = {}
    for f in failures:
        r = f.get('row')
        if r is not None:
            reason = f.get('reason', '')
            error_map[r] = (error_map[r] + '; ' + reason) if r in error_map else reason

    # Write original data rows; Excel row i corresponds to original_rows[i-1]
    for i, row in enumerate(original_rows[1:], start=2):
        while len(row) < 4:
            row.append(None)

        has_error = i in error_map
        for col, val in enumerate(row[:4], 1):
            cell = ws.cell(row=i, column=col, value=str(val) if val is not None else '')
            cell.border = thin
            if has_error:
                cell.font = red_font

        remarks_cell = ws.cell(row=i, column=5, value=error_map.get(i, ''))
        remarks_cell.border = thin
        if has_error:
            remarks_cell.font = red_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def _build_allowance_error_excel(original_rows: list[list], failures: list[dict]) -> str:
    """
    Build an allowance-specific error report that preserves the original file structure.
    Output columns: ID Number | Employee Name | Allowance Type | Amount |
                    Deposited Date | Covered Period | Remarks
    Red header row; erroneous rows rendered in red font.

    original_rows[0] = header row (skipped in output)
    failures entries have 'row' = Excel row number (2-based), 'reason'.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Allowance Errors'

    def _side():
        return Side(style='thin', color='FF000000')

    thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    hdr_fill = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')
    red_font = Font(color='FFFF0000')
    hdr_font = Font(bold=True, color='FFFFFFFF')

    display_headers = ['ID Number', 'Employee Name', 'Allowance Type', 'Amount',
                       'Deposited Date', 'Covered Period', 'Remarks']
    col_widths = [14, 28, 24, 14, 16, 22, 60]

    for col, (h, w) in enumerate(zip(display_headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = thin
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = w

    # Build error map: excel_row_num → combined reason string
    error_map: dict[int, str] = {}
    for f in failures:
        r = f.get('row')
        if r is not None:
            reason = f.get('reason', '')
            error_map[r] = (error_map[r] + '; ' + reason) if r in error_map else reason

    for i, row in enumerate(original_rows[1:], start=2):
        while len(row) < 6:
            row.append(None)

        has_error = i in error_map
        for col, val in enumerate(row[:6], 1):
            cell = ws.cell(row=i, column=col, value=str(val) if val is not None else '')
            cell.border = thin
            if has_error:
                cell.font = red_font

        remarks_cell = ws.cell(row=i, column=7, value=error_map.get(i, ''))
        remarks_cell.border = thin
        if has_error:
            remarks_cell.font = red_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def _build_savings_error_excel(original_rows: list[list], failures: list[dict]) -> str:
    """
    Build a savings-specific error report that preserves the original file structure.
    Output columns: ID Number | Employee Name | Savings Type | Savings | Remarks
    Red header row; erroneous rows rendered in red font.

    original_rows[0] = header row (skipped in output)
    failures entries have 'row' = Excel row number (2-based), 'reason'.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Savings Errors'

    def _side():
        return Side(style='thin', color='FF000000')

    thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    hdr_fill = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')
    red_font = Font(color='FFFF0000')
    hdr_font = Font(bold=True, color='FFFFFFFF')

    display_headers = ['ID Number', 'Employee Name', 'Savings Type', 'Savings', 'Remarks']
    col_widths      = [14, 28, 24, 14, 60]

    for col, (h, w) in enumerate(zip(display_headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = thin
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = w

    # Build error map: excel_row_num → combined reason string
    error_map: dict[int, str] = {}
    for f in failures:
        r = f.get('row')
        if r is not None:
            reason = f.get('reason', '')
            error_map[r] = (error_map[r] + '; ' + reason) if r in error_map else reason

    for i, row in enumerate(original_rows[1:], start=2):
        while len(row) < 4:
            row.append(None)

        has_error = i in error_map
        for col, val in enumerate(row[:4], 1):
            cell = ws.cell(row=i, column=col, value=str(val) if val is not None else '')
            cell.border = thin
            if has_error:
                cell.font = red_font

        remarks_cell = ws.cell(row=i, column=5, value=error_map.get(i, ''))
        remarks_cell.border = thin
        if has_error:
            remarks_cell.font = red_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def _build_loan_error_excel(original_rows: list[list], failures: list[dict]) -> str:
    """
    Build a loan-specific error report that preserves the original file structure.
    Output columns: ID Number | Employee Name | Loan Type | Principal Balance |
                    Monthly Deduction | Remarks
    Red header row; erroneous rows rendered in red font.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Loan Errors'

    def _side():
        return Side(style='thin', color='FF000000')

    thin = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    hdr_fill = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')
    red_font = Font(color='FFFF0000')
    hdr_font = Font(bold=True, color='FFFFFFFF')

    display_headers = ['ID Number', 'Employee Name', 'Loan Type', 'Principal Balance', 'Monthly Deduction', 'Remarks']
    col_widths = [14, 28, 24, 20, 20, 60]

    for col, (h, w) in enumerate(zip(display_headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = w

    error_map: dict[int, str] = {}
    for f in failures:
        r = f.get('row')
        if r is not None:
            reason = f.get('reason', '')
            error_map[r] = (error_map[r] + '; ' + reason) if r in error_map else reason

    for i, row in enumerate(original_rows[1:], start=2):
        while len(row) < 5:
            row.append(None)

        has_error = i in error_map
        for col, val in enumerate(row[:5], 1):
            cell = ws.cell(row=i, column=col, value=str(val) if val is not None else '')
            cell.border = thin
            if has_error:
                cell.font = red_font

        remarks_cell = ws.cell(row=i, column=6, value=error_map.get(i, ''))
        remarks_cell.border = thin
        if has_error:
            remarks_cell.font = red_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _parse_decimal(value, field: str, row: int) -> tuple[Decimal | None, dict | None]:
    """Parse a cell value as a positive Decimal. Returns (decimal, None) or (None, error_dict).
    Comma-formatted values (e.g. '1,500.00') are accepted."""
    try:
        d = Decimal(str(value).replace(',', '')).quantize(Decimal('0.01'))
        if d <= 0:
            return None, {'row': row, 'field': field, 'reason': f'{field} must be greater than 0.'}
        return d, None
    except (InvalidOperation, TypeError, ValueError):
        return None, {'row': row, 'field': field, 'reason': f'{field} is not a valid number.'}


def _parse_non_negative_decimal(value, field: str, row: int) -> tuple[Decimal | None, dict | None]:
    """Parse a cell value as a non-negative Decimal (>=0). Comma-formatted values accepted."""
    try:
        d = Decimal(str(value).replace(',', '')).quantize(Decimal('0.01'))
        if d < 0:
            return None, {'row': row, 'field': field, 'reason': f'{field} must be 0 or greater.'}
        return d, None
    except (InvalidOperation, TypeError, ValueError):
        return None, {'row': row, 'field': field, 'reason': f'{field} is not a valid number.'}


def _lookup_employee(idnumber: str, row: int):
    """Return (employee, None) or (None, error_dict). Rejects privileged users."""
    from userLogin.models import loginCredentials
    try:
        emp = loginCredentials.objects.get(idnumber=idnumber)
    except loginCredentials.DoesNotExist:
        return None, {'row': row, 'field': 'idnumber', 'reason': f'No employee found with ID "{idnumber}".'}
    if emp.admin or emp.accounting or emp.hr:
        return None, {'row': row, 'field': 'idnumber', 'reason': f'Employee "{idnumber}" is a privileged user (admin/accounting/hr) and cannot have finance records.'}
    return emp, None


def _read_xlsx(file_obj) -> tuple[list[list], str | None]:
    """Open an uploaded xlsx file and return (rows_of_values, error_message_or_None)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows, None
    except Exception as exc:
        return [], f'Could not read xlsx file: {exc}'


def _queue_finance_notifications(
    *,
    notification_type: str,
    title: str,
    module: str,
    user_messages: dict,
) -> None:
    """Create one notification per user after the surrounding transaction commits."""
    if not user_messages:
        return

    def _emit() -> None:
        Notification.objects.bulk_create([
            Notification(
                recipient=user,
                notification_scope='specific_user',
                notification_type=notification_type,
                title=title,
                message=message,
                module=module,
            )
            for user, message in user_messages.items()
        ])

    transaction.on_commit(_emit)


# ── OJT Payslip PDF builder ───────────────────────────────────────────────────

def _build_ojt_payslip_pdf(obj, user, line_name: str = '—') -> bytes:
    """Return raw PDF bytes for an OJT payslip, mirroring OJTPayslipViewModal."""
    from pathlib import Path as _Path
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as _rl_canvas
    from reportlab.lib.colors import HexColor

    buf = io.BytesIO()
    W, H = A4  # 595.28 x 841.89 pt
    c = _rl_canvas.Canvas(buf, pagesize=A4)

    # ── Try to register a Unicode-capable font (supports ₱) ──────────────────
    body_font = 'Helvetica'
    bold_font = 'Helvetica-Bold'
    peso = 'PHP '
    try:
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase import pdfmetrics
        _font_pairs = [
            (r'C:\Windows\Fonts\arial.ttf',   r'C:\Windows\Fonts\arialbd.ttf'),
            (r'C:\Windows\Fonts\calibri.ttf', r'C:\Windows\Fonts\calibrib.ttf'),
        ]
        for _reg, _reg_bold in _font_pairs:
            import os as _os
            if _os.path.exists(_reg):
                pdfmetrics.registerFont(TTFont('_OJTBody', _reg))
                body_font = '_OJTBody'
                peso = '₱'  # ₱
                if _os.path.exists(_reg_bold):
                    pdfmetrics.registerFont(TTFont('_OJTBold', _reg_bold))
                    bold_font = '_OJTBold'
                break
    except Exception:
        pass

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _n(v):
        try:
            return f"{float(v):,.2f}"
        except Exception:
            return '0.00'

    def _fmt_period(start, end):
        if not start or not end:
            return '—'
        MONTHS = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                  'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        sm = MONTHS[start.month - 1]
        em = MONTHS[end.month - 1]
        sd, ed, sy, ey = start.day, end.day, start.year, end.year
        if start.month == end.month and sy == ey:
            return f"{sm} {sd}-{ed},{sy}"
        if sy == ey:
            return f"{sm} {sd} - {em} {ed},{sy}"
        return f"{sm} {sd},{sy} - {em} {ed},{ey}"

    # ── Employee data ─────────────────────────────────────────────────────────
    lastname  = (getattr(user, 'lastname',  '') or '').strip()
    firstname = (getattr(user, 'firstname', '') or '').strip()
    if lastname and firstname:
        full_name = f"{lastname}, {firstname}".upper()
    elif lastname or firstname:
        full_name = (lastname or firstname).upper()
    else:
        full_name = (getattr(user, 'idnumber', '') or str(user.pk)).upper()

    id_number = getattr(user, 'idnumber', '') or ''
    period    = _fmt_period(obj.period_start, obj.period_end)

    # ── Layout constants ──────────────────────────────────────────────────────
    ML        = 40          # left/right margin
    MT        = 30          # top margin
    CW        = W - 2 * ML  # content width = 515.28 pt
    blue      = HexColor('#1a3db5')
    gray_bg   = HexColor('#f0f0f0')
    dark_gray = HexColor('#555555')
    border_c  = HexColor('#aaaaaa')
    cell_c    = HexColor('#cccccc')
    ROW_H     = 14
    HDR_H     = 16

    y = H - MT  # current y-cursor (descends as content is drawn)

    # ── Company header (logo + name + address) ────────────────────────────────
    LOGO_SIZE = 52
    logo_path = _Path(__file__).resolve().parent.parent.parent.parent / 'public' / 'ryonanlogo.png'
    if logo_path.exists():
        try:
            c.drawImage(
                str(logo_path), ML, y - LOGO_SIZE,
                width=LOGO_SIZE, height=LOGO_SIZE, mask='auto',
            )
        except Exception:
            pass

    text_cx = (ML + LOGO_SIZE + 6 + W - ML) / 2  # center between logo-right and right margin
    c.setFont(bold_font, 11)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(text_cx, y - 16, 'RYONAN ELECTRIC PHILIPPINES CORPORATION')
    c.setFont(body_font, 8)
    c.setFillColor(dark_gray)
    c.drawCentredString(text_cx, y - 29, '105 East Main Avenue, Special Export Processing Zone')
    c.drawCentredString(text_cx, y - 39, 'Laguna, Technopark, Binan, Laguna')

    y -= LOGO_SIZE + 8

    # ── Horizontal rule ───────────────────────────────────────────────────────
    c.setStrokeColor(border_c)
    c.line(ML, y, W - ML, y)
    y -= 4

    # ── Employee info (2-column grid) ─────────────────────────────────────────
    LBL_SZ = 8
    VAL_SZ = 9
    GAP    = 13
    col1_x = ML
    col2_x = ML + CW / 2

    for (lbl1, val1), (lbl2, val2) in [
        (('ID Number:', id_number), ('Period Covered:', period)),
        (('Employee Name:', full_name), ('Line:', line_name or '—')),
    ]:
        y -= GAP
        c.setFont(body_font, LBL_SZ)
        c.setFillColor(dark_gray)
        c.drawString(col1_x, y, lbl1)
        c.drawString(col2_x, y, lbl2)
        y -= GAP
        c.setFont(bold_font, VAL_SZ)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(col1_x, y, val1)
        c.drawString(col2_x, y, val2)
        y -= 4

    y -= 6

    # ── Two-column tables ─────────────────────────────────────────────────────
    TBL_GAP = 8
    TW      = (CW - TBL_GAP) / 2
    left_x  = ML
    right_x = ML + TW + TBL_GAP
    LBL_W   = TW * 0.62
    VAL_W   = TW - LBL_W
    PAD     = 3

    def _draw_table(tx, title, rows, start_y):
        cy = start_y
        # Header
        c.setFillColor(gray_bg)
        c.rect(tx, cy - HDR_H, TW, HDR_H, fill=1, stroke=0)
        c.setStrokeColor(border_c)
        c.rect(tx, cy - HDR_H, TW, HDR_H, fill=0, stroke=1)
        c.setFont(bold_font, 9)
        c.setFillColorRGB(0, 0, 0)
        c.drawCentredString(tx + TW / 2, cy - HDR_H + 5, title)
        cy -= HDR_H

        for label, val, is_bold in rows:
            fn = bold_font if is_bold else body_font
            c.setStrokeColor(cell_c)
            # Label cell
            c.rect(tx, cy - ROW_H, LBL_W, ROW_H, fill=0, stroke=1)
            c.setFont(fn, 8)
            c.setFillColor(blue)
            c.drawString(tx + PAD, cy - ROW_H + 4, label)
            # Value cell
            c.rect(tx + LBL_W, cy - ROW_H, VAL_W, ROW_H, fill=0, stroke=1)
            c.setFillColorRGB(0, 0, 0)
            c.drawRightString(tx + TW - PAD, cy - ROW_H + 4, val)
            cy -= ROW_H

        return cy  # bottom edge of table

    left_rows = [
        ('REGULAR # of Days Work',     _n(obj.regular_day),        False),
        ('ALLOWANCE/DAY',              _n(obj.allowance_day),       False),
        ('Total:',                     _n(obj.total_allowance),     True),
        ('REG ND ALLOWANCE',           _n(obj.nd_allowance),        False),
        ('GRAND TOTAL',                _n(obj.grand_total),         False),
        ('BASIC ALLOW.SCHOOL SHARE',   _n(obj.basic_school_share),  False),
        ('BASIC ALLOW. OJT SHARE',     _n(obj.basic_ojt_share),     False),
        ('DEDUCTION',                  _n(obj.deduction),           False),
        ('NET BASIC ALLOW. OJT SHARE', _n(obj.net_ojt_share),       True),
    ]
    right_rows = [
        ('RICE ALLOWANCE',             _n(obj.rice_allowance),      False),
        ('Reg OT ALLOWANCE',           _n(obj.ot_allowance),        False),
        ('REG ND OT ALLOWANCE',        _n(obj.nd_ot_allowance),     False),
        ('SPECIAL HOLIDAY',            _n(obj.special_holiday),     False),
        ('LEGAL HOLIDAY',              _n(obj.legal_holiday),       False),
        ('SAT-OFF ALLOWANCE',          _n(obj.satoff_allowance),    False),
        ('RD OT',                      _n(obj.rd_ot),               False),
        ('PERFECT ATTENDANCE',         _n(obj.perfect_attendance),  False),
        ('ADJUSTMENT',                 _n(obj.adjustment),          False),
        ('DEDUCTION 2',                _n(obj.deduction_2),         False),
        ('NET OJT OT PAY ALLOWANCE',   _n(obj.ot_pay_allowance),    True),
    ]

    tbl_start = y
    left_bot  = _draw_table(left_x,  'Regular Day', left_rows,  tbl_start)
    right_bot = _draw_table(right_x, 'Allowances',  right_rows, tbl_start)

    y = min(left_bot, right_bot) - 10

    # ── Total Allowance box ───────────────────────────────────────────────────
    BOX_H = 24
    c.setStrokeColor(border_c)
    c.rect(ML, y - BOX_H, CW, BOX_H, fill=0, stroke=1)
    c.setFont(bold_font, 11)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(W / 2, y - BOX_H + 8, f"TOTAL ALLOWANCE:  {peso}{_n(obj.total_allow)}")

    c.save()
    buf.seek(0)
    return buf.read()


# ── Views ─────────────────────────────────────────────────────────────────────

class FinanceTypeListView(APIView):
    """GET /api/finance/admin/types — all 4 type lists."""
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err
        return Response({
            'allowance_types': AllowanceTypeSerializer(AllowanceType.objects.all(), many=True).data,
            'loan_types':      LoanTypeSerializer(LoanType.objects.all(), many=True).data,
            'savings_types':   SavingsTypeSerializer(SavingsType.objects.all(), many=True).data,
            'payslip_types':   PayslipTypeSerializer(PayslipType.objects.all(), many=True).data,
        })


class FinanceChartView(APIView):
    """
    GET /api/finance/admin/chart
    Params: view (fiscal|monthly|weekly), year, month, week_start (ISO date)
    Returns counts of Loan, Allowance, Savings records per time bucket.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        view_mode = request.GET.get('view', 'monthly')
        today = datetime.date.today()

        if view_mode == 'fiscal':
            year = int(request.GET.get('year', today.year))
            # Fiscal year: July → June
            fy_start = datetime.date(year, 7, 1)
            months = []
            for i in range(12):
                m = (fy_start.month - 1 + i) % 12 + 1
                y = fy_start.year + ((fy_start.month - 1 + i) // 12)
                months.append((y, m))
            data = []
            for (y, m) in months:
                label = datetime.date(y, m, 1).strftime('%b %Y')
                loans_c = Loan.objects.filter(created_at__year=y, created_at__month=m).count()
                allow_c = Allowance.objects.filter(created_at__year=y, created_at__month=m).count()
                sav_c   = Savings.objects.filter(created_at__year=y, created_at__month=m).count()
                data.append({'label': label, 'loans': loans_c, 'allowances': allow_c, 'savings': sav_c})
            return Response({'view': view_mode, 'fy_start': str(fy_start), 'data': data})

        elif view_mode == 'weekly':
            try:
                week_start = datetime.date.fromisoformat(request.GET.get('week_start', str(today)))
            except (ValueError, TypeError):
                week_start = today
            # Align to Monday of the given week
            week_start = week_start - datetime.timedelta(days=week_start.weekday())
            data = []
            for i in range(7):
                day = week_start + datetime.timedelta(days=i)
                label = day.strftime('%a %d')
                loans_c = Loan.objects.filter(created_at__date=day).count()
                allow_c = Allowance.objects.filter(created_at__date=day).count()
                sav_c   = Savings.objects.filter(created_at__date=day).count()
                data.append({'label': label, 'loans': loans_c, 'allowances': allow_c, 'savings': sav_c})
            return Response({'view': view_mode, 'week_start': str(week_start), 'data': data})

        else:  # monthly (default)
            year = int(request.GET.get('year', today.year))
            month = int(request.GET.get('month', today.month))
            _, days_in_month = calendar.monthrange(year, month)
            data = []
            for day_num in range(1, days_in_month + 1):
                day = datetime.date(year, month, day_num)
                label = str(day_num)
                loans_c = Loan.objects.filter(created_at__date=day).count()
                allow_c = Allowance.objects.filter(created_at__date=day).count()
                sav_c   = Savings.objects.filter(created_at__date=day).count()
                data.append({'label': label, 'loans': loans_c, 'allowances': allow_c, 'savings': sav_c})
            return Response({'view': view_mode, 'year': year, 'month': month, 'data': data})


class FinanceEmployeeListView(APIView):
    """
    GET /api/finance/admin/employees
    Paginated list of non-privileged employees annotated with finance counts/totals.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        from userLogin.models import loginCredentials
        from userProfile.models import workInformation

        qs = loginCredentials.objects.filter(
            admin=False, accounting=False, hr=False,
        ).prefetch_related(
            Prefetch(
                'workinformation_set',
                queryset=workInformation.objects.select_related('department', 'line').order_by('-created_at'),
            )
        )

        # Annotate with finance counts and totals
        qs = qs.annotate(
            loans_count=Count('loans', distinct=True),
            loans_total=Sum('loans__principal_amount'),
            loans_balance=Sum('loans__current_balance'),
            allowances_count=Count('allowances', distinct=True),
            allowances_total=Sum('allowances__amount'),
            savings_count=Count('savings', distinct=True),
            savings_total=Sum('savings__amount'),
            deductions_count=Count('deductions', distinct=True),
            deductions_total=Sum('deductions__amount'),
            payslips_count=Count('payslips', distinct=True),
        )

        # Search
        search_q = request.GET.get('search', '').strip()
        if search_q:
            qs = qs.filter(
                Q(idnumber__icontains=search_q) |
                Q(firstname__icontains=search_q) |
                Q(lastname__icontains=search_q)
            )

        # Column filters (single-select for dept/line, multi for idnumbers)
        department_id_raw = request.GET.get('department_id', '').strip()
        if department_id_raw and department_id_raw.isdigit():
            qs = qs.filter(Exists(
                workInformation.objects.filter(employee=OuterRef('pk'), department_id=int(department_id_raw))
            ))

        line_id_raw = request.GET.get('line_id', '').strip()
        if line_id_raw and line_id_raw.isdigit():
            qs = qs.filter(Exists(
                workInformation.objects.filter(employee=OuterRef('pk'), line_id=int(line_id_raw))
            ))

        idnumbers_raw = request.GET.get('idnumbers', '').strip()
        if idnumbers_raw:
            idnumber_list = [x.strip() for x in idnumbers_raw.split(',') if x.strip()]
            if idnumber_list:
                qs = qs.filter(idnumber__in=idnumber_list)

        # Sort
        sort_by  = request.GET.get('sort_by',  'lastname')
        sort_dir = request.GET.get('sort_dir', 'asc')
        allowed_sorts = {
            'idnumber', 'firstname', 'lastname',
            'department', 'line',
            'loans_count', 'loans_total', 'loans_balance',
            'allowances_count', 'allowances_total',
            'savings_count', 'savings_total',
            'deductions_count', 'deductions_total',
            'payslips_count',
        }
        if sort_by not in allowed_sorts:
            sort_by = 'lastname'
        order_prefix = '' if sort_dir == 'asc' else '-'
        if sort_by == 'department':
            qs = qs.order_by(f'{order_prefix}workinformation__department__name')
        elif sort_by == 'line':
            qs = qs.order_by(f'{order_prefix}workinformation__line__name')
        else:
            qs = qs.order_by(f'{order_prefix}{sort_by}')

        # Pagination
        try:
            page = max(1, int(request.GET.get('page', 1)))
        except (ValueError, TypeError):
            page = 1
        page_size = 10
        total = qs.count()
        start = (page - 1) * page_size
        qs_page = list(qs[start: start + page_size])

        serializer = FinanceEmployeeRowSerializer(qs_page, many=True)
        return Response({
            'results':     serializer.data,
            'count':       total,
            'page':        page,
            'page_size':   page_size,
            'total_pages': max(1, -(-total // page_size)),
        })


class FinanceEmployeeDetailView(APIView):
    """
    GET /api/finance/admin/employees/<idnumber>/records
    Returns all finance records (loans, allowances, savings, payslips) for a
    specific non-privileged employee.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, idnumber: str) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        from userLogin.models import loginCredentials
        try:
            emp = loginCredentials.objects.get(idnumber=idnumber)
        except loginCredentials.DoesNotExist:
            return Response({'detail': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        if emp.admin or emp.accounting or emp.hr:
            return Response({'detail': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        from .models import OJTPayslipData

        loans        = Loan.objects.filter(employee=emp).order_by('-created_at')
        allowances   = Allowance.objects.filter(employee=emp).order_by('-created_at')
        savings      = Savings.objects.filter(employee=emp).order_by('-created_at')
        payslips     = Payslip.objects.filter(employee=emp).order_by('-created_at')
        ojt_payslips = OJTPayslipData.objects.filter(employee=emp).order_by('-period_start', '-created_at')

        ojt_data = [
            {
                'id':                obj.pk,
                'period_start':      obj.period_start.isoformat() if obj.period_start else None,
                'period_end':        obj.period_end.isoformat()   if obj.period_end   else None,
                'regular_day':       str(obj.regular_day),
                'allowance_day':     str(obj.allowance_day),
                'total_allowance':   str(obj.total_allowance),
                'nd_allowance':      str(obj.nd_allowance),
                'grand_total':       str(obj.grand_total),
                'basic_school_share': str(obj.basic_school_share),
                'basic_ojt_share':   str(obj.basic_ojt_share),
                'deduction':         str(obj.deduction),
                'net_ojt_share':     str(obj.net_ojt_share),
                'rice_allowance':    str(obj.rice_allowance),
                'ot_allowance':      str(obj.ot_allowance),
                'nd_ot_allowance':   str(obj.nd_ot_allowance),
                'special_holiday':   str(obj.special_holiday),
                'legal_holiday':     str(obj.legal_holiday),
                'satoff_allowance':  str(obj.satoff_allowance),
                'rd_ot':             str(obj.rd_ot),
                'adjustment':        str(obj.adjustment),
                'deduction_2':       str(obj.deduction_2),
                'ot_pay_allowance':  str(obj.ot_pay_allowance),
                'total_allow':       str(obj.total_allow),
                'perfect_attendance': str(obj.perfect_attendance),
                'holiday_date':      obj.holiday_date,
                'rd_ot_date':        obj.rd_ot_date,
                'created_at':        obj.created_at.isoformat(),
            }
            for obj in ojt_payslips
        ]

        return Response({
            'loans':        LoanSerializer(loans, many=True).data,
            'allowances':   AllowanceSerializer(allowances, many=True).data,
            'savings':      SavingsSerializer(savings, many=True).data,
            'payslips':     PayslipSerializer(payslips, many=True, context={'request': request}).data,
            'ojt_payslips': ojt_data,
        })


class FinanceImportView(APIView):
    """
    POST /api/finance/admin/import?record_type=<type>
    Accepts multipart xlsx upload and imports records per record_type.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        record_type = request.GET.get('record_type', '').strip().lower()
        valid_types = {'allowance', 'loan', 'deduction', 'savings'}
        if record_type not in valid_types:
            return Response(
                {'detail': f'record_type must be one of: {", ".join(sorted(valid_types))}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        rows, read_err = _read_xlsx(file_obj)
        if read_err:
            return Response({'detail': read_err}, status=status.HTTP_400_BAD_REQUEST)

        if len(rows) < 2:
            return Response({'detail': 'File has no data rows (need at least one header + one data row).'}, status=status.HTTP_400_BAD_REQUEST)

        handler = {
            'allowance':  self._import_allowances,
            'loan':       self._import_loans,
            'deduction':  self._import_deductions,
            'savings':    self._import_savings,
        }[record_type]

        # cutoff_date is required for deduction imports; validated here before handler
        cutoff_date = None
        if record_type == 'deduction':
            import datetime
            cutoff_raw = request.POST.get('cutoff_date', '').strip()
            if not cutoff_raw:
                return Response({'detail': 'cutoff_date is required for deduction imports.'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                cutoff_date = datetime.date.fromisoformat(cutoff_raw)
            except (ValueError, TypeError):
                return Response({'detail': 'cutoff_date must be a valid ISO date (YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)

        if record_type == 'allowance':
            # _import_allowances returns (summary_dict, failures) instead of (int, failures)
            allowance_summary, failures = self._import_allowances(rows)
            imported = sum(allowance_summary.values()) if allowance_summary else 0
        elif record_type == 'deduction':
            imported, failures = self._import_deductions(rows, cutoff_date)
        else:
            imported, failures = handler(rows)
        if failures:
            if record_type == 'deduction':
                error_b64 = _build_deduction_error_excel(rows, failures)
            elif record_type == 'allowance':
                error_b64 = _build_allowance_error_excel(rows, failures)
            elif record_type == 'loan':
                error_b64 = _build_loan_error_excel(rows, failures)
            elif record_type == 'savings':
                error_b64 = _build_savings_error_excel(rows, failures)
            else:
                error_b64 = _build_error_excel(failures)
        else:
            error_b64 = None
        response_data: dict = {
            'imported':         imported,
            'failed':           len(failures),
            'error_report_b64': error_b64,
        }
        if record_type == 'allowance' and not failures:
            response_data['allowance_summary'] = allowance_summary
        return Response(response_data)

    # ── Allowance import ──────────────────────────────────────────────────────
    # Columns: idnumber*, employee_name[ref only], allowance_type*(name), amount*,
    #          deposited_date (MM/DD/YYYY, nullable), covered_period (nullable text)
    # All-or-nothing: entire file is pre-validated before any data is written.

    def _import_allowances(self, rows) -> tuple[int, list[dict]]:
        failures: list[dict] = []
        type_map = {t.name.lower(): t for t in AllowanceType.objects.all()}
        operations: list[tuple] = []  # (employee, a_type, amount, deposited_date, covered_period)

        # ── Phase 1: pre-validate every row ──────────────────────────────────
        for row_idx, row in enumerate(rows[1:], start=2):
            while len(row) < 6:
                row.append(None)

            idnumber       = str(row[0]).strip() if row[0] is not None else ''
            # row[1] = employee name — for reference only, not validated
            type_name      = str(row[2]).strip() if row[2] is not None else ''
            amount_raw     = row[3]
            dep_date_raw   = row[4]
            covered_period = str(row[5]).strip() if row[5] is not None else ''

            if not idnumber:
                failures.append({'row': row_idx, 'field': 'ID Number', 'reason': 'ID Number is required.'})
                continue

            employee, emp_err = _lookup_employee(idnumber, row_idx)
            if emp_err:
                failures.append(emp_err)
                continue

            a_type = type_map.get(type_name.lower())
            if not a_type:
                failures.append({'row': row_idx, 'field': 'Allowance Type', 'reason': f'Allowance type "{type_name}" not found.'})
                continue

            # Amount: non-negative for all; percentage types may be 0
            amount, amt_err = _parse_non_negative_decimal(amount_raw, 'Amount', row_idx)
            if amt_err:
                failures.append(amt_err)
                continue

            # Deposited Date: MM/DD/YYYY, nullable
            # openpyxl returns date-formatted cells as Python datetime/date objects,
            # not strings.  Accept all three forms: native date, native datetime,
            # and string (MM/DD/YYYY or ISO YYYY-MM-DD as fallback).
            deposited_date = None
            if dep_date_raw is not None:
                if isinstance(dep_date_raw, datetime.datetime):
                    deposited_date = dep_date_raw.date()
                elif isinstance(dep_date_raw, datetime.date):
                    deposited_date = dep_date_raw
                else:
                    dep_str = str(dep_date_raw).strip()
                    if dep_str and dep_str.lower() not in ('none', 'n/a', ''):
                        parsed = False
                        # Primary: MM/DD/YYYY (user-visible format)
                        try:
                            deposited_date = datetime.datetime.strptime(dep_str, '%m/%d/%Y').date()
                            parsed = True
                        except ValueError:
                            pass
                        # Fallback: YYYY-MM-DD (ISO / openpyxl text cells)
                        if not parsed:
                            try:
                                deposited_date = datetime.date.fromisoformat(dep_str[:10])
                                parsed = True
                            except ValueError:
                                pass
                        if not parsed:
                            failures.append({
                                'row': row_idx,
                                'field': 'Deposited Date',
                                'reason': (
                                    f'Deposited Date must be in MM/DD/YYYY format '
                                    f'(e.g. 04/30/2026). Got: "{dep_str}".'
                                ),
                            })
                            continue

            operations.append((employee, a_type, amount, deposited_date, covered_period))

        # ── Phase 2: reject entire batch if any row failed ────────────────────
        if failures:
            return {}, failures

        # ── Phase 3: apply all allowances in a single atomic transaction ──────
        #
        # GLOBAL PRE-DELETION — scoped to affected employees, covers ALL their
        # allowance records that satisfy either rule (not limited to the types
        # present in the uploaded file):
        #
        #   Rule 1: allowance_type.replace_on_upload = True
        #           → delete every record for those employees where the type has
        #             replace_on_upload=True, regardless of which types are uploaded.
        #
        #   Rule 2: allowance_type.replace_on_upload = False AND deposited_date IS NOT NULL
        #           → delete every dated transaction record for those employees,
        #             regardless of which types are uploaded.
        #             Null-deposited balance rows always survive.

        # Collect deduped employee PKs from the validated operations
        affected_employee_pks: set[int] = {emp.pk for emp, *_ in operations}

        summary = {'updated': 0, 'added': 0, 'replaced': 0}

        with transaction.atomic():
            # ── Global pre-deletion ────────────────────────────────────────
            if affected_employee_pks:
                emp_ids = list(affected_employee_pks)

                # Rule 1: delete ALL records for affected employees whose type
                #         has replace_on_upload=True (any deposited_date value)
                Allowance.objects.filter(
                    employee_id__in=emp_ids,
                    allowance_type__replace_on_upload=True,
                ).delete()

                # Rule 2: delete all DATED records for affected employees whose
                #         type has replace_on_upload=False
                Allowance.objects.filter(
                    employee_id__in=emp_ids,
                    allowance_type__replace_on_upload=False,
                    deposited_date__isnull=False,
                ).delete()

            # ── Inserts / accumulations ────────────────────────────────────
            for employee, a_type, amount, deposited_date, covered_period in operations:

                if a_type.replace_on_upload or deposited_date is not None:
                    # replace=True (any date), OR replace=False + deposited_date present:
                    # pre-deletion already cleared the slate — insert as new record.
                    Allowance.objects.create(
                        employee=employee,
                        allowance_type=a_type,
                        amount=amount,
                        deposited_date=deposited_date,
                        covered_period=covered_period,
                    )
                    summary['replaced'] += 1

                else:
                    # replace=False, no deposited_date:
                    # pre-deletion cleared dated rows — accumulate onto balance record.
                    existing = Allowance.objects.filter(
                        employee=employee,
                        allowance_type=a_type,
                        deposited_date__isnull=True,
                    ).first()
                    if existing:
                        existing.amount = existing.amount + amount
                        if covered_period:
                            existing.covered_period = covered_period
                        existing.save(update_fields=['amount', 'covered_period'])
                        summary['updated'] += 1
                    else:
                        Allowance.objects.create(
                            employee=employee,
                            allowance_type=a_type,
                            amount=amount,
                            deposited_date=None,
                            covered_period=covered_period,
                        )
                        summary['added'] += 1

        return summary, []

    # ── Loan import ───────────────────────────────────────────────────────────
    # Columns: idnumber*, employee_name[ignored], loan_type*(name), principal_balance*, monthly_deduction

    def _import_loans(self, rows) -> tuple[int, list[dict]]:
        imported = 0
        failures = []
        type_map = {t.name.lower(): t for t in LoanType.objects.all()}
        operations: list[tuple] = []  # (employee, loan_type, principal, monthly_deduction)
        user_loan_types: dict = {}

        # ── Phase 1: pre-validate every row ───────────────────────────────────
        for row_idx, row in enumerate(rows[1:], start=2):
            while len(row) < 5:
                row.append(None)

            idnumber         = str(row[0]).strip() if row[0] is not None else ''
            # row[1] = employee name — for reference only, not validated
            type_name        = str(row[2]).strip() if row[2] is not None else ''
            amount_raw       = row[3]
            monthly_ded_raw  = row[4]

            if not idnumber:
                failures.append({'row': row_idx, 'field': 'idnumber', 'reason': 'idnumber is required.'})
                continue

            employee, emp_err = _lookup_employee(idnumber, row_idx)
            if emp_err:
                failures.append(emp_err)
                continue

            l_type = type_map.get(type_name.lower())
            if not l_type:
                failures.append({'row': row_idx, 'field': 'loan_type', 'reason': f'Loan type "{type_name}" not found.'})
                continue

            principal, amt_err = _parse_decimal(amount_raw, 'principal_amount', row_idx)
            if amt_err:
                failures.append(amt_err)
                continue

            # Parse optional monthly deduction
            monthly_deduction = None
            if monthly_ded_raw is not None and str(monthly_ded_raw).strip() != '':
                monthly_deduction, md_err = _parse_decimal(monthly_ded_raw, 'monthly_deduction', row_idx)
                if md_err:
                    failures.append(md_err)
                    continue

            if not l_type.stackable:
                active_exists = Loan.objects.filter(
                    employee=employee,
                    loan_type=l_type,
                    current_balance__gt=Decimal('0'),
                ).exists()
                if active_exists:
                    failures.append({
                        'row': row_idx,
                        'field': 'loan_type',
                        'reason': (
                            f'Employee "{idnumber}" already has an active loan of type '
                            f'"{l_type.name}" (stackable=False). Settle the balance first.'
                        ),
                    })
                    continue

            operations.append((employee, l_type, principal, monthly_deduction))
            bucket = user_loan_types.setdefault(employee, set())
            bucket.add(l_type.name)

        # ── Phase 2: reject entire batch if any row failed ────────────────────
        if failures:
            return 0, failures

        # ── Phase 3: apply all loans in one atomic transaction ─────────────────
        with transaction.atomic():
            for employee, l_type, principal, monthly_deduction in operations:
                if l_type.stackable:
                    existing = Loan.objects.filter(
                        employee=employee,
                        loan_type=l_type,
                        current_balance__gt=Decimal('0'),
                    ).order_by('-created_at').first()

                    if existing is not None:
                        loan_to_update = Loan.objects.select_for_update().get(pk=existing.pk)
                        loan_to_update.principal_amount += principal
                        loan_to_update.current_balance += principal
                        loan_to_update.seen = False
                        if monthly_deduction is not None:
                            loan_to_update.monthly_deduction = monthly_deduction
                        loan_to_update.save(update_fields=['principal_amount', 'current_balance', 'monthly_deduction', 'seen', 'updated_at'])
                    else:
                        Loan.objects.create(
                            employee=employee,
                            loan_type=l_type,
                            principal_amount=principal,
                            current_balance=principal,
                            monthly_deduction=monthly_deduction,
                            seen=False,
                        )
                else:
                    Loan.objects.create(
                        employee=employee,
                        loan_type=l_type,
                        principal_amount=principal,
                        current_balance=principal,
                        monthly_deduction=monthly_deduction,
                        seen=False,
                    )
                imported += 1

            user_messages = {
                employee: (
                    'New loan record(s) were uploaded to your account: '
                    + ', '.join(sorted(type_names))
                    + '.'
                )
                for employee, type_names in user_loan_types.items()
                if type_names
            }
            _queue_finance_notifications(
                notification_type='finance_loan_uploaded',
                title='New Loan Record Uploaded',
                module='finance',
                user_messages=user_messages,
            )

        return imported, []

    # ── Deduction import ──────────────────────────────────────────────────────
    # Columns: idnumber*, employee_name[ignored], loan_type*(name), deduction_amount*
    # All-or-nothing: every row is validated first; if ANY row fails, the
    # entire batch is rejected and nothing is written to the database.

    def _import_deductions(self, rows, cutoff_date) -> tuple[int, list[dict]]:
        failures  = []
        type_map  = {t.name.lower(): t for t in LoanType.objects.all()}
        operations: list[tuple] = []  # (employee, loan_pk, amount, loan_type_name)
        user_loan_types: dict = {}

        # ── Phase 1: pre-validate every row ───────────────────────────────────
        for row_idx, row in enumerate(rows[1:], start=2):
            while len(row) < 4:
                row.append(None)

            idnumber   = str(row[0]).strip() if row[0] is not None else ''
            # row[1] = employee name — for reference only, not validated
            type_name  = str(row[2]).strip() if row[2] is not None else ''
            amount_raw = row[3]

            if not idnumber:
                failures.append({'row': row_idx, 'field': 'ID Number', 'reason': 'ID Number is required.'})
                continue

            employee, emp_err = _lookup_employee(idnumber, row_idx)
            if emp_err:
                failures.append(emp_err)
                continue

            l_type = type_map.get(type_name.lower())
            if not l_type:
                failures.append({'row': row_idx, 'field': 'Loan Type', 'reason': f'Loan type "{type_name}" not found.'})
                continue

            amount, amt_err = _parse_decimal(amount_raw, 'Deduction', row_idx)
            if amt_err:
                failures.append(amt_err)
                continue

            loan = Loan.objects.filter(
                employee=employee, loan_type=l_type, current_balance__gt=Decimal('0')
            ).order_by('-created_at').first()
            if loan is None:
                failures.append({
                    'row': row_idx,
                    'field': 'Loan Type',
                    'reason': f'No active loan of type "{l_type.name}" found for employee "{idnumber}".',
                })
                continue

            if amount > loan.current_balance:
                failures.append({
                    'row': row_idx,
                    'field': 'Deduction',
                    'reason': (
                        f'Deduction of {amount} exceeds current balance of '
                        f'{loan.current_balance} for loan type "{l_type.name}".'
                    ),
                })
                continue

            operations.append((employee, loan.pk, amount, l_type.name))
            bucket = user_loan_types.setdefault(employee, set())
            bucket.add(l_type.name)

        # ── Phase 2: reject entire batch if any row failed ────────────────────
        if failures:
            return 0, failures

        # ── Phase 3: apply all deductions in a single atomic transaction ──────
        imported = 0
        with transaction.atomic():
            for employee, loan_pk, amount, _loan_type_name in operations:
                loan_obj = Loan.objects.select_for_update().get(pk=loan_pk)
                loan_obj.current_balance -= amount
                loan_obj.save(update_fields=['current_balance', 'updated_at'])
                Deduction.objects.create(
                    employee=employee,
                    loan=loan_obj,
                    amount=amount,
                    description='',
                    cutoff_date=cutoff_date,
                )
                imported += 1

            user_messages = {
                employee: (
                    'New loan deduction(s) were uploaded for: '
                    + ', '.join(sorted(type_names))
                    + '.'
                )
                for employee, type_names in user_loan_types.items()
                if type_names
            }
            _queue_finance_notifications(
                notification_type='finance_deduction_uploaded',
                title='Loan Deduction Uploaded',
                module='finance',
                user_messages=user_messages,
            )

        return imported, []

    # ── Savings import ────────────────────────────────────────────────────────
    # Columns: idnumber*, employee_name[ref only], savings_type*(name), savings*
    # All-or-nothing: entire file is pre-validated before any data is written.
    # withdraw=False (default): amount is accumulated onto the employee's existing
    #   balance row for the same savings type (created if absent).
    # withdraw=True records are standalone transactions set through other means.

    def _import_savings(self, rows) -> tuple[int, list[dict]]:
        failures: list[dict] = []
        type_map  = {t.name.lower(): t for t in SavingsType.objects.all()}
        operations: list[tuple] = []  # (employee, s_type, amount)

        # ── Phase 1: pre-validate every row ──────────────────────────────────
        for row_idx, row in enumerate(rows[1:], start=2):
            while len(row) < 4:
                row.append(None)

            idnumber   = str(row[0]).strip() if row[0] is not None else ''
            # row[1] = employee name — for reference only, not validated
            type_name  = str(row[2]).strip() if row[2] is not None else ''
            amount_raw = row[3]

            if not idnumber:
                failures.append({'row': row_idx, 'field': 'ID Number', 'reason': 'ID Number is required.'})
                continue

            employee, emp_err = _lookup_employee(idnumber, row_idx)
            if emp_err:
                failures.append(emp_err)
                continue

            if not type_name:
                failures.append({'row': row_idx, 'field': 'Savings Type', 'reason': 'Savings Type is required.'})
                continue

            s_type = type_map.get(type_name.lower())
            if not s_type:
                failures.append({'row': row_idx, 'field': 'Savings Type', 'reason': f'Savings type "{type_name}" not found.'})
                continue

            amount, amt_err = _parse_non_negative_decimal(amount_raw, 'Savings', row_idx)
            if amt_err:
                failures.append(amt_err)
                continue

            operations.append((employee, s_type, amount))

        # ── Phase 2: reject entire batch if any row failed ────────────────────
        if failures:
            return 0, failures

        # ── Phase 3: apply all savings in a single atomic transaction ─────────
        imported = 0
        with transaction.atomic():
            for employee, s_type, amount in operations:
                # Accumulate onto the existing withdraw=False balance row, or create one.
                existing = Savings.objects.filter(
                    employee=employee,
                    savings_type=s_type,
                    withdraw=False,
                ).first()
                if existing:
                    locked = Savings.objects.select_for_update().get(pk=existing.pk)
                    locked.amount += amount
                    locked.save(update_fields=['amount'])
                else:
                    Savings.objects.create(
                        employee=employee,
                        savings_type=s_type,
                        amount=amount,
                        withdraw=False,
                    )
                imported += 1

        return imported, []



class FinancePayslipUploadView(APIView):
    """
    POST /api/finance/admin/payslip-upload
    Multipart fields: idnumber, payslip_type (name), period_start (ISO),
    period_end (ISO), description (optional), file (PDF ≤ 5 MB).
    """
    permission_classes = [IsAuthenticated]

    def post(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        idnumber    = request.data.get('idnumber', '').strip()
        type_name   = request.data.get('payslip_type', '').strip()
        start_raw   = request.data.get('period_start', '').strip()
        end_raw     = request.data.get('period_end', '').strip()
        description = request.data.get('description', '').strip()
        file_obj    = request.FILES.get('file')

        errors: dict[str, str] = {}
        if not idnumber:
            errors['idnumber'] = 'idnumber is required.'
        if not type_name:
            errors['payslip_type'] = 'payslip_type is required.'
        if not start_raw:
            errors['period_start'] = 'period_start is required.'
        if not end_raw:
            errors['period_end'] = 'period_end is required.'
        if not file_obj:
            errors['file'] = 'A PDF file is required.'
        elif not file_obj.name.lower().endswith('.pdf'):
            errors['file'] = 'Only PDF files are accepted.'
        elif file_obj.size > 5 * 1024 * 1024:
            errors['file'] = 'File must not exceed 5 MB.'

        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        employee, emp_err = _lookup_employee(idnumber, 0)
        if emp_err:
            return Response({'idnumber': emp_err['reason']}, status=status.HTTP_400_BAD_REQUEST)

        try:
            p_type = PayslipType.objects.get(name__iexact=type_name)
        except PayslipType.DoesNotExist:
            return Response(
                {'payslip_type': f'Payslip type "{type_name}" not found.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            period_start = datetime.date.fromisoformat(start_raw)
        except ValueError:
            return Response({'period_start': f'Invalid date: "{start_raw}".'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            period_end = datetime.date.fromisoformat(end_raw)
        except ValueError:
            return Response({'period_end': f'Invalid date: "{end_raw}".'}, status=status.HTTP_400_BAD_REQUEST)

        if period_end < period_start:
            return Response(
                {'period_end': 'period_end must not be before period_start.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Duplicate check — same employee + type + period
        if Payslip.objects.filter(
            employee=employee,
            payslip_type=p_type,
            period_start=period_start,
            period_end=period_end,
        ).exists():
            return Response(
                {
                    'detail': (
                        f'A payslip for this employee already exists with type '
                        f'"{p_type.name}" and period {period_start} to {period_end}. '
                        f'Upload skipped to prevent duplication.'
                    )
                },
                status=status.HTTP_409_CONFLICT,
            )

        payslip = Payslip.objects.create(
            employee=employee,
            payslip_type=p_type,
            period_start=period_start,
            period_end=period_end,
            file=file_obj,
            description=description,
        )
        from .serializers import PayslipSerializer
        return Response(
            PayslipSerializer(payslip, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
        )


class OfficeFinanceRateView(APIView):
    """
    GET  /api/finance/admin/office-rates      — list all offices with rates.
    GET  /api/finance/admin/office-rates/<pk> — rates for one office (by office pk).
    PUT  /api/finance/admin/office-rates/<pk> — create-or-update rates.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int | None = None) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        if pk is not None:
            try:
                rate = OfficeFinanceRate.objects.select_related('office').get(office_id=pk)
            except OfficeFinanceRate.DoesNotExist:
                return Response({'detail': 'No rate record for this office.'}, status=status.HTTP_404_NOT_FOUND)
            return Response(OfficeFinanceRateSerializer(rate).data)

        rates = OfficeFinanceRate.objects.select_related('office').all()
        return Response(OfficeFinanceRateSerializer(rates, many=True).data)

    def put(self, request, pk: int | None = None) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        if pk is None:
            return Response({'detail': 'Office pk is required for PUT.'}, status=status.HTTP_400_BAD_REQUEST)

        from generalsettings.models import Office
        try:
            office = Office.objects.get(pk=pk)
        except Office.DoesNotExist:
            return Response({'detail': 'Office not found.'}, status=status.HTTP_404_NOT_FOUND)

        rate, _ = OfficeFinanceRate.objects.get_or_create(office=office)
        serializer = OfficeFinanceRateSerializer(rate, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class FinanceExportView(APIView):
    """
    GET /api/finance/admin/export
    Params: record_type (allowance|loan|deduction|savings|payslip|all), date_from, date_to
    Returns an xlsx file.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> HttpResponse:
        err = _require_accounting_admin(request)
        if err:
            return err

        record_type = request.GET.get('record_type', 'all').strip().lower()
        valid_types = {'allowance', 'loan', 'deduction', 'savings', 'payslip', 'all'}
        if record_type not in valid_types:
            return Response(
                {'detail': f'record_type must be one of: {", ".join(sorted(valid_types))}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        today = datetime.date.today()
        try:
            date_from = datetime.date.fromisoformat(request.GET.get('date_from', str(today)))
        except (ValueError, TypeError):
            date_from = today
        try:
            date_to = datetime.date.fromisoformat(request.GET.get('date_to', str(today)))
        except (ValueError, TypeError):
            date_to = today
        if date_to < date_from:
            date_from, date_to = date_to, date_from

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        def _side():
            return Side(style='thin', color='FF000000')

        thin = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        hdr_fill = PatternFill(start_color='FF1E3A5F', end_color='FF1E3A5F', fill_type='solid')

        def _hdr_cell(ws, row, col, value):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = hdr_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='center')
            return cell

        def _data_cell(ws, row, col, value):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin
            return cell

        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        # Non-privileged employee IDs
        from userLogin.models import loginCredentials
        non_priv_ids = loginCredentials.objects.filter(
            admin=False, accounting=False, hr=False
        ).values_list('pk', flat=True)

        def _filter_date(qs, date_field='created_at'):
            return qs.filter(
                **{f'{date_field}__date__gte': date_from, f'{date_field}__date__lte': date_to}
            )

        sheets_to_build = (
            ['allowance', 'loan', 'deduction', 'savings', 'payslip']
            if record_type == 'all'
            else [record_type]
        )

        for sheet_type in sheets_to_build:
            ws = wb.create_sheet(title=sheet_type.capitalize())

            if sheet_type == 'allowance':
                headers = ['Employee ID', 'First Name', 'Last Name', 'Allowance Type', 'Amount', 'Description', 'Created At']
                for col, h in enumerate(headers, 1):
                    _hdr_cell(ws, 1, col, h)
                qs = _filter_date(
                    Allowance.objects.select_related('employee', 'allowance_type')
                    .filter(employee_id__in=non_priv_ids).order_by('created_at')
                )
                for r, obj in enumerate(qs, 2):
                    _data_cell(ws, r, 1, obj.employee.idnumber)
                    _data_cell(ws, r, 2, obj.employee.firstname)
                    _data_cell(ws, r, 3, obj.employee.lastname)
                    _data_cell(ws, r, 4, obj.allowance_type.name)
                    _data_cell(ws, r, 5, float(obj.amount))
                    _data_cell(ws, r, 6, obj.description)
                    _data_cell(ws, r, 7, obj.created_at.strftime('%B %d, %Y'))
                col_widths = [14, 16, 16, 22, 14, 30, 20]

            elif sheet_type == 'loan':
                headers = ['Employee ID', 'First Name', 'Last Name', 'Loan Type', 'Principal', 'Balance', 'Description', 'Reference No.', 'Created At']
                for col, h in enumerate(headers, 1):
                    _hdr_cell(ws, 1, col, h)
                qs = _filter_date(
                    Loan.objects.select_related('employee', 'loan_type')
                    .filter(employee_id__in=non_priv_ids).order_by('created_at')
                )
                for r, obj in enumerate(qs, 2):
                    _data_cell(ws, r, 1, obj.employee.idnumber)
                    _data_cell(ws, r, 2, obj.employee.firstname)
                    _data_cell(ws, r, 3, obj.employee.lastname)
                    _data_cell(ws, r, 4, obj.loan_type.name)
                    _data_cell(ws, r, 5, float(obj.principal_amount))
                    _data_cell(ws, r, 6, float(obj.current_balance))
                    _data_cell(ws, r, 7, obj.description)
                    _data_cell(ws, r, 8, obj.reference_number)
                    _data_cell(ws, r, 9, obj.created_at.strftime('%B %d, %Y'))
                col_widths = [14, 16, 16, 22, 14, 14, 30, 18, 20]

            elif sheet_type == 'deduction':
                headers = ['Employee ID', 'First Name', 'Last Name', 'Loan Type', 'Amount', 'Description', 'Created At']
                for col, h in enumerate(headers, 1):
                    _hdr_cell(ws, 1, col, h)
                qs = _filter_date(
                    Deduction.objects.select_related('employee', 'loan', 'loan__loan_type')
                    .filter(employee_id__in=non_priv_ids).order_by('created_at')
                )
                for r, obj in enumerate(qs, 2):
                    _data_cell(ws, r, 1, obj.employee.idnumber)
                    _data_cell(ws, r, 2, obj.employee.firstname)
                    _data_cell(ws, r, 3, obj.employee.lastname)
                    _data_cell(ws, r, 4, obj.loan.loan_type.name)
                    _data_cell(ws, r, 5, float(obj.amount))
                    _data_cell(ws, r, 6, obj.description)
                    _data_cell(ws, r, 7, obj.created_at.strftime('%B %d, %Y'))
                col_widths = [14, 16, 16, 22, 14, 30, 20]

            elif sheet_type == 'savings':
                headers = ['Employee ID', 'First Name', 'Last Name', 'Savings Type', 'Amount', 'Description', 'Created At']
                for col, h in enumerate(headers, 1):
                    _hdr_cell(ws, 1, col, h)
                qs = _filter_date(
                    Savings.objects.select_related('employee', 'savings_type')
                    .filter(employee_id__in=non_priv_ids).order_by('created_at')
                )
                for r, obj in enumerate(qs, 2):
                    _data_cell(ws, r, 1, obj.employee.idnumber)
                    _data_cell(ws, r, 2, obj.employee.firstname)
                    _data_cell(ws, r, 3, obj.employee.lastname)
                    _data_cell(ws, r, 4, obj.savings_type.name)
                    _data_cell(ws, r, 5, float(obj.amount))
                    _data_cell(ws, r, 6, obj.description)
                    _data_cell(ws, r, 7, obj.created_at.strftime('%B %d, %Y'))
                col_widths = [14, 16, 16, 22, 14, 30, 20]

            else:  # payslip
                headers = ['Employee ID', 'First Name', 'Last Name', 'Payslip Type', 'Period Start', 'Period End', 'File URL', 'Description', 'Created At']
                for col, h in enumerate(headers, 1):
                    _hdr_cell(ws, 1, col, h)
                qs = (
                    Payslip.objects.select_related('employee', 'payslip_type')
                    .filter(
                        employee_id__in=non_priv_ids,
                        period_start__gte=date_from,
                        period_start__lte=date_to,
                    ).order_by('period_start')
                )
                base_url = request.build_absolute_uri('/')[:-1]
                for r, obj in enumerate(qs, 2):
                    file_url = f'{base_url}{obj.file.url}' if obj.file else ''
                    _data_cell(ws, r, 1, obj.employee.idnumber)
                    _data_cell(ws, r, 2, obj.employee.firstname)
                    _data_cell(ws, r, 3, obj.employee.lastname)
                    _data_cell(ws, r, 4, obj.payslip_type.name)
                    _data_cell(ws, r, 5, str(obj.period_start))
                    _data_cell(ws, r, 6, str(obj.period_end))
                    _data_cell(ws, r, 7, file_url)
                    _data_cell(ws, r, 8, obj.description)
                    _data_cell(ws, r, 9, obj.created_at.strftime('%B %d, %Y'))
                col_widths = [14, 16, 16, 22, 14, 14, 50, 30, 20]

            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'finance_{record_type}_{date_from}_{date_to}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class FinancePrincipalBalanceTemplateView(APIView):
    """
    GET /api/finance/admin/template/principal-balance
    Returns a clean xlsx template for bulk principal-balance import.
    Column C (Loan Type) contains a dropdown data-validation listing all
    active LoanType names so users can pick from a list in Excel.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> HttpResponse:
        err = _require_accounting_admin(request)
        if err:
            return err

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        loan_type_names = list(LoanType.objects.values_list('name', flat=True))

        def _side():
            return Side(style='thin', color='FF000000')

        thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        hdr_fill = PatternFill(start_color='FF2845D6', end_color='FF2845D6', fill_type='solid')

        headers    = ['ID Number', 'Employee Name', 'Loan Type', 'Principal Balance', 'Monthly Deduction']
        col_widths = [14, 28, 24, 22, 22]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Principal Balance'

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill      = hdr_fill
            cell.border    = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 18

        # Column C: dropdown of all active loan types
        if loan_type_names:
            formula = '"' + ','.join(loan_type_names) + '"'
            dv = DataValidation(
                type='list',
                formula1=formula,
                sqref='C2:C5000',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error      = 'Please select a valid loan type from the dropdown list.'
            dv.errorTitle = 'Invalid Loan Type'
            dv.prompt     = 'Select a loan type from the list.'
            dv.promptTitle = 'Loan Type'
            ws.add_data_validation(dv)

        # Example data row (Row 2) — italic grey so users can recognise it as a sample
        example_type   = loan_type_names[0] if loan_type_names else 'Loan Type Name'
        example_values = ['10001', 'Firstname Lastname', example_type, '10,000.00', '1,500.00']
        for col, val in enumerate(example_values, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.border = thin
            cell.font   = Font(color='FF888888', italic=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="principal-balance_template.xlsx"'
        return response


class FinanceDeductionTemplateView(APIView):
    """
    GET /api/finance/admin/template/deduction
    Returns a deduction import template .xlsx.
    Columns: ID Number | Loan Type | Deduction
    Column B has a dropdown validation listing all active LoanType names.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> HttpResponse:
        err = _require_accounting_admin(request)
        if err:
            return err

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        loan_type_names = list(LoanType.objects.values_list('name', flat=True))

        def _side():
            return Side(style='thin', color='FF000000')

        thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        hdr_fill = PatternFill(start_color='FF2845D6', end_color='FF2845D6', fill_type='solid')

        headers    = ['ID Number', 'Employee Name', 'Loan Type', 'Deduction']
        col_widths = [14, 28, 24, 20]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Deductions'

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill      = hdr_fill
            cell.border    = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 18

        # Column C: dropdown of all active loan types
        if loan_type_names:
            formula = '"' + ','.join(loan_type_names) + '"'
            dv = DataValidation(
                type='list',
                formula1=formula,
                sqref='C2:C5000',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error       = 'Please select a valid loan type from the dropdown list.'
            dv.errorTitle  = 'Invalid Loan Type'
            dv.prompt      = 'Select a loan type from the list.'
            dv.promptTitle = 'Loan Type'
            ws.add_data_validation(dv)

        # Example data row (italic grey)
        example_type   = loan_type_names[0] if loan_type_names else 'Loan Type Name'
        example_values = ['10001', 'Firstname Lastname', example_type, '1,500.00']
        for col, val in enumerate(example_values, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.border = thin
            cell.font   = Font(color='FF888888', italic=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="deductions_template.xlsx"'
        return response


class FinanceAllowanceTemplateView(APIView):
    """
    GET /api/finance/admin/template/allowance
    Returns an allowance import template .xlsx.
    Columns: ID Number | Employee Name | Allowance Type | Amount |
             Deposited Date | Covered Period
    Column C has a dropdown validation listing all active AllowanceType names.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> HttpResponse:
        err = _require_accounting_admin(request)
        if err:
            return err

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        allowance_type_names = list(AllowanceType.objects.values_list('name', flat=True))

        def _side():
            return Side(style='thin', color='FF000000')

        thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        hdr_fill = PatternFill(start_color='FF2845D6', end_color='FF2845D6', fill_type='solid')

        headers    = ['ID Number', 'Employee Name', 'Allowance Type', 'Amount', 'Deposited Date', 'Covered Period']
        col_widths = [14, 28, 24, 14, 16, 22]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Allowances'

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill      = hdr_fill
            cell.border    = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 18

        # Column C: dropdown of all active allowance types
        if allowance_type_names:
            formula = '"' + ','.join(allowance_type_names) + '"'
            dv = DataValidation(
                type='list',
                formula1=formula,
                sqref='C2:C5000',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error       = 'Please select a valid allowance type from the dropdown list.'
            dv.errorTitle  = 'Invalid Allowance Type'
            dv.prompt      = 'Select an allowance type from the list.'
            dv.promptTitle = 'Allowance Type'
            ws.add_data_validation(dv)

        # Example data row (italic grey)
        example_type   = allowance_type_names[0] if allowance_type_names else 'Allowance Type Name'
        example_values = ['10001', 'Firstname Lastname', example_type, '1,500.00', '04/30/2026', 'April 2026']
        for col, val in enumerate(example_values, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.border = thin
            cell.font   = Font(color='FF888888', italic=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="allowances_template.xlsx"'
        return response


class FinanceSavingsTemplateView(APIView):
    """
    GET /api/finance/admin/template/savings
    Returns a savings import template .xlsx.
    Columns: ID Number | Employee Name | Savings Type | Savings
    Column C has a dropdown validation listing all active SavingsType names.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> HttpResponse:
        err = _require_accounting_admin(request)
        if err:
            return err

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        savings_type_names = list(SavingsType.objects.values_list('name', flat=True))

        def _side():
            return Side(style='thin', color='FF000000')

        thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        hdr_fill = PatternFill(start_color='FF2845D6', end_color='FF2845D6', fill_type='solid')

        headers    = ['ID Number', 'Employee Name', 'Savings Type', 'Savings']
        col_widths = [14, 28, 24, 14]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Savings'

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill      = hdr_fill
            cell.border    = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 18

        # Column C: dropdown of all active savings types
        if savings_type_names:
            formula = '"' + ','.join(savings_type_names) + '"'
            dv = DataValidation(
                type='list',
                formula1=formula,
                sqref='C2:C5000',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error       = 'Please select a valid savings type from the dropdown list.'
            dv.errorTitle  = 'Invalid Savings Type'
            dv.prompt      = 'Select a savings type from the list.'
            dv.promptTitle = 'Savings Type'
            ws.add_data_validation(dv)

        # Example data row (italic grey)
        example_type   = savings_type_names[0] if savings_type_names else 'Savings Type Name'
        example_values = ['10001', 'Firstname Lastname', example_type, '1,500.00']
        for col, val in enumerate(example_values, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.border = thin
            cell.font   = Font(color='FF888888', italic=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="savings_template.xlsx"'
        return response


class FinanceOJTPayslipTemplateView(APIView):
    """
    GET /api/finance/admin/template/ojt-payslip
    Returns a clean xlsx template for bulk OJT payslip import.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> HttpResponse:
        err = _require_accounting_admin(request)
        if err:
            return err

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        def _side():
            return Side(style='thin', color='FF000000')

        thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        hdr_fill = PatternFill(start_color='FF2845D6', end_color='FF2845D6', fill_type='solid')

        headers = [
            'ID Number',
            'Regular Day',
            'Allowance Day',
            'Total Allowance',
            'ND Allowance',
            'Grand Total',
            'Basic School Share',
            'Basic OJT Share',
            'Deduction',
            'Net OJT Share',
            'Rice Allowance',
            'OT Allowance',
            'ND OT Allowance',
            'Special Holiday',
            'Legal Holiday',
            'Satoff Allowance',
            'RD OT',
            'Adjustment',
            'Deduction 2',
            'OT Pay Allowance',
            'Total Allow',
            'Holiday Date',
            'RD OT Date',
            'Perfect Attendance',
        ]
        col_widths = [18] * len(headers)

        wb = Workbook()
        ws = wb.active
        ws.title = 'OJT Payslip'

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill      = hdr_fill
            cell.border    = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 18

        example_values = [
            '960921', '22', '5000', '110000', '1500',
            '111500', '55750', '55750', '0', '55750',
            '0', '0', '0', '0', '0',
            '0', '0', '0', '0', '0',
            '0', '01/15/2026', '01/22/2026', '0',
        ]
        for col, val in enumerate(example_values, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.border = thin
            cell.font   = Font(color='FF888888', italic=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ojt_payslip_template.xlsx"'
        return response


def _format_ojt_cutoff(start: 'datetime.date', end: 'datetime.date') -> str:
    """Format a date range into the OJT cut-off string.

    Same month & year  : "Apr 15 - 30, 2026"
    Different month, same year  : "Apr 30 - May 08, 2026"
    Different month & year : "Dec 24, 2026 - Jan 05, 2027"
    """
    MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    sm, em = MONTHS[start.month - 1], MONTHS[end.month - 1]
    sd, ed = start.day, end.day
    sy, ey = start.year, end.year
    if start.month == end.month and sy == ey:
        return f"{sm} {sd} - {ed}, {sy}"
    elif sy == ey:
        return f"{sm} {sd} - {em} {ed:02d}, {sy}"
    else:
        return f"{sm} {sd}, {sy} - {em} {ed:02d}, {ey}"


def _parse_ojt_decimal(value, field: str, row: int) -> 'tuple[Decimal | None, dict | None]':
    """Parse a cell as a non-negative Decimal for OJT import.
    None / blank / N/A → Decimal('0').
    Non-numeric text → error.
    """
    if value is None:
        return Decimal('0'), None
    s = str(value).strip()
    if s == '' or s.lower() in ('none', 'n/a', '-'):
        return Decimal('0'), None
    try:
        d = Decimal(s.replace(',', '')).quantize(Decimal('0.01'))
        if d < 0:
            return None, {'row': row, 'field': field, 'reason': f'{field} must be 0 or greater.'}
        return d, None
    except (InvalidOperation, TypeError, ValueError):
        return None, {'row': row, 'field': field, 'reason': f'{field} must be a numeric value (letters are not allowed).'}


def _build_ojt_error_excel(original_rows: list[list], failures: list[dict]) -> str:
    """Build an OJT-specific error report preserving all original columns + a Remarks column."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OJT_HEADERS = [
        'ID Number', 'Regular Day', 'Allowance Day', 'Total Allowance', 'ND Allowance',
        'Grand Total', 'Basic School Share', 'Basic OJT Share', 'Deduction', 'Net OJT Share',
        'Rice Allowance', 'OT Allowance', 'ND OT Allowance', 'Special Holiday', 'Legal Holiday',
        'Satoff Allowance', 'RD OT', 'Adjustment', 'Deduction 2', 'OT Pay Allowance',
        'Total Allow', 'Holiday Date', 'RD OT Date', 'Perfect Attendance', 'Remarks',
    ]
    NUM_DATA_COLS = 24

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'OJT Import Errors'

    def _side():
        return Side(style='thin', color='FF000000')

    thin     = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    hdr_fill = PatternFill(start_color='FFCC0000', end_color='FFCC0000', fill_type='solid')
    red_font = Font(color='FFFF0000')
    hdr_font = Font(bold=True, color='FFFFFFFF')

    for col, h in enumerate(OJT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = thin
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 20 if col == len(OJT_HEADERS) else 18

    error_map: dict[int, str] = {}
    for f in failures:
        r = f.get('row')
        if r is not None:
            reason = f.get('reason', '')
            error_map[r] = (error_map[r] + '; ' + reason) if r in error_map else reason

    for i, row in enumerate(original_rows[1:], start=2):
        while len(row) < NUM_DATA_COLS:
            row.append(None)
        has_error = i in error_map
        for col, val in enumerate(row[:NUM_DATA_COLS], 1):
            cell = ws.cell(row=i, column=col, value=str(val) if val is not None else '')
            cell.border = thin
            if has_error:
                cell.font = red_font
        remarks_cell = ws.cell(row=i, column=NUM_DATA_COLS + 1, value=error_map.get(i, ''))
        remarks_cell.border = thin
        if has_error:
            remarks_cell.font = red_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


class FinanceOJTPayslipImportView(APIView):
    """
    POST /api/finance/admin/import/ojt-payslip
    Accepts a single xlsx file + period_start + period_end (YYYY-MM-DD).
    Validates all rows before writing (all-or-nothing).
    Returns { imported, failed, error_report_b64 }.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request) -> Response:
        import datetime as _dt
        err = _require_accounting_admin(request)
        if err:
            return err

        # ── Parse period dates ─────────────────────────────────────────────────
        period_start_raw = request.POST.get('period_start', '').strip()
        period_end_raw   = request.POST.get('period_end',   '').strip()
        if not period_start_raw or not period_end_raw:
            return Response({'detail': 'period_start and period_end are required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            period_start = _dt.date.fromisoformat(period_start_raw)
            period_end   = _dt.date.fromisoformat(period_end_raw)
        except (ValueError, TypeError):
            return Response({'detail': 'period_start and period_end must be valid ISO dates (YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)
        if period_end < period_start:
            period_start, period_end = period_end, period_start

        # ── Read file ──────────────────────────────────────────────────────────
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        rows, read_err = _read_xlsx(file_obj)
        if read_err:
            return Response({'detail': read_err}, status=status.HTTP_400_BAD_REQUEST)
        if len(rows) < 2:
            return Response({'detail': 'File has no data rows (need at least header + one data row).'}, status=status.HTTP_400_BAD_REQUEST)

        # Column order (0-based): matches template exactly
        # 0:id, 1:regular_day, 2:allowance_day, 3:total_allowance, 4:nd_allowance,
        # 5:grand_total, 6:basic_school_share, 7:basic_ojt_share, 8:deduction,
        # 9:net_ojt_share, 10:rice_allowance, 11:ot_allowance, 12:nd_ot_allowance,
        # 13:special_holiday, 14:legal_holiday, 15:satoff_allowance, 16:rd_ot,
        # 17:adjustment, 18:deduction_2, 19:ot_pay_allowance, 20:total_allow,
        # 21:holiday_date, 22:rd_ot_date, 23:perfect_attendance
        NUMERIC_FIELDS = [
            (1,  'Regular Day',       'regular_day'),
            (2,  'Allowance Day',     'allowance_day'),
            (3,  'Total Allowance',   'total_allowance'),
            (4,  'ND Allowance',      'nd_allowance'),
            (5,  'Grand Total',       'grand_total'),
            (6,  'Basic School Share','basic_school_share'),
            (7,  'Basic OJT Share',   'basic_ojt_share'),
            (8,  'Deduction',         'deduction'),
            (9,  'Net OJT Share',     'net_ojt_share'),
            (10, 'Rice Allowance',    'rice_allowance'),
            (11, 'OT Allowance',      'ot_allowance'),
            (12, 'ND OT Allowance',   'nd_ot_allowance'),
            (13, 'Special Holiday',   'special_holiday'),
            (14, 'Legal Holiday',     'legal_holiday'),
            (15, 'Satoff Allowance',  'satoff_allowance'),
            (16, 'RD OT',             'rd_ot'),
            (17, 'Adjustment',        'adjustment'),
            (18, 'Deduction 2',       'deduction_2'),
            (19, 'OT Pay Allowance',  'ot_pay_allowance'),
            (20, 'Total Allow',       'total_allow'),
            (23, 'Perfect Attendance','perfect_attendance'),
        ]

        failures: list[dict] = []
        operations: list[dict] = []

        # ── Phase 1: validate every row ────────────────────────────────────────
        for row_idx, row in enumerate(rows[1:], start=2):
            while len(row) < 24:
                row.append(None)

            idnumber = str(row[0]).strip() if row[0] is not None else ''
            if not idnumber:
                failures.append({'row': row_idx, 'field': 'ID Number', 'reason': 'ID Number is required.'})
                continue

            employee, emp_err = _lookup_employee(idnumber, row_idx)
            if emp_err:
                failures.append(emp_err)
                continue

            row_data: dict = {'employee': employee}
            row_ok = True

            for col_idx, label, model_field in NUMERIC_FIELDS:
                val, parse_err = _parse_ojt_decimal(row[col_idx], label, row_idx)
                if parse_err:
                    failures.append(parse_err)
                    row_ok = False
                else:
                    row_data[model_field] = val

            if not row_ok:
                continue

            # AlphaNumeric text fields — store as-is; None → empty string
            def _text(v) -> str:
                if v is None:
                    return ''
                s = str(v).strip()
                return '' if s.lower() in ('none', 'n/a') else s

            row_data['holiday_date']  = _text(row[21])
            row_data['rd_ot_date']    = _text(row[22])
            row_data['period_start']  = period_start
            row_data['period_end']    = period_end
            operations.append(row_data)

        # ── Phase 2: all-or-nothing ────────────────────────────────────────────
        if failures:
            error_b64 = _build_ojt_error_excel(rows, failures)
            return Response({'imported': 0, 'failed': len(failures), 'error_report_b64': error_b64})

        # ── Phase 3: bulk create in atomic transaction ─────────────────────────
        from .models import OJTPayslipData
        with transaction.atomic():
            objs = [OJTPayslipData(**data) for data in operations]
            OJTPayslipData.objects.bulk_create(objs)

        return Response({'imported': len(objs), 'failed': 0, 'error_report_b64': None})


class FinanceEmployeeFilterOptionsView(APIView):
    """
    GET /api/finance/admin/employee-filters
    Returns departments, lines, and idnumbers available for column-level
    filtering of the employee list.  Only reflects non-privileged employees
    (admin=False, accounting=False, hr=False).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        from userLogin.models import loginCredentials
        from generalsettings.models import Department, Line

        qs = loginCredentials.objects.filter(admin=False, accounting=False, hr=False)

        dept_ids = (
            qs.filter(workinformation__department__isnull=False)
            .values_list('workinformation__department_id', flat=True)
            .distinct()
        )
        line_ids = (
            qs.filter(workinformation__line__isnull=False)
            .values_list('workinformation__line_id', flat=True)
            .distinct()
        )

        departments = list(
            Department.objects.filter(id__in=dept_ids).order_by('name').values('id', 'name')
        )
        lines = list(
            Line.objects.filter(id__in=line_ids).order_by('name').values('id', 'name')
        )
        idnumbers = list(qs.order_by('idnumber').values_list('idnumber', flat=True))

        return Response({'departments': departments, 'lines': lines, 'idnumbers': idnumbers})


# ── OJT Payslip delete ────────────────────────────────────────────────────────

class FinanceOJTPayslipDeleteView(APIView):
    """
    DELETE /api/finance/admin/ojt-payslips/<pk>
    Permanently deletes an OJT payslip data record.
    """
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk: int) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err
        from .models import OJTPayslipData
        try:
            obj = OJTPayslipData.objects.get(pk=pk)
        except OJTPayslipData.DoesNotExist:
            return Response({'detail': 'OJT payslip not found.'}, status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Payslip delete ─────────────────────────────────────────────────────────────

class FinancePayslipDeleteView(APIView):
    """
    DELETE /api/finance/admin/payslips/<pk>
    Permanently deletes a payslip record (and its associated file).
    """
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk: int) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err
        try:
            payslip = Payslip.objects.get(pk=pk)
        except Payslip.DoesNotExist:
            return Response({'detail': 'Payslip not found.'}, status=status.HTTP_404_NOT_FOUND)
        # Remove the file from storage before deleting the DB row
        if payslip.file:
            payslip.file.delete(save=False)
        payslip.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Loan deductions ────────────────────────────────────────────────────────────

class FinanceLoanDeductionsView(APIView):
    """
    GET /api/finance/admin/loans/<pk>/deductions
    Returns all deduction records for a specific loan.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err
        try:
            loan = Loan.objects.get(pk=pk)
        except Loan.DoesNotExist:
            return Response({'detail': 'Loan not found.'}, status=status.HTTP_404_NOT_FOUND)

        deductions = Deduction.objects.filter(loan=loan).order_by('-created_at').values(
            'id', 'amount', 'description', 'cutoff_date', 'created_at',
        )
        return Response({
            'loan_id':          loan.pk,
            'loan_type_name':   loan.loan_type.name,
            'principal_amount': str(loan.principal_amount),
            'current_balance':  str(loan.current_balance),
            'deductions':       list(deductions),
        })


# ── Savings withdraw ───────────────────────────────────────────────────────────

class FinanceSavingsWithdrawView(APIView):
    """
    POST /api/finance/admin/savings/<pk>/withdraw
    Marks the given savings record as withdrawn (withdraw=True).
    No new row is created — the existing record is updated in place.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err
        try:
            savings = Savings.objects.get(pk=pk)
        except Savings.DoesNotExist:
            return Response({'detail': 'Savings record not found.'}, status=status.HTTP_404_NOT_FOUND)

        if savings.withdraw:
            return Response({'detail': 'This savings record has already been withdrawn.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            savings_row = Savings.objects.select_for_update().get(pk=pk)
            savings_row.withdraw = True
            savings_row.save(update_fields=['withdraw'])
            _queue_finance_notifications(
                notification_type='finance_savings_withdrawn',
                title='Savings Withdrawal Recorded',
                module='finance',
                user_messages={
                    savings_row.employee: (
                        f'A withdrawal of \u20b1{savings_row.amount} was recorded from '
                        f'{savings_row.savings_type.name} savings.'
                    )
                },
            )

        return Response(
            SavingsSerializer(savings_row).data,
            status=status.HTTP_200_OK,
        )


# ── Loan system settings ───────────────────────────────────────────────────────

class LoanSettingsView(APIView):
    """
    GET  /api/finance/admin/loan-settings  — retrieve current settings
    PUT  /api/finance/admin/loan-settings  — update settings
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err
        return Response(LoanSettingsSerializer(LoanSettings.get_settings()).data)

    def put(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err
        with transaction.atomic():
            settings_obj = LoanSettings.objects.select_for_update().get_or_create(pk=1)[0]
            ser = LoanSettingsSerializer(settings_obj, data=request.data, partial=True)
            if not ser.is_valid():
                return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
            ser.save()
        return Response(ser.data)


# ── Palette for auto-assigning colors ─────────────────────────────────────────

_COLOR_PALETTE = [
    '#2845D6', '#10B981', '#F59E0B', '#EF4444',
    '#8B5CF6', '#EC4899', '#14B8A6', '#F97316',
    '#06B6D4', '#84CC16', '#0EA5E9', '#A855F7',
]

_TYPE_MODEL_MAP = {
    'allowance': (AllowanceType, AllowanceTypeSerializer),
    'loan':      (LoanType,      LoanTypeSerializer),
    'savings':   (SavingsType,   SavingsTypeSerializer),
    'payslip':   (PayslipType,   PayslipTypeSerializer),
}


class FinanceTypeCRUDView(APIView):
    """
    POST   /api/finance/admin/types/create          — create a new type
    PUT    /api/finance/admin/types/<pk>            — rename / recolor a type
    DELETE /api/finance/admin/types/<pk>            — delete a type

    Body fields
    -----------
    type_category  : 'allowance' | 'loan' | 'savings' | 'payslip'  (required)
    name           : str (required for POST)
    color          : '#rrggbb' hex string (optional; auto-assigned on POST)
    replace_on_upload / stackable : bool (optional, type-specific)
    """
    permission_classes = [IsAuthenticated]

    def _get_model(self, category: str):
        entry = _TYPE_MODEL_MAP.get(category)
        if entry is None:
            return None, None
        return entry

    def _next_color(self, Model) -> str:
        """Pick the next palette color not currently used by any existing row."""
        used = set(Model.objects.values_list('color', flat=True))
        for c in _COLOR_PALETTE:
            if c not in used:
                return c
        # All palette colors are used — cycle based on count
        return _COLOR_PALETTE[Model.objects.count() % len(_COLOR_PALETTE)]

    @transaction.atomic
    def post(self, request) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        category = (request.data.get('type_category') or '').strip().lower()
        Model, Serializer = self._get_model(category)
        if Model is None:
            return Response({'detail': 'Invalid type_category.'}, status=status.HTTP_400_BAD_REQUEST)

        name = (request.data.get('name') or '').strip()
        if not name:
            return Response({'detail': 'name is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(name) > 100:
            return Response({'detail': 'name must not exceed 100 characters.'}, status=status.HTTP_400_BAD_REQUEST)

        if Model.objects.filter(name__iexact=name).exists():
            return Response({'detail': f'A {category} type with this name already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        color = (request.data.get('color') or '').strip()
        if not color:
            color = self._next_color(Model)

        kwargs: dict = {'name': name, 'color': color}
        if category == 'allowance':
            kwargs['replace_on_upload'] = bool(request.data.get('replace_on_upload', False))
            kwargs['percentage']        = bool(request.data.get('percentage', False))
        elif category == 'loan':
            kwargs['stackable'] = bool(request.data.get('stackable', False))

        obj = Model.objects.create(**kwargs)
        return Response(Serializer(obj).data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def put(self, request, pk: int | None = None) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        if pk is None:
            return Response({'detail': 'pk required.'}, status=status.HTTP_400_BAD_REQUEST)

        category = (request.data.get('type_category') or '').strip().lower()
        Model, Serializer = self._get_model(category)
        if Model is None:
            return Response({'detail': 'Invalid type_category.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            obj = Model.objects.get(pk=pk)
        except Model.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        name = (request.data.get('name') or '').strip()
        if name:
            if len(name) > 100:
                return Response({'detail': 'name must not exceed 100 characters.'}, status=status.HTTP_400_BAD_REQUEST)
            if Model.objects.filter(name__iexact=name).exclude(pk=pk).exists():
                return Response({'detail': f'A {category} type with this name already exists.'}, status=status.HTTP_400_BAD_REQUEST)
            obj.name = name

        color = (request.data.get('color') or '').strip()
        if color:
            obj.color = color

        if category == 'allowance' and 'replace_on_upload' in request.data:
            obj.replace_on_upload = bool(request.data['replace_on_upload'])
        if category == 'allowance' and 'percentage' in request.data:
            obj.percentage = bool(request.data['percentage'])
        if category == 'loan' and 'stackable' in request.data:
            obj.stackable = bool(request.data['stackable'])

        obj.save()
        return Response(Serializer(obj).data)

    @transaction.atomic
    def delete(self, request, pk: int | None = None) -> Response:
        err = _require_accounting_admin(request)
        if err:
            return err

        if pk is None:
            return Response({'detail': 'pk required.'}, status=status.HTTP_400_BAD_REQUEST)

        category = (request.data.get('type_category') or '').strip().lower()
        Model, _ = self._get_model(category)
        if Model is None:
            return Response({'detail': 'Invalid type_category.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            obj = Model.objects.get(pk=pk)
        except Model.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── User-facing finance views ─────────────────────────────────────────────────

def _build_payslip_email(payslip, employee) -> tuple[str, str]:
    """Build subject and HTML body for a payslip email notification."""
    period_start = payslip.period_start.strftime('%B %d, %Y')
    period_end   = payslip.period_end.strftime('%B %d, %Y')
    period_label = f'{period_start} \u2013 {period_end}'
    full_name    = (
        f'{employee.firstname or ""} {employee.lastname or ""}'.strip()
        or employee.idnumber
    )
    subject = f'Your Payslip for {period_label} is Ready'
    html_body = f"""\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Payslip Notification</title>
</head>
<body style="margin:0;padding:0;background-color:#f4f6f8;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" border="0"
         style="background-color:#f4f6f8;padding:32px 16px;">
    <tr>
      <td align="center">
        <table width="600" cellpadding="0" cellspacing="0" border="0"
               style="max-width:600px;width:100%;background-color:#ffffff;
                      border-radius:12px;overflow:hidden;
                      box-shadow:0 4px 16px rgba(0,0,0,0.08);">

          <!-- Header -->
          <tr>
            <td style="background-color:#2845D6;padding:28px 32px;text-align:center;">
              <h1 style="margin:0;color:#ffffff;font-size:22px;
                         font-weight:700;letter-spacing:-0.3px;">REPConnect</h1>
              <p style="margin:6px 0 0;color:rgba(255,255,255,0.80);
                        font-size:13px;">HR &amp; Finance System</p>
            </td>
          </tr>

          <!-- Body -->
          <tr>
            <td style="padding:32px 32px 24px;">
              <p style="margin:0 0 16px;font-size:15px;color:#262626;">Dear <strong>{full_name}</strong>,</p>

              <p style="margin:0 0 16px;font-size:14px;color:#3a3a3a;line-height:1.7;">
                We would like to inform you that your payslip for the cut-off period
                <strong>{period_label}</strong> is now available.
                A copy has been securely attached to this email for your reference and review.
              </p>

              <p style="margin:0 0 24px;font-size:14px;color:#3a3a3a;line-height:1.7;">
                Please take a moment to verify the details. If you have any questions or require
                clarification regarding your payslip, you may coordinate directly with your
                <strong>HR or Finance department</strong> for assistance.
              </p>

              <p style="margin:0 0 8px;font-size:14px;color:#3a3a3a;">Thank you.</p>
              <p style="margin:0;font-size:14px;color:#3a3a3a;">
                Sincerely,<br />
                <strong>REPConnect</strong><br />
                <span style="color:#6b7280;">HR &amp; Finance System</span>
              </p>
            </td>
          </tr>

          <!-- Notice block -->
          <tr>
            <td style="padding:0 32px 24px;">
              <table width="100%" cellpadding="0" cellspacing="0" border="0"
                     style="background-color:#fef9ec;border:1px solid #fde68a;
                            border-radius:8px;">
                <tr>
                  <td style="padding:12px 16px;font-size:12px;color:#92400e;line-height:1.6;">
                    &#9888;&nbsp; <em>This is a system-generated email.
                    Please do not reply to this message.</em>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Security Reminder -->
          <tr>
            <td style="padding:0 32px 32px;">
              <table width="100%" cellpadding="0" cellspacing="0" border="0"
                     style="border-top:1px solid #e5e7eb;">
                <tr>
                  <td style="padding-top:20px;font-size:11px;color:#6b7280;line-height:1.7;">
                    <strong style="color:#374151;">Security Reminder:</strong><br />
                    Please remain vigilant against phishing and malicious emails.
                    Do not click on links from suspicious or unknown senders, and avoid
                    downloading attachments with unfamiliar or unexpected file formats.
                    Always verify the sender and ensure the email is legitimate before
                    taking any action.
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Footer -->
          <tr>
            <td style="background-color:#f4f6f8;padding:16px 32px;text-align:center;
                       border-top:1px solid #e5e7eb;">
              <p style="margin:0;font-size:11px;color:#9ca3af;">
                &copy; REPConnect HR &amp; Finance System &bull; All rights reserved.
              </p>
            </td>
          </tr>

        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""
    return subject, html_body


def _build_ojt_payslip_email(obj, employee, period_label: str) -> tuple[str, str]:
    """Build subject and HTML body for an OJT payslip email notification."""
    full_name = (
        f'{employee.firstname or ""} {employee.lastname or ""}'.strip()
        or employee.idnumber
    )
    subject = f'Your OJT Payslip for {period_label} is Ready'
    html_body = f"""\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>OJT Payslip Notification</title>
</head>
<body style="margin:0;padding:0;background-color:#f4f6f8;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" border="0"
         style="background-color:#f4f6f8;padding:32px 16px;">
    <tr>
      <td align="center">
        <table width="600" cellpadding="0" cellspacing="0" border="0"
               style="max-width:600px;width:100%;background-color:#ffffff;
                      border-radius:12px;overflow:hidden;
                      box-shadow:0 4px 16px rgba(0,0,0,0.08);">

          <!-- Header -->
          <tr>
            <td style="background-color:#2845D6;padding:28px 32px;text-align:center;">
              <h1 style="margin:0;color:#ffffff;font-size:22px;
                         font-weight:700;letter-spacing:-0.3px;">REPConnect</h1>
              <p style="margin:6px 0 0;color:rgba(255,255,255,0.80);
                        font-size:13px;">HR &amp; Finance System</p>
            </td>
          </tr>

          <!-- Body -->
          <tr>
            <td style="padding:32px 32px 24px;">
              <p style="margin:0 0 16px;font-size:15px;color:#262626;">Dear <strong>{full_name}</strong>,</p>

              <p style="margin:0 0 16px;font-size:14px;color:#3a3a3a;line-height:1.7;">
                We would like to inform you that your OJT payslip for the period
                <strong>{period_label}</strong> is now available.
                A copy has been securely attached to this email for your reference and review.
              </p>

              <p style="margin:0 0 24px;font-size:14px;color:#3a3a3a;line-height:1.7;">
                Please take a moment to verify the details. If you have any questions or require
                clarification regarding your OJT payslip, you may coordinate directly with your
                <strong>HR or Finance department</strong> for assistance.
              </p>

              <p style="margin:0 0 8px;font-size:14px;color:#3a3a3a;">Thank you.</p>
              <p style="margin:0;font-size:14px;color:#3a3a3a;">
                Sincerely,<br />
                <strong>REPConnect</strong><br />
                <span style="color:#6b7280;">HR &amp; Finance System</span>
              </p>
            </td>
          </tr>

          <!-- Notice block -->
          <tr>
            <td style="padding:0 32px 24px;">
              <table width="100%" cellpadding="0" cellspacing="0" border="0"
                     style="background-color:#fef9ec;border:1px solid #fde68a;
                            border-radius:8px;">
                <tr>
                  <td style="padding:12px 16px;font-size:12px;color:#92400e;line-height:1.6;">
                    &#9888;&nbsp; <em>This is a system-generated email.
                    Please do not reply to this message.</em>
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Security Reminder -->
          <tr>
            <td style="padding:0 32px 32px;">
              <table width="100%" cellpadding="0" cellspacing="0" border="0"
                     style="border-top:1px solid #e5e7eb;">
                <tr>
                  <td style="padding-top:20px;font-size:11px;color:#6b7280;line-height:1.7;">
                    <strong style="color:#374151;">Security Reminder:</strong><br />
                    Please remain vigilant against phishing and malicious emails.
                    Do not click on links from suspicious or unknown senders, and avoid
                    downloading attachments with unfamiliar or unexpected file formats.
                    Always verify the sender and ensure the email is legitimate before
                    taking any action.
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- Footer -->
          <tr>
            <td style="background-color:#f4f6f8;padding:16px 32px;text-align:center;
                       border-top:1px solid #e5e7eb;">
              <p style="margin:0;font-size:11px;color:#9ca3af;">
                &copy; REPConnect HR &amp; Finance System &bull; All rights reserved.
              </p>
            </td>
          </tr>

        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""
    return subject, html_body


class UserFinanceLoanSettingsView(APIView):
    """
    GET /api/finance/my/loan-settings
    Returns current loan deduction frequency settings (read-only for all authenticated users).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        return Response(LoanSettingsSerializer(LoanSettings.get_settings()).data)


class UserFinanceRecordsView(APIView):
    """
    GET /api/finance/my/records
    Returns all finance records that belong to the authenticated user.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        user       = request.user
        loans      = list(Loan.objects.filter(employee=user).order_by('-created_at'))
        allowances = Allowance.objects.filter(employee=user).order_by('-created_at')
        savings    = Savings.objects.filter(employee=user).order_by('-created_at')
        payslips   = Payslip.objects.filter(employee=user).order_by('-created_at')
        from .models import OJTPayslipData
        ojt_payslips = OJTPayslipData.objects.filter(employee=user).order_by('-period_start', '-created_at')
        ojt_data = [
            {
                'id':                obj.pk,
                'period_start':      obj.period_start.isoformat() if obj.period_start else None,
                'period_end':        obj.period_end.isoformat()   if obj.period_end   else None,
                'regular_day':       str(obj.regular_day),
                'allowance_day':     str(obj.allowance_day),
                'total_allowance':   str(obj.total_allowance),
                'nd_allowance':      str(obj.nd_allowance),
                'grand_total':       str(obj.grand_total),
                'basic_school_share': str(obj.basic_school_share),
                'basic_ojt_share':   str(obj.basic_ojt_share),
                'deduction':         str(obj.deduction),
                'net_ojt_share':     str(obj.net_ojt_share),
                'rice_allowance':    str(obj.rice_allowance),
                'ot_allowance':      str(obj.ot_allowance),
                'nd_ot_allowance':   str(obj.nd_ot_allowance),
                'special_holiday':   str(obj.special_holiday),
                'legal_holiday':     str(obj.legal_holiday),
                'satoff_allowance':  str(obj.satoff_allowance),
                'rd_ot':             str(obj.rd_ot),
                'adjustment':        str(obj.adjustment),
                'deduction_2':       str(obj.deduction_2),
                'ot_pay_allowance':  str(obj.ot_pay_allowance),
                'total_allow':       str(obj.total_allow),
                'perfect_attendance': str(obj.perfect_attendance),
                'holiday_date':      obj.holiday_date,
                'rd_ot_date':        obj.rd_ot_date,
                'sent':              obj.sent,
                'created_at':        obj.created_at.isoformat(),
            }
            for obj in ojt_payslips
        ]

        response_data = {
            'loans':        LoanSerializer(loans,      many=True).data,
            'allowances':   AllowanceSerializer(allowances, many=True).data,
            'savings':      SavingsSerializer(savings,  many=True).data,
            'payslips':     PayslipSerializer(payslips, many=True, context={'request': request}).data,
            'ojt_payslips': ojt_data,
        }
        # Mark unseen loans as seen after serializing so the "New" pill is visible on first load
        Loan.objects.filter(employee=user, seen=False).update(seen=True)
        return Response(response_data)


class UserFinanceUnseenCountView(APIView):
    """
    GET /api/finance/my/unseen-count
    Returns sidebar badge count from unseen loans + unsent payslips.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        from .models import OJTPayslipData
        user = request.user
        unseen_loans      = Loan.objects.filter(employee=user, seen=False).count()
        unsent_payslips   = Payslip.objects.filter(employee=user, sent=False).count()
        unsent_ojt        = OJTPayslipData.objects.filter(employee=user, sent=False).count()
        return Response({'count': unseen_loans + unsent_payslips + unsent_ojt})


class UserFinanceLoanDeductionsView(APIView):
    """
    GET /api/finance/my/loans/<pk>/deductions
    Returns deduction records for a loan that belongs to the authenticated user.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int) -> Response:
        try:
            loan = Loan.objects.get(pk=pk, employee=request.user)
        except Loan.DoesNotExist:
            return Response({'detail': 'Loan not found.'}, status=status.HTTP_404_NOT_FOUND)

        deductions = Deduction.objects.filter(loan=loan).order_by('-created_at').values(
            'id', 'amount', 'description', 'cutoff_date', 'created_at',
        )
        return Response({
            'loan_id':          loan.pk,
            'loan_type_name':   loan.loan_type.name,
            'principal_amount': str(loan.principal_amount),
            'current_balance':  str(loan.current_balance),
            'deductions':       list(deductions),
        })


class UserFinancePayslipSendEmailView(APIView):
    """
    POST /api/finance/my/payslips/<pk>/send-email
    Sends the payslip PDF to the authenticated user's registered email address.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int) -> Response:
        try:
            payslip = Payslip.objects.select_related('employee', 'payslip_type').get(
                pk=pk, employee=request.user,
            )
        except Payslip.DoesNotExist:
            return Response({'detail': 'Payslip not found.'}, status=status.HTTP_404_NOT_FOUND)

        if not payslip.file:
            return Response(
                {'detail': 'This payslip has no attached file.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from generalsettings.models import EmailConfiguration
        try:
            config = EmailConfiguration.objects.get(pk=1)
        except EmailConfiguration.DoesNotExist:
            return Response(
                {'detail': 'Email configuration has not been set up.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        requested_recipient = str(request.data.get('recipient_email', '') or '').strip().lower()
        personal_email = (request.user.email or '').strip().lower()

        work_email = ''
        try:
            personal_info = request.user.personal_info
            work_email = (personal_info.work_email or '').strip().lower()
        except Exception:
            work_email = ''

        allowed_recipients = {email for email in [personal_email, work_email] if email}
        if requested_recipient and requested_recipient not in allowed_recipients:
            return Response(
                {'detail': 'Selected email is not allowed for this account.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        recipient_email = requested_recipient or personal_email
        if not recipient_email:
            return Response(
                {'detail': 'No email address is on file for your account.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        subject, html_body = _build_payslip_email(payslip, request.user)

        import smtplib
        from email.mime.application import MIMEApplication
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        msg = MIMEMultipart('mixed')
        from_addr = f'{config.from_name} <{config.username}>' if config.from_name else config.username
        msg['From']    = from_addr
        msg['To']      = recipient_email
        msg['Subject'] = subject

        alt = MIMEMultipart('alternative')
        alt.attach(MIMEText(html_body, 'html'))
        msg.attach(alt)

        original_filename = (payslip.file.name.split('/')[-1] if payslip.file.name else 'payslip.pdf')
        try:
            with payslip.file.open('rb') as pdf_file:
                part = MIMEApplication(pdf_file.read(), Name=original_filename)
            part['Content-Disposition'] = f'attachment; filename="{original_filename}"'
            msg.attach(part)
        except Exception as exc:
            return Response(
                {'detail': f'Could not read payslip file: {exc}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            port = config.smtp_port
            if port == 465:
                use_ssl, use_tls = True, False
            elif port == 587:
                use_ssl, use_tls = False, True
            else:
                use_ssl, use_tls = bool(config.use_ssl), bool(config.use_tls)

            _cls = smtplib.SMTP_SSL if use_ssl else smtplib.SMTP
            with _cls(config.smtp_host, port, timeout=15) as smtp:
                if use_tls and not use_ssl:
                    smtp.starttls()
                smtp.login(config.username, config.password)
                smtp.sendmail(config.username, recipient_email, msg.as_string())
        except Exception as exc:
            return Response(
                {'detail': f'Email sending failed: {exc}'},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        payslip.sent = True
        payslip.save(update_fields=['sent'])

        return Response({'detail': 'Payslip sent to your email successfully.'})


class UserFinanceOJTPayslipSendEmailView(APIView):
    """
    POST /api/finance/my/ojt-payslips/<pk>/send-email
    Sends OJT payslip data as an HTML email to the authenticated user.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int) -> Response:
        from .models import OJTPayslipData
        try:
            obj = OJTPayslipData.objects.get(pk=pk, employee=request.user)
        except OJTPayslipData.DoesNotExist:
            return Response({'detail': 'OJT payslip not found.'}, status=status.HTTP_404_NOT_FOUND)

        from generalsettings.models import EmailConfiguration
        try:
            config = EmailConfiguration.objects.get(pk=1)
        except EmailConfiguration.DoesNotExist:
            return Response(
                {'detail': 'Email configuration has not been set up.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        requested_recipient = str(request.data.get('recipient_email', '') or '').strip().lower()
        personal_email = (request.user.email or '').strip().lower()
        work_email = ''
        try:
            work_email = (request.user.personal_info.work_email or '').strip().lower()
        except Exception:
            pass

        allowed_recipients = {e for e in [personal_email, work_email] if e}
        if requested_recipient and requested_recipient not in allowed_recipients:
            return Response(
                {'detail': 'Selected email is not allowed for this account.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        recipient_email = requested_recipient or personal_email
        if not recipient_email:
            return Response(
                {'detail': 'No email address is on file for your account.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = request.user

        # ── Resolve line name ─────────────────────────────────────────────────
        line_name = '—'
        try:
            from userProfile.models import workInformation
            wi = workInformation.objects.select_related('line').filter(
                employee=user,
            ).order_by('-created_at').first()
            if wi and wi.line:
                line_name = wi.line.name or '—'
        except Exception:
            pass

        def _fmt_period(start, end):
            if not start or not end:
                return '—'
            MONTHS = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
            sm, em = MONTHS[start.month - 1], MONTHS[end.month - 1]
            sd, ed, sy, ey = start.day, end.day, start.year, end.year
            if start.month == end.month and sy == ey:
                return f"{sm} {sd}-{ed},{sy}"
            if sy == ey:
                return f"{sm} {sd} - {em} {ed},{sy}"
            return f"{sm} {sd},{sy} - {em} {ed},{ey}"

        period = _fmt_period(obj.period_start, obj.period_end)

        # ── Generate PDF attachment ───────────────────────────────────────────
        pdf_bytes: bytes | None = None
        try:
            pdf_bytes = _build_ojt_payslip_pdf(obj, user, line_name=line_name)
        except Exception:
            pass  # non-fatal — email still sends without attachment

        safe_period = period.replace(' ', '_').replace(',', '').replace('/', '_')
        pdf_filename = f"ojt_payslip_{safe_period}.pdf"

        # ── Build email (same style as regular payslip) ───────────────────────
        subject, html_body = _build_ojt_payslip_email(obj, user, period_label=period)

        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.application import MIMEApplication

        msg = MIMEMultipart('mixed')
        from_addr = f'{config.from_name} <{config.username}>' if config.from_name else config.username
        msg['From']    = from_addr
        msg['To']      = recipient_email
        msg['Subject'] = subject

        # HTML body wrapped in alternative part
        alt_part = MIMEMultipart('alternative')
        alt_part.attach(MIMEText(html_body, 'html'))
        msg.attach(alt_part)

        # PDF attachment (if generation succeeded)
        if pdf_bytes:
            pdf_part = MIMEApplication(pdf_bytes, _subtype='pdf')
            pdf_part.add_header('Content-Disposition', 'attachment', filename=pdf_filename)
            msg.attach(pdf_part)

        try:
            port = config.smtp_port
            if port == 465:
                use_ssl, use_tls = True, False
            elif port == 587:
                use_ssl, use_tls = False, True
            else:
                use_ssl, use_tls = bool(config.use_ssl), bool(config.use_tls)

            _cls = smtplib.SMTP_SSL if use_ssl else smtplib.SMTP
            with _cls(config.smtp_host, port, timeout=15) as smtp:
                if use_tls and not use_ssl:
                    smtp.starttls()
                smtp.login(config.username, config.password)
                smtp.sendmail(config.username, recipient_email, msg.as_string())
        except Exception as exc:
            return Response(
                {'detail': f'Email sending failed: {exc}'},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        obj.sent = True
        obj.save(update_fields=['sent'])

        return Response({'detail': 'OJT payslip sent to your email successfully.'})
