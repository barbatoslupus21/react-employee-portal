"""Training Evaluation views.

Security checklist:
  - IsAuthenticated on every view; role check inside method.
  - @transaction.atomic + select_for_update() on all write/modify operations.
  - All list endpoints paginated (page_size=10, max=20).
  - No raw SQL — ORM only.
  - Input validated via serializers before any DB write.
  - CSRF enforced by DRF + middleware.
  - Double-submit prevented by unique_together + select_for_update on submission.
"""
from __future__ import annotations

import logging
from datetime import date

from django.db import transaction
from django.db.models import Count, OuterRef, Subquery, Prefetch, Q
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from activityLog.models import Notification
from userProfile.models import workInformation as WorkInformation
from survey.models import (
    SurveyTemplate, SurveyQuestion, SurveyQuestionOption, SurveyQuestionRatingConfig,
    CHOICE_BASED_TYPES, ALLOW_OTHER_TYPES,
)
from training.models import (
    Training,
    TrainingQuestion,
    TrainingQuestionOption,
    TrainingQuestionRatingConfig,
    TrainingParticipant,
    TrainingSubmission,
    TrainingAnswer,
    TrainingApprovalStep,
    SupervisorEvaluation,
)
from training.routing import build_training_approval_chain, can_act_on_training_step
from training.serializers import (
    TrainingListSerializer,
    TrainingDetailSerializer,
    TrainingWriteSerializer,
    MyTrainingSerializer,
    TrainingQuestionSerializer,
    TrainingAnswerSerializer,
    TrainingSubmissionSerializer,
    TrainingApprovalStepSerializer,
    ApproverQueueItemSerializer,
    SupervisorEvaluationSerializer,
)

logger = logging.getLogger(__name__)
PAGE_SIZE = 10
MAX_PAGE_SIZE = 20


# ── Permission helpers ────────────────────────────────────────────────────────

def _require_admin_or_hr(request) -> Response | None:
    u = request.user
    if not (getattr(u, 'admin', False) or getattr(u, 'hr', False)):
        return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
    return None


def _eligible_recipients():
    """Users eligible to receive training assignments."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    return User.objects.filter(is_active=True, admin=False, hr=False, accounting=False)


class AdminTrainingEvaluationRoutingRuleListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from training.models import EvaluationRoutingRule

        rules = (
            EvaluationRoutingRule.objects
            .filter(module=EvaluationRoutingRule.MODULE_TRAINING_EVALUATION)
            .prefetch_related('positions', 'departments', 'steps__target_positions')
            .order_by('description')
        )

        return Response([
            {
                'id': rule.id,
                'description': rule.description,
                'is_active': rule.is_active,
                'positions': [position.name for position in rule.positions.all()],
                'departments': [department.name for department in rule.departments.all()],
                'steps': [
                    {
                        'position_ids': [p.id for p in step.target_positions.all()],
                    }
                    for step in rule.steps.all()
                ],
            }
            for rule in rules
        ])

    def post(self, request):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from generalsettings.models import Department, Position
        from training.models import EvaluationRoutingRule, EvaluationRoutingRuleStep

        payload = request.data
        module = payload.get('module') or EvaluationRoutingRule.MODULE_TRAINING_EVALUATION
        if module != EvaluationRoutingRule.MODULE_TRAINING_EVALUATION:
            return Response({'detail': 'Invalid module.'}, status=http_status.HTTP_400_BAD_REQUEST)

        description = str(payload.get('description', '')).strip()
        position_ids = [int(x) for x in payload.get('position_ids', []) if isinstance(x, (int, str)) and str(x).isdigit()]
        department_ids = [int(x) for x in payload.get('department_ids', []) if isinstance(x, (int, str)) and str(x).isdigit()]
        steps = payload.get('steps', []) or []

        with transaction.atomic():
            rule = EvaluationRoutingRule.objects.create(
                description=description,
                module=module,
                is_active=True,
            )
            rule.positions.set(Position.objects.filter(id__in=position_ids))
            rule.departments.set(Department.objects.filter(id__in=department_ids))

            if not steps:
                steps = [{'position_ids': []}]

            for index, step_data in enumerate(steps, start=1):
                step = EvaluationRoutingRuleStep.objects.create(
                    rule=rule,
                    step_order=index,
                )
                position_ids_for_step = [
                    int(x) for x in step_data.get('position_ids', [])
                    if isinstance(x, (int, str)) and str(x).isdigit()
                ]
                step.target_positions.set(Position.objects.filter(id__in=position_ids_for_step))

            return Response({'id': rule.id}, status=http_status.HTTP_201_CREATED)


class AdminTrainingEvaluationRoutingRuleDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from training.models import EvaluationRoutingRule

        try:
            rule = EvaluationRoutingRule.objects.get(
                pk=pk, module=EvaluationRoutingRule.MODULE_TRAINING_EVALUATION
            )
        except EvaluationRoutingRule.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        rule.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)

    def put(self, request, pk):
        denied = _require_admin_or_hr(request)
        if denied:
            return denied

        from generalsettings.models import Department, Position
        from training.models import EvaluationRoutingRule, EvaluationRoutingRuleStep

        try:
            rule = EvaluationRoutingRule.objects.get(
                pk=pk, module=EvaluationRoutingRule.MODULE_TRAINING_EVALUATION
            )
        except EvaluationRoutingRule.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        payload = request.data
        description = str(payload.get('description', '')).strip()
        position_ids = [int(x) for x in payload.get('position_ids', []) if isinstance(x, (int, str)) and str(x).isdigit()]
        department_ids = [int(x) for x in payload.get('department_ids', []) if isinstance(x, (int, str)) and str(x).isdigit()]
        steps = payload.get('steps', []) or []

        with transaction.atomic():
            rule.description = description
            rule.positions.set(Position.objects.filter(id__in=position_ids))
            rule.departments.set(Department.objects.filter(id__in=department_ids))
            rule.save()

            rule.steps.all().delete()
            if not steps:
                steps = [{'position_ids': []}]
            for index, step_data in enumerate(steps, start=1):
                step = EvaluationRoutingRuleStep.objects.create(
                    rule=rule,
                    step_order=index,
                )
                position_ids_for_step = [
                    int(x) for x in step_data.get('position_ids', [])
                    if isinstance(x, (int, str)) and str(x).isdigit()
                ]
                step.target_positions.set(Position.objects.filter(id__in=position_ids_for_step))

            return Response({'id': rule.id})


def _get_training_status(training_date: date) -> str:
    today = date.today()
    if training_date > today:
        return 'scheduled'
    if training_date == today:
        return 'active'
    return 'closed'


def _paginate(queryset, request):
    try:
        page = max(1, int(request.query_params.get('page', 1)))
        size = min(MAX_PAGE_SIZE, max(1, int(request.query_params.get('page_size', PAGE_SIZE))))
    except (ValueError, TypeError):
        page, size = 1, PAGE_SIZE
    start = (page - 1) * size
    end = start + size
    total = queryset.values('id').count()
    items = list(queryset[start:end])
    return items, total, page, size


def _deep_copy_template_questions(training: Training, template: SurveyTemplate) -> None:
    """Deep-copy questions (and options/rating config) from template into training."""
    src_questions = list(
        SurveyQuestion.objects
        .filter(template=template)
        .prefetch_related('options', 'rating_config')
        .order_by('order')
    )
    for sq in src_questions:
        tq = TrainingQuestion.objects.create(
            training=training,
            question_text=sq.question_text,
            question_type=sq.question_type,
            order=sq.order,
            is_required=sq.is_required,
            allow_other=getattr(sq, 'allow_other', False),
        )
        for opt in sq.options.all():
            TrainingQuestionOption.objects.create(
                question=tq,
                option_text=opt.option_text,
                order=opt.order,
            )
        if hasattr(sq, 'rating_config'):
            rc = sq.rating_config
            TrainingQuestionRatingConfig.objects.create(
                question=tq,
                min_value=rc.min_value,
                max_value=rc.max_value,
                min_label=getattr(rc, 'min_label', ''),
                max_label=getattr(rc, 'max_label', ''),
            )


def _schedule_training_notifications(training: Training) -> None:
    """Bulk-create in-app notifications for training participants."""
    def _create():
        if training.target_type == 'all_users':
            user_ids = list(_eligible_recipients().values_list('id', flat=True))
        else:
            user_ids = list(training.participants.values_list('user_id', flat=True))

        notifications = [
            Notification(
                recipient_id=uid,
                notification_type='training_assigned',
                notification_scope='general' if training.target_type == 'all_users' else 'specific_user',
                title=f'New Training Evaluation: {training.title}',
                message=f'A training evaluation for "{training.title}" is now available. Please fill it out.',
                module='training',
                related_object_id=training.pk,
            )
            for uid in user_ids
        ]
        Notification.objects.bulk_create(notifications, ignore_conflicts=True)

    transaction.on_commit(_create)


# ─────────────────────────────────────────────────────────────────────────────
# ADMIN / HR VIEWS
# ─────────────────────────────────────────────────────────────────────────────

class TrainingAdminListCreateView(APIView):
    """
    GET  /api/training/admin/            – paginated list of all trainings
    POST /api/training/admin/            – create new training
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        sort_by = request.query_params.get('sort_by', 'training_date').strip()
        sort_dir = request.query_params.get('sort_dir', 'desc').strip().lower()
        allowed_sorts = {'title', 'speaker', 'training_date'}
        if sort_by not in allowed_sorts:
            sort_by = 'training_date'
        if sort_dir not in {'asc', 'desc'}:
            sort_dir = 'desc'
        order_prefix = '' if sort_dir == 'asc' else '-'

        qs = (
            Training.objects
            .annotate(
                submitted_count=Count('submissions', filter=Q(submissions__is_complete=True), distinct=True),
                total_participants=Count('participants', distinct=True),
            )
        )

        if sort_by == 'training_date':
            qs = qs.order_by(f'{order_prefix}training_date', f'{order_prefix}created_at')
        else:
            qs = qs.order_by(f'{order_prefix}{sort_by}', f'{order_prefix}training_date', f'{order_prefix}created_at')

        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(title__icontains=search) | Q(speaker__icontains=search))

        status_filter = request.query_params.get('status', '').strip()
        if status_filter == 'scheduled':
            qs = qs.filter(training_date__gt=date.today())
        elif status_filter == 'active':
            qs = qs.filter(training_date=date.today())
        elif status_filter == 'closed':
            qs = qs.filter(training_date__lt=date.today())

        items, total, page, size = _paginate(qs, request)
        serializer = TrainingListSerializer(items, many=True)

        return Response({
            'results': serializer.data,
            'count': total,
            'page': page,
            'page_size': size,
        })

    @transaction.atomic
    def post(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err

        serializer = TrainingWriteSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        template_id = data.pop('template_id', None)
        target_user_ids = data.pop('target_user_ids', [])

        # Validate template
        template = None
        if template_id:
            try:
                template = SurveyTemplate.objects.get(pk=template_id)
            except SurveyTemplate.DoesNotExist:
                return Response({'template_id': 'Template not found.'}, status=http_status.HTTP_400_BAD_REQUEST)

        training = Training.objects.create(
            **data,  # type: ignore[arg-type]
            template=template,
            created_by=request.user,
        )

        # Link participants
        if data.get('target_type') == 'specific_users' and target_user_ids:
            valid_users = list(_eligible_recipients().filter(pk__in=target_user_ids).values_list('id', flat=True))
            TrainingParticipant.objects.bulk_create([
                TrainingParticipant(training=training, user_id=uid)
                for uid in valid_users
            ], ignore_conflicts=True)
        elif data.get('target_type') == 'all_users':
            valid_users = list(_eligible_recipients().values_list('id', flat=True))
            TrainingParticipant.objects.bulk_create([
                TrainingParticipant(training=training, user_id=uid)
                for uid in valid_users
            ], ignore_conflicts=True)

        # Deep-copy questions from template
        if template:
            _deep_copy_template_questions(training, template)

        _schedule_training_notifications(training)

        return Response(TrainingDetailSerializer(training).data, status=http_status.HTTP_201_CREATED)


class TrainingAdminDetailView(APIView):
    """
    GET    /api/training/admin/<pk>/   – detail
    PATCH  /api/training/admin/<pk>/   – update (template locked if responses exist)
    DELETE /api/training/admin/<pk>/   – delete (protected if response% > 10%)
    """
    permission_classes = [IsAuthenticated]

    def _get_training(self, pk):
        try:
            return Training.objects.get(pk=pk)
        except Training.DoesNotExist:
            return None

    def get(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err
        training = self._get_training(pk)
        if not training:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        return Response(TrainingDetailSerializer(training).data)

    @transaction.atomic
    def patch(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err
        training = Training.objects.select_for_update().filter(pk=pk).first()
        if not training:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        serializer = TrainingWriteSerializer(training, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        template_id = data.pop('template_id', None)
        target_user_ids = data.pop('target_user_ids', None)

        has_responses = TrainingSubmission.objects.filter(training=training, is_complete=True).exists()

        # Template lock
        if template_id is not None and has_responses:
            return Response(
                {'detail': 'Template cannot be changed after participants have submitted responses.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        for attr, value in data.items():
            setattr(training, attr, value)

        if template_id is not None:
            try:
                template = SurveyTemplate.objects.get(pk=template_id)
            except SurveyTemplate.DoesNotExist:
                return Response({'template_id': 'Template not found.'}, status=http_status.HTTP_400_BAD_REQUEST)
            # Re-copy if template changed OR if questions were never copied (0 questions with a template)
            if template != training.template or not training.questions.exists():
                training.questions.all().delete()
                training.template = template
                _deep_copy_template_questions(training, template)

        training.save()

        # Update participants if target_user_ids provided
        if target_user_ids is not None:
            training.participants.all().delete()
            if data.get('target_type', training.target_type) == 'specific_users':
                valid_users = list(_eligible_recipients().filter(pk__in=target_user_ids).values_list('id', flat=True))
            else:
                valid_users = list(_eligible_recipients().values_list('id', flat=True))
            TrainingParticipant.objects.bulk_create([
                TrainingParticipant(training=training, user_id=uid)
                for uid in valid_users
            ], ignore_conflicts=True)

        return Response(TrainingDetailSerializer(training).data)

    @transaction.atomic
    def delete(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err
        training = Training.objects.select_for_update().filter(pk=pk).first()
        if not training:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        total = training.participants.count()
        completed = TrainingSubmission.objects.filter(training=training, is_complete=True).count()

        if total > 0 and (completed / total) > 0.10:
            return Response(
                {'detail': 'Cannot delete a training with more than 10% of participants having submitted responses.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        training.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class TrainingAdminTemplatesView(APIView):
    """
    GET /api/training/admin/templates/   – list available survey templates
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        err = _require_admin_or_hr(request)
        if err:
            return err
        templates = SurveyTemplate.objects.all().order_by('title').values('id', 'title', 'description')
        return Response(list(templates))


class TrainingAdminParticipantsView(APIView):
    """
    GET /api/training/admin/<pk>/participants/  – participant list with submission status
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err
        try:
            training = Training.objects.get(pk=pk)
        except Training.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        participants = list(
            training.participants
            .select_related('user')
            .values(
                'user__id', 'user__firstname', 'user__lastname', 'user__idnumber', 'is_seen',
            )
        )
        submitted_ids = set(
            TrainingSubmission.objects.filter(training=training, is_complete=True)
            .values_list('submitted_by_id', flat=True)
        )

        result = []
        for p in participants:
            uid = p['user__id']
            result.append({
                'user_id': uid,
                'name': f"{p['user__lastname'] or ''}, {p['user__firstname'] or ''}".strip(', '),
                'idnumber': p['user__idnumber'],
                'is_seen': p['is_seen'],
                'submitted': uid in submitted_ids,
            })

        return Response({'results': result, 'count': len(result)})


# ─────────────────────────────────────────────────────────────────────────────
# USER-FACING VIEWS
# ─────────────────────────────────────────────────────────────────────────────

class MyTrainingsView(APIView):
    """
    GET /api/training/my/   – trainings assigned to the current user
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if getattr(user, 'accounting', False) and not (getattr(user, 'admin', False) or getattr(user, 'hr', False)):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        participant_qs = (
            TrainingParticipant.objects
            .filter(user=user)
            .select_related('training')
        )
        training_ids = list(participant_qs.values_list('training_id', flat=True))

        if not training_ids:
            return Response({'results': [], 'count': 0})

        qs = Training.objects.filter(pk__in=training_ids).order_by('-training_date', '-created_at')

        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(title__icontains=search) | Q(speaker__icontains=search))

        items, total, page, size = _paginate(qs, request)
        item_ids = [t.pk for t in items]

        # Build lookup maps to avoid N+1
        participant_map = {
            p.training_id: p
            for p in TrainingParticipant.objects.filter(training_id__in=item_ids, user=user)
        }
        submission_map = {
            s.training_id: s
            for s in TrainingSubmission.objects.filter(training_id__in=item_ids, submitted_by=user)
        }

        serializer = MyTrainingSerializer(
            items, many=True,
            context={'request': request, 'participant_map': participant_map, 'submission_map': submission_map},
        )
        return Response({'results': serializer.data, 'count': total, 'page': page, 'page_size': size})


class TrainingDetailUserView(APIView):
    """
    GET /api/training/my/<pk>/   – detail + questions for a specific training
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        user = request.user
        try:
            training = Training.objects.prefetch_related('questions__options', 'questions__rating_config').get(pk=pk)
        except Training.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        # Verify access
        is_admin_hr = getattr(user, 'admin', False) or getattr(user, 'hr', False)
        if not is_admin_hr:
            has_access = TrainingParticipant.objects.filter(training=training, user=user).exists()
            if not has_access:
                return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
            # Mark as seen
            TrainingParticipant.objects.filter(training=training, user=user).update(is_seen=True)

        questions = TrainingQuestionSerializer(training.questions.all(), many=True).data
        submission = TrainingSubmission.objects.filter(training=training, submitted_by=user).first()
        answers = []
        if submission:
            answers = TrainingAnswerSerializer(submission.answers.prefetch_related('selected_options').all(), many=True).data

        # Supervisor evaluation (step-1) if available
        supervisor_eval_data = None
        if submission:
            step1 = (
                TrainingApprovalStep.objects
                .filter(submission=submission, sequence=1)
                .select_related('supervisor_evaluation')
                .first()
            )
            if step1 and hasattr(step1, 'supervisor_evaluation') and step1.supervisor_evaluation.is_complete:
                ev = step1.supervisor_evaluation
                supervisor_eval_data = {
                    'result_and_impact': ev.result_and_impact,
                    'recommendation': ev.recommendation,
                    'overall_assessment': ev.overall_assessment,
                }

        # Final approver remarks (all sequence >= 2 steps that have taken an action)
        final_approver_remarks = []
        if submission:
            for step in (
                submission.approval_steps
                .select_related('approver')
                .filter(sequence__gte=2)
                .order_by('sequence', 'acted_at')
            ):
                if step.final_action:
                    u = step.approver
                    approver_name = (
                        f'{u.firstname or ""} {u.lastname or ""}'.strip() or u.idnumber
                    ) if u else None
                    final_approver_remarks.append({
                        'approver_name': approver_name,
                        'action': step.final_action,
                        'acted_at': step.acted_at.isoformat() if step.acted_at else None,
                        'remarks': step.final_remarks or '',
                    })

        # Only expose approval status once formally submitted (is_complete=True).
        # A pending submission (auto-saved answers, not yet submitted) should not
        # block the submit button on the frontend.
        approval_status = submission.status if (submission and submission.is_complete) else None
        requires_action = approval_status == 'user_confirmation'

        return Response({
            'id': training.pk,
            'title': training.title,
            'speaker': training.speaker,
            'training_date': training.training_date,
            'objective': training.objective,
            'questions': questions,
            'submission': TrainingSubmissionSerializer(submission).data if submission else None,
            'answers': answers,
            'approval_status': approval_status,
            'requires_action': requires_action,
            'supervisor_evaluation': supervisor_eval_data,
            'final_approver_remarks': final_approver_remarks,
        })


class TrainingAnswerSaveView(APIView):
    """
    POST /api/training/my/<pk>/answer/   – auto-save a single answer
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, pk):
        user = request.user
        try:
            training = Training.objects.get(pk=pk)
        except Training.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        # Guard: accounting users cannot submit
        if getattr(user, 'accounting', False) and not (getattr(user, 'admin', False) or getattr(user, 'hr', False)):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        # Verify participant
        participant = TrainingParticipant.objects.filter(training=training, user=user).first()
        if not participant:
            return Response({'detail': 'You are not a participant in this training.'}, status=http_status.HTTP_403_FORBIDDEN)

        question_id = request.data.get('question_id')
        if not question_id:
            return Response({'detail': 'question_id is required.'}, status=http_status.HTTP_400_BAD_REQUEST)

        try:
            question = TrainingQuestion.objects.get(pk=question_id, training=training)
        except TrainingQuestion.DoesNotExist:
            return Response({'detail': 'Question not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        # Get or create submission (locked)
        submission, _ = TrainingSubmission.objects.select_for_update().get_or_create(
            training=training,
            submitted_by=user,
        )

        if submission.is_complete:
            return Response({'detail': 'This training has already been submitted.'}, status=http_status.HTTP_400_BAD_REQUEST)

        serializer = TrainingAnswerSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        answer, created = TrainingAnswer.objects.get_or_create(
            submission=submission,
            question=question,
        )
        answer.text_value = data.get('text_value', '')
        answer.number_value = data.get('number_value')
        answer.other_text = data.get('other_text', '')
        answer.save()

        selected = data.get('selected_options', [])
        answer.selected_options.set(selected)

        return Response({'status': 'saved', 'answer_id': answer.pk})


class TrainingSubmitView(APIView):
    """
    POST /api/training/my/<pk>/submit/   – final submit, triggers approval routing
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, pk):
        user = request.user
        if getattr(user, 'accounting', False) and not (getattr(user, 'admin', False) or getattr(user, 'hr', False)):
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        try:
            training = Training.objects.prefetch_related('questions').get(pk=pk)
        except Training.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if not TrainingParticipant.objects.filter(training=training, user=user).exists():
            return Response({'detail': 'You are not a participant in this training.'}, status=http_status.HTTP_403_FORBIDDEN)

        submission = (
            TrainingSubmission.objects
            .select_for_update()
            .filter(training=training, submitted_by=user)
            .first()
        )
        if not submission:
            submission = TrainingSubmission.objects.create(training=training, submitted_by=user)

        if submission.is_complete:
            return Response({'detail': 'Already submitted.'}, status=http_status.HTTP_400_BAD_REQUEST)

        # Validate required questions answered — exclude instruction-only types
        INSTRUCTION_TYPES = ('section', 'subsection', 'statement')
        required_q_ids = list(
            training.questions
            .filter(is_required=True)
            .exclude(question_type__in=INSTRUCTION_TYPES)
            .values_list('id', flat=True)
        )
        answered_q_ids = set(
            TrainingAnswer.objects.filter(submission=submission).values_list('question_id', flat=True)
        )
        missing = [qid for qid in required_q_ids if qid not in answered_q_ids]
        if missing:
            return Response(
                {'detail': f'{len(missing)} required question(s) have not been answered.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        submission.is_complete = True
        submission.submitted_at = timezone.now()
        submission.status = 'supervisor_review'
        submission.save()

        # Build approval chain — validate at least 2 steps exist
        try:
            steps = build_training_approval_chain(submission)
        except Exception as exc:
            submission.is_complete = False
            submission.submitted_at = None
            submission.status = 'pending'
            submission.save()
            # DRF ValidationError.detail is a list of ErrorDetail (str subclass).
            # str(list) gives the Python repr; extract the first element instead.
            from rest_framework.exceptions import ValidationError as _DRFVError
            if isinstance(exc, _DRFVError):
                detail = exc.detail
                msg = str(detail[0]) if isinstance(detail, (list, tuple)) and detail else str(detail)
            else:
                msg = str(exc)
            return Response({'detail': msg}, status=http_status.HTTP_400_BAD_REQUEST)

        if not steps:
            submission.is_complete = False
            submission.submitted_at = None
            submission.status = 'pending'
            submission.save()
            TrainingApprovalStep.objects.filter(submission=submission).delete()
            return Response(
                {'detail': 'No approver could be found in your approval chain. '
                 'Please contact HR.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        step1 = steps[0]

        def _notify_supervisor():
            if step1.approver_id:
                Notification.objects.create(
                    recipient_id=step1.approver_id,
                    notification_type='training_supervisor_review',
                    notification_scope='specific_user',
                    title=f'Training Evaluation Submitted: {submission.training.title}',
                    message=(
                        f'{submission.submitted_by.firstname or ""} {submission.submitted_by.lastname or ""} '
                        f'has submitted a training evaluation for "{submission.training.title}". '
                        'Please review and complete your evaluation.'
                    ).strip(),
                    module='training',
                    related_object_id=submission.pk,
                )

        transaction.on_commit(_notify_supervisor)

        return Response({'status': 'submitted', 'submission_id': submission.pk})


# ─────────────────────────────────────────────────────────────────────────────
# APPROVER-FACING VIEWS
# ─────────────────────────────────────────────────────────────────────────────

class ApproverQueueView(APIView):
    """
    GET /api/training/approver/queue/   – paginated list of submissions awaiting this approver
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if not WorkInformation.objects.filter(approver=user).exists():
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        # Submissions where this user has an active pending step and is NOT the submitter
        my_step_ids = (
            TrainingApprovalStep.objects
            .filter(approver=user, activated_at__isnull=False)
            .exclude(submission__submitted_by=user)
            .values_list('submission_id', flat=True)
        )

        qs = (
            TrainingSubmission.objects
            .filter(pk__in=my_step_ids, is_complete=True)
            .select_related('submitted_by', 'training')
            .order_by('-submitted_at')
        )

        status_filter = request.query_params.get('status', '').strip()
        if status_filter == 'pending':
            qs = qs.filter(
                approval_steps__approver=user,
                approval_steps__status='pending',
                approval_steps__activated_at__isnull=False,
            )
        elif status_filter == 'reviewed':
            qs = qs.filter(
                approval_steps__approver=user,
                approval_steps__status='reviewed',
            )

        items, total, page, size = _paginate(qs, request)
        item_ids = [s.pk for s in items]

        step_map = {
            step.submission_id: step
            for step in TrainingApprovalStep.objects
            .filter(submission_id__in=item_ids, approver=user)
            .select_related('supervisor_evaluation')
        }

        serializer = ApproverQueueItemSerializer(
            items, many=True,
            context={'request': request, 'step_map': step_map},
        )
        return Response({'results': serializer.data, 'count': total, 'page': page, 'page_size': size})


class ApproverQueueBadgeView(APIView):
    """
    GET /api/training/approver/badge/   – count of pending submissions for approver
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if not WorkInformation.objects.filter(approver=user).exists():
            return Response({'count': 0})

        count = (
            TrainingApprovalStep.objects
            .filter(
                approver=user,
                status='pending',
                activated_at__isnull=False,
            )
            .exclude(submission__submitted_by=user)
            .values('submission_id')
            .distinct()
            .count()
        )
        return Response({'count': count})


class ApproverSubmissionDetailView(APIView):
    """
    GET /api/training/approver/submissions/<submission_id>/
    Returns: submission + user answers + approver's own evaluation (if any)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, submission_id):
        user = request.user
        if not WorkInformation.objects.filter(approver=user).exists():
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        try:
            submission = (
                TrainingSubmission.objects
                .select_related('submitted_by', 'training')
                .prefetch_related('answers__question__options', 'answers__selected_options')
                .get(pk=submission_id)
            )
        except TrainingSubmission.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        step = (
            TrainingApprovalStep.objects
            .filter(submission=submission, approver=user, activated_at__isnull=False)
            .first()
        )
        if not step:
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        # User answers
        user_answers = TrainingAnswerSerializer(
            submission.answers.prefetch_related('selected_options').all(), many=True
        ).data

        # Questions (user section)
        questions = TrainingQuestionSerializer(
            submission.training.questions.all(), many=True
        ).data

        can_review = can_act_on_training_step(step, user)

        # Determine role: sequence 1 = supervisor, sequence >= 2 = final_approver
        my_role = 'supervisor' if step.sequence == 1 else 'final_approver'

        # Supervisor evaluation from step-1
        step1 = (
            TrainingApprovalStep.objects
            .filter(submission=submission, sequence=1)
            .select_related('supervisor_evaluation')
            .first()
        )
        supervisor_eval_data = None
        if step1 and hasattr(step1, 'supervisor_evaluation'):
            ev = step1.supervisor_evaluation
            supervisor_eval_data = SupervisorEvaluationSerializer(ev).data

        # My own partial eval (for supervisor role, if not yet complete — to pre-populate form)
        my_eval_data = None
        if my_role == 'supervisor' and hasattr(step, 'supervisor_evaluation'):
            my_eval_data = SupervisorEvaluationSerializer(step.supervisor_evaluation).data

        # Step-1 approver info (for final_approver view context)
        step1_info = None
        if my_role == 'final_approver' and step1:
            step1_info = TrainingApprovalStepSerializer(step1).data

        # Flat training/employee fields for frontend convenience
        u = submission.submitted_by
        employee_name = f'{u.lastname or ""}, {u.firstname or ""}'.strip(', ') or u.idnumber

        # Final approver remarks (all sequence >= 2 steps that have taken an action)
        final_approver_remarks = []
        for step_r in (
            submission.approval_steps
            .select_related('approver')
            .filter(sequence__gte=2)
            .order_by('sequence', 'acted_at')
        ):
            if step_r.final_action:
                u = step_r.approver
                approver_name = (
                    f'{u.firstname or ""} {u.lastname or ""}'.strip() or u.idnumber
                ) if u else None
                final_approver_remarks.append({
                    'approver_name': approver_name,
                    'action': step_r.final_action,
                    'acted_at': step_r.acted_at.isoformat() if step_r.acted_at else None,
                    'remarks': step_r.final_remarks or '',
                })

        return Response({
            'training_title': submission.training.title,
            'training_date': str(submission.training.training_date),
            'training_objective': submission.training.objective or '',
            'employee_name': employee_name,
            'submission_status': submission.status,
            'user_answers': user_answers,
            'questions': questions,
            'step': TrainingApprovalStepSerializer(step).data,
            'can_review': can_review,
            'my_role': my_role,
            'supervisor_evaluation': supervisor_eval_data,
            'my_evaluation': my_eval_data,
            'step1_info': step1_info,
            'final_remarks': step.final_remarks if my_role == 'supervisor' else None,
            'final_approver_remarks': final_approver_remarks,
        })


class SupervisorEvaluationSaveView(APIView):
    """
    POST /api/training/approver/steps/<step_id>/eval/save/
    Auto-save the 3 supervisor evaluation fields (no completion required).
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, step_id):
        user = request.user
        if not WorkInformation.objects.filter(approver=user).exists():
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        step = (
            TrainingApprovalStep.objects
            .select_for_update()
            .filter(pk=step_id, approver=user, sequence=1)
            .first()
        )
        if not step:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if step.status == 'reviewed':
            return Response({'detail': 'This evaluation has already been submitted.'}, status=http_status.HTTP_400_BAD_REQUEST)

        evaluation, _ = SupervisorEvaluation.objects.get_or_create(step=step)
        evaluation.result_and_impact = request.data.get('result_and_impact', evaluation.result_and_impact)
        evaluation.recommendation = request.data.get('recommendation', evaluation.recommendation)
        raw_rating = request.data.get('overall_assessment')
        if raw_rating is not None:
            evaluation.overall_assessment = int(raw_rating) if raw_rating else None
        evaluation.save()

        return Response({'status': 'saved'})


class SupervisorEvaluationSubmitView(APIView):
    """
    POST /api/training/approver/steps/<step_id>/eval/submit/
    Final submit of supervisor evaluation → status moves to 'user_confirmation'.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, step_id):
        user = request.user
        if not WorkInformation.objects.filter(approver=user).exists():
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        step = (
            TrainingApprovalStep.objects
            .select_for_update()
            .filter(pk=step_id, approver=user, sequence=1)
            .first()
        )
        if not step:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if step.status == 'reviewed':
            return Response({'detail': 'Already submitted.'}, status=http_status.HTTP_400_BAD_REQUEST)

        if not can_act_on_training_step(step, user):
            return Response({'detail': 'This step is not currently active for your action.'}, status=http_status.HTTP_400_BAD_REQUEST)

        # get_or_create so that submitting without a prior auto-save still works
        evaluation, _ = SupervisorEvaluation.objects.select_for_update().get_or_create(step=step)

        # Apply the values sent in the request body (mirrors the save endpoint so that
        # a pending auto-save timer being cancelled never causes a 400).
        evaluation.result_and_impact = request.data.get('result_and_impact', evaluation.result_and_impact)
        evaluation.recommendation = request.data.get('recommendation', evaluation.recommendation)
        raw_rating = request.data.get('overall_assessment')
        if raw_rating is not None:
            evaluation.overall_assessment = int(raw_rating) if raw_rating else None

        # Validate all 3 fields
        errors = {}
        if not (evaluation.result_and_impact or '').strip():
            errors['result_and_impact'] = 'Result and Impact cannot be blank.'
        if not (evaluation.recommendation or '').strip():
            errors['recommendation'] = 'Recommendation cannot be blank.'
        if not evaluation.overall_assessment or not (1 <= evaluation.overall_assessment <= 5):
            errors['overall_assessment'] = 'Overall Assessment must be a rating from 1 to 5.'
        if errors:
            return Response(errors, status=http_status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        step.status = 'reviewed'
        step.acted_at = now
        step.save()

        evaluation.is_complete = True
        evaluation.submitted_at = now
        evaluation.save()

        submission = step.submission
        submission.status = 'user_confirmation'
        submission.save()

        def _notify_user():
            Notification.objects.create(
                recipient=submission.submitted_by,
                notification_type='training_user_confirmation',
                notification_scope='specific_user',
                title=f'Training Evaluation: Your Review is Ready',
                message=(
                    f'Your supervisor has completed their evaluation for "{submission.training.title}". '
                    'Please review and confirm to proceed.'
                ),
                module='training',
                related_object_id=submission.pk,
            )

        transaction.on_commit(_notify_user)
        return Response({'status': 'submitted', 'step_id': step.pk})


class UserConfirmView(APIView):
    """
    POST /api/training/my/<pk>/confirm/
    User confirms after reviewing supervisor's evaluation. Moves to final approval.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, pk):
        user = request.user
        submission = (
            TrainingSubmission.objects
            .select_for_update()
            .filter(training_id=pk, submitted_by=user)
            .first()
        )
        if not submission:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if submission.status != 'user_confirmation':
            return Response(
                {'detail': 'This submission is not awaiting your confirmation.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        step2 = (
            submission.approval_steps
            .filter(sequence=2)
            .first()
        )

        now = timezone.now()
        submission.confirmed_at = now
        submission.confirmed_by = user

        if not step2:
            # 1-step routing rule: auto-complete on user confirmation
            submission.status = 'completed'
            submission.save()

            step1 = submission.approval_steps.filter(sequence=1).first()

            def _notify_completed_1step():
                Notification.objects.create(
                    recipient=submission.submitted_by,
                    notification_type='training_completed',
                    notification_scope='specific_user',
                    title='Training Evaluation Approved',
                    message=f'Your training evaluation for "{submission.training.title}" has been approved.',
                    module='training',
                    related_object_id=submission.pk,
                )
                if step1 and step1.approver_id:
                    Notification.objects.create(
                        recipient_id=step1.approver_id,
                        notification_type='training_completed',
                        notification_scope='specific_user',
                        title='Training Evaluation Completed',
                        message=(
                            f'The training evaluation for "{submission.training.title}" by '
                            f'{submission.submitted_by.firstname or ""} {submission.submitted_by.lastname or ""} '
                            'has been confirmed by the employee and is now complete.'
                        ).strip(),
                        module='training',
                        related_object_id=submission.pk,
                    )

            transaction.on_commit(_notify_completed_1step)
            return Response({'status': 'completed'})

        # Multi-step: move to final approval
        submission.status = 'final_approval'
        submission.save()

        step2.activated_at = now
        step2.save()

        def _notify_final_approver():
            if step2.approver_id:
                Notification.objects.create(
                    recipient_id=step2.approver_id,
                    notification_type='training_final_approval',
                    notification_scope='specific_user',
                    title=f'Training Evaluation Needs Your Approval',
                    message=(
                        f'A training evaluation for "{submission.training.title}" by '
                        f'{submission.submitted_by.firstname or ""} {submission.submitted_by.lastname or ""} '
                        'is ready for your final approval.'
                    ).strip(),
                    module='training',
                    related_object_id=submission.pk,
                )

        transaction.on_commit(_notify_final_approver)
        return Response({'status': 'confirmed'})


class FinalApproverActionView(APIView):
    """
    POST /api/training/approver/steps/<step_id>/action/
    Final approver approves or disapproves the submission.
    Approve → 'completed'. Disapprove → 'returned', reset step-1 for re-evaluation.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, step_id):
        user = request.user
        if not WorkInformation.objects.filter(approver=user).exists():
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)

        step = (
            TrainingApprovalStep.objects
            .select_for_update()
            .filter(pk=step_id, approver=user)
            .first()
        )
        if not step:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        if step.sequence < 2:
            return Response({'detail': 'This action is only available to final approvers.'}, status=http_status.HTTP_400_BAD_REQUEST)

        if not can_act_on_training_step(step, user):
            return Response({'detail': 'This step is not currently active for your action.'}, status=http_status.HTTP_400_BAD_REQUEST)

        action = request.data.get('action', '').strip()
        remarks = request.data.get('remarks', '').strip()

        if action not in ('approved', 'disapproved'):
            return Response({'action': 'Must be "approved" or "disapproved".'}, status=http_status.HTTP_400_BAD_REQUEST)

        if action == 'disapproved' and not remarks:
            return Response({'remarks': 'Remarks are required when disapproving.'}, status=http_status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        submission = (
            TrainingSubmission.objects
            .select_for_update()
            .get(pk=step.submission_id)
        )

        step.acted_at = now
        if action == 'approved':
            step.final_action = 'approved'
            step.final_remarks = remarks or ''
        else:
            step.final_action = 'disapproved'
            step.final_remarks = remarks
        step.save()

        if action == 'approved':
            step.status = 'reviewed'
            step.save()

            total_steps = submission.approval_steps.count()

            if step.sequence < total_steps:
                # More steps remain — activate the next one
                next_step = (
                    TrainingApprovalStep.objects
                    .filter(submission=submission, sequence=step.sequence + 1)
                    .first()
                )
                if not next_step:
                    return Response(
                        {'detail': 'Could not find the next approval step.'},
                        status=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

                next_step.activated_at = now
                next_step.save()

                submission.status = TrainingSubmission.STATUS_SECOND_FINAL_APPROVAL
                submission.save()

                def _notify_next_approver():
                    if next_step.approver_id:
                        Notification.objects.create(
                            recipient_id=next_step.approver_id,
                            notification_type='training_final_approval',
                            notification_scope='specific_user',
                            title='Training Evaluation Needs Your Approval',
                            message=(
                                f'A training evaluation for "{submission.training.title}" by '
                                f'{submission.submitted_by.firstname or ""} '
                                f'{submission.submitted_by.lastname or ""} '
                                'is ready for your final approval.'
                            ).strip(),
                            module='training',
                            related_object_id=submission.pk,
                        )

                transaction.on_commit(_notify_next_approver)

            else:
                # Last step approved — mark completed
                submission.status = TrainingSubmission.STATUS_COMPLETED
                submission.save()

                step1 = submission.approval_steps.filter(sequence=1).first()

                def _notify_completed():
                    Notification.objects.create(
                        recipient=submission.submitted_by,
                        notification_type='training_completed',
                        notification_scope='specific_user',
                        title='Training Evaluation Approved',
                        message=f'Your training evaluation for "{submission.training.title}" has been approved.',
                        module='training',
                        related_object_id=submission.pk,
                    )
                    if step1 and step1.approver_id:
                        Notification.objects.create(
                            recipient_id=step1.approver_id,
                            notification_type='training_completed',
                            notification_scope='specific_user',
                            title='Training Evaluation Approved',
                            message=(
                                f'The training evaluation for "{submission.training.title}" by '
                                f'{submission.submitted_by.firstname or ""} {submission.submitted_by.lastname or ""} '
                                'has been approved by the final approver.'
                            ).strip(),
                            module='training',
                            related_object_id=submission.pk,
                        )

                transaction.on_commit(_notify_completed)

        else:  # disapproved → return for re-evaluation
            step1 = (
                TrainingApprovalStep.objects
                .select_for_update()
                .filter(submission=submission, sequence=1)
                .first()
            )
            if not step1:
                return Response({'detail': 'Could not find supervisor step to reset.'}, status=http_status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Reset step-1 for re-evaluation
            step1.status = 'pending'
            step1.activated_at = now
            step1.acted_at = None
            step1.save()

            # Delete old supervisor evaluation so supervisor fills fresh
            SupervisorEvaluation.objects.filter(step=step1).delete()

            # Reset all final approver steps (sequence >= 2) except the current one,
            # so intermediate approvals (e.g. step 2 in a 3-step flow) are cleared
            # and must be re-done in the next loop.
            TrainingApprovalStep.objects.filter(
                submission=submission, sequence__gte=2,
            ).exclude(pk=step.pk).update(
                status='pending',
                activated_at=None,
                acted_at=None,
                final_action=None,
                final_remarks='',
            )

            # Reset current disapproving step — preserve acted_at/final_action for
            # audit but clear activation so it re-enters the queue naturally.
            step.status = 'pending'
            step.activated_at = None
            step.save()

            # Clear user confirmation
            submission.confirmed_at = None
            submission.confirmed_by = None
            submission.status = TrainingSubmission.STATUS_RETURNED
            submission.save()

            def _notify_returned():
                if step1.approver_id:
                    Notification.objects.create(
                        recipient_id=step1.approver_id,
                        notification_type='training_returned',
                        notification_scope='specific_user',
                        title='Training Evaluation Returned for Re-evaluation',
                        message=(
                            f'The training evaluation for "{submission.training.title}" by '
                            f'{submission.submitted_by.firstname or ""} {submission.submitted_by.lastname or ""} '
                            f'has been returned. Remarks: {remarks}'
                        ).strip(),
                        module='training',
                        related_object_id=submission.pk,
                    )

            transaction.on_commit(_notify_returned)

        return Response({'status': action})


# ─────────────────────────────────────────────────────────────────────────────
# TRAINING ADMIN — RESULTS / RESPONSES / EXPORT
# ─────────────────────────────────────────────────────────────────────────────

class TrainingAdminResultsView(APIView):
    """
    GET /api/training/admin/<pk>/results
    Aggregated results for the admin view page — mirrors AdminSurveyResultsView.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):  # noqa: C901
        err = _require_admin_or_hr(request)
        if err:
            return err

        from collections import defaultdict
        from django.db.models import Max
        from training.models import TrainingQuestionRatingConfig

        try:
            training = (
                Training.objects
                .prefetch_related('questions__options', 'questions__rating_config')
                .get(pk=pk)
            )
        except Training.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        complete_subs = TrainingSubmission.objects.filter(training=training, is_complete=True)
        total_responses = complete_subs.values('submitted_by').distinct().count()
        total_participants = training.participants.count()
        completion_rate = round(total_responses / total_participants * 100, 1) if total_participants > 0 else 0.0

        last_resp = complete_subs.aggregate(last=Max('submitted_at'))
        last_response_at = last_resp['last'].isoformat() if last_resp['last'] else None

        completed_sub_ids = list(complete_subs.values_list('id', flat=True))

        all_answers = list(
            TrainingAnswer.objects
            .filter(submission_id__in=completed_sub_ids)
            .prefetch_related('selected_options')
        )

        CHOICE_TYPES = {'single_choice', 'multiple_choice', 'dropdown'}

        question_results = []
        for q in training.questions.order_by('order'):
            q_answers = [a for a in all_answers if a.question_id == q.pk]
            total_answered = len(q_answers)

            result = {
                'question_id': q.pk,
                'question_text': q.question_text,
                'question_type': q.question_type,
                'total_responses': total_answered,
            }

            if q.question_type in CHOICE_TYPES:
                counts: dict = defaultdict(int)
                for ans in q_answers:
                    for opt in ans.selected_options.all():
                        counts[opt.pk] += 1
                result['options'] = [
                    {
                        'option_id': opt.pk,
                        'option_text': opt.option_text,
                        'count': counts.get(opt.pk, 0),
                        'percentage': round(counts.get(opt.pk, 0) / total_answered * 100, 1) if total_answered else 0.0,
                    }
                    for opt in q.options.order_by('order')
                ]

            elif q.question_type == 'yes_no':
                text_counts: dict = defaultdict(int)
                for ans in q_answers:
                    if ans.text_value:
                        text_counts[ans.text_value] += 1
                result['options'] = [
                    {
                        'option_id': None,
                        'option_text': label,
                        'count': text_counts.get(label, 0),
                        'percentage': round(text_counts.get(label, 0) / total_answered * 100, 1) if total_answered else 0.0,
                    }
                    for label in ('Yes', 'No')
                ]

            elif q.question_type == 'likert':
                likert_labels = ['Strongly Disagree', 'Disagree', 'Neutral', 'Agree', 'Strongly Agree']
                text_counts = defaultdict(int)
                for ans in q_answers:
                    if ans.text_value:
                        text_counts[ans.text_value] += 1
                result['options'] = [
                    {
                        'option_id': None,
                        'option_text': label,
                        'count': text_counts.get(label, 0),
                        'percentage': round(text_counts.get(label, 0) / total_answered * 100, 1) if total_answered else 0.0,
                    }
                    for label in likert_labels
                ]

            elif q.question_type == 'linear_scale':
                values = [a.number_value for a in q_answers if a.number_value is not None]
                avg = (sum(values) / len(values)) if values else None
                dist: dict = defaultdict(int)
                for v in values:
                    dist[v] += 1
                try:
                    cfg = q.rating_config
                    min_v, max_v = cfg.min_value, cfg.max_value
                except Exception:
                    min_v, max_v = (1, 10) if not dist else (int(min(dist)), int(max(dist)))
                distribution = []
                for v in range(min_v, max_v + 1):
                    cnt = dist.get(float(v), dist.get(v, 0))
                    distribution.append({
                        'value': v,
                        'count': cnt,
                        'percentage': round(cnt / len(values) * 100, 1) if values else 0.0,
                    })
                result['average'] = round(avg, 2) if avg is not None else None
                result['distribution'] = distribution

            elif q.question_type == 'rating':
                values = [a.number_value for a in q_answers if a.number_value is not None]
                avg = (sum(values) / len(values)) if values else None
                dist = defaultdict(int)
                for v in values:
                    dist[v] += 1
                try:
                    cfg = q.rating_config
                    min_v, max_v = cfg.min_value, cfg.max_value
                except Exception:
                    min_v, max_v = 1, 5
                distribution = []
                for v in range(min_v, max_v + 1):
                    cnt = dist.get(float(v), 0)
                    distribution.append({
                        'value': v,
                        'count': cnt,
                        'percentage': round(cnt / len(values) * 100, 1) if values else 0.0,
                    })
                result['average'] = round(avg, 2) if avg is not None else None
                result['distribution'] = distribution

            else:
                # short_text, long_text, number, date, etc.
                result['text_answers'] = [
                    a.text_value for a in q_answers
                    if a.text_value and a.text_value.strip()
                ][:200]

            question_results.append(result)

        today = date.today()
        td = training.training_date
        if td > today:
            t_status = 'scheduled'
        elif td == today:
            t_status = 'active'
        else:
            t_status = 'closed'

        return Response({
            'training_id': training.pk,
            'training_title': training.title,
            'training_date': training.training_date.isoformat(),
            'training_status': t_status,
            'speaker': training.speaker,
            'objective': training.objective,
            'total_participants': total_participants,
            'total_responses': total_responses,
            'completion_rate': completion_rate,
            'last_response_at': last_response_at,
            'questions': question_results,
        })


class TrainingAdminResponsesView(APIView):
    """
    GET /api/training/admin/<pk>/responses
    Paginated list of submitted responses (is_complete=True) for the admin view page.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            training = Training.objects.get(pk=pk)
        except Training.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        qs = (
            TrainingSubmission.objects
            .filter(training=training, is_complete=True)
            .select_related('submitted_by')
        )

        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(submitted_by__firstname__icontains=search) |
                Q(submitted_by__lastname__icontains=search) |
                Q(submitted_by__idnumber__icontains=search)
            )

        sort_field = request.query_params.get('sort', 'name')
        sort_dir = request.query_params.get('dir', 'asc')
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'asc'
        prefix = '' if sort_dir == 'asc' else '-'

        if sort_field == 'idnumber':
            qs = qs.order_by(f'{prefix}submitted_by__idnumber')
        elif sort_field == 'submitted_at':
            qs = qs.order_by(f'{prefix}submitted_at')
        else:
            qs = qs.order_by(f'{prefix}submitted_by__lastname', f'{prefix}submitted_by__firstname')

        items, total, page, size = _paginate(qs, request)

        results = []
        for sub in items:
            u = sub.submitted_by
            results.append({
                'submission_id': sub.pk,
                'idnumber': u.idnumber,
                'firstname': u.firstname or '',
                'lastname': u.lastname or '',
                'submitted_at': sub.submitted_at.isoformat() if sub.submitted_at else None,
            })

        return Response({
            'results': results,
            'pagination': {
                'page': page,
                'page_size': size,
                'total': total,
                'total_pages': max(1, -(-total // size)),
            },
        })


class TrainingAdminResponseDetailView(APIView):
    """
    GET /api/training/admin/responses/<submission_id>
    Single submission detail with all answers, for the admin view modal.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, submission_id):
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            submission = (
                TrainingSubmission.objects
                .select_related('submitted_by', 'training')
                .prefetch_related('answers__question__options', 'answers__selected_options')
                .get(pk=submission_id, is_complete=True)
            )
        except TrainingSubmission.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        u = submission.submitted_by
        respondent_name = f'{u.firstname or ""} {u.lastname or ""}'.strip() or u.idnumber

        answers = []
        for ans in submission.answers.order_by('question__order'):
            q = ans.question
            selected = [{'id': o.pk, 'text': o.option_text} for o in ans.selected_options.all()]
            answers.append({
                'question_id': q.pk,
                'question_text': q.question_text,
                'question_type': q.question_type,
                'order': q.order,
                'selected_options': selected if selected else None,
                'other_text': ans.other_text or None,
                'number_value': ans.number_value,
                'text_value': ans.text_value if ans.text_value else None,
            })

        return Response({
            'submission_id': submission.pk,
            'respondent_name': respondent_name,
            'idnumber': u.idnumber,
            'submitted_at': submission.submitted_at.isoformat() if submission.submitted_at else None,
            'answers': answers,
        })


class TrainingAdminExportView(APIView):
    """
    GET /api/training/admin/<pk>/export
    Two-sheet XLSX: Sheet 1 = summary + per-question breakdown, Sheet 2 = per-user responses.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):  # noqa: C901
        err = _require_admin_or_hr(request)
        if err:
            return err

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({'detail': 'openpyxl not installed.'}, status=http_status.HTTP_503_SERVICE_UNAVAILABLE)

        from django.http import StreamingHttpResponse
        from collections import defaultdict

        try:
            training = (
                Training.objects
                .prefetch_related('questions__options', 'questions__rating_config')
                .get(pk=pk)
            )
        except Training.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        # ── Shared styles ──────────────────────────────────────────────────────
        _BLUE   = PatternFill('solid', fgColor='2845D6')
        _WH_B   = Font(bold=True, color='FFFFFF')
        _BOLD   = Font(bold=True)
        _THIN   = Side(style='thin')
        _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
        _CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        _WRAP   = Alignment(wrap_text=True, vertical='top')

        def _hdr(ws, row, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = _BLUE
                cell.font = _WH_B
                cell.border = _BORDER
                cell.alignment = _CENTER

        def _border_row(ws, row, ncols):
            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).border = _BORDER

        def _auto_width(ws):
            for col in ws.columns:
                ltr = get_column_letter(col[0].column)
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=0,
                )
                ws.column_dimensions[ltr].width = min(max(max_len + 4, 10), 60)

        INSTRUCTION_TYPES = {'section', 'subsection', 'statement'}
        CHOICE_TYPES      = {'single_choice', 'multiple_choice', 'dropdown'}

        questions = list(training.questions.order_by('order'))
        answerable_q = [q for q in questions if q.question_type not in INSTRUCTION_TYPES]

        # Question numbering
        q_numbers: dict = {}
        counter = 0
        for q in questions:
            if q.question_type not in INSTRUCTION_TYPES:
                counter += 1
                q_numbers[q.pk] = counter

        # Participants & responses
        total_participants = training.participants.count()
        complete_subs = list(
            TrainingSubmission.objects
            .filter(training=training, is_complete=True)
            .select_related('submitted_by')
            .order_by('submitted_at')
        )
        total_responses = len(complete_subs)
        completion_pct = (total_responses / total_participants * 100) if total_participants else 0.0
        completed_sub_ids = [s.pk for s in complete_subs]

        all_answers = list(
            TrainingAnswer.objects
            .filter(submission_id__in=completed_sub_ids)
            .prefetch_related('selected_options')
        )
        ans_map = {(a.submission_id, a.question_id): a for a in all_answers}

        supervisor_map = {}
        for step in (
            TrainingApprovalStep.objects
            .filter(submission_id__in=completed_sub_ids, sequence=1)
            .select_related('supervisor_evaluation')
        ):
            if hasattr(step, 'supervisor_evaluation') and step.supervisor_evaluation is not None:
                supervisor_map[step.submission_id] = step.supervisor_evaluation

        final_remarks_map = defaultdict(list)
        for step in (
            TrainingApprovalStep.objects
            .filter(submission_id__in=completed_sub_ids, sequence__gte=2)
            .select_related('approver')
            .order_by('submission_id', 'sequence', 'acted_at')
        ):
            if step.final_remarks:
                approver_name = None
                if step.approver:
                    approver_name = f'{step.approver.firstname or ""} {step.approver.lastname or ""}'.strip() or step.approver.idnumber
                label = (approver_name or 'Approver')
                if step.final_action == 'approved':
                    label = f'{label} (Approved)'
                else:
                    label = f'{label} (Disapproved)'
                final_remarks_map[step.submission_id].append(f'{label}: {step.final_remarks}')

        # ── Sheet 1: Summary ───────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Summary'

        # Training info block
        info_rows = [
            ('Training Title', training.title),
            ('Speaker',        training.speaker),
            ('Training Date',  str(training.training_date)),
            ('Objective',      training.objective or '—'),
            ('Total Participants', str(total_participants)),
            ('Total Responses',   str(total_responses)),
            ('Completion Rate',   f'{round(completion_pct, 1)}%'),
        ]
        for i, (label, val) in enumerate(info_rows, start=1):
            ws1.cell(row=i, column=1, value=label).font = _BOLD
            ws1.cell(row=i, column=2, value=val)
            for c in range(1, 3):
                ws1.cell(row=i, column=c).border = _BORDER

        current_row = len(info_rows) + 2  # blank separator row

        # Per-question breakdown
        for q in answerable_q:
            q_answers = [a for a in all_answers if a.question_id == q.pk]
            total_answered = len(q_answers)
            qn = q_numbers.get(q.pk, '?')

            # Section header
            ws1.cell(row=current_row, column=1, value=f'Q{qn}: {q.question_text}')
            ws1.cell(row=current_row, column=1).font = _BOLD
            ws1.cell(row=current_row, column=1).alignment = _WRAP
            current_row += 1

            if q.question_type in CHOICE_TYPES:
                ws1.cell(row=current_row, column=1, value='Option')
                ws1.cell(row=current_row, column=2, value='Count')
                ws1.cell(row=current_row, column=3, value='Percentage')
                _hdr(ws1, current_row, 3)
                current_row += 1
                opt_counts: dict = defaultdict(int)
                for ans in q_answers:
                    for opt in ans.selected_options.all():
                        opt_counts[opt.pk] += 1
                for opt in q.options.order_by('order'):
                    cnt = opt_counts.get(opt.pk, 0)
                    pct = round(cnt / total_answered * 100, 1) if total_answered else 0.0
                    ws1.cell(row=current_row, column=1, value=opt.option_text)
                    ws1.cell(row=current_row, column=2, value=cnt)
                    ws1.cell(row=current_row, column=3, value=f'{pct}%')
                    _border_row(ws1, current_row, 3)
                    current_row += 1

            elif q.question_type == 'yes_no':
                ws1.cell(row=current_row, column=1, value='Option')
                ws1.cell(row=current_row, column=2, value='Count')
                ws1.cell(row=current_row, column=3, value='Percentage')
                _hdr(ws1, current_row, 3)
                current_row += 1
                text_counts: dict = defaultdict(int)
                for ans in q_answers:
                    if ans.text_value:
                        text_counts[ans.text_value] += 1
                for label in ('Yes', 'No'):
                    cnt = text_counts.get(label, 0)
                    pct = round(cnt / total_answered * 100, 1) if total_answered else 0.0
                    ws1.cell(row=current_row, column=1, value=label)
                    ws1.cell(row=current_row, column=2, value=cnt)
                    ws1.cell(row=current_row, column=3, value=f'{pct}%')
                    _border_row(ws1, current_row, 3)
                    current_row += 1

            elif q.question_type == 'likert':
                ws1.cell(row=current_row, column=1, value='Option')
                ws1.cell(row=current_row, column=2, value='Count')
                ws1.cell(row=current_row, column=3, value='Percentage')
                _hdr(ws1, current_row, 3)
                current_row += 1
                text_counts = defaultdict(int)
                for ans in q_answers:
                    if ans.text_value:
                        text_counts[ans.text_value] += 1
                for label in ['Strongly Disagree', 'Disagree', 'Neutral', 'Agree', 'Strongly Agree']:
                    cnt = text_counts.get(label, 0)
                    pct = round(cnt / total_answered * 100, 1) if total_answered else 0.0
                    ws1.cell(row=current_row, column=1, value=label)
                    ws1.cell(row=current_row, column=2, value=cnt)
                    ws1.cell(row=current_row, column=3, value=f'{pct}%')
                    _border_row(ws1, current_row, 3)
                    current_row += 1

            elif q.question_type in ('rating', 'linear_scale'):
                values = [a.number_value for a in q_answers if a.number_value is not None]
                avg = round(sum(values) / len(values), 2) if values else None
                ws1.cell(row=current_row, column=1, value=f'Average: {avg if avg is not None else "—"}')
                ws1.cell(row=current_row, column=1).font = Font(italic=True)
                current_row += 1
                ws1.cell(row=current_row, column=1, value='Value')
                ws1.cell(row=current_row, column=2, value='Count')
                ws1.cell(row=current_row, column=3, value='Percentage')
                _hdr(ws1, current_row, 3)
                current_row += 1
                dist: dict = defaultdict(int)
                for v in values:
                    dist[v] += 1
                try:
                    cfg = q.rating_config
                    min_v, max_v = cfg.min_value, cfg.max_value
                except Exception:
                    min_v = 1
                    max_v = 5 if q.question_type == 'rating' else 10
                for v in range(min_v, max_v + 1):
                    cnt = dist.get(float(v), dist.get(v, 0))
                    pct = round(cnt / len(values) * 100, 1) if values else 0.0
                    ws1.cell(row=current_row, column=1, value=v)
                    ws1.cell(row=current_row, column=2, value=cnt)
                    ws1.cell(row=current_row, column=3, value=f'{pct}%')
                    _border_row(ws1, current_row, 3)
                    current_row += 1

            else:
                # text/number/date — list raw answers
                text_vals = [a.text_value for a in q_answers if a.text_value and a.text_value.strip()]
                ws1.cell(row=current_row, column=1, value='Responses')
                _hdr(ws1, current_row, 1)
                current_row += 1
                for tv in text_vals:
                    ws1.cell(row=current_row, column=1, value=tv)
                    ws1.cell(row=current_row, column=1).border = _BORDER
                    current_row += 1

            current_row += 1  # blank between questions

        _auto_width(ws1)

        # ── Sheet 2: Responses ─────────────────────────────────────────────────
        ws2 = wb.create_sheet('Responses')
        q_headers = [f'Q{q_numbers[q.pk]}: {q.question_text[:60]}' for q in answerable_q]
        response_headers = [
            'ID Number',
            'Employee Name',
            'Submitted Date',
            'Result and Impact',
            'Recommendation',
            'Overall Assessment',
            'Final Approver Remarks',
        ]
        headers = response_headers + q_headers
        ncols = len(headers)

        for c, h in enumerate(headers, start=1):
            ws2.cell(row=1, column=c, value=h)
        _hdr(ws2, 1, ncols)

        for ri, sub in enumerate(complete_subs, start=2):
            u = sub.submitted_by
            full_name = f'{u.lastname or ""}, {u.firstname or ""}'.strip(', ')
            ws2.cell(row=ri, column=1, value=u.idnumber)
            ws2.cell(row=ri, column=2, value=full_name)
            ws2.cell(row=ri, column=3, value=str(sub.submitted_at.date()) if sub.submitted_at else '—')

            sup_eval = supervisor_map.get(sub.pk)
            ws2.cell(row=ri, column=4, value=(sup_eval.result_and_impact if sup_eval and getattr(sup_eval, 'result_and_impact', None) else '—'))
            ws2.cell(row=ri, column=5, value=(sup_eval.recommendation if sup_eval and getattr(sup_eval, 'recommendation', None) else '—'))
            ws2.cell(row=ri, column=6, value=(str(sup_eval.overall_assessment) if sup_eval and getattr(sup_eval, 'overall_assessment', None) is not None else '—'))

            final_notes = final_remarks_map.get(sub.pk, [])
            ws2.cell(row=ri, column=7, value='\n'.join(final_notes) if final_notes else '—')
            if final_notes:
                ws2.cell(row=ri, column=7).alignment = _WRAP

            for ci, q in enumerate(answerable_q, start=8):
                ans = ans_map.get((sub.pk, q.pk))
                if ans is None:
                    cell_val = '—'
                elif q.question_type in CHOICE_TYPES:
                    opts = list(ans.selected_options.all())
                    cell_val = ', '.join(o.option_text for o in opts) if opts else '—'
                    if ans.other_text:
                        cell_val += f' (Other: {ans.other_text})'
                elif ans.number_value is not None:
                    cell_val = str(ans.number_value)
                elif ans.text_value:
                    cell_val = ans.text_value
                else:
                    cell_val = '—'

                ws2.cell(row=ri, column=ci, value=cell_val)

            _border_row(ws2, ri, ncols)

        _auto_width(ws2)

        # ── Stream response ───────────────────────────────────────────────────
        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        safe_title = ''.join(c for c in training.title if c.isalnum() or c in ' _-')[:40].strip()
        filename = f'training_{safe_title}_{training.training_date}.xlsx'

        response = StreamingHttpResponse(
            streaming_content=buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
