import calendar
import datetime
import logging
from datetime import timedelta

from django.conf import settings
from django.contrib.auth import authenticate, get_user_model
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Min, OuterRef, Q, Subquery
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework_simplejwt.tokens import RefreshToken

from .models import LoginAttempt
from .serializers import EmployeeAdminSerializer, LoginSerializer, UserSerializer

logger = logging.getLogger(__name__)
User = get_user_model()

MAX_FAILED_ATTEMPTS = 5
LOGIN_WINDOW_MINUTES = 15

_ACCESS_COOKIE = getattr(settings, 'JWT_ACCESS_COOKIE', 'access_token')
_REFRESH_COOKIE = getattr(settings, 'JWT_REFRESH_COOKIE', 'refresh_token')
_REFRESH_PATH = '/api/auth/token/refresh'


def _get_client_ip(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '127.0.0.1')


def _cookie_secure():
    return not settings.DEBUG


def _set_auth_cookies(response, access, refresh):
    jwt = settings.SIMPLE_JWT
    access_max = int(jwt['ACCESS_TOKEN_LIFETIME'].total_seconds())
    refresh_max = int(jwt['REFRESH_TOKEN_LIFETIME'].total_seconds())
    secure = _cookie_secure()

    response.set_cookie(
        key=_ACCESS_COOKIE,
        value=access,
        max_age=access_max,
        httponly=True,
        secure=secure,
        samesite='Strict',
        path='/',
    )
    response.set_cookie(
        key=_REFRESH_COOKIE,
        value=refresh,
        max_age=refresh_max,
        httponly=True,
        secure=secure,
        samesite='Strict',
        path=_REFRESH_PATH,
    )
    return response


def _clear_auth_cookies(response):
    response.delete_cookie(_ACCESS_COOKIE, path='/')
    response.delete_cookie(_REFRESH_COOKIE, path=_REFRESH_PATH)
    return response


@method_decorator(ensure_csrf_cookie, name='dispatch')
class CsrfCookieView(APIView):
    """
    GET /api/auth/csrf
    Seeds the csrftoken cookie for the SPA.  Must be called once on app
    load so the frontend can read the token and attach it as X-CSRFToken
    on all subsequent non-safe requests.  Requires no authentication.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        return Response({'detail': 'CSRF cookie set.'})


@method_decorator(csrf_protect, name='dispatch')
class LoginView(APIView):
    """
    POST /api/auth/login/
    Authenticates with username + password.
    Returns user data and sets JWT tokens in HttpOnly cookies.
    Enforces max 5 failed attempts per IP per 15-minute window.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        ip = _get_client_ip(request)

        # Rate-limit check
        window_start = timezone.now() - timedelta(minutes=LOGIN_WINDOW_MINUTES)
        recent_failures = LoginAttempt.objects.filter(
            ip_address=ip,
            created_at__gte=window_start,
            was_successful=False,
        ).count()

        if recent_failures >= MAX_FAILED_ATTEMPTS:
            logger.warning('login rate-limit hit ip=%s failures=%d', ip, recent_failures)
            return Response(
                {'detail': 'Too many failed attempts. Please wait 15 minutes and try again.'},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
            )

        # Input validation
        serializer = LoginSerializer(data=request.data)
        if not serializer.is_valid():
            LoginAttempt.objects.create(ip_address=ip, was_successful=False)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Authenticate
        user = authenticate(
            request,
            username=serializer.validated_data['username'],
            password=serializer.validated_data['password'],
        )

        if user is None:
            LoginAttempt.objects.create(ip_address=ip, was_successful=False)
            logger.warning(
                'failed login username=%s ip=%s',
                serializer.validated_data['username'],
                ip,
            )
            return Response(
                {'detail': 'Invalid credentials.', 'code': 'invalid_credentials'},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        if user.locked:
            LoginAttempt.objects.create(ip_address=ip, user=user, was_successful=False)
            logger.warning(
                'locked account login attempt username=%s ip=%s',
                serializer.validated_data['username'],
                ip,
            )
            return Response(
                {
                    'detail': 'Your account is locked. Please proceed to HR for unlocking.',
                    'code': 'account_locked',
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        if not user.active:
            LoginAttempt.objects.create(ip_address=ip, user=user, was_successful=False)
            logger.warning(
                'inactive account login attempt username=%s ip=%s',
                serializer.validated_data['username'],
                ip,
            )
            return Response(
                {
                    'detail': 'Your account has been deactivated. Please contact HR.',
                    'code': 'account_inactive',
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        # Success
        LoginAttempt.objects.create(ip_address=ip, user=user, was_successful=True)
        logger.info('successful login user_id=%d ip=%s', user.pk, ip)

        refresh = RefreshToken.for_user(user)
        response = Response(
            {'detail': 'Login successful.', 'user': UserSerializer(user).data},
            status=status.HTTP_200_OK,
        )
        return _set_auth_cookies(response, str(refresh.access_token), str(refresh))


@method_decorator(csrf_protect, name='dispatch')
class LogoutView(APIView):
    """
    POST /api/auth/logout/
    Blacklists the refresh token and clears auth cookies.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        raw_refresh = request.COOKIES.get(_REFRESH_COOKIE)
        if raw_refresh:
            try:
                RefreshToken(raw_refresh).blacklist()
            except TokenError:
                pass
        response = Response({'detail': 'Logged out successfully.'}, status=status.HTTP_200_OK)
        return _clear_auth_cookies(response)


@method_decorator(csrf_protect, name='dispatch')
class TokenRefreshView(APIView):
    """
    POST /api/auth/token/refresh/
    Issues a new access token using the HttpOnly refresh cookie.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        raw_refresh = request.COOKIES.get(_REFRESH_COOKIE)
        if not raw_refresh:
            return Response(
                {'detail': 'Refresh token not found.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )
        try:
            refresh = RefreshToken(raw_refresh)
            new_access = str(refresh.access_token)
            new_refresh = str(refresh)
        except TokenError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_401_UNAUTHORIZED)

        response = Response({'detail': 'Token refreshed.'}, status=status.HTTP_200_OK)
        return _set_auth_cookies(response, new_access, new_refresh)


class MeView(APIView):
    """GET /api/auth/me/ — returns the currently authenticated user."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(UserSerializer(request.user).data)


class UserListView(APIView):
    """
    GET /api/auth/users
    Returns a list of active users for member selection in events.
    Supports optional ?q= search by name or idnumber.
    Requires authentication.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        query = request.GET.get('q', '').strip()
        qs = User.objects.filter(active=True).exclude(
            Q(admin=True) & (Q(hr=True) | Q(accounting=True))
        ).order_by('firstname', 'lastname')

        if query:
            qs = qs.filter(
                Q(firstname__icontains=query)
                | Q(lastname__icontains=query)
                | Q(idnumber__icontains=query)
            )[:50]
        else:
            qs = qs[:200]

        data = [
            {
                'id': u.pk,
                'idnumber': u.idnumber,
                'firstname': u.firstname,
                'lastname': u.lastname,
                'avatar': (
                    request.build_absolute_uri(u.avatar.url)
                    if u.avatar and u.avatar.name
                    else None
                ),
            }
            for u in qs
        ]
        return Response(data)


# ── Employee admin helpers ─────────────────────────────────────────────────────

_EMP_ADMIN_REQUIRED = {'admin', 'hr'}

def _require_employee_admin(request) -> Response | None:
    """Return a 403 Response if the requesting user is not admin or hr."""
    u = request.user
    if not (u.admin or u.hr):
        return Response(
            {'detail': 'You do not have permission to perform this action.'},
            status=status.HTTP_403_FORBIDDEN,
        )
    return None


def _employee_queryset():
    """Base queryset: all non-privileged users (not admin, hr, or accounting)."""
    return User.objects.exclude(
        Q(admin=True) | Q(hr=True) | Q(accounting=True)
    )


# ── GET /api/auth/admin/employees ─────────────────────────────────────────────

class EmployeeAdminListView(APIView):
    """
    GET /api/auth/admin/employees
    Paginated list of non-privileged employees for the Employees admin page.
    Query params:
      q             — search by firstname, lastname, or idnumber
      department_id — filter by department
      line_id       — filter by line
      status        — 'active' | 'inactive' | 'locked' | 'unlocked'
      idnumbers     — comma-separated list of ID numbers to include
      page          — page number (default 1, page_size 10)
      sort          — idnumber | lastname | department | line | employment_type | active
      dir           — asc | desc
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        from userProfile.models import workInformation
        from django.db.models import Prefetch

        q             = request.GET.get('q', '').strip()
        dept_id       = request.GET.get('department_id', '').strip()
        line_id       = request.GET.get('line_id', '').strip()
        status_filter = request.GET.get('status', '').strip()
        page_num      = max(1, int(request.GET.get('page', 1) or 1))
        sort_field    = request.GET.get('sort', 'lastname').strip()
        sort_dir      = request.GET.get('dir', 'asc').strip()

        qs = _employee_queryset()

        if q:
            qs = qs.filter(
                Q(firstname__icontains=q)
                | Q(lastname__icontains=q)
                | Q(idnumber__icontains=q)
            )

        if status_filter == 'active':
            qs = qs.filter(active=True, locked=False)
        elif status_filter == 'inactive':
            qs = qs.filter(active=False)
        elif status_filter == 'locked':
            qs = qs.filter(locked=True)
        elif status_filter == 'unlocked':
            qs = qs.filter(locked=False)

        idnumbers_raw = request.GET.get('idnumbers', '').strip()
        if idnumbers_raw:
            idnumber_list = [x.strip() for x in idnumbers_raw.split(',') if x.strip()]
            if idnumber_list:
                qs = qs.filter(idnumber__in=idnumber_list)

        # Filter by department or line via workInformation FK
        if dept_id:
            qs = qs.filter(workinformation__department_id=dept_id).distinct()
        if line_id:
            qs = qs.filter(workinformation__line_id=line_id).distinct()

        # ── Sorting ────────────────────────────────────────────────────────────
        order_prefix = '' if sort_dir == 'asc' else '-'
        allowed_sorts = {'idnumber', 'lastname', 'department', 'line', 'employment_type', 'active'}
        if sort_field not in allowed_sorts:
            sort_field = 'lastname'

        # For related-field sorts use a scalar Subquery so no new JOIN rows appear.
        wi_latest = workInformation.objects.filter(employee=OuterRef('pk')).order_by('-created_at')

        if sort_field == 'department':
            qs = qs.annotate(
                _sort_key=Subquery(wi_latest.values('department__name')[:1])
            ).order_by(f'{order_prefix}_sort_key', 'lastname', 'firstname')
        elif sort_field == 'line':
            qs = qs.annotate(
                _sort_key=Subquery(wi_latest.values('line__name')[:1])
            ).order_by(f'{order_prefix}_sort_key', 'lastname', 'firstname')
        elif sort_field == 'employment_type':
            qs = qs.annotate(
                _sort_key=Subquery(wi_latest.values('employment_type__name')[:1])
            ).order_by(f'{order_prefix}_sort_key', 'lastname', 'firstname')
        elif sort_field == 'active':
            qs = qs.order_by(f'{order_prefix}active', f'{order_prefix}locked', 'lastname', 'firstname')
        else:
            qs = qs.order_by(f'{order_prefix}{sort_field}', 'lastname', 'firstname', 'idnumber')

        # Prefetch the most-recent workInformation record per employee
        wi_qs = workInformation.objects.select_related(
            'department', 'line', 'employment_type'
        ).order_by('-created_at')

        qs = qs.prefetch_related(
            Prefetch('workinformation_set', queryset=wi_qs, to_attr='_work_records')
        )

        paginator = Paginator(qs, 10)
        page_obj  = paginator.get_page(page_num)

        serializer = EmployeeAdminSerializer(page_obj.object_list, many=True)
        return Response({
            'results':  serializer.data,
            'count':    paginator.count,
            'num_pages': paginator.num_pages,
            'page':     page_num,
        })


# ── PATCH /api/auth/admin/employees/<pk>/status ───────────────────────────────

@method_decorator(csrf_protect, name='dispatch')
class EmployeeAdminStatusView(APIView):
    """
    PATCH /api/auth/admin/employees/<pk>/status
    Body: { "action": "activate" | "deactivate" | "lock" | "unlock" }
    Applies the requested status change to the target employee.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk: int) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        # Idempotency
        idem_key = request.headers.get('X-Idempotency-Key', '').strip()
        if idem_key:
            cached = cache.get(f'idem:emp_status:{idem_key}')
            if cached is not None:
                return Response(cached, status=status.HTTP_200_OK)

        action = request.data.get('action', '').strip()
        if action not in ('activate', 'deactivate', 'lock', 'unlock'):
            return Response(
                {'detail': 'Invalid action. Must be activate, deactivate, lock, or unlock.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Prevent self-targeting
        if int(pk) == request.user.pk:
            return Response(
                {'detail': 'You cannot change the status of your own account.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            target = User.objects.select_for_update().get(pk=pk)
        except User.DoesNotExist:
            return Response({'detail': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Re-verify target is not privileged (not trusting frontend state)
        if target.admin or target.hr or target.accounting:
            return Response(
                {'detail': 'Cannot modify the status of privileged accounts.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        if action == 'activate':
            target.active = True
        elif action == 'deactivate':
            target.active = False
        elif action == 'lock':
            target.locked = True
        elif action == 'unlock':
            target.locked = False

        target.save(update_fields=['active', 'locked'])

        result = {
            'id':     target.pk,
            'active': target.active,
            'locked': target.locked,
        }
        if idem_key:
            cache.set(f'idem:emp_status:{idem_key}', result, timeout=60 * 60 * 24)

        return Response(result)


# ── GET /api/auth/admin/employees/chart ───────────────────────────────────────

class EmployeeAdminChartView(APIView):
    """
    GET /api/auth/admin/employees/chart
    Returns employee counts per time bucket read from EmployeeSnapshot records.

    Query params
    ============
    view        fiscal | monthly | weekly  (default: fiscal)
    year        4-digit year              (default: current year)
    month       1-12                      (monthly view, default: current month)
    week_start  ISO date YYYY-MM-DD       (weekly view, default: Monday of today's week)

    Metrics per bucket: total, ojt, regular, probationary, female, male.

    For each bucket the last available snapshot within that bucket is used.
    Buckets with no snapshot record return all-zero counts.
    Fiscal year runs July → June.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        from .models import EmployeeSnapshot

        view_mode = request.GET.get('view', 'fiscal')
        today     = datetime.date.today()

        def _empty():
            return {'total': 0, 'regular': 0, 'probationary': 0, 'ojt': 0, 'male': 0, 'female': 0}

        def _snap_dict(snap):
            if snap is None:
                return _empty()
            return {
                'total':        snap.total,
                'regular':      snap.regular,
                'probationary': snap.probationary,
                'ojt':          snap.ojt,
                'male':         snap.male,
                'female':       snap.female,
            }

        if view_mode == 'fiscal':
            year     = int(request.GET.get('year', today.year))
            fy_start = datetime.date(year, 7, 1)
            fy_end   = datetime.date(year + 1, 6, 30)
            snap_map = {
                s.snapshot_date: s
                for s in EmployeeSnapshot.objects.filter(
                    snapshot_date__range=(fy_start, fy_end)
                )
            }
            data = []
            for i in range(12):
                m = (fy_start.month - 1 + i) % 12 + 1
                y = fy_start.year + ((fy_start.month - 1 + i) // 12)
                label = datetime.date(y, m, 1).strftime('%b %Y')
                # Prefer the last snapshot of the month
                _, last_day = calendar.monthrange(y, m)
                best = None
                for d in range(last_day, 0, -1):
                    candidate = datetime.date(y, m, d)
                    if candidate in snap_map:
                        best = snap_map[candidate]
                        break
                data.append({'label': label, **_snap_dict(best)})
            return Response({'view': view_mode, 'fy_start': str(fy_start), 'data': data})

        elif view_mode == 'weekly':
            try:
                week_start = datetime.date.fromisoformat(
                    request.GET.get('week_start', str(today))
                )
            except (ValueError, TypeError):
                week_start = today
            # Normalise to Monday
            week_start = week_start - datetime.timedelta(days=week_start.weekday())
            week_end   = week_start + datetime.timedelta(days=6)
            snap_map = {
                s.snapshot_date: s
                for s in EmployeeSnapshot.objects.filter(
                    snapshot_date__range=(week_start, week_end)
                )
            }
            data = []
            for i in range(7):
                day = week_start + datetime.timedelta(days=i)
                data.append({'label': day.strftime('%a %d'), **_snap_dict(snap_map.get(day))})
            return Response({'view': view_mode, 'week_start': str(week_start), 'data': data})

        else:  # monthly
            year  = int(request.GET.get('year', today.year))
            month = int(request.GET.get('month', today.month))
            _, days_in_month = calendar.monthrange(year, month)
            snap_map = {
                s.snapshot_date: s
                for s in EmployeeSnapshot.objects.filter(
                    snapshot_date__year=year,
                    snapshot_date__month=month,
                )
            }
            data = []
            for day_num in range(1, days_in_month + 1):
                day = datetime.date(year, month, day_num)
                data.append({'label': str(day_num), **_snap_dict(snap_map.get(day))})
            return Response({'view': view_mode, 'year': year, 'month': month, 'data': data})


# ── POST /api/auth/admin/employees/snapshot ───────────────────────────────────

@method_decorator(csrf_protect, name='dispatch')
class EmployeeAdminSnapshotTriggerView(APIView):
    """
    POST /api/auth/admin/employees/snapshot
    Manually triggers snapshot creation / backfill.

    Body (all optional)
    ===================
    date          YYYY-MM-DD   Snapshot for a specific date (default: today).
    backfill_from YYYY-MM-DD   Fill every missing snapshot from this date to today.

    When ``backfill_from`` is provided, ``date`` is ignored.
    All writes are idempotent (update_or_create).
    Requires admin or HR role.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        from .utils import take_snapshot
        from .models import EmployeeSnapshot

        today         = datetime.date.today()
        backfill_from = request.data.get('backfill_from')
        date_param    = request.data.get('date')

        if backfill_from:
            try:
                start = datetime.date.fromisoformat(str(backfill_from))
            except ValueError:
                return Response(
                    {'detail': 'Invalid backfill_from date. Expected YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if start > today:
                return Response(
                    {'detail': 'backfill_from cannot be in the future.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            existing = set(
                EmployeeSnapshot.objects
                .filter(snapshot_date__gte=start, snapshot_date__lte=today)
                .values_list('snapshot_date', flat=True)
            )
            missing = [
                start + datetime.timedelta(days=i)
                for i in range((today - start).days + 1)
                if (start + datetime.timedelta(days=i)) not in existing
            ]

            created_dates: list[str] = []
            errors: list[dict]       = []
            for d in missing:
                try:
                    take_snapshot(d)
                    created_dates.append(str(d))
                except Exception as exc:
                    logger.error('Snapshot backfill failed for %s: %s', d, exc)
                    errors.append({'date': str(d), 'error': str(exc)})

            return Response({'created': created_dates, 'errors': errors})

        # Single-date mode
        if date_param:
            try:
                target = datetime.date.fromisoformat(str(date_param))
            except ValueError:
                return Response(
                    {'detail': 'Invalid date. Expected YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            target = today

        try:
            snap, created = take_snapshot(target)
            return Response({
                'date':         str(target),
                'created':      created,
                'total':        snap.total,
                'regular':      snap.regular,
                'probationary': snap.probationary,
                'ojt':          snap.ojt,
                'male':         snap.male,
                'female':       snap.female,
            })
        except Exception as exc:
            logger.error('Snapshot trigger failed for %s: %s', target, exc)
            return Response({'detail': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ── GET /api/auth/admin/employees/filters ─────────────────────────────────────

class EmployeeAdminFilterOptionsView(APIView):
    """
    GET /api/auth/admin/employees/filters
    Returns all departments and lines for the filter popovers.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        from generalsettings.models import Department, Line

        departments = list(Department.objects.values('id', 'name').order_by('name'))
        lines = list(Line.objects.values('id', 'name', 'department_id').order_by('name'))
        return Response({'departments': departments, 'lines': lines})


# ── Password-reset notification helper ───────────────────────────────────────

def _send_password_reset_notification(target_pk: int, reset_by) -> None:
    """
    Dispatch an in-app notification to the affected user after the transaction
    commits so the notification is never written if the reset is rolled back.
    """
    def _create() -> None:
        try:
            from activityLog.models import Notification
            Notification.objects.create(
                recipient_id=target_pk,
                notification_type='password_reset',
                title='Password Reset by Administrator',
                message=(
                    'Your account password has been reset by an administrator. '
                    'Please log in using the default password format (Repco_{ID Number}) '
                    'and change it immediately to secure your account. '
                    'Do not share your password with anyone.'
                ),
                module='',
            )
        except Exception:
            logger.exception(
                'Failed to create password reset notification for user_id=%d', target_pk
            )

    try:
        transaction.on_commit(_create)
    except Exception:
        _create()


# ── POST /api/auth/admin/employees/<pk>/reset-password ────────────────────────

@method_decorator(csrf_protect, name='dispatch')
class EmployeeAdminPasswordResetView(APIView):
    """
    POST /api/auth/admin/employees/<pk>/reset-password
    Resets the target employee's password to the default format: Repco_{idnumber}.
    Requires admin or HR role.
    The new password value is never returned in any response.
    Accepts X-Idempotency-Key for safe replay.
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, pk: int) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        idem_key = request.headers.get('X-Idempotency-Key', '').strip()
        if idem_key:
            cached = cache.get(f'idem:emp_pwreset:{idem_key}')
            if cached is not None:
                return Response(cached, status=status.HTTP_200_OK)

        if int(pk) == request.user.pk:
            return Response(
                {'detail': 'You cannot reset your own password via this endpoint.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            target = User.objects.select_for_update().get(pk=pk)
        except User.DoesNotExist:
            return Response({'detail': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        if target.admin or target.hr or target.accounting:
            return Response(
                {'detail': 'Cannot reset the password of privileged accounts.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        target.set_password(f'Repco_{target.idnumber}')
        target.change_password = True
        target.save(update_fields=['password', 'change_password'])

        _send_password_reset_notification(target.pk, request.user)

        result = {'id': target.pk, 'idnumber': target.idnumber, 'reset': True}
        if idem_key:
            cache.set(f'idem:emp_pwreset:{idem_key}', result, timeout=60 * 60 * 24)

        logger.info(
            'admin password reset user_id=%d idnumber=%s by=%d',
            target.pk, target.idnumber, request.user.pk,
        )
        return Response(result)


# ── POST /api/auth/admin/employees/import ─────────────────────────────────────

@method_decorator(csrf_protect, name='dispatch')
class EmployeeAdminImportView(APIView):
    """
    POST /api/auth/admin/employees/import  (multipart/form-data, field: file)

    Accepts XLSX, XLS, or CSV files.  All 8 columns are required in every row.

    Expected columns (row 1 = header, ignored):
      A: ID Number   B: First Name   C: Last Name   D: Email
      E: Department  F: Line         G: Employment Type  H: Date Hired (MM/DD/YYYY)

    Behaviour
    ---------
    Phase 1 – Validation: every row is checked. If *any* error exists:
      • An XLSX error report is generated where invalid cells are highlighted
        in red font and a 'Remarks' column describes the specific errors per row.
      • HTTP 422 is returned with the error report as a file download.
      • No records are created (all-or-nothing).

    Phase 2 – Creation (only reached when all rows pass):
      • Every user is created inside a single atomic transaction.
      • Password is auto-set to Repco_{idnumber} — never in the payload.
      • If department/line/employment_type are provided a workInformation record
        is created for the user.
      • Returns { "imported": N }.

    Requires admin or HR role.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        ext = uploaded.name.rsplit('.', 1)[-1].lower() if '.' in uploaded.name else ''
        if ext not in ('xlsx', 'xls', 'csv'):
            return Response(
                {'detail': 'Unsupported file type. Upload XLSX, XLS, or CSV.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Parse file into list-of-lists ───────────────────────────────────
        raw_rows: list[list[str]] = []
        try:
            if ext == 'csv':
                import csv as _csv
                import io
                content = uploaded.read().decode('utf-8-sig')
                reader = _csv.reader(io.StringIO(content))
                raw_rows = [row for row in reader]
            else:
                import openpyxl
                wb = openpyxl.load_workbook(uploaded, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    raw_rows.append([str(c).strip() if c is not None else '' for c in row])
        except Exception as exc:
            logger.error('Employee import parse error: %s', exc)
            return Response(
                {'detail': f'Could not parse file: {exc}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(raw_rows) < 2:
            return Response(
                {'detail': 'File must contain at least one data row after the header.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data_rows = raw_rows[1:]  # skip header

        # ── Prefetch master-data lookup tables ──────────────────────────────
        from generalsettings.models import Department, EmploymentType, Line
        dept_lookup: dict[str, int]   = {d.name.lower(): d.pk for d in Department.objects.all()}
        line_lookup: dict[str, int]   = {l.name.lower(): l.pk for l in Line.objects.all()}
        etype_lookup: dict[str, int]  = {e.name.lower(): e.pk for e in EmploymentType.objects.all()}
        existing_ids: set[str]        = set(User.objects.values_list('idnumber', flat=True))

        # ── Column indices ──────────────────────────────────────────────────
        COL_IDNUMBER  = 0
        COL_FIRSTNAME = 1
        COL_LASTNAME  = 2
        COL_EMAIL     = 3
        COL_DEPT      = 4
        COL_LINE      = 5
        COL_ETYPE     = 6
        COL_HIRED     = 7
        NCOLS         = 8

        import re as _re
        _EMAIL_RE = _re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', _re.IGNORECASE)

        def _parse_date(raw: str):
            """Try MM/DD/YYYY then YYYY-MM-DD. Returns date or None."""
            for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y'):
                try:
                    return datetime.datetime.strptime(raw.strip(), fmt).date()
                except ValueError:
                    continue
            return None

        # ── Validate every row ──────────────────────────────────────────────
        # row_errors[i] = {col_index: error_message}
        row_errors: list[dict[int, str]] = []
        seen_ids: set[str] = set()
        valid_rows: list[dict] = []

        for row_idx, raw in enumerate(data_rows):
            # Pad to at least NCOLS columns
            padded = list(raw) + [''] * max(0, NCOLS - len(raw))
            errs: dict[int, str] = {}

            idnumber    = padded[COL_IDNUMBER ].strip()
            firstname   = padded[COL_FIRSTNAME].strip()
            lastname    = padded[COL_LASTNAME ].strip()
            email       = padded[COL_EMAIL    ].strip()
            dept_raw    = padded[COL_DEPT     ].strip()
            line_raw    = padded[COL_LINE     ].strip()
            etype_raw   = padded[COL_ETYPE    ].strip()
            hired_raw   = padded[COL_HIRED    ].strip()

            # Required fields
            if not idnumber:
                errs[COL_IDNUMBER] = 'ID Number is required.'
            elif idnumber in existing_ids or idnumber in seen_ids:
                errs[COL_IDNUMBER] = f'ID Number "{idnumber}" already exists.'
            else:
                seen_ids.add(idnumber)

            if not firstname:
                errs[COL_FIRSTNAME] = 'First Name is required.'

            if not lastname:
                errs[COL_LASTNAME] = 'Last Name is required.'

            if not email:
                errs[COL_EMAIL] = 'Email is required.'
            elif not _EMAIL_RE.match(email):
                errs[COL_EMAIL] = f'"{email}" is not a valid email address.'

            dept_id: int | None = None
            if not dept_raw:
                errs[COL_DEPT] = 'Department is required.'
            elif dept_raw.lower() not in dept_lookup:
                errs[COL_DEPT] = f'Department "{dept_raw}" does not exist.'
            else:
                dept_id = dept_lookup[dept_raw.lower()]

            line_id: int | None = None
            if line_raw:
                if line_raw.lower() not in line_lookup:
                    errs[COL_LINE] = f'Line "{line_raw}" does not exist.'
                else:
                    line_id = line_lookup[line_raw.lower()]

            etype_id: int | None = None
            if etype_raw:
                if etype_raw.lower() not in etype_lookup:
                    errs[COL_ETYPE] = f'Employment Type "{etype_raw}" does not exist.'
                else:
                    etype_id = etype_lookup[etype_raw.lower()]

            hired_date = None
            if hired_raw:
                hired_date = _parse_date(hired_raw)
                if hired_date is None:
                    errs[COL_HIRED] = f'Date Hired "{hired_raw}" must be MM/DD/YYYY.'

            row_errors.append(errs)
            if not errs:
                valid_rows.append({
                    'idnumber':  idnumber,
                    'firstname': firstname,
                    'lastname':  lastname,
                    'email':     email,
                    'dept_id':   dept_id,
                    'line_id':   line_id,
                    'etype_id':  etype_id,
                    'hired':     hired_date,
                })

        # ── Build error report if any row has errors ────────────────────────
        if any(row_errors):
            import io
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter

            RED_FONT    = Font(color='FF0000', bold=False)
            GREY_FONT   = Font(color='888888')
            HEADER_FONT = Font(bold=True, color='FF000000')
            HEADER_FILL = PatternFill(fill_type='solid', fgColor='FFDBE5FF')
            BORDER_SIDE = Side(style='thin', color='FF000000')
            TABLE_BORDER = Border(
                left=BORDER_SIDE,
                right=BORDER_SIDE,
                top=BORDER_SIDE,
                bottom=BORDER_SIDE,
            )

            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = 'Import Errors'

            # Header row (original headers + Remarks)
            hdr = ['ID Number', 'First Name', 'Last Name', 'Email',
                   'Department', 'Line', 'Employment Type', 'Date Hired', 'Remarks']
            ws_out.append(hdr)
            for col_idx, _ in enumerate(hdr, start=1):
                cell = ws_out.cell(row=1, column=col_idx)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = TABLE_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx, (raw, errs) in enumerate(zip(data_rows, row_errors)):
                padded = list(raw) + [''] * max(0, NCOLS - len(raw))
                remarks = '; '.join(errs.values())
                out_row = padded[:NCOLS] + [remarks]
                ws_out.append(out_row)
                excel_row = row_idx + 2  # +2: 1-based + header offset
                for col_idx in range(1, NCOLS + 2):
                    cell = ws_out.cell(row=excel_row, column=col_idx)
                    cell.border = TABLE_BORDER
                    if col_idx - 1 in errs:
                        cell.font = RED_FONT
                if errs:
                    # Remarks cell always red when row has errors
                    ws_out.cell(row=excel_row, column=NCOLS + 1).font = RED_FONT
                else:
                    # Valid rows: grey remarks cell (empty)
                    ws_out.cell(row=excel_row, column=NCOLS + 1).font = GREY_FONT

            # Auto-size columns based on their content
            for col_idx, _ in enumerate(hdr, start=1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws_out[col_letter]:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                ws_out.column_dimensions[col_letter].width = min(max(max_length + 2, 14), 40)

            buf = io.BytesIO()
            wb_out.save(buf)
            buf.seek(0)

            from django.http import HttpResponse
            resp = HttpResponse(
                buf.getvalue(),
                status=422,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            resp['Content-Disposition'] = 'attachment; filename="import_errors.xlsx"'
            resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return resp

        # ── Create all records atomically ───────────────────────────────────
        from userProfile.models import workInformation
        from generalsettings.models import Office, Shift

        # Find a default Office and Shift for workInformation (required FKs).
        default_office = Office.objects.first()
        default_shift  = Shift.objects.first()

        if not default_office or not default_shift:
            return Response(
                {'detail': 'No Office or Shift found in the system. Please create at least one before importing.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        with transaction.atomic():
            for row in valid_rows:
                idnumber = row['idnumber']
                user = User(
                    idnumber  = idnumber,
                    username  = idnumber,
                    firstname = row['firstname'],
                    lastname  = row['lastname'],
                    email     = row['email'],
                    active    = True,
                    locked    = False,
                )
                user.set_password(f'Repco_{idnumber}')
                user.save()

                if row['dept_id']:
                    workInformation.objects.create(
                        employee          = user,
                        office            = default_office,
                        shift             = default_shift,
                        department_id     = row['dept_id'],
                        line_id           = row['line_id'],
                        employment_type_id= row['etype_id'],
                        date_hired        = row['hired'],
                    )
                created_count += 1

        logger.info(
            'employee import: %d records created by user_id=%d',
            created_count, request.user.pk,
        )
        return Response({'imported': created_count})


# ── POST /api/auth/admin/employees/export ─────────────────────────────────────

@method_decorator(csrf_protect, name='dispatch')
class EmployeeAdminExportView(APIView):
    """
    POST /api/auth/admin/employees/export
    Generate and download an Excel report.

    Body parameters
    ===============
    type        personal_info | work_info | summary | all
    view        fiscal | monthly | weekly         (summary only)
    year        4-digit int                       (fiscal + monthly)
    month       1-12                              (monthly)
    week_start  YYYY-MM-DD                        (weekly)
    """
    permission_classes = [IsAuthenticated]

    def post(self, request) -> Response:
        err = _require_employee_admin(request)
        if err:
            return err

        export_type = str(request.data.get('type', '')).strip()
        if export_type not in ('personal_info', 'work_info', 'summary', 'all'):
            return Response(
                {'detail': 'Invalid export type. Choose: personal_info, work_info, summary, all.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        import io
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        # ── Shared style helpers ──────────────────────────────────────────────
        _thin   = Side(style='thin', color='FF000000')
        _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
        _hdr_font = Font(bold=True, color='FFFFFFFF')
        _hdr_fill = PatternFill(fill_type='solid', fgColor='FF2845D6')
        _center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
        _left     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        def _style_header(ws, headers):
            ws.append(headers)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font      = _hdr_font
                cell.fill      = _hdr_fill
                cell.border    = _border
                cell.alignment = _center

        def _style_row(ws, row_idx, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border    = _border
                cell.alignment = _left

        def _auto_width(ws, headers):
            for c, h in enumerate(headers, start=1):
                col_letter = get_column_letter(c)
                max_len    = len(str(h))
                for row in ws.iter_rows(min_col=c, max_col=c, min_row=2):
                    for cell in row:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 50)

        def _fmt_date(d):
            """Format date as 'April 14, 2026'."""
            if d is None:
                return ''
            months = [
                'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December',
            ]
            return f'{months[d.month - 1]} {d.day}, {d.year}'

        def _fmt_address(block_lot, street, barangay, city, province, country):
            parts = [p for p in [block_lot, street, barangay, city, province, country] if p]
            return ', '.join(parts)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)          # discard default blank sheet

        # ── Personal Information + Emergency Contact sheets ───────────────────
        if export_type in ('personal_info', 'all'):
            from userProfile.models import (
                PersonalInformation, PresentAddress, ProvincialAddress,
                EmergencyContact, FamilyBackground,
            )

            active_users = (
                User.objects
                .filter(active=True, admin=False, hr=False)
                .select_related(
                    'personal_info',
                    'present_address',
                    'provincial_address',
                    'emergency_contact',
                    'family_background',
                )
                .prefetch_related('children')
                .order_by('lastname', 'firstname')
            )

            if not active_users.exists():
                return Response(
                    {'detail': 'No active employees found to export.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

            ws_pi = wb.create_sheet('Personal Information')
            pi_headers = [
                'ID Number', 'Employee Name', 'Nickname', 'Email Address',
                'Birthdate', 'Birth Place', 'Contact Number',
                'Present Address', 'Provincial Address',
                "Mother's Name", "Father's Name", 'Spouse Name', 'Children',
            ]
            _style_header(ws_pi, pi_headers)

            for ri, user in enumerate(active_users, start=2):
                pi   = getattr(user, 'personal_info',       None)
                pa   = getattr(user, 'present_address',     None)
                prov = getattr(user, 'provincial_address',  None)
                fb   = getattr(user, 'family_background',   None)

                # Employee Name
                fn     = user.firstname or ''
                ln     = user.lastname  or ''
                mn     = (pi.middle_name if pi and pi.middle_name else '').strip()
                name   = f"{ln}, {fn} {mn}".strip().rstrip(',')

                # Email
                emails = [e for e in [user.email or '', pi.work_email if pi else ''] if e]
                email_str = ', '.join(emails)

                # Present / Provincial address
                if pa:
                    present_str = _fmt_address(pa.block_lot, pa.street, pa.barangay, pa.city, pa.province, pa.country)
                else:
                    present_str = ''

                if prov:
                    if prov.same_as_present and pa:
                        prov_str = _fmt_address(pa.block_lot, pa.street, pa.barangay, pa.city, pa.province, pa.country)
                    else:
                        prov_str = _fmt_address(prov.block_lot, prov.street, prov.barangay, prov.city, prov.province, prov.country)
                else:
                    prov_str = ''

                children_str = ', '.join(c.name for c in user.children.all())

                ws_pi.append([
                    user.idnumber,
                    name,
                    pi.nickname        if pi else '',
                    email_str,
                    _fmt_date(pi.birth_date if pi else None),
                    pi.birth_place     if pi else '',
                    pi.contact_number  if pi else '',
                    present_str,
                    prov_str,
                    fb.mother_name     if fb else '',
                    fb.father_name     if fb else '',
                    fb.spouse_name     if fb else '',
                    children_str,
                ])
                _style_row(ws_pi, ri, len(pi_headers))

            _auto_width(ws_pi, pi_headers)
            ws_pi.freeze_panes = 'A2'

            # ── Emergency Contact sheet ───────────────────────────────────────
            ws_ec = wb.create_sheet('Emergency Contact')
            ec_headers = ['ID Number', 'Employee Name', 'Contact Name', 'Relationship', 'Contact Number', 'Address']
            _style_header(ws_ec, ec_headers)

            for ri, user in enumerate(active_users, start=2):
                ec  = getattr(user, 'emergency_contact', None)
                fn  = user.firstname or ''
                ln  = user.lastname  or ''
                nm  = f"{ln}, {fn}".strip().rstrip(',')
                ws_ec.append([
                    user.idnumber,
                    nm,
                    ec.name           if ec else '',
                    ec.relationship   if ec else '',
                    ec.contact_number if ec else '',
                    ec.address        if ec else '',
                ])
                _style_row(ws_ec, ri, len(ec_headers))

            _auto_width(ws_ec, ec_headers)
            ws_ec.freeze_panes = 'A2'

        # ── Work Information sheet ────────────────────────────────────────────
        if export_type in ('work_info', 'all'):
            from userProfile.models import workInformation as WI

            qs = (
                WI.objects
                .filter(employee__active=True, employee__admin=False, employee__hr=False)
                .select_related(
                    'employee', 'department', 'line', 'office', 'shift',
                    'position', 'employment_type',
                )
                .order_by('employee__lastname', 'employee__firstname', '-created_at')
            )

            # Latest WI record per employee
            seen_ids: set[int] = set()
            work_rows = []
            for wi in qs:
                if wi.employee_id not in seen_ids:
                    seen_ids.add(wi.employee_id)
                    work_rows.append(wi)

            if not work_rows and export_type == 'work_info':
                return Response(
                    {'detail': 'No work information records found.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

            ws_wi = wb.create_sheet('Work Information')
            wi_headers = [
                'ID Number', 'Employee Name', 'Department', 'Line', 'Office', 'Shift',
                'Position', 'Employment Type', 'Date Hired',
                'TIN Number', 'SSS Number', 'HDMF Number', 'PhilHealth Number', 'Bank Account',
            ]
            _style_header(ws_wi, wi_headers)

            for ri, wi in enumerate(work_rows, start=2):
                u   = wi.employee
                fn  = u.firstname or ''
                ln  = u.lastname  or ''
                nm  = f"{ln}, {fn}".strip().rstrip(',')
                ws_wi.append([
                    u.idnumber,
                    nm,
                    wi.department.name       if wi.department       else '',
                    wi.line.name             if wi.line             else '',
                    wi.office.name           if wi.office           else '',
                    wi.shift.name            if wi.shift            else '',
                    wi.position.name         if wi.position         else '',
                    wi.employment_type.name  if wi.employment_type  else '',
                    f"{wi.date_hired.strftime('%B')} {wi.date_hired.day}, {wi.date_hired.year}" if wi.date_hired else '',
                    wi.tin_number,
                    wi.sss_number,
                    wi.hdmf_number,
                    wi.philhealth_number,
                    wi.bank_account,
                ])
                _style_row(ws_wi, ri, len(wi_headers))

            _auto_width(ws_wi, wi_headers)
            ws_wi.freeze_panes = 'A2'

        # ── Summary sheet with 3D charts ──────────────────────────────────────
        if export_type == 'summary':
            from .models import EmployeeSnapshot
            from openpyxl.chart import BarChart3D, Reference
            from openpyxl.utils import get_column_letter

            view_mode = str(request.data.get('view', 'fiscal')).strip()
            today     = datetime.date.today()

            def _empty_row(label):
                return {'label': label, 'total': 0, 'ojt': 0, 'probationary': 0, 'regular': 0, 'male': 0, 'female': 0}

            def _snap_row(label, snap):
                if snap is None:
                    return _empty_row(label)
                return {
                    'label':        label,
                    'total':        snap.total,
                    'ojt':          snap.ojt,
                    'probationary': snap.probationary,
                    'regular':      snap.regular,
                    'male':         snap.male,
                    'female':       snap.female,
                }

            if view_mode == 'fiscal':
                year     = int(request.data.get('year', today.year))
                fy_start = datetime.date(year, 7, 1)
                fy_end   = datetime.date(year + 1, 6, 30)
                snap_map = {s.snapshot_date: s for s in EmployeeSnapshot.objects.filter(snapshot_date__range=(fy_start, fy_end))}
                chart_data = []
                for i in range(12):
                    m      = (fy_start.month - 1 + i) % 12 + 1
                    y_curr = fy_start.year + ((fy_start.month - 1 + i) // 12)
                    label  = datetime.date(y_curr, m, 1).strftime('%b %Y')
                    _, last_day = calendar.monthrange(y_curr, m)
                    best = None
                    for d in range(last_day, 0, -1):
                        cand = datetime.date(y_curr, m, d)
                        if cand in snap_map:
                            best = snap_map[cand]
                            break
                    chart_data.append(_snap_row(label, best))

            elif view_mode == 'weekly':
                try:
                    ws_start = datetime.date.fromisoformat(str(request.data.get('week_start', str(today))))
                except (ValueError, TypeError):
                    ws_start = today
                ws_start  = ws_start - datetime.timedelta(days=ws_start.weekday())
                ws_end    = ws_start + datetime.timedelta(days=6)
                snap_map  = {s.snapshot_date: s for s in EmployeeSnapshot.objects.filter(snapshot_date__range=(ws_start, ws_end))}
                chart_data = [
                    _snap_row((ws_start + datetime.timedelta(days=i)).strftime('%a %d'), snap_map.get(ws_start + datetime.timedelta(days=i)))
                    for i in range(7)
                ]

            else:  # monthly
                year  = int(request.data.get('year',  today.year))
                month = int(request.data.get('month', today.month))
                _, days = calendar.monthrange(year, month)
                snap_map   = {s.snapshot_date: s for s in EmployeeSnapshot.objects.filter(snapshot_date__year=year, snapshot_date__month=month)}
                chart_data = [
                    _snap_row(str(d), snap_map.get(datetime.date(year, month, d)))
                    for d in range(1, days + 1)
                ]

            if all(row['total'] == 0 for row in chart_data):
                return Response(
                    {'detail': 'No snapshot data found for the selected period.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

            ws_sum = wb.create_sheet('Summary')

            # ── Data table ────────────────────────────────────────────────────
            # Charts are placed side-by-side in ONE row (A1 / I1 / Q1).
            # height=5.3 cm ≈ 200 px ≈ ~10 rows; reserve 12 rows for comfort.
            CHART_ROWS = 12
            DATA_ROW   = CHART_ROWS + 2           # data table starts at row 14
            tbl_headers  = ['Period', 'Total', 'OJT', 'Probationary', 'Regular', 'Male', 'Female']
            n_rows        = len(chart_data)
            data_last_row = DATA_ROW + n_rows

            for c_idx, h in enumerate(tbl_headers, start=1):
                cell = ws_sum.cell(row=DATA_ROW, column=c_idx)
                cell.value     = h
                cell.font      = _hdr_font
                cell.fill      = _hdr_fill
                cell.border    = _border
                cell.alignment = _center

            for r_offset, row in enumerate(chart_data, start=1):
                r = DATA_ROW + r_offset
                vals = [row['label'], row['total'], row['ojt'], row['probationary'], row['regular'], row['male'], row['female']]
                for c_idx, v in enumerate(vals, start=1):
                    cell = ws_sum.cell(row=r, column=c_idx)
                    cell.value     = v
                    cell.border    = _border
                    cell.alignment = _left

            for c_idx, h in enumerate(tbl_headers, start=1):
                cl      = get_column_letter(c_idx)
                max_len = len(h)
                for row in ws_sum.iter_rows(min_row=DATA_ROW + 1, min_col=c_idx, max_col=c_idx):
                    for cell in row:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                ws_sum.column_dimensions[cl].width = min(max(max_len + 2, 12), 30)

            # ── 3D Bar Charts — ONE row, side-by-side (A1 / I1 / Q1) ─────────
            from openpyxl.chart.legend import Legend

            cats = Reference(ws_sum, min_col=1, min_row=DATA_ROW + 1, max_row=data_last_row)

            CHART_W = 15   # cm
            CHART_H = 5.3  # cm ≈ 200 px

            # Colorful series palettes (hex without #)
            COLOR_TOTAL     = ['4BACC6']                      # teal-blue
            COLOR_EMP_TYPES = ['4472C4', 'ED7D31', '70AD47']  # blue / orange / green
            COLOR_GENDER    = ['5B9BD5', 'FF6B9D']            # blue / pink

            def _make_legend_bottom():
                lgnd = Legend()
                lgnd.position = 'b'
                return lgnd

            def _color_series(chart, palette):
                for i, color in enumerate(palette):
                    if i < len(chart.series):
                        chart.series[i].graphicalProperties.solidFill = color

            # Chart 1 — Total Employees  (A1)
            c1             = BarChart3D()
            c1.type        = 'col'
            c1.title       = 'Total Employees'
            c1.grouping    = 'clustered'
            c1.width       = CHART_W
            c1.height      = CHART_H
            c1.style       = 11
            c1.varyColors  = True
            c1.x_axis.title = 'Period'
            c1.y_axis.title = 'Headcount'
            c1.legend       = _make_legend_bottom()
            d1 = Reference(ws_sum, min_col=2, max_col=2, min_row=DATA_ROW, max_row=data_last_row)
            c1.add_data(d1, titles_from_data=True)
            c1.set_categories(cats)
            ws_sum.add_chart(c1, 'A1')

            # Chart 2 — By Employment Type  (I1)
            c2              = BarChart3D()
            c2.type         = 'col'
            c2.title        = 'Employees by Employment Type'
            c2.grouping     = 'clustered'
            c2.width        = CHART_W
            c2.height       = CHART_H
            c2.style        = 11
            c2.x_axis.title = 'Period'
            c2.y_axis.title = 'Headcount'
            c2.legend       = _make_legend_bottom()
            d2 = Reference(ws_sum, min_col=3, max_col=5, min_row=DATA_ROW, max_row=data_last_row)
            c2.add_data(d2, titles_from_data=True)
            c2.set_categories(cats)
            _color_series(c2, COLOR_EMP_TYPES)
            ws_sum.add_chart(c2, 'I1')

            # Chart 3 — By Gender  (Q1)
            c3              = BarChart3D()
            c3.type         = 'col'
            c3.title        = 'Employees by Gender'
            c3.grouping     = 'clustered'
            c3.width        = CHART_W
            c3.height       = CHART_H
            c3.style        = 11
            c3.x_axis.title = 'Period'
            c3.y_axis.title = 'Headcount'
            c3.legend       = _make_legend_bottom()
            d3 = Reference(ws_sum, min_col=6, max_col=7, min_row=DATA_ROW, max_row=data_last_row)
            c3.add_data(d3, titles_from_data=True)
            c3.set_categories(cats)
            _color_series(c3, COLOR_GENDER)
            ws_sum.add_chart(c3, 'Q1')

        if not wb.sheetnames:
            return Response({'detail': 'Nothing to export.'}, status=status.HTTP_404_NOT_FOUND)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        type_slug = {
            'personal_info': 'personal_information',
            'work_info':     'work_information',
            'summary':       'summary',
            'all':           'all',
        }.get(export_type, export_type)
        filename = f'employees_{type_slug}_{datetime.date.today().isoformat()}.xlsx'

        from django.http import HttpResponse
        resp = HttpResponse(
            buf.getvalue(),
            status=200,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition']     = f'attachment; filename="{filename}"'
        resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp
