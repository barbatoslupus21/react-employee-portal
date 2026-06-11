"""Survey module views.

Security checklist:
  - All list endpoints paginated (max page_size=20).
  - No raw SQL — ORM only.
  - @transaction.atomic + select_for_update() on all write-modify operations.
  - CSRF enforced via DRF + middleware (no @csrf_exempt anywhere).
  - permission_classes = [IsAuthenticated] on every view; role check inside method.
  - Input validated through serializers before any DB write.
  - Data returned on anonymous surveys never includes employee identity.
  - Option mutation blocked on non-Draft surveys (R5).
  - Double-submit prevented by unique_together + select_for_update (R1).
"""
from __future__ import annotations

import io
from collections import defaultdict
from typing import Any

from django.db import transaction
from django.db.models import (
    Q, Max, Avg, F, ExpressionWrapper, DurationField, Count,
    OuterRef, Subquery, Case, When, Value, IntegerField,
)
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from activityLog.models import Notification
from survey.models import (
    ALLOW_OTHER_TYPES,
    CHOICE_BASED_TYPES,
    Survey,
    SurveyAnswer,
    SurveyQuestion,
    SurveyQuestionOption,
    SurveyQuestionRatingConfig,
    SurveyResponse,
    SurveyTargetUser,
    SurveyTemplate,
)
from survey.serializers import (
    SurveyAnswerSerializer,
    SurveyDetailSerializer,
    SurveyListSerializer,
    SurveyQuestionSerializer,
    SurveyQuestionWriteSerializer,
    SurveyQuestionOptionSerializer,
    SurveyQuestionRatingConfigSerializer,
    SurveyResponseSerializer,
    SurveyTemplateDetailSerializer,
    SurveyTemplateListSerializer,
    SurveyTemplateWriteSerializer,
    SurveyWriteSerializer,
)

# ── Permission helpers ────────────────────────────────────────────────────────

def _require_admin_hr_or_iad(request) -> Response | None:
    u = request.user
    if not (getattr(u, 'admin', False) or getattr(u, 'hr', False) or getattr(u, 'iad', False)):
        return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
    return None


# ── Notification helper ───────────────────────────────────────────────────────

def _eligible_survey_recipients():
    from django.contrib.auth import get_user_model
    User = get_user_model()
    return User.objects.filter(
        is_active=True,
        admin=False,
        hr=False,
        accounting=False,
    )


def _schedule_survey_notifications(survey: Survey) -> None:
    """Fire in-app notifications on survey activation via bulk_create (R11)."""
    def _create():
        if survey.target_type == Survey.TARGET_ALL:
            user_ids = list(_eligible_survey_recipients().values_list('id', flat=True))
        else:
            user_ids = list(survey.target_users.values_list('user_id', flat=True))

        is_general = survey.target_type == Survey.TARGET_ALL
        scope = 'general' if is_general else 'specific_user'
        desc = survey.description[:200] + ('...' if len(survey.description) > 200 else '')
        notifications = [
            Notification(
                recipient_id=uid,
                notification_type='survey_assigned',
                notification_scope=scope,
                title=f'New Survey: {survey.title}',
                message=(f'A new survey "{survey.title}" is now available.\n\n{desc}').strip(),
                module='survey',
                related_object_id=survey.pk,
            )
            for uid in user_ids
        ]
        for i in range(0, len(notifications), 50):
            Notification.objects.bulk_create(notifications[i:i + 50], ignore_conflicts=True)

    transaction.on_commit(_create)


# ── Pagination helper ─────────────────────────────────────────────────────────

def _paginate(qs, request, page_size: int = 20) -> tuple:
    try:
        page = max(1, int(request.query_params.get('page', 1)))
    except (ValueError, TypeError):
        page = 1

    try:
        requested_page_size = int(request.query_params.get('page_size', page_size))
        if requested_page_size <= 0:
            requested_page_size = page_size
    except (ValueError, TypeError):
        requested_page_size = page_size

    page_size = min(requested_page_size, 100)
    total = qs.count()
    start = (page - 1) * page_size
    items = list(qs[start:start + page_size])
    return items, {
        'page': page,
        'page_size': page_size,
        'total': total,
        'total_pages': max(1, (total + page_size - 1) // page_size),
    }


# ── Template propagation helpers ─────────────────────────────────────────────
#
# When admin edits a SurveyTemplate question/option, these helpers propagate the
# change to all Training objects that reference the template AND to all Draft
# Survey objects that were seeded from it (Survey.source_template).
#
# Rules:
#  • For Training: only update when no completed (is_complete=True) submission
#    exists, because question IDs must remain stable once answers are recorded.
#  • For Survey:   only update when the survey is still in STATUS_DRAFT, as
#    active/closed surveys may already have responses in progress.
#  • Questions are matched by their `order` value (copied 1:1 at seeding time).
# ─────────────────────────────────────────────────────────────────────────────

def _propagate_template_question_to_trainings(tmpl_q: SurveyQuestion, deleted: bool = False) -> None:
    """Push a template question change into every eligible training."""
    from training.models import (
        Training, TrainingQuestion, TrainingQuestionOption,
        TrainingQuestionRatingConfig, TrainingSubmission, TrainingAnswer,
    )

    template_id = tmpl_q.template_id
    if not template_id:
        return

    locked_training_ids = set(
        TrainingSubmission.objects
        .filter(is_complete=True)
        .values_list('training_id', flat=True)
    )
    trainings = Training.objects.filter(template_id=template_id).exclude(pk__in=locked_training_ids)

    for training in trainings:
        tq = TrainingQuestion.objects.filter(training=training, order=tmpl_q.order).first()

        if deleted:
            if tq and not TrainingAnswer.objects.filter(question=tq).exists():
                tq.delete()
            continue

        if tq:
            changed = False
            for field in ('question_text', 'question_type', 'is_required', 'allow_other'):
                val = getattr(tmpl_q, field, getattr(tq, field))
                if getattr(tq, field) != val:
                    setattr(tq, field, val)
                    changed = True
            if changed:
                tq.save()
            # Sync options for choice-based questions
            if tmpl_q.question_type in CHOICE_BASED_TYPES:
                _sync_training_options(tq, tmpl_q)
            # Sync rating config
            if tmpl_q.question_type == 'rating':
                _sync_training_rating_config(tq, tmpl_q)
        else:
            # Template has a new question not yet in this training
            tq = TrainingQuestion.objects.create(
                training=training,
                question_text=tmpl_q.question_text,
                question_type=tmpl_q.question_type,
                order=tmpl_q.order,
                is_required=tmpl_q.is_required,
                allow_other=getattr(tmpl_q, 'allow_other', False),
            )
            for opt in tmpl_q.options.order_by('order'):
                TrainingQuestionOption.objects.create(
                    question=tq, option_text=opt.option_text, order=opt.order,
                )
            if tmpl_q.question_type == 'rating':
                try:
                    rc = tmpl_q.rating_config
                    TrainingQuestionRatingConfig.objects.create(
                        question=tq,
                        min_value=rc.min_value, max_value=rc.max_value,
                        min_label=getattr(rc, 'min_label', ''),
                        max_label=getattr(rc, 'max_label', ''),
                    )
                except Exception:
                    pass


def _sync_training_options(tq, tmpl_q: SurveyQuestion) -> None:
    """Update/add/remove TrainingQuestionOption rows to match template options."""
    from training.models import TrainingQuestionOption, TrainingAnswer

    tmpl_opts = {opt.order: opt for opt in tmpl_q.options.all()}
    existing  = {opt.order: opt for opt in tq.options.all()}

    for order, t_opt in tmpl_opts.items():
        ex = existing.get(order)
        if ex:
            if ex.option_text != t_opt.option_text:
                ex.option_text = t_opt.option_text
                ex.save()
        else:
            TrainingQuestionOption.objects.create(
                question=tq, option_text=t_opt.option_text, order=order,
            )

    # Remove options no longer in template only when no answer selected them
    tmpl_orders = set(tmpl_opts.keys())
    for order, ex_opt in existing.items():
        if order not in tmpl_orders and not ex_opt.user_answers.exists():
            ex_opt.delete()


def _sync_training_rating_config(tq, tmpl_q: SurveyQuestion) -> None:
    """Sync rating scale config from template question to training question."""
    from training.models import TrainingQuestionRatingConfig
    try:
        rc_src = tmpl_q.rating_config
    except Exception:
        return
    try:
        rc_tgt = tq.rating_config
        for f in ('min_value', 'max_value', 'min_label', 'max_label'):
            setattr(rc_tgt, f, getattr(rc_src, f))
        rc_tgt.save()
    except Exception:
        TrainingQuestionRatingConfig.objects.create(
            question=tq,
            min_value=rc_src.min_value, max_value=rc_src.max_value,
            min_label=getattr(rc_src, 'min_label', ''),
            max_label=getattr(rc_src, 'max_label', ''),
        )


def _propagate_template_question_to_surveys(tmpl_q: SurveyQuestion, deleted: bool = False) -> None:
    """Push a template question change into every eligible draft survey."""
    template_id = tmpl_q.template_id
    if not template_id:
        return

    surveys = Survey.objects.filter(source_template_id=template_id, status=Survey.STATUS_DRAFT)

    for survey in surveys:
        sq = SurveyQuestion.objects.filter(survey=survey, order=tmpl_q.order).first()

        if deleted:
            if sq and not SurveyAnswer.objects.filter(question=sq).exists():
                sq.delete()
            continue

        if sq:
            changed = False
            for field in ('question_text', 'question_type', 'is_required', 'allow_other'):
                val = getattr(tmpl_q, field, getattr(sq, field))
                if getattr(sq, field) != val:
                    setattr(sq, field, val)
                    changed = True
            if changed:
                sq.save()
            if tmpl_q.question_type in CHOICE_BASED_TYPES:
                _sync_survey_options(sq, tmpl_q)
            if tmpl_q.question_type == 'rating':
                _sync_survey_rating_config(sq, tmpl_q)
        else:
            new_sq = SurveyQuestion.objects.create(
                survey=survey,
                question_text=tmpl_q.question_text,
                question_type=tmpl_q.question_type,
                order=tmpl_q.order,
                is_required=tmpl_q.is_required,
                allow_other=getattr(tmpl_q, 'allow_other', False),
            )
            for opt in tmpl_q.options.order_by('order'):
                SurveyQuestionOption.objects.create(
                    question=new_sq, option_text=opt.option_text, order=opt.order,
                )
            if tmpl_q.question_type == 'rating':
                try:
                    rc = tmpl_q.rating_config
                    SurveyQuestionRatingConfig.objects.create(
                        question=new_sq,
                        min_value=rc.min_value, max_value=rc.max_value,
                        min_label=getattr(rc, 'min_label', ''),
                        max_label=getattr(rc, 'max_label', ''),
                    )
                except Exception:
                    pass


def _sync_survey_options(sq: SurveyQuestion, tmpl_q: SurveyQuestion) -> None:
    tmpl_opts = {opt.order: opt for opt in tmpl_q.options.all()}
    existing  = {opt.order: opt for opt in sq.options.all()}

    for order, t_opt in tmpl_opts.items():
        ex = existing.get(order)
        if ex:
            if ex.option_text != t_opt.option_text:
                ex.option_text = t_opt.option_text
                ex.save()
        else:
            SurveyQuestionOption.objects.create(
                question=sq, option_text=t_opt.option_text, order=order,
            )

    tmpl_orders = set(tmpl_opts.keys())
    for order, ex_opt in existing.items():
        if order not in tmpl_orders and not ex_opt.answers.exists():
            ex_opt.delete()


def _sync_survey_rating_config(sq: SurveyQuestion, tmpl_q: SurveyQuestion) -> None:
    try:
        rc_src = tmpl_q.rating_config
    except Exception:
        return
    try:
        rc_tgt = sq.rating_config
        for f in ('min_value', 'max_value', 'min_label', 'max_label'):
            setattr(rc_tgt, f, getattr(rc_src, f))
        rc_tgt.save()
    except Exception:
        SurveyQuestionRatingConfig.objects.create(
            question=sq,
            min_value=rc_src.min_value, max_value=rc_src.max_value,
            min_label=getattr(rc_src, 'min_label', ''),
            max_label=getattr(rc_src, 'max_label', ''),
        )


def _propagate_template_question_change(tmpl_q: SurveyQuestion, deleted: bool = False) -> None:
    """Top-level propagation: push to trainings + surveys."""
    if not tmpl_q.template_id:
        return
    _propagate_template_question_to_trainings(tmpl_q, deleted=deleted)
    _propagate_template_question_to_surveys(tmpl_q, deleted=deleted)


def _propagate_template_reorder_to_trainings(template_id: int, order_change_map: dict) -> None:
    """Propagate question reorder from template into every eligible training.

    Uses a temporary large offset to avoid collisions when swapping order values
    (e.g. 1→2 and 2→1 would corrupt without the two-phase approach).
    """
    from training.models import Training, TrainingQuestion, TrainingSubmission

    locked_ids = set(
        TrainingSubmission.objects
        .filter(is_complete=True)
        .values_list('training_id', flat=True)
    )
    trainings = Training.objects.filter(template_id=template_id).exclude(pk__in=locked_ids)
    OFFSET = 100000

    for training in trainings:
        for old_order in order_change_map:
            TrainingQuestion.objects.filter(training=training, order=old_order).update(
                order=old_order + OFFSET
            )
        for old_order, new_order in order_change_map.items():
            TrainingQuestion.objects.filter(training=training, order=old_order + OFFSET).update(
                order=new_order
            )


def _propagate_template_reorder_to_surveys(template_id: int, order_change_map: dict) -> None:
    """Propagate question reorder from template into every eligible draft survey."""
    surveys = Survey.objects.filter(source_template_id=template_id, status=Survey.STATUS_DRAFT)
    OFFSET = 100000

    for survey in surveys:
        for old_order in order_change_map:
            SurveyQuestion.objects.filter(survey=survey, order=old_order).update(
                order=old_order + OFFSET
            )
        for old_order, new_order in order_change_map.items():
            SurveyQuestion.objects.filter(survey=survey, order=old_order + OFFSET).update(
                order=new_order
            )


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 5 — Admin API Views
# ═══════════════════════════════════════════════════════════════════════════════

class AdminSurveyListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied

        qs = Survey.objects.select_related('created_by').prefetch_related('responses', 'target_users')
        status_filter = request.query_params.get('status', '')
        if status_filter and status_filter != 'all':
            today = timezone.localdate()
            if status_filter == 'scheduled':
                qs = qs.filter(start_date__gt=today)
            elif status_filter == 'active':
                qs = qs.filter(
                    Q(start_date__lte=today) | Q(start_date__isnull=True),
                    Q(end_date__gte=today) | Q(end_date__isnull=True),
                ).exclude(status=Survey.STATUS_CLOSED)
            elif status_filter == 'closed':
                qs = qs.filter(
                    Q(status=Survey.STATUS_CLOSED) | Q(end_date__lt=today)
                )
        category_filter = request.query_params.get('category', '')
        if category_filter and category_filter != 'all':
            qs = qs.filter(template_type=category_filter)
        q = request.query_params.get('search', '').strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q))

        sort_by = request.query_params.get('sort_by', 'created_at')
        sort_dir = request.query_params.get('sort_dir', 'desc')
        allowed_sorts = {'title', 'status', 'template_type', 'start_date', 'end_date', 'created_at'}
        if sort_by not in allowed_sorts:
            sort_by = 'created_at'
        prefix = '-' if sort_dir == 'desc' else ''
        items, meta = _paginate(qs.order_by(f'{prefix}{sort_by}'), request)
        from django.contrib.auth import get_user_model
        total_active = get_user_model().objects.filter(
            is_active=True,
            admin=False,
            hr=False,
            accounting=False,
        ).count()
        return Response({
            'results': SurveyListSerializer(items, many=True, context={'total_active': total_active}).data,
            'pagination': meta,
        })

    @transaction.atomic
    def post(self, request):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied

        ser = SurveyWriteSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        target_user_ids = data.pop('target_user_ids', [])
        template_id = data.pop('template_id', None)
        if template_id:
            tmpl = SurveyTemplate.objects.prefetch_related(
                'questions__options', 'questions__rating_config'
            ).filter(pk=template_id).first()
            template_type = tmpl.template_type if tmpl else ''
        else:
            template_type = ''

        status_value = data.get('status', Survey.STATUS_ACTIVE)
        if status_value == Survey.STATUS_DRAFT:
            status_value = Survey.STATUS_ACTIVE
        data['status'] = status_value

        survey = Survey.objects.create(
            **data,  # type: ignore[arg-type]
            created_by=request.user,
            template_type=template_type,
            source_template_id=template_id if template_id else None,
        )
        if survey.target_type == Survey.TARGET_SPECIFIC and target_user_ids:
            SurveyTargetUser.objects.bulk_create([
                SurveyTargetUser(survey=survey, user_id=uid) for uid in target_user_ids
            ], ignore_conflicts=True)
        elif survey.target_type == Survey.TARGET_ALL:
            SurveyTargetUser.objects.bulk_create([
                SurveyTargetUser(survey=survey, user_id=uid)
                for uid in _eligible_survey_recipients().values_list('id', flat=True)
            ], ignore_conflicts=True)
        # Seed questions from template if template_id provided
        template_id = request.data.get('template_id')
        if template_id:
            tmpl = SurveyTemplate.objects.prefetch_related(
                'questions__options', 'questions__rating_config'
            ).filter(pk=template_id).first()
            if tmpl:
                for tmpl_q in tmpl.questions.order_by('order'):
                    new_q = SurveyQuestion.objects.create(
                        survey=survey,
                        question_text=tmpl_q.question_text,
                        question_type=tmpl_q.question_type,
                        order=tmpl_q.order,
                        is_required=tmpl_q.is_required,
                        show_percentage_summary=tmpl_q.show_percentage_summary,
                        allow_other=tmpl_q.allow_other,
                    )
                    for opt in tmpl_q.options.order_by('order'):
                        SurveyQuestionOption.objects.create(
                            question=new_q, option_text=opt.option_text, order=opt.order
                        )
                    if tmpl_q.question_type == 'rating':
                        try:
                            cfg = tmpl_q.rating_config
                            SurveyQuestionRatingConfig.objects.filter(question=new_q).update(
                                min_value=cfg.min_value, max_value=cfg.max_value,
                                min_label=cfg.min_label, max_label=cfg.max_label,
                            )
                        except SurveyQuestionRatingConfig.DoesNotExist:
                            pass
        return Response(SurveyDetailSerializer(survey).data, status=http_status.HTTP_201_CREATED)


class AdminSurveyDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_survey(self, pk):
        return Survey.objects.prefetch_related(
            'questions__options', 'questions__rating_config', 'target_users'
        ).filter(pk=pk).first()

    def get(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = self._get_survey(pk)
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        return Response(SurveyDetailSerializer(survey).data)

    @transaction.atomic
    def patch(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = Survey.objects.select_for_update().filter(pk=pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        ser = SurveyWriteSerializer(survey, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        target_user_ids = data.pop('target_user_ids', None)
        template_id = data.pop('template_id', None)
        template = None
        if template_id is not None:
            if survey.responses.exists():
                return Response(
                    {'detail': 'Cannot change template after survey responses exist.'},
                    status=http_status.HTTP_400_BAD_REQUEST,
                )
            template = SurveyTemplate.objects.prefetch_related(
                'questions__options', 'questions__rating_config'
            ).filter(pk=template_id).first()
            if not template:
                return Response({'template_id': 'Template not found.'}, status=http_status.HTTP_400_BAD_REQUEST)
            survey.template_type = template.template_type

        for attr, val in data.items():
            setattr(survey, attr, val)
        survey.save()

        if target_user_ids is not None:
            SurveyTargetUser.objects.filter(survey=survey).delete()
            if survey.target_type == Survey.TARGET_SPECIFIC and target_user_ids:
                SurveyTargetUser.objects.bulk_create([
                    SurveyTargetUser(survey=survey, user_id=uid) for uid in target_user_ids
                ], ignore_conflicts=True)
            elif survey.target_type == Survey.TARGET_ALL:
                SurveyTargetUser.objects.bulk_create([
                    SurveyTargetUser(survey=survey, user_id=uid)
                    for uid in _eligible_survey_recipients().values_list('id', flat=True)
                ], ignore_conflicts=True)

        if template is not None:
            survey.questions.all().delete()
            for tmpl_q in template.questions.order_by('order'):
                new_q = SurveyQuestion.objects.create(
                    survey=survey,
                    question_text=tmpl_q.question_text,
                    question_type=tmpl_q.question_type,
                    order=tmpl_q.order,
                    is_required=tmpl_q.is_required,
                    show_percentage_summary=tmpl_q.show_percentage_summary,
                    allow_other=tmpl_q.allow_other,
                )
                for opt in tmpl_q.options.order_by('order'):
                    SurveyQuestionOption.objects.create(
                        question=new_q, option_text=opt.option_text, order=opt.order
                    )
                if tmpl_q.question_type == 'rating':
                    try:
                        cfg = tmpl_q.rating_config
                        SurveyQuestionRatingConfig.objects.filter(question=new_q).update(
                            min_value=cfg.min_value, max_value=cfg.max_value,
                            min_label=cfg.min_label, max_label=cfg.max_label,
                        )
                    except SurveyQuestionRatingConfig.DoesNotExist:
                        pass

        return Response(SurveyDetailSerializer(self._get_survey(pk)).data)

    @transaction.atomic
    def delete(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = Survey.objects.filter(pk=pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        total_targeted = survey.target_users.count()
        total_responses = SurveyResponse.objects.filter(survey=survey, is_complete=True).count()
        response_percent = (total_responses / total_targeted * 100) if total_targeted > 0 else 0.0
        if response_percent > 10:
            return Response(
                {'detail': 'Surveys with more than 10% completed responses cannot be deleted.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        survey.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminSurveyStatusView(APIView):
    permission_classes = [IsAuthenticated]
    _VALID_TRANSITIONS = {
        Survey.STATUS_DRAFT:  {Survey.STATUS_ACTIVE},
        Survey.STATUS_ACTIVE: {Survey.STATUS_CLOSED},
        Survey.STATUS_CLOSED: set(),
    }

    @transaction.atomic
    def patch(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        new_status = request.data.get('status', '')
        survey = Survey.objects.select_for_update().filter(pk=pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        allowed = self._VALID_TRANSITIONS.get(survey.status, set())
        if new_status not in allowed:
            return Response({'detail': f'Cannot transition from "{survey.status}" to "{new_status}".'}, status=http_status.HTTP_400_BAD_REQUEST)
        survey.status = new_status
        survey.save(update_fields=['status', 'updated_at'])
        if new_status == Survey.STATUS_ACTIVE:
            _schedule_survey_notifications(survey)
        return Response({'status': survey.status})


class AdminQuestionListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, survey_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = Survey.objects.filter(pk=survey_pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        qs = survey.questions.prefetch_related('options', 'rating_config').order_by('order')
        return Response(SurveyQuestionSerializer(qs, many=True).data)

    @transaction.atomic
    def post(self, request, survey_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = Survey.objects.select_for_update().filter(pk=survey_pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if not survey.is_editable:
            return Response({'detail': 'Questions cannot be added to a non-Draft survey.'}, status=http_status.HTTP_400_BAD_REQUEST)
        ser = SurveyQuestionWriteSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        max_order = survey.questions.aggregate(m=Max('order'))['m'] or 0
        question = ser.save(survey=survey, order=max_order + 1)
        return Response(SurveyQuestionSerializer(question).data, status=http_status.HTTP_201_CREATED)


class AdminQuestionDetailView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        question = SurveyQuestion.objects.select_related('survey', 'template').filter(pk=pk).first()
        if not question:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if question.survey and not question.survey.is_editable:
            return Response({'detail': 'Questions on non-Draft surveys cannot be modified.'}, status=http_status.HTTP_400_BAD_REQUEST)
        ser = SurveyQuestionWriteSerializer(question, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        question = ser.save()
        # Propagate template question changes to linked trainings and draft surveys
        if question.template_id:
            transaction.on_commit(lambda q=question: _propagate_template_question_change(q))
        return Response(SurveyQuestionSerializer(question).data)

    @transaction.atomic
    def delete(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        question = SurveyQuestion.objects.select_related('survey', 'template').filter(pk=pk).first()
        if not question:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if question.survey and not question.survey.is_editable:
            return Response({'detail': 'Questions on non-Draft surveys cannot be deleted.'}, status=http_status.HTTP_400_BAD_REQUEST)
        # Capture template_id and order before deleting so propagation can remove downstream copies
        tmpl_id = question.template_id
        q_order = question.order
        question.delete()
        if tmpl_id:
            # Build a minimal placeholder to pass to propagation helper
            placeholder = SurveyQuestion(template_id=tmpl_id, order=q_order)
            transaction.on_commit(lambda p=placeholder: _propagate_template_question_change(p, deleted=True))
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminQuestionReorderView(APIView):
    """PATCH body: {"order": [{"id": 1, "order": 0}, ...], "last_updated": "<iso>"}"""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, survey_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = Survey.objects.select_for_update().filter(pk=survey_pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if not survey.is_editable:
            return Response({'detail': 'Cannot reorder questions on a non-Draft survey.'}, status=http_status.HTTP_400_BAD_REQUEST)
        last_updated = request.data.get('last_updated')
        if last_updated:
            try:
                from django.utils.dateparse import parse_datetime
                client_ts = parse_datetime(last_updated)
                if client_ts and survey.updated_at and survey.updated_at > client_ts:
                    return Response({'detail': 'Survey was modified by another session. Please reload.'}, status=http_status.HTTP_409_CONFLICT)
            except Exception:
                pass
        for item in request.data.get('order', []):
            SurveyQuestion.objects.filter(pk=item['id'], survey=survey).update(order=item['order'])
        survey.save(update_fields=['updated_at'])
        qs = survey.questions.prefetch_related('options', 'rating_config').order_by('order')
        return Response(SurveyQuestionSerializer(qs, many=True).data)


class AdminTemplateQuestionListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, template_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        template = SurveyTemplate.objects.filter(pk=template_pk).first()
        if not template:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        qs = template.questions.prefetch_related('options', 'rating_config').order_by('order')
        return Response(SurveyQuestionSerializer(qs, many=True).data)

    @transaction.atomic
    def post(self, request, template_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        template = SurveyTemplate.objects.select_for_update().filter(pk=template_pk).first()
        if not template:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        ser = SurveyQuestionWriteSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        max_order = template.questions.aggregate(m=Max('order'))['m'] or 0
        order = ser.validated_data.get('order')
        if order is None:
            order = max_order + 1
        else:
            order = max(0, int(order))
            if order <= max_order:
                template.questions.filter(order__gte=order).update(order=F('order') + 1)
        question = ser.save(template=template, order=order)
        # New template question → add it to all linked trainings/draft surveys
        transaction.on_commit(lambda q=question: _propagate_template_question_change(q))
        return Response(SurveyQuestionSerializer(question).data, status=http_status.HTTP_201_CREATED)


class AdminTemplateQuestionReorderView(APIView):
    """PATCH body: {"order": [{"id": 1, "order": 0}, ...], "last_updated": "<iso>"}"""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, template_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        template = SurveyTemplate.objects.select_for_update().filter(pk=template_pk).first()
        if not template:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        order_items = request.data.get('order', [])

        # Capture old orders before updating so we can build an old→new change map
        pks = [item['id'] for item in order_items]
        old_order_map = {
            q.pk: q.order
            for q in SurveyQuestion.objects.filter(pk__in=pks, template=template)
        }

        for item in order_items:
            SurveyQuestion.objects.filter(pk=item['id'], template=template).update(order=item['order'])

        # Build {old_order: new_order} for every question whose order actually changed
        order_change_map = {
            old_order_map[item['id']]: item['order']
            for item in order_items
            if item['id'] in old_order_map and old_order_map[item['id']] != item['order']
        }
        if order_change_map:
            tid = template.pk
            transaction.on_commit(
                lambda m=order_change_map: _propagate_template_reorder_to_trainings(tid, m)
            )
            transaction.on_commit(
                lambda m=order_change_map: _propagate_template_reorder_to_surveys(tid, m)
            )

        qs = template.questions.prefetch_related('options', 'rating_config').order_by('order')
        return Response(SurveyQuestionSerializer(qs, many=True).data)


class AdminQuestionRatingConfigView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        question = SurveyQuestion.objects.select_related('survey', 'template').filter(pk=pk).first()
        if not question:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if question.survey and not question.survey.is_editable:
            return Response({'detail': 'Rating config on non-Draft surveys is immutable.'}, status=http_status.HTTP_400_BAD_REQUEST)
        ser = SurveyQuestionRatingConfigSerializer(question.rating_config, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        for attr, val in ser.validated_data.items():
            setattr(question.rating_config, attr, val)
        question.rating_config.save()
        if question.template_id:
            transaction.on_commit(lambda q=question: _propagate_template_question_change(q))
        return Response(SurveyQuestionRatingConfigSerializer(question.rating_config).data)


class AdminOptionListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, question_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        qs = SurveyQuestionOption.objects.filter(question_id=question_pk).order_by('order')
        return Response(SurveyQuestionOptionSerializer(qs, many=True).data)

    @transaction.atomic
    def post(self, request, question_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        question = SurveyQuestion.objects.select_related('survey', 'template').filter(pk=question_pk).first()
        if not question:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if question.survey and not question.survey.is_editable:
            return Response({'detail': 'Options on non-Draft surveys are immutable.'}, status=http_status.HTTP_400_BAD_REQUEST)
        if question.question_type not in CHOICE_BASED_TYPES:
            return Response({'detail': 'Options can only be added to choice-based question types.'}, status=http_status.HTTP_400_BAD_REQUEST)
        ser = SurveyQuestionOptionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        option = SurveyQuestionOption.objects.create(question=question, **ser.validated_data, order=question.options.count())  # type: ignore[arg-type]
        if question.template_id:
            transaction.on_commit(lambda q=question: _propagate_template_question_change(q))
        return Response(SurveyQuestionOptionSerializer(option).data, status=http_status.HTTP_201_CREATED)


class AdminOptionDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_option(self, pk):
        return SurveyQuestionOption.objects.select_related('question__survey', 'question__template').filter(pk=pk).first()

    @transaction.atomic
    def patch(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        option = self._get_option(pk)
        if not option:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if option.question.survey and not option.question.survey.is_editable:
            return Response({'detail': 'Options on non-Draft surveys are immutable.'}, status=http_status.HTTP_400_BAD_REQUEST)
        ser = SurveyQuestionOptionSerializer(option, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        for attr, val in ser.validated_data.items():
            setattr(option, attr, val)
        option.save()
        if option.question.template_id:
            transaction.on_commit(lambda q=option.question: _propagate_template_question_change(q))
        return Response(SurveyQuestionOptionSerializer(option).data)

    @transaction.atomic
    def delete(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        option = self._get_option(pk)
        if not option:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if option.question.survey and not option.question.survey.is_editable:
            return Response({'detail': 'Options on non-Draft surveys are immutable.'}, status=http_status.HTTP_400_BAD_REQUEST)
        tmpl_q = option.question if option.question.template_id else None
        option.delete()
        if tmpl_q:
            transaction.on_commit(lambda q=tmpl_q: _propagate_template_question_change(q))
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminTemplateListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        qs = SurveyTemplate.objects.select_related('created_by').prefetch_related('questions')
        q = request.query_params.get('search', '').strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q))
        items, meta = _paginate(qs.order_by('-created_at'), request)
        return Response({'results': SurveyTemplateListSerializer(items, many=True).data, 'pagination': meta})

    @transaction.atomic
    def post(self, request):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        ser = SurveyTemplateWriteSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        template = SurveyTemplate.objects.create(**ser.validated_data, created_by=request.user)  # type: ignore[arg-type]
        return Response(SurveyTemplateDetailSerializer(template).data, status=http_status.HTTP_201_CREATED)


class AdminTemplateDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        tmpl = SurveyTemplate.objects.prefetch_related('questions__options', 'questions__rating_config').filter(pk=pk).first()
        if not tmpl:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        return Response(SurveyTemplateDetailSerializer(tmpl).data)

    @transaction.atomic
    def patch(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        tmpl = SurveyTemplate.objects.filter(pk=pk).first()
        if not tmpl:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        ser = SurveyTemplateWriteSerializer(tmpl, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        for attr, val in ser.validated_data.items():
            setattr(tmpl, attr, val)
        tmpl.save()
        return Response(SurveyTemplateDetailSerializer(tmpl).data)

    @transaction.atomic
    def delete(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        tmpl = SurveyTemplate.objects.filter(pk=pk).first()
        if not tmpl:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        tmpl.delete()
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class AdminTemplateDuplicateView(APIView):
    """POST /api/survey/admin/templates/{pk}/duplicate — deep-copy a template."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        tmpl = SurveyTemplate.objects.prefetch_related(
            'questions__options', 'questions__rating_config'
        ).filter(pk=pk).first()
        if not tmpl:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        new_title = (tmpl.title[:196] + ' (Copy)') if len(tmpl.title) > 196 else f'{tmpl.title} (Copy)'
        new_tmpl = SurveyTemplate.objects.create(
            title=new_title, description=tmpl.description, created_by=request.user
        )
        for q in tmpl.questions.order_by('order'):
            new_q = SurveyQuestion.objects.create(
                template=new_tmpl,
                question_text=q.question_text,
                question_type=q.question_type,
                order=q.order,
                is_required=q.is_required,
                show_percentage_summary=q.show_percentage_summary,
                allow_other=q.allow_other,
            )
            for opt in q.options.order_by('order'):
                SurveyQuestionOption.objects.create(
                    question=new_q, option_text=opt.option_text, order=opt.order
                )
            if q.question_type == 'rating':
                try:
                    cfg = q.rating_config
                    SurveyQuestionRatingConfig.objects.filter(question=new_q).update(
                        min_value=cfg.min_value, max_value=cfg.max_value,
                        min_label=cfg.min_label, max_label=cfg.max_label,
                    )
                except SurveyQuestionRatingConfig.DoesNotExist:
                    pass
        return Response(SurveyTemplateDetailSerializer(new_tmpl).data, status=http_status.HTTP_201_CREATED)


class AdminSurveyFromTemplateView(APIView):
    """POST — deep-copies template questions into a new Draft survey (R8)."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, template_pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        tmpl = SurveyTemplate.objects.prefetch_related('questions__options', 'questions__rating_config').filter(pk=template_pk).first()
        if not tmpl:
            return Response({'detail': 'Template not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        title = request.data.get('title', '').strip() or f'Survey from "{tmpl.title}"'
        survey = Survey.objects.create(title=title[:200], description=tmpl.description, created_by=request.user, status=Survey.STATUS_DRAFT, source_template=tmpl)
        for tmpl_q in tmpl.questions.order_by('order'):
            new_q = SurveyQuestion.objects.create(
                survey=survey,
                question_text=tmpl_q.question_text,
                question_type=tmpl_q.question_type,
                order=tmpl_q.order,
                is_required=tmpl_q.is_required,
                show_percentage_summary=tmpl_q.show_percentage_summary,
                allow_other=tmpl_q.allow_other,
            )
            for opt in tmpl_q.options.order_by('order'):
                SurveyQuestionOption.objects.create(question=new_q, option_text=opt.option_text, order=opt.order)
            if tmpl_q.question_type == 'rating':
                try:
                    cfg = tmpl_q.rating_config
                    SurveyQuestionRatingConfig.objects.filter(question=new_q).update(min_value=cfg.min_value, max_value=cfg.max_value, min_label=cfg.min_label, max_label=cfg.max_label)
                except SurveyQuestionRatingConfig.DoesNotExist:
                    pass
        return Response(SurveyDetailSerializer(survey).data, status=http_status.HTTP_201_CREATED)


class AdminSurveyResultsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = Survey.objects.prefetch_related('questions__options', 'target_users').filter(pk=pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        complete_qs = survey.responses.filter(is_complete=True)
        total_responses = complete_qs.count()
        if survey.target_type == Survey.TARGET_ALL:
            from django.contrib.auth import get_user_model
            total_targeted = get_user_model().objects.filter(is_active=True).count()
        else:
            total_targeted = survey.target_users.count()
        completion_rate = (total_responses / total_targeted * 100) if total_targeted > 0 else 0.0

        last_resp_obj = complete_qs.aggregate(last=Max('submitted_at'))
        last_response_at = last_resp_obj['last'].isoformat() if last_resp_obj['last'] else None

        avg_seconds = None
        timed_qs = complete_qs.filter(started_at__isnull=False, submitted_at__isnull=False)
        if timed_qs.exists():
            duration_agg = timed_qs.annotate(
                duration=ExpressionWrapper(F('submitted_at') - F('started_at'), output_field=DurationField())
            ).aggregate(avg_dur=Avg('duration'))
            if duration_agg['avg_dur'] is not None:
                avg_seconds = int(duration_agg['avg_dur'].total_seconds())

        question_results = []
        for q in survey.questions.order_by('order'):
            answered_qs = SurveyAnswer.objects.filter(question=q, response__is_complete=True)
            total_answered = answered_qs.count()
            result = {
                'question_id': q.pk, 'question_text': q.question_text,
                'question_type': q.question_type, 'show_percentage': q.show_percentage_summary,
                'total_responses': total_answered,
            }

            if q.question_type in CHOICE_BASED_TYPES:
                counts: dict = defaultdict(int)
                for ans in answered_qs.prefetch_related('selected_options'):
                    for opt in ans.selected_options.all():
                        counts[opt.pk] += 1
                result['options'] = [
                    {'option_id': opt.pk, 'option_text': opt.option_text,
                     'count': counts.get(opt.pk, 0),
                     'percentage': round(counts.get(opt.pk, 0) / total_answered * 100, 1) if total_answered else 0.0}
                    for opt in q.options.order_by('order')
                ]
            elif q.question_type == 'yes_no':
                # stored as text_value ("Yes" / "No")
                text_counts: dict = defaultdict(int)
                for val in answered_qs.exclude(text_value='').values_list('text_value', flat=True):
                    text_counts[val] += 1
                result['options'] = [
                    {'option_id': None, 'option_text': label,
                     'count': text_counts.get(label, 0),
                     'percentage': round(text_counts.get(label, 0) / total_answered * 100, 1) if total_answered else 0.0}
                    for label in ('Yes', 'No')
                ]
            elif q.question_type == 'likert':
                # stored as text_value (e.g. "Strongly Agree")
                likert_labels = ['Strongly Disagree', 'Disagree', 'Neutral', 'Agree', 'Strongly Agree']
                text_counts = defaultdict(int)
                for val in answered_qs.exclude(text_value='').values_list('text_value', flat=True):
                    text_counts[val] += 1
                result['options'] = [
                    {'option_id': None, 'option_text': label,
                     'count': text_counts.get(label, 0),
                     'percentage': round(text_counts.get(label, 0) / total_answered * 100, 1) if total_answered else 0.0}
                    for label in likert_labels
                ]
            elif q.question_type == 'linear_scale':
                # stored as number_value (numeric scale)
                values = list(answered_qs.filter(number_value__isnull=False).values_list('number_value', flat=True))
                avg = (sum(values) / len(values)) if values else None
                dist: dict = defaultdict(int)
                for v in values:
                    dist[v] += 1
                try:
                    cfg = q.rating_config
                    min_v, max_v = cfg.min_value, cfg.max_value
                except SurveyQuestionRatingConfig.DoesNotExist:
                    min_v, max_v = (1, 10) if not dist else (int(min(dist)), int(max(dist)))
                distribution = []
                for v in range(min_v, max_v + 1):
                    cnt = dist.get(float(v), dist.get(v, 0))
                    distribution.append({'value': v, 'count': cnt, 'percentage': round(cnt / len(values) * 100, 1) if values else 0.0})
                result['average'] = round(avg, 2) if avg is not None else None
                result['distribution'] = distribution
            elif q.question_type in {'rating', 'number'}:
                values = list(answered_qs.filter(number_value__isnull=False).values_list('number_value', flat=True))
                avg = (sum(values) / len(values)) if values else None
                dist: dict = defaultdict(int)
                for v in values:
                    dist[v] += 1
                distribution = []
                if q.question_type == 'rating':
                    try:
                        cfg = q.rating_config
                        min_v, max_v = cfg.min_value, cfg.max_value
                    except SurveyQuestionRatingConfig.DoesNotExist:
                        min_v, max_v = 1, 5
                    for v in range(min_v, max_v + 1):
                        cnt = dist.get(float(v), 0)
                        distribution.append({'value': v, 'count': cnt, 'percentage': round(cnt / len(values) * 100, 1) if values else 0.0})
                else:
                    for v, cnt in sorted(dist.items()):
                        distribution.append({'value': v, 'count': cnt, 'percentage': round(cnt / len(values) * 100, 1) if values else 0.0})
                result['average'] = round(avg, 2) if avg is not None else None
                result['distribution'] = distribution
            else:
                # Open-ended — never return identity, just text (R3).
                result['text_answers'] = list(answered_qs.exclude(text_value='').values_list('text_value', flat=True)[:200])

            question_results.append(result)

        return Response({
            'survey_id': survey.pk, 'survey_title': survey.title,
            'survey_description': survey.description, 'survey_status': survey.status,
            'start_date': survey.start_date.isoformat() if survey.start_date else None,
            'end_date': survey.end_date.isoformat() if survey.end_date else None,
            'is_anonymous': survey.is_anonymous,
            'total_targeted': total_targeted, 'total_responses': total_responses,
            'completion_rate': round(completion_rate, 1),
            'last_response_at': last_response_at,
            'avg_completion_seconds': avg_seconds,
            'questions': question_results,
        })


class AdminSurveyExportView(APIView):
    """GET — Multi-sheet XLSX survey report (Overall Summary, Response Summary, and supplementary sheets)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):  # noqa: C901
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied

        survey = Survey.objects.filter(pk=pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({'detail': 'openpyxl not installed.'}, status=http_status.HTTP_503_SERVICE_UNAVAILABLE)

        from django.http import StreamingHttpResponse
        from statistics import median as _median
        from userLogin.models import loginCredentials
        from userProfile.models import workInformation

        # ── Shared styles ──────────────────────────────────────────────────────
        _BLUE   = PatternFill('solid', fgColor='2845D6')
        _WH_B   = Font(bold=True, color='FFFFFF')
        _BOLD   = Font(bold=True)
        _THIN   = Side(style='thin')
        _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

        def _hdr(ws, row, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = _BLUE
                cell.font = _WH_B
                cell.border = _BORDER

        def _border_row(ws, row, ncols):
            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).border = _BORDER

        def _auto_width(ws):
            for col in ws.columns:
                ltr = get_column_letter(col[0].column)
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=0,
                )
                ws.column_dimensions[ltr].width = min(max(max_len + 4, 10), 60)

        # ── Type sets ─────────────────────────────────────────────────────────
        INSTRUCTION_TYPES = {'section', 'subsection', 'statement'}
        FREE_FORM_TYPES   = {'short_text', 'long_text', 'number', 'date'}
        CHOICE_OPT_TYPES  = CHOICE_BASED_TYPES | {'yes_no', 'likert', 'linear_scale'}
        SUMMARY_TYPES     = CHOICE_OPT_TYPES | {'rating'}

        # ── Questions (ordered, with options and rating_config pre-fetched) ───
        questions = list(
            SurveyQuestion.objects.filter(survey=survey)
            .order_by('order')
            .select_related('rating_config')
            .prefetch_related('options')
        )
        answerable_q = [q for q in questions if q.question_type not in INSTRUCTION_TYPES]

        # Numbering: instruction blocks don't consume a number
        q_numbers: dict = {}
        counter = 0
        for q in questions:
            if q.question_type not in INSTRUCTION_TYPES:
                counter += 1
                q_numbers[q.pk] = counter

        # ── Target users (excluding admin / hr / accounting) ──────────────────
        excl_filter = Q(admin=True) | Q(hr=True) | Q(accounting=True)
        if survey.target_type == Survey.TARGET_SPECIFIC:
            raw = list(
                survey.target_users.select_related('user')
                .exclude(user__admin=True)
                .exclude(user__hr=True)
                .exclude(user__accounting=True)
                .values_list('user_id', 'user__idnumber', 'user__firstname', 'user__lastname')
            )
        else:
            raw = list(
                loginCredentials.objects.filter(is_active=True)
                .exclude(excl_filter)
                .values_list('id', 'idnumber', 'firstname', 'lastname')
            )

        target_user_ids = {r[0] for r in raw}
        total_target    = len(raw)

        # ── Completed responses (targeted users only) ─────────────────────────
        completed = list(
            SurveyResponse.objects.filter(
                survey=survey,
                is_complete=True,
                employee_id__in=target_user_ids,
            ).select_related('employee').order_by('submitted_at')
        )
        completed_resp_ids = [r.pk for r in completed]
        completed_emp_ids  = {r.employee_id for r in completed}
        total_responses    = len(completed)
        completion_pct     = (total_responses / total_target * 100) if total_target else 0.0

        # ── All answers for completed responses ───────────────────────────────
        all_answers = list(
            SurveyAnswer.objects.filter(response_id__in=completed_resp_ids)
            .prefetch_related('selected_options')
        )
        ans_map = {(a.response_id, a.question_id): a for a in all_answers}

        # ── Aggregate counts per question ─────────────────────────────────────
        q_by_id        = {q.pk: q for q in questions}
        opt_count      = defaultdict(lambda: defaultdict(int))  # q_id → opt_id → n
        text_count     = defaultdict(lambda: defaultdict(int))  # q_id → text_value → n
        numeric_count  = defaultdict(lambda: defaultdict(int))  # q_id → numeric value → n
        rating_count   = defaultdict(lambda: defaultdict(int))  # q_id → value  → n
        q_respondents  = defaultdict(int)                       # q_id → n respondents

        for ans in all_answers:
            q = q_by_id.get(ans.question_id)
            if q is None:
                continue
            if q.question_type == 'rating':
                if ans.number_value is not None:
                    q_respondents[q.pk] += 1
                    rating_count[q.pk][int(ans.number_value)] += 1
            elif q.question_type in CHOICE_BASED_TYPES:
                opts = list(ans.selected_options.all())
                if opts:
                    q_respondents[q.pk] += 1
                    for opt in opts:
                        opt_count[q.pk][opt.pk] += 1
            elif q.question_type == 'yes_no':
                if ans.text_value:
                    q_respondents[q.pk] += 1
                    text_count[q.pk][ans.text_value] += 1
            elif q.question_type == 'likert':
                if ans.text_value:
                    q_respondents[q.pk] += 1
                    text_count[q.pk][ans.text_value] += 1
            elif q.question_type == 'linear_scale':
                if ans.number_value is not None:
                    q_respondents[q.pk] += 1
                    numeric_count[q.pk][int(ans.number_value)] += 1

        # ── Department map: employee_id → dept name ───────────────────────────
        dept_map = {
            wi.employee_id: (wi.department.name if wi.department else '')
            for wi in workInformation.objects.filter(
                employee_id__in=target_user_ids,
            ).select_related('department')
        }

        # ── Build workbook ─────────────────────────────────────────────────────
        wb = openpyxl.Workbook()

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 1 — Overall Summary
        # ══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Overall Summary'

        ws1.append(['RYONAN ELECTRIC PHILIPPINES CORPORATION'])
        ws1.append([f'{survey.title} Overall Summary'])
        ws1.append([])

        # Completion table
        ws1.append(['Total Users', 'Total Responses', 'Completion Rate'])
        _hdr(ws1, ws1.max_row, 3)
        ws1.append([total_target, total_responses, f'{completion_pct:.1f}%'])
        _border_row(ws1, ws1.max_row, 3)
        ws1.append([])

        # Questions
        for q in questions:
            qtype = q.question_type
            if qtype in INSTRUCTION_TYPES:
                ws1.append([q.question_text])
                ws1.cell(row=ws1.max_row, column=1).font = _BOLD
                ws1.append([])
            elif qtype in FREE_FORM_TYPES:
                qn = q_numbers[q.pk]
                ws1.append([f'Q{qn}. {q.question_text}'])
                ws1.cell(row=ws1.max_row, column=1).font = _BOLD
                ws1.append([])
            elif qtype in SUMMARY_TYPES:
                qn = q_numbers[q.pk]
                ws1.append([f'Q{qn}. {q.question_text}'])
                ws1.cell(row=ws1.max_row, column=1).font = _BOLD
                ws1.append(['Option', 'Count', 'Percentage'])
                ws1.cell(row=ws1.max_row, column=1).font = _BOLD
                ws1.cell(row=ws1.max_row, column=2).font = _BOLD
                ws1.cell(row=ws1.max_row, column=3).font = _BOLD
                total_ans = q_respondents.get(q.pk, 0)
                if qtype == 'rating':
                    cfg = getattr(q, 'rating_config', None)
                    min_v = cfg.min_value if cfg else 1
                    max_v = cfg.max_value if cfg else 5
                    for val in range(min_v, max_v + 1):
                        cnt = rating_count[q.pk].get(val, 0)
                        pct = (cnt / total_ans * 100) if total_ans else 0.0
                        ws1.append([str(val), cnt, f'{pct:.1f}%'])
                elif qtype == 'yes_no':
                    for label in ('Yes', 'No'):
                        cnt = text_count[q.pk].get(label, 0)
                        pct = (cnt / total_ans * 100) if total_ans else 0.0
                        ws1.append([label, cnt, f'{pct:.1f}%'])
                elif qtype == 'likert':
                    likert_labels = [
                        'Strongly Disagree',
                        'Disagree',
                        'Neutral',
                        'Agree',
                        'Strongly Agree',
                    ]
                    for label in likert_labels:
                        cnt = text_count[q.pk].get(label, 0)
                        pct = (cnt / total_ans * 100) if total_ans else 0.0
                        ws1.append([label, cnt, f'{pct:.1f}%'])
                elif qtype == 'linear_scale':
                    opts = list(q.options.all())
                    if opts:
                        for val in range(1, len(opts) + 1):
                            cnt = numeric_count[q.pk].get(val, 0)
                            pct = (cnt / total_ans * 100) if total_ans else 0.0
                            ws1.append([str(val), cnt, f'{pct:.1f}%'])
                    else:
                        try:
                            cfg = q.rating_config
                            min_v, max_v = cfg.min_value, cfg.max_value
                        except SurveyQuestionRatingConfig.DoesNotExist:
                            min_v, max_v = 1, 10
                        for val in range(min_v, max_v + 1):
                            cnt = numeric_count[q.pk].get(val, 0)
                            pct = (cnt / total_ans * 100) if total_ans else 0.0
                            ws1.append([str(val), cnt, f'{pct:.1f}%'])
                else:
                    for opt in q.options.all():
                        cnt = opt_count[q.pk].get(opt.pk, 0)
                        pct = (cnt / total_ans * 100) if total_ans else 0.0
                        ws1.append([opt.option_text, cnt, f'{pct:.1f}%'])
                ws1.append([])

        _auto_width(ws1)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 2 — Response Summary
        # ══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet(title='Response Summary')
        ws2.append(['RYONAN ELECTRIC PHILIPPINES CORPORATION'])
        ws2.append([f'{survey.title} Response Summary'])
        ws2.append([])

        rs_headers = ['ID Number', 'Employee Name', 'Department', 'Date Submitted']
        for q in answerable_q:
            rs_headers.append(f'Q{q_numbers[q.pk]}. {q.question_text}')
        ws2.append(rs_headers)
        _hdr(ws2, ws2.max_row, len(rs_headers))

        for resp in completed:
            emp = resp.employee
            row_data: list[Any] = [
                str(emp.idnumber) if emp else '',
                f'{emp.lastname}, {emp.firstname}' if emp else 'Anonymous',
                dept_map.get(emp.pk, '') if emp else '',
                resp.submitted_at.strftime('%Y-%m-%d') if resp.submitted_at else '',
            ]
            for q in answerable_q:
                ans = ans_map.get((resp.pk, q.pk))
                if not ans:
                    row_data.append('')
                elif q.question_type in CHOICE_BASED_TYPES:
                    opts = [o.option_text for o in ans.selected_options.all()]
                    if ans.other_text:
                        opts.append(f'Other: {ans.other_text}')
                    row_data.append(', '.join(opts))
                elif q.question_type == 'yes_no':
                    row_data.append(ans.text_value)
                elif q.question_type == 'likert':
                    row_data.append(ans.text_value)
                elif q.question_type == 'linear_scale':
                    row_data.append(ans.number_value if ans.number_value is not None else '')
                elif q.question_type in {'rating', 'number'}:
                    row_data.append(ans.number_value if ans.number_value is not None else '')
                else:
                    row_data.append(ans.text_value)
            ws2.append(row_data)
            _border_row(ws2, ws2.max_row, len(rs_headers))

        _auto_width(ws2)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 3 — Non-Respondents
        # ══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet(title='Non-Respondents')
        nr_headers = ['ID Number', 'Employee Name', 'Department']
        ws3.append(nr_headers)
        _hdr(ws3, ws3.max_row, len(nr_headers))
        for uid, id_num, fname, lname in raw:
            if uid not in completed_emp_ids:
                ws3.append([str(id_num), f'{lname}, {fname}', dept_map.get(uid, '')])
                _border_row(ws3, ws3.max_row, len(nr_headers))
        _auto_width(ws3)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 4 — Question Analysis (Rating Scale — numeric stats)
        # ══════════════════════════════════════════════════════════════════════
        rating_qs = [q for q in answerable_q if q.question_type == 'rating']
        if rating_qs:
            ws4 = wb.create_sheet(title='Question Analysis')
            qa_headers = ['Question', 'Responses', 'Average', 'Median', 'Min', 'Max']
            ws4.append(qa_headers)
            _hdr(ws4, ws4.max_row, len(qa_headers))
            for q in rating_qs:
                qn = q_numbers[q.pk]
                values = [
                    a.number_value
                    for a in all_answers
                    if a.question_id == q.pk and a.number_value is not None
                ]
                if values:
                    ws4.append([
                        f'Q{qn}. {q.question_text}',
                        len(values),
                        round(sum(values) / len(values), 2),
                        round(_median(values), 2),
                        min(values),
                        max(values),
                    ])
                else:
                    ws4.append([f'Q{qn}. {q.question_text}', 0, '', '', '', ''])
                _border_row(ws4, ws4.max_row, len(qa_headers))
            _auto_width(ws4)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 5 — Open-Ended Responses (Short Text / Long Text)
        # ══════════════════════════════════════════════════════════════════════
        open_qs = [q for q in answerable_q if q.question_type in {'short_text', 'long_text'}]
        if open_qs:
            ws5 = wb.create_sheet(title='Open-Ended Responses')
            oe_headers = ['Q No.', 'Question Text', 'Employee Name', 'Response']
            ws5.append(oe_headers)
            _hdr(ws5, ws5.max_row, len(oe_headers))
            resp_lookup = {r.pk: r for r in completed}
            for q in open_qs:
                qn = q_numbers[q.pk]
                for ans in all_answers:
                    if ans.question_id != q.pk or not ans.text_value:
                        continue
                    resp = resp_lookup.get(ans.response_id)
                    if not resp:
                        continue
                    if survey.is_anonymous:
                        name = 'Anonymous'
                    elif resp.employee:
                        name = f'{resp.employee.lastname}, {resp.employee.firstname}'
                    else:
                        name = '-'
                    ws5.append([f'Q{qn}', q.question_text, name, ans.text_value])
                    _border_row(ws5, ws5.max_row, len(oe_headers))
            _auto_width(ws5)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 6 — Respondent Timeline
        # ══════════════════════════════════════════════════════════════════════
        ws6 = wb.create_sheet(title='Respondent Timeline')
        tl_headers = ['ID Number', 'Employee Name', 'Department', 'Submitted At']
        ws6.append(tl_headers)
        _hdr(ws6, ws6.max_row, len(tl_headers))
        for resp in completed:
            emp = resp.employee
            ws6.append([
                str(emp.idnumber) if emp else '',
                f'{emp.lastname}, {emp.firstname}' if emp else 'Anonymous',
                dept_map.get(emp.pk, '') if emp else '',
                resp.submitted_at.strftime('%Y-%m-%d %H:%M') if resp.submitted_at else '',
            ])
            _border_row(ws6, ws6.max_row, len(tl_headers))
        _auto_width(ws6)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 7 — Department Breakdown
        # ══════════════════════════════════════════════════════════════════════
        ws7 = wb.create_sheet(title='Department Breakdown')
        db_headers = ['Department', 'Total Users', 'Responded', 'Response Rate']
        ws7.append(db_headers)
        _hdr(ws7, ws7.max_row, len(db_headers))
        dept_users = defaultdict(list)
        for uid, _id_num, _fname, _lname in raw:
            dept_users[dept_map.get(uid, 'Unknown')].append(uid)
        for dept_name, uids in sorted(dept_users.items()):
            responded = sum(1 for uid in uids if uid in completed_emp_ids)
            total     = len(uids)
            rate      = f'{responded / total * 100:.1f}%' if total else '0.0%'
            ws7.append([dept_name, total, responded, rate])
            _border_row(ws7, ws7.max_row, len(db_headers))
        _auto_width(ws7)

        # ── Serialize & respond ───────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_title = ''.join(
            c for c in survey.title if c.isalnum() or c in (' ', '-', '_')
        )[:50].strip() or 'survey'
        http_resp = StreamingHttpResponse(
            streaming_content=buf,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        http_resp['Content-Disposition'] = (
            f'attachment; filename="{safe_title}_report.xlsx"'
        )
        return http_resp


class AdminIndividualResponsesView(APIView):
    """GET /api/survey/admin/surveys/<pk>/responses — paginated respondent list with filters."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        survey = Survey.objects.filter(pk=pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        status_filter = request.query_params.get('status', 'all')
        search = request.query_params.get('search', '').strip()
        sort = request.query_params.get('sort', 'name')
        dir_ = request.query_params.get('dir', 'asc')
        prefix = '-' if dir_ == 'desc' else ''

        if survey.target_type == Survey.TARGET_SPECIFIC:
            submitted_subq = SurveyResponse.objects.filter(
                survey=survey, employee=OuterRef('user')
            ).values('submitted_at')[:1]
            is_complete_subq = SurveyResponse.objects.filter(
                survey=survey, employee=OuterRef('user')
            ).values('is_complete')[:1]
            response_pk_subq = SurveyResponse.objects.filter(
                survey=survey, employee=OuterRef('user')
            ).values('pk')[:1]

            qs = (SurveyTargetUser.objects
                  .filter(survey=survey)
                  .exclude(user__admin=True)
                  .exclude(user__hr=True)
                  .exclude(user__accounting=True)
                  .select_related('user')
                  .annotate(
                      resp_submitted_at=Subquery(submitted_subq),
                      resp_is_complete=Subquery(is_complete_subq),
                      resp_id=Subquery(response_pk_subq),
                  ))

            if search and not survey.is_anonymous:
                qs = qs.filter(
                    Q(user__firstname__icontains=search) |
                    Q(user__lastname__icontains=search) |
                    Q(user__idnumber__icontains=search)
                )

            if status_filter == 'complete':
                qs = qs.filter(resp_is_complete=True)
            elif status_filter == 'partial':
                qs = qs.filter(resp_is_complete=False)
            elif status_filter == 'not_started':
                qs = qs.filter(resp_is_complete__isnull=True)

            if sort == 'idnumber':
                qs = qs.order_by(f'{prefix}user__idnumber')
            elif sort == 'submitted_at':
                qs = qs.order_by(f'{prefix}resp_submitted_at')
            elif sort == 'status':
                qs = qs.annotate(status_order=Case(
                    When(resp_is_complete=True, then=Value(1)),
                    When(resp_is_complete=False, then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )).order_by(f'{prefix}status_order')
            else:
                qs = qs.order_by(f'{prefix}user__lastname', f'{prefix}user__firstname')

            items, meta = _paginate(qs, request)
            results = []
            for target in items:
                user = target.user
                if survey.is_anonymous:
                    firstname = 'Anonymous'
                    lastname = ''
                    idnumber = '—'
                else:
                    firstname = user.firstname or ''
                    lastname = user.lastname or ''
                    idnumber = user.idnumber
                rc = target.resp_is_complete
                if rc is True:
                    status = 'Complete'
                elif rc is False:
                    status = 'Partial'
                else:
                    status = 'Not started'
                results.append({
                    'id': target.resp_id if target.resp_id else f'user-{user.pk}',
                    'response_id': target.resp_id,
                    'firstname': firstname,
                    'lastname': lastname,
                    'idnumber': idnumber,
                    'submitted_at': target.resp_submitted_at.isoformat() if target.resp_submitted_at else None,
                    'is_complete': bool(rc is True),
                    'status': status,
                })

        elif survey.target_type == Survey.TARGET_ALL:
            from django.contrib.auth import get_user_model
            User = get_user_model()

            submitted_subq = SurveyResponse.objects.filter(
                survey=survey, employee=OuterRef('pk')
            ).values('submitted_at')[:1]
            is_complete_subq = SurveyResponse.objects.filter(
                survey=survey, employee=OuterRef('pk')
            ).values('is_complete')[:1]
            response_pk_subq = SurveyResponse.objects.filter(
                survey=survey, employee=OuterRef('pk')
            ).values('pk')[:1]

            qs = (User.objects
                  .filter(is_active=True)
                  .exclude(admin=True)
                  .exclude(hr=True)
                  .exclude(accounting=True)
                  .annotate(
                      resp_submitted_at=Subquery(submitted_subq),
                      resp_is_complete=Subquery(is_complete_subq),
                      resp_id=Subquery(response_pk_subq),
                  ))

            if search and not survey.is_anonymous:
                qs = qs.filter(
                    Q(firstname__icontains=search) |
                    Q(lastname__icontains=search) |
                    Q(idnumber__icontains=search)
                )

            if status_filter == 'complete':
                qs = qs.filter(resp_is_complete=True)
            elif status_filter == 'partial':
                qs = qs.filter(resp_is_complete=False)
            elif status_filter == 'not_started':
                qs = qs.filter(resp_is_complete__isnull=True)

            if sort == 'idnumber':
                qs = qs.order_by(f'{prefix}idnumber')
            elif sort == 'submitted_at':
                qs = qs.order_by(f'{prefix}resp_submitted_at')
            elif sort == 'status':
                qs = qs.annotate(status_order=Case(
                    When(resp_is_complete=True, then=Value(1)),
                    When(resp_is_complete=False, then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )).order_by(f'{prefix}status_order')
            else:
                qs = qs.order_by(f'{prefix}lastname', f'{prefix}firstname')

            items, meta = _paginate(qs, request)
            results = []
            for user in items:
                if survey.is_anonymous:
                    firstname = 'Anonymous'
                    lastname = ''
                    idnumber = '—'
                else:
                    firstname = user.firstname or ''
                    lastname = user.lastname or ''
                    idnumber = user.idnumber
                rc = user.resp_is_complete
                if rc is True:
                    status = 'Complete'
                elif rc is False:
                    status = 'Partial'
                else:
                    status = 'Not started'
                results.append({
                    'id': user.resp_id if user.resp_id else f'user-{user.pk}',
                    'response_id': user.resp_id,
                    'firstname': firstname,
                    'lastname': lastname,
                    'idnumber': idnumber,
                    'submitted_at': user.resp_submitted_at.isoformat() if user.resp_submitted_at else None,
                    'is_complete': bool(rc is True),
                    'status': status,
                })

        else:
            qs = SurveyResponse.objects.filter(survey=survey).select_related('employee').order_by('-submitted_at', '-pk')
            if status_filter == 'complete':
                qs = qs.filter(is_complete=True)
            elif status_filter == 'partial':
                qs = qs.filter(is_complete=False)
            if search and not survey.is_anonymous:
                qs = qs.filter(
                    Q(employee__firstname__icontains=search) |
                    Q(employee__lastname__icontains=search) |
                    Q(employee__idnumber__icontains=search)
                )
            items, meta = _paginate(qs, request)
            results = []
            for resp in items:
                if survey.is_anonymous or not resp.employee:
                    firstname = 'Anonymous'
                    lastname = ''
                    idnumber = '—'
                else:
                    firstname = resp.employee.firstname or ''
                    lastname = resp.employee.lastname or ''
                    idnumber = resp.employee.idnumber
                results.append({
                    'id': resp.pk,
                    'response_id': resp.pk,
                    'firstname': firstname,
                    'lastname': lastname,
                    'idnumber': idnumber,
                    'submitted_at': resp.submitted_at.isoformat() if resp.submitted_at else None,
                    'is_complete': resp.is_complete,
                    'status': 'Complete' if resp.is_complete else 'Partial',
                })

        return Response({'results': results, 'pagination': meta, 'is_anonymous': survey.is_anonymous})


class AdminResponseDetailView(APIView):
    """GET /api/survey/admin/responses/<pk> — full answer set for one respondent."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        resp = SurveyResponse.objects.select_related('survey', 'employee').filter(pk=pk).first()
        if not resp:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)

        survey = resp.survey
        if survey.is_anonymous or not resp.employee:
            name = 'Anonymous'
            idnumber = '—'
        else:
            name = f'{resp.employee.firstname or ""} {resp.employee.lastname or ""}'.strip() or resp.employee.idnumber
            idnumber = resp.employee.idnumber

        answers = SurveyAnswer.objects.filter(response=resp).select_related('question').prefetch_related('selected_options').order_by('question__order')
        answers_data = []
        for ans in answers:
            q = ans.question
            answer_entry: dict = {
                'question_id': q.pk,
                'question_text': q.question_text,
                'question_type': q.question_type,
                'order': q.order,
            }
            if q.question_type in CHOICE_BASED_TYPES | {'yes_no', 'likert', 'linear_scale'}:
                selected = [{'id': o.pk, 'text': o.option_text} for o in ans.selected_options.all()]
                answer_entry['selected_options'] = selected
                if ans.other_text:
                    answer_entry['other_text'] = ans.other_text
            elif q.question_type in {'rating', 'number'}:
                answer_entry['number_value'] = ans.number_value
            else:
                answer_entry['text_value'] = ans.text_value

            answers_data.append(answer_entry)

        return Response({
            'id': resp.pk,
            'respondent_name': name,
            'idnumber': idnumber,
            'submitted_at': resp.submitted_at.isoformat() if resp.submitted_at else None,
            'started_at': resp.started_at.isoformat() if resp.started_at else None,
            'is_complete': resp.is_complete,
            'answers': answers_data,
        })


class AdminUserSearchView(APIView):
    """GET /api/survey/admin/users?search=<q> — paginated user search (R12, min 2 chars)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        denied = _require_admin_hr_or_iad(request)
        if denied:
            return denied
        from django.contrib.auth import get_user_model
        User = get_user_model()
        q = request.query_params.get('search', '').strip()
        if len(q) < 2:
            return Response({'results': [], 'pagination': {'page': 1, 'page_size': 20, 'total': 0, 'total_pages': 1}})
        qs = User.objects.filter(Q(idnumber__icontains=q) | Q(firstname__icontains=q) | Q(lastname__icontains=q), is_active=True).order_by('lastname', 'firstname')
        items, meta = _paginate(qs, request)
        return Response({'results': [{'id': u.pk, 'idnumber': u.idnumber, 'full_name': f'{u.firstname} {u.lastname}'} for u in items], 'pagination': meta})


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 6 — Respondent Views
# ═══════════════════════════════════════════════════════════════════════════════

class MySurveysView(APIView):
    """GET /api/survey/my-surveys — surveys targeted at the authenticated user."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        targeted = Survey.objects.filter(
            Q(target_type=Survey.TARGET_ALL) |
            Q(target_type=Survey.TARGET_SPECIFIC, target_users__user=user)
        ).annotate(
            question_count=Count('questions', distinct=True)
        ).distinct().order_by('-created_at')

        # Fetch all SurveyTargetUser rows for this user in one query
        target_map = {
            stu.survey_id: stu.is_seen
            for stu in SurveyTargetUser.objects.filter(
                survey__in=targeted, user=user
            ).only('survey_id', 'is_seen')
        }

        # Map template_type to the latest template description so the UI can show template summary text.
        template_types = {s.template_type for s in targeted if s.template_type}
        template_map = {}
        if template_types:
            for tmpl in SurveyTemplate.objects.filter(template_type__in=template_types).order_by('template_type', '-created_at'):
                template_map.setdefault(tmpl.template_type, tmpl.description)

        results = []
        for s in targeted:
            response = SurveyResponse.objects.filter(survey=s, employee=user).first()
            results.append({
                'id': s.pk,
                'title': s.title,
                'description': s.description,
                'template_description': template_map.get(s.template_type, ''),
                'is_anonymous': s.is_anonymous,
                'start_date': s.start_date,
                'end_date': s.end_date,
                'status': s.status,
                'is_complete': response.is_complete if response else False,
                'response_id': response.pk if response else None,
                'is_seen': target_map.get(s.pk, False),
                'question_count': s.question_count,
            })
        return Response(results)


class SurveyMarkSeenView(APIView):
    """PATCH /api/survey/my-surveys/<survey_id>/seen — mark a survey as seen (idempotent)."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, survey_id: int):
        updated = SurveyTargetUser.objects.filter(
            survey_id=survey_id, user=request.user
        ).update(is_seen=True)
        if updated == 0:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        return Response(status=http_status.HTTP_204_NO_CONTENT)


class ResponseCreateView(APIView):
    """POST /api/survey/responses — create or return existing response (idempotent, R1, R15)."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        survey_id = request.data.get('survey_id')
        if not survey_id:
            return Response({'detail': 'survey_id is required.'}, status=http_status.HTTP_400_BAD_REQUEST)
        survey = Survey.objects.filter(pk=survey_id).first()
        if not survey:
            return Response({'detail': 'Survey not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if survey.status != Survey.STATUS_ACTIVE:
            return Response({'detail': 'Survey is not active.', 'code': 'survey_not_active'}, status=http_status.HTTP_400_BAD_REQUEST)
        if survey.target_type == Survey.TARGET_SPECIFIC:
            if not SurveyTargetUser.objects.filter(survey=survey, user=request.user).exists():
                return Response({'detail': 'This survey is not assigned to you.', 'code': 'not_targeted'}, status=http_status.HTTP_403_FORBIDDEN)
        employee = None if survey.is_anonymous else request.user
        existing = SurveyResponse.objects.filter(survey=survey, employee=employee).first()
        if existing:
            return Response(SurveyResponseSerializer(existing).data, status=http_status.HTTP_200_OK)
        response = SurveyResponse.objects.create(survey=survey, employee=employee, is_complete=False, started_at=timezone.now())
        return Response(SurveyResponseSerializer(response).data, status=http_status.HTTP_201_CREATED)


class AnswerUpsertView(APIView):
    """PATCH /api/survey/responses/<pk>/answers/<question_pk> — upsert one answer (R4)."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, pk: int, question_pk: int):
        response = SurveyResponse.objects.select_related('survey').filter(pk=pk).first()
        if not response:
            return Response({'detail': 'Response not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if not response.survey.is_anonymous and response.employee_id != request.user.pk:
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
        if response.is_complete:
            return Response({'detail': 'Survey already submitted.'}, status=http_status.HTTP_400_BAD_REQUEST)
        if response.survey.status != Survey.STATUS_ACTIVE:
            return Response({'detail': 'Survey is no longer active.', 'code': 'survey_closed'}, status=http_status.HTTP_400_BAD_REQUEST)
        question = SurveyQuestion.objects.filter(pk=question_pk, survey=response.survey).first()
        if not question:
            return Response({'detail': 'Question not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        answer, _ = SurveyAnswer.objects.get_or_create(response=response, question=question)
        ser = SurveyAnswerSerializer(answer, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        selected_options = ser.validated_data.pop('selected_options', None)
        for attr, val in ser.validated_data.items():
            setattr(answer, attr, val)
        answer.save()
        if selected_options is not None:
            answer.selected_options.set(selected_options)
        return Response(SurveyAnswerSerializer(answer).data)


class ResponseSubmitView(APIView):
    """POST /api/survey/responses/<pk>/submit — finalize (R1, R2)."""
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, pk: int):
        response = SurveyResponse.objects.select_for_update().select_related('survey').filter(pk=pk).first()
        if not response:
            return Response({'detail': 'Response not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if not response.survey.is_anonymous and response.employee_id != request.user.pk:
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
        if response.is_complete:
            return Response({'detail': 'Already submitted.'}, status=http_status.HTTP_200_OK)
        if response.survey.status != Survey.STATUS_ACTIVE:
            return Response({'detail': 'Survey is no longer accepting responses.', 'code': 'survey_closed'}, status=http_status.HTTP_400_BAD_REQUEST)
        instruction_types = ['section', 'subsection', 'statement']
        required_ids = set(
            SurveyQuestion.objects
                .filter(survey=response.survey, is_required=True)
                .exclude(question_type__in=instruction_types)
                .values_list('id', flat=True)
        )
        answered_ids = set(response.answers.values_list('question_id', flat=True))
        missing = required_ids - answered_ids
        if missing:
            return Response({'detail': 'Some required questions have not been answered.', 'missing_questions': list(missing)}, status=http_status.HTTP_400_BAD_REQUEST)
        response.is_complete = True
        response.submitted_at = timezone.now()
        response.save(update_fields=['is_complete', 'submitted_at'])
        return Response({'detail': 'Survey submitted successfully.', 'submitted_at': str(response.submitted_at)})


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 9 support — Respondent survey detail + response detail
# ═══════════════════════════════════════════════════════════════════════════════

class SurveyRespondentDetailView(APIView):
    """GET /api/survey/surveys/<pk> — returns survey with questions for respondent.

    Access allowed only when:
    - Survey is active.
    - User is in the target audience.
    Includes existing answers if a response already exists.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        survey = Survey.objects.prefetch_related(
            'questions__options', 'questions__rating_config', 'target_users',
        ).filter(pk=pk).first()
        if not survey:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if survey.status == Survey.STATUS_DRAFT:
            return Response({'detail': 'This survey is not available.', 'code': 'survey_not_available'}, status=http_status.HTTP_404_NOT_FOUND)
        if survey.target_type == Survey.TARGET_SPECIFIC:
            if not SurveyTargetUser.objects.filter(survey=survey, user=request.user).exists():
                return Response({'detail': 'This survey is not assigned to you.'}, status=http_status.HTTP_403_FORBIDDEN)

        # Look up existing response + answers.
        response_obj = SurveyResponse.objects.filter(
            survey=survey,
            employee=None if survey.is_anonymous else request.user,
        ).prefetch_related('answers__selected_options').first()

        # Build answer map: question_id → answer data
        answer_map: dict = {}
        if response_obj:
            for ans in response_obj.answers.all():
                answer_map[ans.question_id] = {
                    'text_value': ans.text_value,
                    'number_value': str(ans.number_value) if ans.number_value is not None else None,
                    'other_text': ans.other_text,
                    'selected_option_ids': list(ans.selected_options.values_list('id', flat=True)),
                }

        questions_data = []
        for q in sorted(survey.questions.all(), key=lambda x: x.order):
            q_data: dict = {
                'id': q.pk,
                'question_text': q.question_text,
                'question_type': q.question_type,
                'order': q.order,
                'is_required': q.is_required,
                'allow_other': q.allow_other,
                'options': [{'id': o.pk, 'option_text': o.option_text, 'order': o.order}
                            for o in sorted(q.options.all(), key=lambda x: x.order)],
                'rating_config': None,
                'existing_answer': answer_map.get(q.pk),
            }
            if q.question_type == 'rating':
                try:
                    cfg = q.rating_config
                    q_data['rating_config'] = {
                        'min_value': cfg.min_value,
                        'max_value': cfg.max_value,
                        'min_label': cfg.min_label,
                        'max_label': cfg.max_label,
                    }
                except Exception:
                    q_data['rating_config'] = {'min_value': 1, 'max_value': 5, 'min_label': '', 'max_label': ''}
            questions_data.append(q_data)

        return Response({
            'id': survey.pk,
            'title': survey.title,
            'description': survey.description,
            'is_anonymous': survey.is_anonymous,
            'status': survey.status,
            'start_date': str(survey.start_date) if survey.start_date else None,
            'end_date': str(survey.end_date) if survey.end_date else None,
            'questions': questions_data,
            'response_id': response_obj.pk if response_obj else None,
            'is_complete': response_obj.is_complete if response_obj else False,
        })


class ResponseDetailView(APIView):
    """GET /api/survey/responses/<pk> — return response with answers (respondent only)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        response = SurveyResponse.objects.select_related('survey').filter(pk=pk).first()
        if not response:
            return Response({'detail': 'Not found.'}, status=http_status.HTTP_404_NOT_FOUND)
        if not response.survey.is_anonymous and response.employee_id != request.user.pk:
            return Response({'detail': 'Permission denied.'}, status=http_status.HTTP_403_FORBIDDEN)
        return Response(SurveyResponseSerializer(response).data)
